"""
Programme d'actualisation et validation automatique du fichier de configuration V2.

Version améliorée avec :
- Validation complète de la structure des colonnes
- Détection et correction automatique des colonnes mal nommées
- Gestion intelligente des colonnes en trop (préfixe EXTRA_)
- Validation du contenu de chaque cellule
- Rapport détaillé et structuré
"""

import sys
import pandas as pd
import numpy as np
from pathlib import Path
from typing import Dict, List, Set, Tuple, Optional, Any
from dataclasses import dataclass, field
from core.config_manager import ConfigManager
import re
import openpyxl
from openpyxl.styles import PatternFill, Font
import difflib


@dataclass
class ValidationResult:
    """Résultat d'une validation."""
    valide: bool
    message: Optional[str] = None
    valeur_corrigee: Optional[Any] = None
    severite: str = "info"  # info, warning, error


@dataclass
class RapportFeuille:
    """Rapport de validation pour une feuille."""
    nom: str
    colonnes_manquantes: List[str] = field(default_factory=list)
    colonnes_ajoutees: List[str] = field(default_factory=list)
    colonnes_renommees: List[Tuple[str, str]] = field(default_factory=list)  # (ancien, nouveau)
    colonnes_extra: List[str] = field(default_factory=list)
    erreurs_contenu: List[str] = field(default_factory=list)
    warnings_contenu: List[str] = field(default_factory=list)
    corrections_contenu: List[str] = field(default_factory=list)
    nb_lignes_valides: int = 0
    nb_lignes_total: int = 0
    structure_modifiee: bool = False


class ColumnValidator:
    """Validateur de colonnes avec règles spécifiques par type."""
    
    @staticmethod
    def valider_semaine(valeur: Any, nb_semaines_max: int = 52) -> ValidationResult:
        """Valide une semaine (doit être un entier entre 1 et nb_semaines_max)."""
        if pd.isna(valeur) or str(valeur).strip() == '':
            return ValidationResult(False, "Semaine vide", severite="error")
        
        try:
            semaine = int(float(valeur))  # float() pour gérer "1.0"
            if 1 <= semaine <= nb_semaines_max:
                return ValidationResult(True, valeur_corrigee=semaine)
            else:
                return ValidationResult(
                    False,
                    f"Semaine {semaine} hors limites (1-{nb_semaines_max})",
                    severite="error"
                )
        except (ValueError, TypeError):
            return ValidationResult(
                False,
                f"Format semaine invalide: '{valeur}' (attendu: entier)",
                severite="error"
            )
    
    @staticmethod
    def valider_capacite(valeur: Any, capacite_max: int = 100) -> ValidationResult:
        """Valide une capacité occupée (entier positif)."""
        if pd.isna(valeur) or str(valeur).strip() == '':
            return ValidationResult(True)  # Capacité optionnelle (défaut = capacité totale)
        
        try:
            capacite = int(float(valeur))
            if capacite < 0:
                return ValidationResult(
                    False,
                    f"Capacité occupée négative: {capacite}",
                    valeur_corrigee=0,
                    severite="error"
                )
            elif capacite > capacite_max:
                return ValidationResult(
                    False,
                    f"Capacité occupée ({capacite}) > capacité max probable ({capacite_max})",
                    severite="warning"
                )
            return ValidationResult(True, valeur_corrigee=capacite)
        except (ValueError, TypeError):
            return ValidationResult(
                False,
                f"Format capacité invalide: '{valeur}' (attendu: entier positif)",
                severite="error"
            )
    
    @staticmethod
    def valider_horaire(valeur: Any) -> ValidationResult:
        """Valide un horaire (format HH:MM, HH:MM:SS, ou HHhMM)."""
        if pd.isna(valeur) or str(valeur).strip() == '':
            return ValidationResult(True)  # Horaire optionnel
        
        valeur_str = str(valeur).strip()
        
        # Formats acceptés: "14:00", "14:00:00", "14h00", "14H00", "14h", "14H"
        patterns = [
            (r'^(\d{1,2}):(\d{2})(?::\d{2})?$', lambda m: f"{int(m.group(1)):02d}:{m.group(2)}"),
            (r'^(\d{1,2})[hH](\d{2})?$', lambda m: f"{int(m.group(1)):02d}:{m.group(2) or '00'}"),
        ]
        
        for pattern, formatter in patterns:
            match = re.match(pattern, valeur_str)
            if match:
                horaire_formate = formatter(match)
                heures, minutes = map(int, horaire_formate.split(':'))
                if 0 <= heures < 24 and 0 <= minutes < 60:
                    return ValidationResult(True, valeur_corrigee=horaire_formate)
        
        return ValidationResult(
            False,
            f"Format horaire invalide: '{valeur}' (attendu: HH:MM ou HHhMM)",
            severite="warning"
        )
    
    @staticmethod
    def valider_institution(valeur: Any, institutions_valides: Set[str]) -> ValidationResult:
        """Valide une institution (doit exister dans la liste)."""
        if pd.isna(valeur) or str(valeur).strip() == '':
            return ValidationResult(False, "Institution vide", severite="error")
        
        valeur_str = str(valeur).strip()
        
        if valeur_str in institutions_valides:
            return ValidationResult(True, valeur_corrigee=valeur_str)
        
        # Recherche floue
        matches = difflib.get_close_matches(valeur_str, institutions_valides, n=1, cutoff=0.6)
        if matches:
            return ValidationResult(
                False,
                f"Institution '{valeur_str}' non trouvée. Vouliez-vous dire '{matches[0]}'?",
                valeur_corrigee=matches[0],
                severite="warning"
            )
        
        return ValidationResult(
            False,
            f"Institution '{valeur_str}' inconnue",
            severite="error"
        )
    
    @staticmethod
    def valider_gymnase(valeur: Any, gymnases_valides: Set[str]) -> ValidationResult:
        """Valide un gymnase (doit exister dans la liste)."""
        if pd.isna(valeur) or str(valeur).strip() == '':
            return ValidationResult(False, "Gymnase vide", severite="error")
        
        valeur_str = str(valeur).strip()
        
        if valeur_str in gymnases_valides:
            return ValidationResult(True, valeur_corrigee=valeur_str)
        
        # Recherche floue
        matches = difflib.get_close_matches(valeur_str, gymnases_valides, n=1, cutoff=0.6)
        if matches:
            return ValidationResult(
                False,
                f"Gymnase '{valeur_str}' non trouvé. Vouliez-vous dire '{matches[0]}'?",
                valeur_corrigee=matches[0],
                severite="warning"
            )
        
        return ValidationResult(
            False,
            f"Gymnase '{valeur_str}' inconnu",
            severite="error"
        )
    
    @staticmethod
    def valider_equipe(valeur: Any, equipes_valides: Set[str]) -> ValidationResult:
        """
        Valide une équipe.
        
        Formats acceptés:
        - 'Institution (numéro)' : ex: "LYON 1 (1)"
        - 'Institution (numéro) [M]' : ex: "LYON 1 (1) [M]"
        - 'Institution (numéro) [F]' : ex: "LYON 1 (1) [F]"
        """
        if pd.isna(valeur) or str(valeur).strip() == '':
            return ValidationResult(False, "Équipe vide", severite="error")
        
        valeur_str = str(valeur).strip()
        
        # Vérifier format: Institution (numéro) optionnellement suivi de [M] ou [F]
        if not re.match(r'^.+\s*\(\d+\)\s*(\s*\[(M|F)\])?\s*$', valeur_str):
            return ValidationResult(
                False,
                f"Format équipe invalide: '{valeur_str}' (attendu: 'Institution (numéro)' ou 'Institution (numéro) [M/F]')",
                severite="warning"
            )
        
        # Vérifier si l'équipe existe dans la liste de référence
        # (qui peut contenir à la fois les versions avec et sans genre)
        if valeur_str in equipes_valides:
            return ValidationResult(True, valeur_corrigee=valeur_str)
        
        return ValidationResult(
            False,
            f"Équipe '{valeur_str}' non trouvée dans la liste",
            severite="error"
        )
    
    @staticmethod
    def valider_texte_libre(valeur: Any) -> ValidationResult:
        """Valide un texte libre (remarques, etc.)."""
        # Toujours valide, pas de contrainte
        return ValidationResult(True)
    
    @staticmethod
    def valider_niveau(valeur: Any) -> ValidationResult:
        """Valide un niveau de gymnase (Haut ou Bas)."""
        if pd.isna(valeur) or str(valeur).strip() == '':
            return ValidationResult(False, "Niveau vide", severite="error")
        
        valeur_str = str(valeur).strip().lower()
        
        if valeur_str in ['haut', 'bas']:
            return ValidationResult(True, valeur_corrigee=valeur_str.capitalize())
        
        return ValidationResult(
            False,
            f"Niveau invalide: '{valeur}' (attendu: 'Haut' ou 'Bas')",
            severite="error"
        )


class ConfigActualisateurV2:
    """Actualise et valide un fichier de configuration avec validation avancée."""
    
    def __init__(self, fichier_path: str):
        self.fichier_path = Path(fichier_path)
        self.config = ConfigManager(str(fichier_path))
        self.rapports_feuilles: Dict[str, RapportFeuille] = {}
        self.validator = ColumnValidator()
        
        # Données de référence
        self.equipes_ref: Set[str] = set()
        self.equipes_par_nom: Dict[str, Set[str]] = {}  # {nom: {genres}}
        self.equipes_toutes_variantes: Set[str] = set()  # Toutes les variantes (avec et sans genre)
        self.equipes_sans_genre: Set[str] = set()  # Noms d'équipes sans [F]/[M] pour Matchs_Fixes
        self.gymnases_ref: Set[str] = set()
        self.institutions_ref: Set[str] = set()
        
        # Mapping des validateurs par colonne
        self.validateurs_colonnes = {
            'Semaine': self.validator.valider_semaine,
            'Heure_Debut': self.validator.valider_horaire,
            'Heure_Fin': self.validator.valider_horaire,
            'Horaire': self.validator.valider_horaire,
            'Gymnase': self.validator.valider_gymnase,
            'Niveau': self.validator.valider_niveau,
            'Remarques': self.validator.valider_texte_libre,
            'Remarque': self.validator.valider_texte_libre,
        }
    
    def actualiser(self) -> bool:
        """Actualise complètement le fichier de configuration."""
        # En-tête principal élégant
        print("\n┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓")
        print("┃                     🔄 ACTUALISATEUR V2                       ┃")
        print("┃              Configuration Excel PyCalendar                   ┃")
        print("┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛\n")
        
        if not self.fichier_path.exists():
            print(f"  ❌ Erreur : Le fichier {self.fichier_path} n'existe pas\n")
            return False
        
        print(f"  📂 Fichier : {self.fichier_path.name}")
        print(f"  📍 Emplacement : {self.fichier_path.parent}\n")
        
        print("  ╔═══════════════════════════════════════════════════════════╗")
        print("  ║               🚀 DÉMARRAGE DE L'ANALYSE                   ║")
        print("  ╚═══════════════════════════════════════════════════════════╝\n")
        
        # Étape 1 : Charger les références
        print("  [1/6] 📊 Chargement des données de référence...", end='', flush=True)
        nb_equipes_avant = len(self.equipes_ref)
        self._charger_references()
        nb_equipes = len(self.equipes_ref)
        nb_gymnases = len(self.gymnases_ref)
        print(f" ✓")
        print(f"        → {nb_equipes} équipe(s), {nb_gymnases} gymnase(s)")
        
        # Étape 2 : Valider et corriger chaque feuille
        print(f"\n  [2/6] 🔍 Validation des feuilles...")
        nb_problemes = self._valider_toutes_feuilles()
        nb_feuilles_valides = len(self.rapports_feuilles)
        if nb_problemes == 0:
            print(f"        ✓ {nb_feuilles_valides} feuille(s) analysée(s)")
        else:
            print(f"        ✓ {nb_feuilles_valides} feuille(s) analysée(s) • {nb_problemes} à corriger")
        
        # Étape 3 : Générer les feuilles manquantes
        print(f"\n  [3/6] 🏗️  Génération des feuilles manquantes...", end='', flush=True)
        nb_generees = self._generer_feuilles_manquantes()
        print(f" ✓")
        if nb_generees > 0:
            print(f"        → {nb_generees} feuille(s) créée(s)")
        else:
            print(f"        → Aucune génération nécessaire")
        
        # Étape 4 : Supprimer tous les exemples
        print(f"\n  [4/6] 🧹 Nettoyage des exemples...", end='', flush=True)
        nb_suppressions = self._supprimer_exemples()
        print(f" ✓")
        if nb_suppressions > 0:
            print(f"        → {nb_suppressions} ligne(s) d'exemple supprimée(s)")
        else:
            print(f"        → Aucun exemple détecté")
        
        # Étape 5 : Pré-remplir Types_Poules
        print(f"\n  [5/6] 🏐 Configuration des types de poules...", end='', flush=True)
        nb_poules_ajoutees = self._preremplir_types_poules()
        print(f" ✓")
        if nb_poules_ajoutees > 0:
            print(f"        → {nb_poules_ajoutees} poule(s) ajoutée(s)")
        else:
            print(f"        → Déjà à jour")
        
        # Étape 6 : Finalisation
        print(f"\n  [6/6] 🎨 Finalisation (formatage & validations)...", end='', flush=True)
        nb_corrections = self._appliquer_corrections()
        self._actualiser_listes_deroulantes()
        print(f" ✓")
        if nb_corrections > 0:
            print(f"        → {nb_corrections} correction(s) appliquée(s)")
        else:
            print(f"        → Structure conforme")
        
        # Afficher le rapport final
        print()
        self._afficher_rapport_final()
        
        return self._a_reussi()
    
    def _charger_references(self):
        """Charge les données de référence depuis les feuilles principales."""
        from core.utils import extraire_genre_depuis_poule, formater_nom_avec_genre
        
        # Charger équipes avec détection des genres
        # Structure: {nom_sans_genre: {genres}} ex: {"LYON 1 (1)": {"M", "F"}}
        self.equipes_par_nom: Dict[str, Set[str]] = {}
        
        df_equipes = self.config.lire_feuille('Equipes')
        if df_equipes is not None and 'Equipe' in df_equipes.columns:
            for idx, row in df_equipes.iterrows():
                equipe = row.get('Equipe')
                if pd.isna(equipe):
                    continue
                    
                equipe_str = str(equipe).strip()
                if not equipe_str:
                    continue
                
                # Ajouter équipe aux références (format simple)
                self.equipes_ref.add(equipe_str)
                
                # Détecter le genre (depuis colonne Genre ou depuis Poule)
                genre = ''
                if 'Genre' in df_equipes.columns:
                    genre_col = row.get('Genre')
                    if pd.notna(genre_col) and str(genre_col).strip() in ['M', 'F']:
                        genre = str(genre_col).strip()
                
                if not genre and 'Poule' in df_equipes.columns:
                    poule = row.get('Poule')
                    if pd.notna(poule):
                        genre = extraire_genre_depuis_poule(str(poule))
                
                # Stocker dans le mapping nom -> genres
                if equipe_str not in self.equipes_par_nom:
                    self.equipes_par_nom[equipe_str] = set()
                if genre:
                    self.equipes_par_nom[equipe_str].add(genre)
                
                # Extraire institution
                match = re.match(r'^(.+?)\s*\(\d+\)\s*$', equipe_str)
                if match:
                    self.institutions_ref.add(match.group(1).strip())
        
        # Charger équipes hors championnat
        df_equipes_hors = self.config.lire_feuille('Equipes_Hors_Championnat')
        if df_equipes_hors is not None and not df_equipes_hors.empty:
            for idx, row in df_equipes_hors.iterrows():
                equipe = row.get('Equipe')
                institution = row.get('Institution')
                genre = row.get('Genre')
                type_championnat = row.get('Type_Championnat')
                
                if pd.isna(equipe):
                    continue
                    
                equipe_str = str(equipe).strip()
                if not equipe_str:
                    continue
                
                # Ajouter équipe hors championnat aux références
                self.equipes_ref.add(equipe_str)
                
                # Détecter le genre depuis la colonne Genre
                genre_str = ''
                if pd.notna(genre) and str(genre).strip() in ['M', 'F']:
                    genre_str = str(genre).strip()
                
                # Stocker dans le mapping nom -> genres
                if equipe_str not in self.equipes_par_nom:
                    self.equipes_par_nom[equipe_str] = set()
                if genre_str:
                    self.equipes_par_nom[equipe_str].add(genre_str)
                
                # Ajouter institution si elle existe
                if pd.notna(institution):
                    institution_str = str(institution).strip()
                    if institution_str:
                        self.institutions_ref.add(institution_str)
        
        # Charger gymnases
        df_gymnases = self.config.lire_feuille('Gymnases')
        if df_gymnases is not None and 'Gymnase' in df_gymnases.columns:
            for gymnase in df_gymnases['Gymnase'].dropna():
                gymnase_str = str(gymnase).strip()
                if gymnase_str:
                    self.gymnases_ref.add(gymnase_str)
        
        # Générer toutes les variantes d'équipes (avec et sans genre)
        self.equipes_toutes_variantes = set(self._generer_variantes_equipes())
        
        # Générer liste des noms sans genre pour Matchs_Fixes
        self.equipes_sans_genre = set(self.equipes_par_nom.keys())
    
    def _valider_toutes_feuilles(self) -> int:
        """Valide et corrige toutes les feuilles définies. Retourne le nombre de feuilles avec problèmes."""
        # Valider TOUTES les feuilles définies dans STRUCTURES
        feuilles_a_valider = list(self.config.STRUCTURES.keys())

        nb_problemes = 0

        for nom_feuille in feuilles_a_valider:
            rapport = self._valider_feuille(nom_feuille)
            self.rapports_feuilles[nom_feuille] = rapport

            # Message concis (seulement si problème)
            if rapport.structure_modifiee or rapport.erreurs_contenu or rapport.warnings_contenu:
                nb_problemes += 1
                print(f"        • {nom_feuille}: ", end='')
                messages = []
                if rapport.colonnes_manquantes:
                    messages.append(f"+{len(rapport.colonnes_manquantes)} col")
                if rapport.colonnes_renommees:
                    messages.append(f"~{len(rapport.colonnes_renommees)} col")
                if rapport.erreurs_contenu:
                    messages.append(f"❌{len(rapport.erreurs_contenu)}")
                if rapport.warnings_contenu:
                    messages.append(f"⚠️{len(rapport.warnings_contenu)}")
                print(", ".join(messages))
        
        return nb_problemes
    
    def _valider_feuille(self, nom_feuille: str) -> RapportFeuille:
        """Valide une feuille et retourne un rapport détaillé."""
        rapport = RapportFeuille(nom=nom_feuille)
        
        df = self.config.lire_feuille(nom_feuille)
        if df is None or df.empty:
            # Feuille vide ou inexistante
            return rapport
        
        structure = self.config.STRUCTURES.get(nom_feuille, {})
        colonnes_attendues = structure.get('colonnes', [])
        
        # 1. Vérifier la structure des colonnes
        self._verifier_structure_colonnes(df, colonnes_attendues, rapport)
        
        # 2. Valider le contenu ligne par ligne
        rapport.nb_lignes_total = len(df)
        self._valider_contenu(df, nom_feuille, rapport)
        
        # 3. Validations spécifiques par feuille
        if nom_feuille == 'Preferences_Gymnases':
            self._valider_preferences_gymnases(df, rapport)
        elif nom_feuille == 'Ententes':
            self._valider_ententes(df, rapport)
        elif nom_feuille == 'Contraintes_Temporelles':
            self._valider_contraintes_temporelles(df, rapport)
        elif nom_feuille == 'Matchs_Fixes':
            self._valider_matchs_fixes(df, rapport)
        elif nom_feuille == 'Equipes_Hors_Championnat':
            self._valider_equipes_hors_championnat(df, rapport)
        
        return rapport 

    def _verifier_structure_colonnes(self, df: pd.DataFrame, 
                                    colonnes_attendues: List[str], 
                                    rapport: RapportFeuille):
        """Vérifie et corrige la structure des colonnes."""
        colonnes_presentes = list(df.columns)
        colonnes_attendues_set = set(colonnes_attendues)
        colonnes_presentes_set = set(colonnes_presentes)
        
        # Colonnes manquantes
        rapport.colonnes_manquantes = list(colonnes_attendues_set - colonnes_presentes_set)
        
        # Colonnes en trop
        colonnes_extra_candidates = colonnes_presentes_set - colonnes_attendues_set
        
        # Essayer de renommer les colonnes mal nommées (similitude)
        colonnes_a_renommer = {}
        colonnes_vraiment_extra = []
        
        for col_extra in colonnes_extra_candidates:
            # Recherche floue parmi les colonnes attendues
            matches = difflib.get_close_matches(col_extra, colonnes_attendues, n=1, cutoff=0.7)
            if matches and matches[0] not in colonnes_presentes_set:
                # Colonne mal nommée, on peut la renommer
                colonnes_a_renommer[col_extra] = matches[0]
                rapport.colonnes_renommees.append((col_extra, matches[0]))
            else:
                # Vraiment une colonne en trop
                colonnes_vraiment_extra.append(col_extra)
        
        rapport.colonnes_extra = colonnes_vraiment_extra
        
        if rapport.colonnes_manquantes or rapport.colonnes_renommees or rapport.colonnes_extra:
            rapport.structure_modifiee = True
    
    def _valider_contenu(self, df: pd.DataFrame, nom_feuille: str, rapport: RapportFeuille):
        """Valide le contenu de chaque cellule."""
        lignes_valides = 0
        
        for idx, row in df.iterrows():
            ligne_num = int(idx) + 2  # +2 pour l'en-tête et index 0-based
            ligne_valide = True
            
            # Ignorer les lignes complètement vides
            if row.isna().all():
                continue
            
            # Valider chaque colonne
            for colonne in df.columns:
                valeur = row[colonne]
                
                # Skip si colonne EXTRA_
                if str(colonne).startswith('EXTRA_'):
                    continue
                
                # Déterminer le validateur approprié
                result = self._valider_cellule(colonne, valeur, nom_feuille)
                
                if not result.valide:
                    ligne_valide = False
                    message = f"Ligne {ligne_num}, colonne '{colonne}': {result.message}"
                    
                    if result.severite == "error":
                        rapport.erreurs_contenu.append(message)
                    elif result.severite == "warning":
                        rapport.warnings_contenu.append(message)
                    
                    if result.valeur_corrigee is not None:
                        rapport.corrections_contenu.append(
                            f"Ligne {ligne_num}, '{colonne}': '{valeur}' → '{result.valeur_corrigee}'"
                        )
            
            if ligne_valide:
                lignes_valides += 1
        
        rapport.nb_lignes_valides = lignes_valides
    
    def _valider_cellule(self, colonne: str, valeur: Any, nom_feuille: str) -> ValidationResult:
        """Valide une cellule selon son type."""
        # Pour Matchs_Fixes, Equipe_1 et Equipe_2 sont validés dans _valider_matchs_fixes
        # car ils nécessitent la colonne Genre pour construire le nom complet
        if nom_feuille == 'Matchs_Fixes' and colonne in ['Equipe_1', 'Equipe_2']:
            # Les équipes peuvent être vides (seulement une validation basique de format si présentes)
            if pd.isna(valeur) or str(valeur).strip() == '':
                return ValidationResult(True)  # Les équipes vides sont autorisées
            valeur_str = str(valeur).strip()
            # Format: Institution (numéro) SANS [F/M]
            if not re.match(r'^.+\s*\(\d+\)\s*$', valeur_str):
                return ValidationResult(
                    False,
                    f"Format équipe invalide: '{valeur_str}' (attendu: 'Institution (numéro)')",
                    severite="warning"
                )
            return ValidationResult(True)
        
        # Validateurs génériques par nom de colonne
        if colonne == 'Semaine':
            return self.validator.valider_semaine(valeur)
        elif colonne in ['Heure_Debut', 'Heure_Fin', 'Horaire']:
            return self.validator.valider_horaire(valeur)
        elif colonne == 'Institution':
            # Pour Equipes_Hors_Championnat, permettre des institutions externes
            if nom_feuille == 'Equipes_Hors_Championnat':
                if pd.isna(valeur) or str(valeur).strip() == '':
                    return ValidationResult(False, "Institution manquante", None, "error")
                # L'institution peut être n'importe quoi, pas de validation stricte
                return ValidationResult(True)
            elif self.institutions_ref:
                return self.validator.valider_institution(valeur, self.institutions_ref)
        elif colonne == 'Gymnase' or colonne.startswith('Gymnase_Pref_'):
            # Validation pour colonne Gymnase ou Gymnase_Pref_N
            # Les colonnes Gymnase_Pref_* peuvent être vides (facultatif)
            if colonne.startswith('Gymnase_Pref_'):
                if pd.isna(valeur) or str(valeur).strip() == '':
                    return ValidationResult(True)  # Les préférences vides sont OK
            
            if self.gymnases_ref:
                return self.validator.valider_gymnase(valeur, self.gymnases_ref)
        elif colonne == 'Equipe':
            # Pour la feuille Equipes_Hors_Championnat, permettre des équipes externes
            if nom_feuille == 'Equipes_Hors_Championnat':
                if pd.isna(valeur) or str(valeur).strip() == '':
                    return ValidationResult(False, "Équipe manquante", None, "error")
                valeur_str = str(valeur).strip()
                # Format attendu: Institution (numéro) [F/M] ou Institution (numéro)
                if not re.match(r'^.+\s*\(\d+\)\s*(?:\s*[FM]\s*)?$', valeur_str):
                    return ValidationResult(
                        False,
                        f"Format équipe invalide: '{valeur_str}' (attendu: 'Institution (numéro)' ou 'Institution (numéro) F/M')",
                        None,
                        "error"
                    )
                return ValidationResult(True)
            # Pour la feuille Equipes, valider contre les équipes chargées (sans variantes genre)
            elif nom_feuille == 'Equipes' and self.equipes_ref:
                return self.validator.valider_equipe(valeur, self.equipes_ref)
            # Pour les autres feuilles, utiliser les variantes avec genre
            elif self.equipes_toutes_variantes:
                return self.validator.valider_equipe(valeur, self.equipes_toutes_variantes)
        elif colonne in ['Institution_Obligatoire', 'Institution_ou_Equipe']:
            if self.institutions_ref:
                return self.validator.valider_institution(valeur, self.institutions_ref)
        elif colonne in ['Institution_1', 'Institution_2']:
            # Validation pour les colonnes d'ententes
            if self.institutions_ref:
                return self.validator.valider_institution(valeur, self.institutions_ref)
        elif colonne in ['Equipe_1', 'Equipe_2']:
            # Validation pour les colonnes de contraintes temporelles
            if nom_feuille == 'Contraintes_Temporelles':
                # Pour Contraintes_Temporelles, les équipes sont sans [F]/[M]
                if pd.isna(valeur) or str(valeur).strip() == '':
                    return ValidationResult(False, f"{colonne} manquante", None, "error")
                valeur_str = str(valeur).strip()
                # Format attendu: Institution (numéro) SANS [F/M]
                if not re.match(r'^.+\s*\(\d+\)\s*$', valeur_str):
                    return ValidationResult(
                        False,
                        f"Format {colonne} invalide: '{valeur_str}' (attendu: 'Institution (numéro)' sans [F]/[M])",
                        None,
                        "error"
                    )
                # Vérifier que l'équipe existe (mais sans genre, donc on ne peut pas valider complètement ici)
                # La validation complète se fait dans _valider_contraintes_temporelles
                return ValidationResult(True)
            elif self.equipes_toutes_variantes:
                return self.validator.valider_equipe(valeur, self.equipes_toutes_variantes)
        elif colonne == 'Genre':
            # Validation pour le genre dans Contraintes_Temporelles
            if nom_feuille == 'Contraintes_Temporelles':
                if pd.isna(valeur) or str(valeur).strip() == '':
                    return ValidationResult(False, f"Genre manquant", None, "error")
                valeur_str = str(valeur).strip().upper()
                if valeur_str not in ['M', 'F']:
                    return ValidationResult(
                        False,
                        f"Genre invalide: '{valeur}' (doit être 'M' ou 'F')",
                        None,
                        "error"
                    )
                return ValidationResult(True)
        elif colonne == 'Type_Contrainte':
            # Validation pour type de contrainte temporelle
            if pd.isna(valeur) or str(valeur).strip() == '':
                return ValidationResult(False, "Le type de contrainte est obligatoire", None, "error")
            type_str = str(valeur).strip()
            if type_str not in ['Avant', 'Apres']:
                return ValidationResult(False, "Le type doit être 'Avant' ou 'Apres'", None, "error")
            return ValidationResult(True)
        elif colonne == 'Type_Championnat':
            # Pour la feuille Equipes_Hors_Championnat, valider le type de championnat
            if nom_feuille == 'Equipes_Hors_Championnat':
                if pd.isna(valeur) or str(valeur).strip() == '':
                    return ValidationResult(False, "Type de championnat manquant", None, "error")
                valeur_str = str(valeur).strip()
                types_valides = ['CFE', 'CFU', 'Autre']
                if valeur_str not in types_valides:
                    return ValidationResult(
                        False,
                        f"Type de championnat '{valeur_str}' invalide. Valeurs acceptées: {', '.join(types_valides)}",
                        None,
                        "error"
                    )
                return ValidationResult(True)
            # Validation pour la pénalité d'entente (doit être un float positif si présent)
            if pd.isna(valeur) or str(valeur).strip() == '':
                return ValidationResult(True)  # Valeur optionnelle
            try:
                penalite = float(valeur)
                if penalite < 0:
                    return ValidationResult(False, "La pénalité doit être positive", 0.0, "error")
                return ValidationResult(True)
            except (ValueError, TypeError):
                return ValidationResult(False, "La pénalité doit être un nombre", None, "error")
        elif colonne == 'Capacite_Occupee':
            return self.validator.valider_capacite(valeur, capacite_max=10)
        elif colonne == 'Remarques':
            return self.validator.valider_texte_libre(valeur)
        
        # Par défaut, accepter
        return ValidationResult(True)
    
    def _valider_preferences_gymnases(self, df: pd.DataFrame, rapport: RapportFeuille):
        """
        Validation spécifique pour la feuille Preferences_Gymnases.
        Ajoute un warning si une institution n'a aucun gymnase préféré.
        """
        if df is None or df.empty:
            return
        
        # Identifier les colonnes de préférences
        colonnes_pref = [col for col in df.columns if col.startswith('Gymnase_Pref_')]
        
        for idx, row in df.iterrows():
            institution = row.get('Institution')
            if pd.isna(institution) or str(institution).strip() == '':
                continue
            
            # Vérifier si au moins un gymnase est renseigné
            a_des_preferences = False
            for col_pref in colonnes_pref:
                valeur = row.get(col_pref)
                if not pd.isna(valeur) and str(valeur).strip() != '':
                    a_des_preferences = True
                    break
            
            if not a_des_preferences:
                ligne_num = int(idx) + 2  # +2 pour l'en-tête et index 0-based
                rapport.warnings_contenu.append(
                    f"Ligne {ligne_num}: Institution '{institution}' n'a aucun gymnase préféré"
                )
    
    def _valider_ententes(self, df: pd.DataFrame, rapport: RapportFeuille):
        """
        Validation spécifique pour la feuille Ententes.
        Détecte les doublons bidirectionnels (LYON 1 ↔ LYON 2 = LYON 2 ↔ LYON 1)
        et vérifie la cohérence des paires.
        """
        if df is None or df.empty:
            return
        
        ententes_vues = {}  # {tuple_sorted: (ligne_num, inst1, inst2)}
        
        for idx, row in df.iterrows():
            ligne_num = int(idx) + 2  # +2 pour l'en-tête et index 0-based
            
            inst1 = row.get('Institution_1')
            inst2 = row.get('Institution_2')
            
            # Ignorer les lignes incomplètes
            if pd.isna(inst1) or pd.isna(inst2):
                continue
            
            inst1_str = str(inst1).strip()
            inst2_str = str(inst2).strip()
            
            if not inst1_str or not inst2_str:
                continue
            
            # Créer une clé bidirectionnelle (ordre alphabétique)
            cle = tuple(sorted([inst1_str, inst2_str]))
            
            # Vérifier les doublons
            if cle in ententes_vues:
                ligne_precedente, prev_inst1, prev_inst2 = ententes_vues[cle]
                rapport.erreurs_contenu.append(
                    f"Ligne {ligne_num}: Entente en doublon avec ligne {ligne_precedente} "
                    f"('{inst1_str}' ↔ '{inst2_str}' déjà définie comme '{prev_inst1}' ↔ '{prev_inst2}')"
                )
            else:
                ententes_vues[cle] = (ligne_num, inst1_str, inst2_str)
    
    def _valider_contraintes_temporelles(self, df: pd.DataFrame, rapport: RapportFeuille):
        """
        Validation spécifique pour la feuille Contraintes_Temporelles.
        Détecte les doublons bidirectionnels, vérifie la cohérence des contraintes.
        """
        if df is None or df.empty:
            return
        
        contraintes_vues = {}  # {tuple_sorted: (ligne_num, eq1, eq2, type, semaine)}
        
        for idx, row in df.iterrows():
            ligne_num = int(idx) + 2
            
            eq1 = row.get('Equipe_1')
            eq2 = row.get('Equipe_2')
            genre = row.get('Genre')
            type_contrainte = row.get('Type_Contrainte')
            semaine = row.get('Semaine')
            
            # Validation des champs obligatoires
            if pd.isna(eq1) or str(eq1).strip() == '':
                rapport.erreurs_contenu.append(f"Ligne {ligne_num}: Équipe_1 manquante")
                continue
            if pd.isna(eq2) or str(eq2).strip() == '':
                rapport.erreurs_contenu.append(f"Ligne {ligne_num}: Équipe_2 manquante")
                continue
            if pd.isna(genre) or str(genre).strip() == '':
                rapport.erreurs_contenu.append(f"Ligne {ligne_num}: Genre manquant")
                continue
            if pd.isna(type_contrainte) or str(type_contrainte).strip() == '':
                rapport.erreurs_contenu.append(f"Ligne {ligne_num}: Type_Contrainte manquant")
                continue
            if pd.isna(semaine):
                rapport.erreurs_contenu.append(f"Ligne {ligne_num}: Semaine manquante")
                continue
            
            eq1_str = str(eq1).strip()
            eq2_str = str(eq2).strip()
            genre_str = str(genre).strip().upper()
            type_str = str(type_contrainte).strip()
            
            # Validation du genre
            if genre_str not in ['M', 'F']:
                rapport.erreurs_contenu.append(
                    f"Ligne {ligne_num}: Genre invalide ('{genre}'), doit être 'M' ou 'F'"
                )
            
            # Validation du format des équipes (sans [F]/[M])
            if not re.match(r'^.+\s*\(\d+\)\s*$', eq1_str):
                rapport.erreurs_contenu.append(
                    f"Ligne {ligne_num}: Format Équipe_1 invalide: '{eq1_str}' "
                    "(attendu: 'Institution (numéro)' sans [F]/[M])"
                )
            if not re.match(r'^.+\s*\(\d+\)\s*$', eq2_str):
                rapport.erreurs_contenu.append(
                    f"Ligne {ligne_num}: Format Équipe_2 invalide: '{eq2_str}' "
                    "(attendu: 'Institution (numéro)' sans [F]/[M])"
                )
            
            # Validation du type de contrainte
            if type_str not in ['Avant', 'Apres']:
                rapport.erreurs_contenu.append(
                    f"Ligne {ligne_num}: Type_Contrainte invalide ('{type_str}'), doit être 'Avant' ou 'Apres'"
                )
            
            # Validation du numéro de semaine
            try:
                semaine_int = int(semaine)
                if semaine_int < 1 or semaine_int > 52:
                    rapport.erreurs_contenu.append(
                        f"Ligne {ligne_num}: Semaine invalide ({semaine_int}), doit être entre 1 et 52"
                    )
                    continue
            except (ValueError, TypeError):
                rapport.erreurs_contenu.append(
                    f"Ligne {ligne_num}: Semaine invalide ('{semaine}'), doit être un nombre"
                )
                continue
            
            # Vérifier que les équipes existent (en combinant équipe + genre)
            eq1_complet = f"{eq1_str} [{genre_str}]"
            eq2_complet = f"{eq2_str} [{genre_str}]"
            
            if eq1_complet not in self.equipes_toutes_variantes:
                rapport.erreurs_contenu.append(
                    f"Ligne {ligne_num}: Équipe '{eq1_complet}' non trouvée dans la liste des équipes"
                )
            if eq2_complet not in self.equipes_toutes_variantes:
                rapport.erreurs_contenu.append(
                    f"Ligne {ligne_num}: Équipe '{eq2_complet}' non trouvée dans la liste des équipes"
                )
            
            # Créer clé bidirectionnelle (équipe1+genre1 + équipe2+genre2)
            cle = tuple(sorted([eq1_complet, eq2_complet]))
            
            # Vérifier doublons (même paire + même type + même semaine = doublon)
            cle_complete = (*cle, type_str, semaine_int)
            
            if cle_complete in contraintes_vues:
                ligne_precedente = contraintes_vues[cle_complete]
                rapport.erreurs_contenu.append(
                    f"Ligne {ligne_num}: Contrainte en doublon avec ligne {ligne_precedente} "
                    f"('{eq1_complet}' ↔ '{eq2_complet}', {type_str} semaine {semaine_int})"
                )
            else:
                contraintes_vues[cle_complete] = ligne_num
            
            # Warning si horaires spécifiés mais invalides
            horaires_possibles = row.get('Horaires_Possibles')
            if pd.notna(horaires_possibles) and str(horaires_possibles).strip():
                horaires_str = str(horaires_possibles).strip()
                # Simple validation: vérifier qu'il y a au moins un horaire avec format approximatif
                if ',' not in horaires_str and 'h' not in horaires_str.lower():
                    rapport.warnings_contenu.append(
                        f"Ligne {ligne_num}: Format d'horaires potentiellement invalide ('{horaires_str}'). "
                        f"Exemple attendu: 'Mercredi 18h00, Vendredi 16h00'"
                    )
    
    def _valider_matchs_fixes(self, df: pd.DataFrame, rapport: RapportFeuille):
        """
        Validation spécifique AVANCÉE pour la feuille Matchs_Fixes.
        - Vérifie format semaine, horaire, score
        - Ajoute des menus déroulants pour équipes, poules, gymnases
        - Détecte et corrige automatiquement la poule si manquante
        - Vérifie cohérence équipes/poule
        """
        if df is None or df.empty:
            return
        
        # Créer un mapping équipe -> poule(s) pour validation
        equipe_vers_poules = {}  # {nom_equipe: [liste_poules]}
        df_equipes = self.config.lire_feuille('Equipes')
        if df_equipes is not None:
            for _, row_eq in df_equipes.iterrows():
                equipe_nom = str(row_eq.get('Equipe', '')).strip()
                poule_nom = str(row_eq.get('Poule', '')).strip()
                if equipe_nom and poule_nom:
                    if equipe_nom not in equipe_vers_poules:
                        equipe_vers_poules[equipe_nom] = []
                    equipe_vers_poules[equipe_nom].append(poule_nom)
        
        # Charger les équipes hors championnat autorisées
        equipes_hors_championnat_autorisees = set()
        df_equipes_hors = self.config.lire_feuille('Equipes_Hors_Championnat')
        if df_equipes_hors is not None:
            for _, row_hors in df_equipes_hors.iterrows():
                equipe_nom = str(row_hors.get('Equipe', '')).strip()
                genre = str(row_hors.get('Genre', '')).strip().upper()
                if equipe_nom:
                    # Ajouter la version sans genre
                    equipes_hors_championnat_autorisees.add(equipe_nom)
                    # Ajouter la version avec genre si spécifié
                    if genre in ['F', 'M']:
                        equipes_hors_championnat_autorisees.add(f"{equipe_nom} [{genre}]")
        
        # Créer mapping poule -> équipes pour recherche inverse
        poule_vers_equipes = {}  # {poule: [liste_equipes]}
        for equipe, poules in equipe_vers_poules.items():
            for poule in poules:
                if poule not in poule_vers_equipes:
                    poule_vers_equipes[poule] = []
                poule_vers_equipes[poule].append(equipe)
        
        for idx, row in df.iterrows():
            ligne_num = int(idx) + 2
            
            eq1 = row.get('Equipe_1')
            eq2 = row.get('Equipe_2')
            genre = row.get('Genre')
            poule = row.get('Poule')
            semaine = row.get('Semaine')
            horaire = row.get('Horaire')
            gymnase = row.get('Gymnase')
            score = row.get('Score')
            type_comp = row.get('Type_Competition')
            
            # Ignorer lignes complètement vides
            if pd.isna(eq1) and pd.isna(eq2):
                continue
            
            # ========== VALIDATION ÉQUIPES ==========
            # Au moins une équipe doit être présente
            eq1_valide = not (pd.isna(eq1) or str(eq1).strip() == '')
            eq2_valide = not (pd.isna(eq2) or str(eq2).strip() == '')
            
            if not eq1_valide and not eq2_valide:
                rapport.erreurs_contenu.append(f"Ligne {ligne_num}: Au moins une équipe doit être spécifiée")
                continue
            
            # Préparer les variables pour la validation
            eq1_str = str(eq1).strip() if eq1_valide else None
            eq2_str = str(eq2).strip() if eq2_valide else None
            genre_str = str(genre).strip().upper() if pd.notna(genre) and str(genre).strip() else None
            
            # Vérifier le format seulement pour les équipes présentes
            format_valide_eq1 = bool(re.match(r'^.+\s*\(\d+\)\s*$', eq1_str)) if eq1_valide and eq1_str else True
            format_valide_eq2 = bool(re.match(r'^.+\s*\(\d+\)\s*$', eq2_str)) if eq2_valide and eq2_str else True
            
            if eq1_valide and not format_valide_eq1:
                rapport.warnings_contenu.append(f"Ligne {ligne_num}: Format Equipe_1 invalide (attendu: Institution (numéro))")
            
            if eq2_valide and not format_valide_eq2:
                rapport.warnings_contenu.append(f"Ligne {ligne_num}: Format Equipe_2 invalide (attendu: Institution (numéro))")
            
            # Passer à la validation d'existence seulement si le format est valide et qu'il y a des équipes
            if not (eq1_valide and format_valide_eq1) and not (eq2_valide and format_valide_eq2):
                continue
            
            # Construire les noms complets avec genre si présent
            if genre_str in ['F', 'M']:
                eq1_complet = f"{eq1_str} [{genre_str}]" if eq1_str else None
                eq2_complet = f"{eq2_str} [{genre_str}]" if eq2_str else None
            else:
                eq1_complet = eq1_str
                eq2_complet = eq2_str
            
            # Vérifier existence des équipes (seulement pour celles qui sont présentes)
            eq1_existe = True  # Par défaut, considérer comme existante si non présente
            eq2_existe = True
            
            if eq1_valide and eq1_str:
                # 1. Chercher le nom de base dans equipes_sans_genre
                eq1_existe = eq1_str in self.equipes_sans_genre
                # 2. Si non trouvé et qu'on a un genre, chercher aussi avec le genre complet dans equipes_toutes_variantes
                if not eq1_existe and genre_str in ['F', 'M']:
                    eq1_existe = eq1_complet in self.equipes_toutes_variantes
            
            if eq2_valide and eq2_str:
                # 1. Chercher le nom de base dans equipes_sans_genre
                eq2_existe = eq2_str in self.equipes_sans_genre
                # 2. Si non trouvé et qu'on a un genre, chercher aussi avec le genre complet dans equipes_toutes_variantes
                if not eq2_existe and genre_str in ['F', 'M']:
                    eq2_existe = eq2_complet in self.equipes_toutes_variantes
            
            # 3. Si toujours pas trouvé, vérifier dans les équipes hors championnat autorisées
            eq1_hors_championnat_autorise = False
            eq2_hors_championnat_autorise = False
            
            if eq1_valide and eq1_str and not eq1_existe:
                eq1_hors_championnat_autorise = eq1_str in equipes_hors_championnat_autorisees or (eq1_complet and eq1_complet in equipes_hors_championnat_autorisees)
                if eq1_hors_championnat_autorise:
                    eq1_existe = True  # Permettre la poursuite de la validation
            
            if eq2_valide and eq2_str and not eq2_existe:
                eq2_hors_championnat_autorise = eq2_str in equipes_hors_championnat_autorisees or (eq2_complet and eq2_complet in equipes_hors_championnat_autorisees)
                if eq2_hors_championnat_autorise:
                    eq2_existe = True  # Permettre la poursuite de la validation
            
            # ÉQUIPES HORS CHAMPIONNAT NON AUTORISÉES : avertissement
            equipes_hors_championnat_non_autorisees = set()
            
            if eq1_valide and eq1_str and not eq1_existe and not eq1_hors_championnat_autorise:
                nom_recherche = eq1_complet if genre_str and eq1_complet else eq1_str
                rapport.warnings_contenu.append(
                    f"Ligne {ligne_num}: Équipe '{nom_recherche}' hors championnat académique (ajouter à la feuille Equipes_Hors_Championnat si autorisée)"
                )
                equipes_hors_championnat_non_autorisees.add(nom_recherche)
                eq1_existe = True  # Permettre la poursuite de la validation pour éviter blocage
            
            if eq2_valide and eq2_str and not eq2_existe and not eq2_hors_championnat_autorise:
                nom_recherche = eq2_complet if genre_str and eq2_complet else eq2_str
                rapport.warnings_contenu.append(
                    f"Ligne {ligne_num}: Équipe '{nom_recherche}' hors championnat académique (ajouter à la feuille Equipes_Hors_Championnat si autorisée)"
                )
                equipes_hors_championnat_non_autorisees.add(nom_recherche)
                eq2_existe = True  # Permettre la poursuite de la validation pour éviter blocage
            
            # Signaler les équipes hors championnat détectées
            if equipes_hors_championnat_non_autorisees:
                rapport.corrections_contenu.append(
                    f"Ligne {ligne_num}: Équipes hors championnat détectées: {', '.join(sorted(equipes_hors_championnat_non_autorisees))}"
                )
            
            # ========== VALIDATION GENRE ==========
            if pd.isna(genre) or str(genre).strip() == '':
                rapport.erreurs_contenu.append(f"Ligne {ligne_num}: Genre manquant")
            elif genre_str not in ['F', 'M']:
                rapport.erreurs_contenu.append(
                    f"Ligne {ligne_num}: Genre invalide ('{genre_str}'), doit être 'F' ou 'M'"
                )
            
            # ========== VALIDATION/AUTO-DÉTECTION POULE ==========
            poule_str = str(poule).strip() if pd.notna(poule) and str(poule).strip() else None
            type_comp_str = str(type_comp).strip() if pd.notna(type_comp) and str(type_comp).strip() else None
            
            # Si pas de type de compétition (= match de championnat régulier académique)
            if not type_comp_str or type_comp_str == 'Acad':
                # La poule est OBLIGATOIRE et doit contenir les deux équipes
                
                if not poule_str:
                    # TENTATIVE DE DÉTECTION AUTOMATIQUE
                    poules_eq1 = equipe_vers_poules.get(eq1_str, [])
                    poules_eq2 = equipe_vers_poules.get(eq2_str, [])
                    
                    # Trouver les poules communes
                    poules_communes = set(poules_eq1) & set(poules_eq2)
                    
                    if len(poules_communes) == 1:
                        poule_detectee = list(poules_communes)[0]
                        rapport.corrections_contenu.append(
                            f"Ligne {ligne_num}: Poule manquante → Auto-détectée '{poule_detectee}' (commune à {eq1_str} et {eq2_str})"
                        )
                        poule_str = poule_detectee
                        # TODO: Appliquer la correction dans le DataFrame
                    elif len(poules_communes) > 1:
                        rapport.erreurs_contenu.append(
                            f"Ligne {ligne_num}: ⚠️ AMBIGUÏTÉ - Poule manquante, {eq1_str} et {eq2_str} sont dans plusieurs poules communes: {', '.join(poules_communes)}"
                        )
                    else:
                        rapport.erreurs_contenu.append(
                            f"Ligne {ligne_num}: ❌ ERREUR CRITIQUE - Poule manquante et {eq1_str} et {eq2_str} ne partagent AUCUNE poule commune ! "
                            f"({eq1_str} dans {poules_eq1 or 'aucune'}, {eq2_str} dans {poules_eq2 or 'aucune'})"
                        )
                else:
                    # Poule fournie → Vérifier qu'elle contient bien les deux équipes
                    if poule_str not in poule_vers_equipes:
                        rapport.erreurs_contenu.append(
                            f"Ligne {ligne_num}: Poule '{poule_str}' inexistante"
                        )
                    else:
                        equipes_dans_poule = poule_vers_equipes[poule_str]
                        
                        if eq1_str not in equipes_dans_poule:
                            rapport.erreurs_contenu.append(
                                f"Ligne {ligne_num}: ❌ INCOHÉRENCE - {eq1_str} n'est PAS dans la poule '{poule_str}' "
                                f"(Équipes de cette poule: {', '.join(equipes_dans_poule)})"
                            )
                        if eq2_str not in equipes_dans_poule:
                            rapport.erreurs_contenu.append(
                                f"Ligne {ligne_num}: ❌ INCOHÉRENCE - {eq2_str} n'est PAS dans la poule '{poule_str}' "
                                f"(Équipes de cette poule: {', '.join(equipes_dans_poule)})"
                            )
            
            # ========== VALIDATION SEMAINE ==========
            if pd.isna(semaine):
                rapport.erreurs_contenu.append(f"Ligne {ligne_num}: Semaine manquante")
            else:
                # Tenter de parser comme nombre (accepter 1.0, "1", etc.)
                try:
                    semaine_float = float(semaine)
                    semaine_int = int(semaine_float)
                    
                    # Vérifier si c'est bien un entier (pas 1.5)
                    if semaine_float != semaine_int:
                        rapport.erreurs_contenu.append(
                            f"Ligne {ligne_num}: Semaine invalide ('{semaine}'), doit être un nombre entier"
                        )
                    elif semaine_int < 1 or semaine_int > 52:
                        rapport.erreurs_contenu.append(
                            f"Ligne {ligne_num}: Semaine hors limites ({semaine_int}), doit être entre 1 et 52"
                        )
                except (ValueError, TypeError):
                    rapport.erreurs_contenu.append(
                        f"Ligne {ligne_num}: Format semaine invalide ('{semaine}'), attendu: nombre entier"
                    )
            
            # ========== VALIDATION HORAIRE ==========
            if pd.isna(horaire) or str(horaire).strip() == '':
                rapport.erreurs_contenu.append(f"Ligne {ligne_num}: Horaire manquant")
            else:
                result = self.validator.valider_horaire(horaire)
                if not result.valide:
                    rapport.erreurs_contenu.append(
                        f"Ligne {ligne_num}: Horaire invalide - {result.message}"
                    )
                elif result.valeur_corrigee and str(result.valeur_corrigee) != str(horaire).strip():
                    rapport.corrections_contenu.append(
                        f"Ligne {ligne_num}: Horaire reformaté: '{horaire}' → '{result.valeur_corrigee}'"
                    )
            
            # ========== VALIDATION GYMNASE ==========
            if pd.isna(gymnase) or str(gymnase).strip() == '':
                rapport.erreurs_contenu.append(f"Ligne {ligne_num}: Gymnase manquant")
            else:
                gymnase_str = str(gymnase).strip()
                if gymnase_str not in self.gymnases_ref:
                    # Recherche floue
                    matches = difflib.get_close_matches(gymnase_str, self.gymnases_ref, n=1, cutoff=0.6)
                    if matches:
                        rapport.erreurs_contenu.append(
                            f"Ligne {ligne_num}: Gymnase '{gymnase_str}' non trouvé. Vouliez-vous dire '{matches[0]}' ?"
                        )
                    else:
                        rapport.erreurs_contenu.append(
                            f"Ligne {ligne_num}: Gymnase '{gymnase_str}' non trouvé dans la feuille Gymnases"
                        )
            
            # ========== VALIDATION SCORE (optionnel mais avec format) ==========
            if pd.notna(score) and str(score).strip():
                score_str = str(score).strip()
                # Formats acceptés: "3-1", "25-23, 25-20", "3-0 (25-20, 25-18, 25-15)", etc.
                # Pattern simple: doit contenir des chiffres et des tirets/virgules
                if not re.search(r'\d+[-,\s]\d+', score_str):
                    rapport.warnings_contenu.append(
                        f"Ligne {ligne_num}: Format de score potentiellement invalide ('{score_str}'). "
                        f"Exemples attendus: '3-1', '25-23, 25-20', '3-0 (25-20, 25-18, 25-15)'"
                    )
            
            # ========== VALIDATION TYPE COMPÉTITION ==========
            types_valides = ['CFE', 'CFU', 'Acad', 'Autre']
            if pd.notna(type_comp) and str(type_comp).strip():
                type_str = str(type_comp).strip()
                if type_str not in types_valides:
                    rapport.warnings_contenu.append(
                        f"Ligne {ligne_num}: Type de compétition '{type_str}' non standard. "
                        f"Valeurs attendues: {', '.join(types_valides)}"
                    )
    
    def _valider_equipes_hors_championnat(self, df: pd.DataFrame, rapport: RapportFeuille):
        """Valide la feuille Equipes_Hors_Championnat.
        
        Cette feuille contient les équipes externes autorisées pour les matchs fixes.
        Contrairement aux autres feuilles, elle permet des équipes et institutions
        qui ne sont pas dans le championnat académique.
        """
        for idx, row in df.iterrows():
            ligne_num = int(idx) + 2
            
            # ========== VALIDATION ÉQUIPE ==========
            equipe = row.get('Equipe')
            if pd.isna(equipe) or str(equipe).strip() == '':
                rapport.erreurs_contenu.append(f"Ligne {ligne_num}: Équipe manquante")
            else:
                equipe_str = str(equipe).strip()
                # Format attendu: Institution (numéro) [F/M] ou Institution (numéro)
                if not re.match(r'^.+\s*\(\d+\)\s*(?:\s*[FM]\s*)?$', equipe_str):
                    rapport.erreurs_contenu.append(
                        f"Ligne {ligne_num}: Format équipe invalide: '{equipe_str}' "
                        "(attendu: 'Institution (numéro)' ou 'Institution (numéro) F/M')"
                    )
            
            # ========== VALIDATION INSTITUTION ==========
            institution = row.get('Institution')
            if pd.isna(institution) or str(institution).strip() == '':
                rapport.erreurs_contenu.append(f"Ligne {ligne_num}: Institution manquante")
            else:
                institution_str = str(institution).strip()
                # L'institution peut être n'importe quoi, pas de validation stricte
                # Juste vérifier que ce n'est pas vide après nettoyage
                if not institution_str:
                    rapport.erreurs_contenu.append(f"Ligne {ligne_num}: Institution vide")
            
            # ========== VALIDATION GENRE ==========
            genre = row.get('Genre')
            if pd.isna(genre) or str(genre).strip() == '':
                rapport.erreurs_contenu.append(f"Ligne {ligne_num}: Genre manquant")
            else:
                genre_str = str(genre).strip()
                if genre_str not in ['M', 'F']:
                    rapport.erreurs_contenu.append(
                        f"Ligne {ligne_num}: Genre invalide ('{genre_str}'), doit être 'F' ou 'M'"
                    )
            
            # ========== VALIDATION TYPE CHAMPIONNAT ==========
            type_championnat = row.get('Type_Championnat')
            if pd.isna(type_championnat) or str(type_championnat).strip() == '':
                rapport.erreurs_contenu.append(f"Ligne {ligne_num}: Type de championnat manquant")
            else:
                type_str = str(type_championnat).strip()
                types_valides = ['CFE', 'CFU', 'Autre']
                if type_str not in types_valides:
                    rapport.erreurs_contenu.append(
                        f"Ligne {ligne_num}: Type de championnat '{type_str}' invalide. "
                        f"Valeurs acceptées: {', '.join(types_valides)}"
                    )
    
    def _generer_feuilles_manquantes(self) -> int:
        """Génère les feuilles manquantes sans aucune ligne d'exemple."""
        statuts = self.config.generer_feuilles_manquantes(conserver_existant=True)
        
        feuilles_creees = [f for f, s in statuts.items() if s == 'créée']
        
        if feuilles_creees:
            # Supprimer immédiatement tout exemple qui aurait pu être créé
            self._supprimer_exemples()
        
        return len(feuilles_creees)
    
    def _generer_feuilles_manquantes_old(self):
        """Ancienne version - génère les feuilles manquantes sans aucune ligne d'exemple."""
        statuts = self.config.generer_feuilles_manquantes(conserver_existant=True)
        
        feuilles_creees = [f for f, s in statuts.items() if s == 'créée']
        
        if feuilles_creees:
            print(f"  ✅ {len(feuilles_creees)} feuilles créées (sans exemples)")
            for f in feuilles_creees:
                print(f"     • {f}")
            
            # Supprimer immédiatement tout exemple qui aurait pu être créé
            self._supprimer_exemples()
        else:
            print("  ✅ Toutes les feuilles existent déjà")
    
    def _supprimer_exemples(self) -> int:
        """Supprime TOUTES les lignes d'exemple de TOUTES les feuilles."""
        wb = openpyxl.load_workbook(self.fichier_path)
        nb_suppressions = 0
        
        for nom_feuille in self.config.STRUCTURES.keys():
            if nom_feuille not in wb.sheetnames:
                continue
            
            df = self.config.lire_feuille(nom_feuille)
            if df is None or len(df) == 0:
                continue
            
            structure = self.config.STRUCTURES.get(nom_feuille, {})
            exemple = structure.get('exemple', {})
            
            if not exemple:
                continue
            
            # Vérifier si la première ligne contient des valeurs d'exemple
            premiere_ligne = df.iloc[0]
            est_exemple = True
            
            # Vérifier chaque colonne
            for col in df.columns:
                if col in exemple:
                    valeur_df = str(premiere_ligne[col]).strip()
                    valeur_exemple = str(exemple[col]).strip()
                    if valeur_df != valeur_exemple:
                        est_exemple = False
                        break
            
            if est_exemple:
                # Supprimer la première ligne (exemple)
                df_nettoye = df.iloc[1:].reset_index(drop=True)
                
                # Supprimer et recréer la feuille
                ws_index = wb.sheetnames.index(nom_feuille)
                del wb[nom_feuille]
                ws = wb.create_sheet(nom_feuille, ws_index)
                
                # Écrire les en-têtes
                for c_idx, col in enumerate(df_nettoye.columns, 1):
                    ws.cell(1, c_idx, col)
                
                # Écrire les données (sans la ligne d'exemple)
                for r_idx, row in enumerate(df_nettoye.itertuples(index=False), 2):
                    for c_idx, value in enumerate(row, 1):
                        if not pd.isna(value):
                            ws.cell(r_idx, c_idx, value)
                
                nb_suppressions += 1
        
        if nb_suppressions > 0:
            wb.save(self.fichier_path)
        
        return nb_suppressions
    
    def _preremplir_types_poules(self) -> int:
        """
        Pré-remplit automatiquement la feuille Types_Poules avec toutes les poules extraites
        de la feuille Equipes. Ajoute une validation par liste déroulante ['Classique', 'Aller-Retour'].
        """
        try:
            # Lire la feuille Equipes pour extraire toutes les poules uniques
            df_equipes = self.config.lire_feuille('Equipes')
            if df_equipes is None or 'Poule' not in df_equipes.columns:
                return 0
            
            # Extraire les poules uniques (non vides)
            poules_uniques = set()
            for poule in df_equipes['Poule'].dropna():
                poule_str = str(poule).strip()
                if poule_str:
                    poules_uniques.add(poule_str)
            
            if not poules_uniques:
                return 0
            
            poules_sorted = sorted(poules_uniques)
            
            # Lire la feuille Types_Poules existante (si elle existe)
            df_types = self.config.lire_feuille('Types_Poules')
            poules_existantes = set()
            
            if df_types is not None and not df_types.empty and 'Poule' in df_types.columns:
                for poule in df_types['Poule'].dropna():
                    poule_str = str(poule).strip()
                    if poule_str:
                        poules_existantes.add(poule_str)
            
            # Déterminer les poules à ajouter (celles qui ne sont pas déjà dans Types_Poules)
            poules_a_ajouter = [p for p in poules_sorted if p not in poules_existantes]
            
            if not poules_a_ajouter:
                # Toujours mettre à jour les validations, même si pas de nouvelles poules
                self._ajouter_validations_types_poules()
                return 0
            
            # Ouvrir le fichier Excel
            wb = openpyxl.load_workbook(self.fichier_path)
            
            # Vérifier que la feuille Types_Poules existe
            if 'Types_Poules' not in wb.sheetnames:
                return 0
            
            ws = wb['Types_Poules']
            
            # Trouver la première ligne vide (après les données existantes)
            derniere_ligne = 1  # Ligne des en-têtes
            for row in ws.iter_rows(min_row=2):
                if row[0].value is None:  # Colonne Poule vide
                    break
                derniere_ligne += 1
            
            # Ajouter les nouvelles poules
            for poule in poules_a_ajouter:
                derniere_ligne += 1
                ws.cell(derniere_ligne, 1, poule)  # Colonne Poule
                ws.cell(derniere_ligne, 2, 'Classique')  # Colonne Type (défaut)
                ws.cell(derniere_ligne, 3, '')  # Colonne Remarques (vide)
            
            wb.save(self.fichier_path)
            
            # Ajouter les validations par liste déroulante
            self._ajouter_validations_types_poules()
            
            return len(poules_a_ajouter)
            
        except Exception as e:
            return 0
    
    def _ajouter_validations_types_poules(self):
        """Ajoute les validations par liste déroulante à la colonne Type de Types_Poules."""
        try:
            from openpyxl.worksheet.datavalidation import DataValidation
            
            wb = openpyxl.load_workbook(self.fichier_path)
            
            if 'Types_Poules' not in wb.sheetnames:
                return
            
            ws = wb['Types_Poules']
            
            # Supprimer les validations existantes
            ws.data_validations.dataValidation = []
            
            # Créer la validation pour la colonne Type (colonne B)
            dv = DataValidation(
                type="list",
                formula1='"Classique,Aller-Retour"',
                allow_blank=False,
                showErrorMessage=True,
                errorTitle="Type invalide",
                error="Veuillez choisir 'Classique' ou 'Aller-Retour'"
            )
            
            # Appliquer à toute la colonne Type (de B2 à B1000)
            dv.add('B2:B1000')
            ws.add_data_validation(dv)
            
            wb.save(self.fichier_path)
            
        except Exception as e:
            print(f"  ⚠️  Erreur lors de l'ajout des validations: {e}")
    
    def _ajouter_validations_matchs_fixes(self):
        """
        Ajoute les validations par liste déroulante pour la feuille Matchs_Fixes:
        - Colonne A (Equipe_1): Équipes SANS le genre entre crochets
        - Colonne B (Equipe_2): Équipes SANS le genre entre crochets
        - Colonne C (Genre): F ou M
        - Colonne D (Poule): Toutes les poules
        - Colonne G (Gymnase): Tous les gymnases
        - Colonne I (Type_Competition): CFE, CFU, Acad, Autre
        """
        try:
            from openpyxl.worksheet.datavalidation import DataValidation
            from openpyxl.utils import quote_sheetname
            
            wb = openpyxl.load_workbook(self.fichier_path)
            
            if 'Matchs_Fixes' not in wb.sheetnames:
                return
            
            ws = wb['Matchs_Fixes']
            
            # Supprimer toutes les validations existantes
            ws.data_validations.dataValidation = []
            
            # ========== CRÉER UNE FEUILLE CACHÉE POUR LES LISTES ==========
            sheet_name_listes = 'Listes_Deroulantes'
            if sheet_name_listes not in wb.sheetnames:
                ws_listes = wb.create_sheet(sheet_name_listes)
                ws_listes.sheet_state = 'hidden'  # Cacher la feuille
            else:
                ws_listes = wb[sheet_name_listes]
                ws_listes.sheet_state = 'hidden'
            
            # ========== EQUIPE_1 et EQUIPE_2 (Colonnes A et B) ==========
            # Utilise la liste des équipes SANS le genre entre crochets
            equipes_base = set()
            for eq in self.equipes_toutes_variantes:
                # Enlever le [F] ou [M] si présent
                eq_sans_genre = eq.replace(' [F]', '').replace(' [M]', '').strip()
                equipes_base.add(eq_sans_genre)
            
            # Ajouter les équipes hors championnat autorisées
            df_equipes_hors = self.config.lire_feuille('Equipes_Hors_Championnat')
            if df_equipes_hors is not None:
                for _, row_hors in df_equipes_hors.iterrows():
                    equipe_nom = str(row_hors.get('Equipe', '')).strip()
                    if equipe_nom:
                        # Enlever le [F] ou [M] si présent dans le nom
                        eq_sans_genre = equipe_nom.replace(' [F]', '').replace(' [M]', '').strip()
                        equipes_base.add(eq_sans_genre)
            
            if equipes_base:
                liste_equipes = sorted(list(equipes_base))
                
                # Écrire la liste dans la feuille cachée (colonne A)
                for i, equipe in enumerate(liste_equipes, start=1):
                    ws_listes.cell(row=i, column=1, value=equipe)
                
                # Créer la validation avec référence à la plage
                nb_equipes = len(liste_equipes)
                formule_equipes = f"{quote_sheetname(sheet_name_listes)}!$A$1:$A${nb_equipes}"
                
                dv_eq1 = DataValidation(
                    type="list",
                    formula1=formule_equipes,
                    allow_blank=True,
                    showErrorMessage=True,
                    errorTitle="Équipe invalide",
                    error="Sélectionnez une équipe de la liste (y compris équipes hors championnat)"
                )
                dv_eq1.add('A2:A1000')
                ws.add_data_validation(dv_eq1)
                
                dv_eq2 = DataValidation(
                    type="list",
                    formula1=formule_equipes,
                    allow_blank=True,
                    showErrorMessage=True,
                    errorTitle="Équipe invalide",
                    error="Sélectionnez une équipe de la liste (y compris équipes hors championnat)"
                )
                dv_eq2.add('B2:B1000')
                ws.add_data_validation(dv_eq2)
            
            # ========== GENRE (Colonne C) - NOUVEAU! ==========
            dv_genre = DataValidation(
                type="list",
                formula1='"F,M"',
                allow_blank=True,
                showErrorMessage=True,
                errorTitle="Genre invalide",
                error="Sélectionnez F ou M"
            )
            dv_genre.add('C2:C1000')
            ws.add_data_validation(dv_genre)
            
            # ========== POULE (Colonne D) ==========
            poules = set()
            df_equipes = self.config.lire_feuille('Equipes')
            if df_equipes is not None:
                for _, row in df_equipes.iterrows():
                    poule = row.get('Poule')
                    if pd.notna(poule) and str(poule).strip():
                        poules.add(str(poule).strip())
            
            if poules:
                liste_poules = sorted(list(poules))
                formule_poules = '"' + ','.join(liste_poules) + '"'
                
                dv_poule = DataValidation(
                    type="list",
                    formula1=formule_poules,
                    allow_blank=True,
                    showErrorMessage=True,
                    errorTitle="Poule invalide",
                    error="Sélectionnez une poule de la liste"
                )
                dv_poule.add('D2:D1000')
                ws.add_data_validation(dv_poule)
            
            # ========== GYMNASE (Colonne G) ==========
            if self.gymnases_ref:
                liste_gymnases = sorted(list(self.gymnases_ref))
                formule_gymnases = '"' + ','.join(liste_gymnases) + '"'
                
                dv_gymnase = DataValidation(
                    type="list",
                    formula1=formule_gymnases,
                    allow_blank=True,
                    showErrorMessage=True,
                    errorTitle="Gymnase invalide",
                    error="Sélectionnez un gymnase de la liste"
                )
                dv_gymnase.add('G2:G1000')
                ws.add_data_validation(dv_gymnase)
            
            # ========== TYPE_COMPETITION (Colonne I) ==========
            dv_type = DataValidation(
                type="list",
                formula1='"CFE,CFU,Acad,Autre"',
                allow_blank=True,
                showErrorMessage=True,
                errorTitle="Type invalide",
                error="Choisissez: CFE, CFU, Acad, Autre"
            )
            dv_type.add('I2:I1000')
            ws.add_data_validation(dv_type)
            
            wb.save(self.fichier_path)
            
        except Exception as e:
            print(f"  ⚠️  Erreur lors de l'ajout des validations Matchs_Fixes: {e}")
    
    def _ajouter_validations_niveaux_gymnases(self):
        """
        Ajoute les validations par liste déroulante pour la feuille Niveaux_Gymnases:
        - Colonne A (Gymnase): Tous les gymnases disponibles
        - Colonne B (Niveau): Haut niveau, Bas niveau
        """
        try:
            from openpyxl.worksheet.datavalidation import DataValidation
            from openpyxl.utils import quote_sheetname
            
            wb = openpyxl.load_workbook(self.fichier_path)
            
            if 'Niveaux_Gymnases' not in wb.sheetnames:
                return
            
            ws = wb['Niveaux_Gymnases']
            
            # Supprimer toutes les validations existantes
            ws.data_validations.dataValidation = []
            
            # ========== CRÉER UNE FEUILLE CACHÉE POUR LES LISTES ==========
            sheet_name_listes = 'Listes_Deroulantes'
            if sheet_name_listes not in wb.sheetnames:
                ws_listes = wb.create_sheet(sheet_name_listes)
                ws_listes.sheet_state = 'hidden'  # Cacher la feuille
            else:
                ws_listes = wb[sheet_name_listes]
                ws_listes.sheet_state = 'hidden'
            
            # ========== GYMNASE (Colonne A) ==========
            # Liste de tous les gymnases
            if not self.gymnases_ref:
                print(f"  ⚠️  Aucun gymnase trouvé pour les validations")
                return
                
            gymnases = sorted(self.gymnases_ref)
            
            # Écrire la liste dans la feuille cachée (utiliser colonne C pour éviter conflit avec colonne A utilisée par Matchs_Fixes)
            col_gymnases = 3  # Colonne C de Listes_Deroulantes
            for i, gymnase in enumerate(gymnases, start=1):
                ws_listes.cell(row=i, column=col_gymnases, value=gymnase)
            
            # Créer la validation pour la colonne Gymnase
            range_gymnases = f"{quote_sheetname(sheet_name_listes)}!$C$1:$C${len(gymnases)}"
            dv_gymnase = DataValidation(
                type="list",
                formula1=range_gymnases,
                allow_blank=False,
                showErrorMessage=True,
                errorTitle="Gymnase invalide",
                error="Choisissez un gymnase dans la liste déroulante"
            )
            dv_gymnase.add('A2:A1000')
            ws.add_data_validation(dv_gymnase)
            
            # ========== NIVEAU (Colonne B) ==========
            # Liste des niveaux (inline, pas besoin de feuille cachée pour 2 valeurs)
            dv_niveau = DataValidation(
                type="list",
                formula1='"Haut niveau,Bas niveau"',
                allow_blank=False,
                showErrorMessage=True,
                errorTitle="Niveau invalide",
                error="Choisissez 'Haut niveau' ou 'Bas niveau'"
            )
            dv_niveau.add('B2:B1000')
            ws.add_data_validation(dv_niveau)
            
            wb.save(self.fichier_path)
            
        except Exception as e:
            print(f"  ⚠️  Erreur lors de l'ajout des validations Niveaux_Gymnases: {e}")
    
    def _appliquer_corrections(self) -> int:
        """Applique toutes les corrections détectées."""
        corrections_appliquees = 0
        
        wb = openpyxl.load_workbook(self.fichier_path)
        
        for nom_feuille, rapport in self.rapports_feuilles.items():
            if not rapport.structure_modifiee and not rapport.corrections_contenu:
                continue
            
            if nom_feuille not in wb.sheetnames:
                continue
            
            df = self.config.lire_feuille(nom_feuille)
            if df is None:
                continue
            
            structure = self.config.STRUCTURES.get(nom_feuille, {})
            colonnes_attendues = structure.get('colonnes', [])
            
            # Renommer les colonnes
            if rapport.colonnes_renommees:
                for ancien, nouveau in rapport.colonnes_renommees:
                    if ancien in df.columns:
                        df.rename(columns={ancien: nouveau}, inplace=True)
                        corrections_appliquees += 1
            
            # Ajouter les colonnes manquantes
            for col_manquante in rapport.colonnes_manquantes:
                df[col_manquante] = None
                corrections_appliquees += 1
            
            # Renommer les colonnes extra avec préfixe EXTRA_
            for col_extra in rapport.colonnes_extra:
                if col_extra in df.columns:
                    nouveau_nom = f"EXTRA_{col_extra}"
                    df.rename(columns={col_extra: nouveau_nom}, inplace=True)
                    corrections_appliquees += 1
            
            # Réorganiser les colonnes : colonnes attendues d'abord, puis EXTRA_
            colonnes_ordonnees = []
            for col in colonnes_attendues:
                if col in df.columns:
                    colonnes_ordonnees.append(col)
            
            # Ajouter les colonnes EXTRA_ à la fin
            for col in df.columns:
                if col.startswith('EXTRA_') and col not in colonnes_ordonnees:
                    colonnes_ordonnees.append(col)
            
            df = df[colonnes_ordonnees]
            
            # Supprimer et recréer la feuille
            ws_index = wb.sheetnames.index(nom_feuille)
            del wb[nom_feuille]
            ws = wb.create_sheet(nom_feuille, ws_index)
            
            # Écrire les données
            for c_idx, col in enumerate(df.columns, 1):
                cell = ws.cell(1, c_idx, col)
                # Mettre en évidence les colonnes EXTRA_
                if str(col).startswith('EXTRA_'):
                    cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
                    cell.font = Font(italic=True, color="FF0000")
            
            for r_idx, row in enumerate(df.itertuples(index=False), 2):
                for c_idx, value in enumerate(row, 1):
                    if not pd.isna(value):
                        ws.cell(r_idx, c_idx, value)
        
        wb.save(self.fichier_path)
        
        return corrections_appliquees
    
    def _generer_variantes_equipes(self) -> List[str]:
        """
        Génère la liste complète des équipes incluant les variantes avec genre.
        
        Règles:
        - Si une équipe existe en un seul genre ou sans genre: pas de suffixe [M]/[F]
        - Si une équipe existe en plusieurs genres (M et F): créer les deux variantes avec [M] et [F]
        
        Returns:
            Liste triée des équipes avec variantes genre si nécessaire
        
        Exemple:
            equipes_par_nom = {
                "LYON 1 (1)": {"M", "F"},  # Existe en 2 genres
                "LYON 2 (1)": {"M"},        # Existe en 1 seul genre
                "LYON 3 (1)": set()         # Pas de genre détecté
            }
            → Retourne: ["LYON 1 (1) [F]", "LYON 1 (1) [M]", "LYON 2 (1)", "LYON 3 (1)"]
        """
        from core.utils import formater_nom_avec_genre
        
        variantes = []
        
        for nom_equipe, genres in self.equipes_par_nom.items():
            if len(genres) > 1:
                # Plusieurs genres: créer une variante pour chaque
                for genre in sorted(genres):  # M avant F par ordre alphabétique
                    variantes.append(formater_nom_avec_genre(nom_equipe, genre))
            else:
                # Un seul genre ou pas de genre: pas de suffixe
                variantes.append(nom_equipe)
        
        return sorted(variantes)
    
    def _actualiser_listes_deroulantes(self):
        """Met à jour les listes déroulantes avec variantes de genres."""
        try:
            # Appliquer le formatage (les variantes seront générées automatiquement
            # par _extraire_liste_valeurs dans config_manager.py)
            self.config._formater_fichier()
            
            # Ajouter explicitement les validations pour Types_Poules (au cas où)
            self._ajouter_validations_types_poules()
            
            # Ajouter les validations pour Matchs_Fixes
            self._ajouter_validations_matchs_fixes()
            
            # Ajouter les validations pour Niveaux_Gymnases
            self._ajouter_validations_niveaux_gymnases()
        except Exception as e:
            pass  # Silencieux si erreur
    
    def _afficher_rapport_final(self):
        """Affiche le rapport final détaillé."""
        # Statistiques globales
        total_erreurs = sum(len(r.erreurs_contenu) for r in self.rapports_feuilles.values())
        total_warnings = sum(len(r.warnings_contenu) for r in self.rapports_feuilles.values())
        total_corrections = sum(len(r.corrections_contenu) for r in self.rapports_feuilles.values())
        
        # Si pas de problème, affichage très concis
        if total_erreurs == 0 and total_warnings == 0:
            print("  ╔═══════════════════════════════════════════════════════════╗")
            print("  ║                  ✅ VALIDATION RÉUSSIE                    ║")
            print("  ╚═══════════════════════════════════════════════════════════╝")
            print(f"\n  📊 {len(self.rapports_feuilles)} feuille(s) validée(s) • Aucun problème détecté")
            print(f"  🎉 Le fichier est prêt à l'emploi !\n")
            return
        
        # Si problèmes, affichage détaillé
        print("\n  ╔═══════════════════════════════════════════════════════════╗")
        if total_erreurs > 0:
            print("  ║              ⚠️  VALIDATION AVEC PROBLÈMES                 ║")
        else:
            print("  ║            ⚠️  VALIDATION AVEC AVERTISSEMENTS              ║")
        print("  ╚═══════════════════════════════════════════════════════════╝\n")
        
        print(f"  📊 Résumé : {len(self.rapports_feuilles)} feuille(s) • " +
              f"{total_erreurs} erreur(s) • {total_warnings} avertissement(s)\n")
        
        # Rapport par feuille (seulement celles avec problèmes)
        feuilles_avec_problemes = [
            (nom, rapport) for nom, rapport in self.rapports_feuilles.items()
            if rapport.erreurs_contenu or rapport.warnings_contenu
        ]
        
        if feuilles_avec_problemes:
            print("  ┌───────────────────────────────────────────────────────────┐")
            print("  │                 DÉTAILS DES PROBLÈMES                     │")
            print("  └───────────────────────────────────────────────────────────┘\n")
            
            for nom_feuille, rapport in feuilles_avec_problemes:
                print(f"  📄 {nom_feuille}:")
                
                if rapport.erreurs_contenu:
                    print(f"     ❌ {len(rapport.erreurs_contenu)} erreur(s):")
                    for err in rapport.erreurs_contenu:  # Afficher TOUTES les erreurs
                        print(f"        • {err}")
                
                if rapport.warnings_contenu:
                    print(f"     ⚠️  {len(rapport.warnings_contenu)} avertissement(s):")
                    for warn in rapport.warnings_contenu:  # Afficher TOUS les avertissements
                        print(f"        • {warn}")
                
                print()
        
        # Message final
        print("  " + "─"*59)
        if total_erreurs > 0:
            print("  ⛔ Corrigez les erreurs avant d'utiliser le fichier")
        else:
            print("  ✓ Le fichier est utilisable (vérifiez les avertissements)")
        print()
    
    def _a_reussi(self) -> bool:
        """Détermine si l'actualisation a réussi."""
        total_erreurs = sum(len(r.erreurs_contenu) for r in self.rapports_feuilles.values())
        return total_erreurs == 0


def actualiser_fichier_v2(fichier: str) -> bool:
    """
    Actualise un fichier de configuration avec validation avancée.
    
    Args:
        fichier: Chemin vers le fichier de configuration
        
    Returns:
        True si l'actualisation a réussi, False sinon
    """
    actualisateur = ConfigActualisateurV2(fichier)
    return actualisateur.actualiser()


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(
        description="Actualise et valide un fichier de configuration (Version 2 - Améliorée)"
    )
    
    parser.add_argument(
        '--fichier',
        default='exemple/config_exemple.xlsx',
        help="Fichier de configuration à actualiser (défaut: exemple/config_exemple.xlsx)"
    )
    
    args = parser.parse_args()
    
    success = actualiser_fichier_v2(args.fichier)
    sys.exit(0 if success else 1)
