"""
Gestionnaire du fichier de configuration central.

Ce module gÃ¨re un fichier Excel unique contenant toutes        'Groupes_Non_Simultaneite': {
            'colonnes': ['Nom_Groupe', 'Entites', 'Remarques'],
            'description': 'Groupes d\'Ã©quipes/institutions Ã  ne pas faire jouer simultanÃ©ment',
            'type': 'manuel',
            'exemple': {
                'Nom_Groupe': 'Grandes Ã‰coles Lyon',
                'Entites': 'ECL, EML, CENTRALE',
                'Remarques': 'Grandes Ã©coles d\'ingÃ©nieurs qui ne doivent pas jouer simultanÃ©ment'
            }
        }tions nÃ©cessaires
Ã  la planification : Ã©quipes, gymnases, indisponibilitÃ©s, prÃ©fÃ©rences, contraintes.

Structure du fichier Excel :
- Equipes : Liste des Ã©quipes avec poule et horaire prÃ©fÃ©rÃ©
- Gymnases : Gymnases avec crÃ©neaux disponibles
- Indispos_Equipes : IndisponibilitÃ©s spÃ©cifiques par Ã©quipe
- Indispos_Institutions : IndisponibilitÃ©s par institution (appliquÃ©es Ã  toutes les Ã©quipes)
- Preferences_Institutions : Lieux prÃ©fÃ©rÃ©s par institution avec classement
- Contraintes_Specifiques : Contraintes particuliÃ¨res (anti-collisions, etc.)
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import logging

logger = logging.getLogger(__name__)


class ConfigManager:
    """Gestionnaire du fichier de configuration central."""
    
    # DÃ©finition des structures de chaque feuille
    STRUCTURES = {
        'Equipes_Hors_Championnat': {
            'colonnes': ['Equipe', 'Institution', 'Genre', 'Type_Championnat', 'Motif', 'Remarques'],
            'description': 'Ã‰quipes autorisÃ©es hors championnat acadÃ©mique pour les matchs fixÃ©s',
            'type': 'manuel',
            'exemple': {
                'Equipe': 'EXTERIEUR (1)',
                'Institution': 'EXTERIEUR',
                'Genre': 'M',
                'Type_Championnat': 'CFE',
                'Motif': 'Match amical',
                'Remarques': 'Ã‰quipe invitÃ©e pour tournoi'
            },
            'notes': [
                'Equipe: Nom complet de l\'Ã©quipe (format: Institution (numÃ©ro))',
                'Institution: Institution de rattachement',
                'Genre: M (masculin) ou F (fÃ©minin)',
                'Type_Championnat: Type de championnat (CFE, CFU, Autre)',
                'Motif: Raison de l\'autorisation (match amical, tournoi, etc.)',
                'Remarques: Informations complÃ©mentaires'
            ]
        },
        'Gymnases': {
            'colonnes': ['Gymnase', 'Adresse', 'Capacite', 'Creneaux'],
            'description': 'Gymnases avec capacitÃ© et crÃ©neaux disponibles (Ã€ REMPLIR MANUELLEMENT)',
            'type': 'manuel',
            'exemple': {
                'Gymnase': 'INSA C',
                'Adresse': '20 Avenue Albert Einstein, 69100 Villeurbanne',
                'Capacite': 2,
                'Creneaux': '09:00, 14:00, 18:00'
            }
        },
        'Indispos_Gymnases': {
            'colonnes': ['Gymnase', 'Semaine', 'Horaire_Debut', 'Horaire_Fin', 'Capacite_Occupee', 'Remarques'],
            'description': 'IndisponibilitÃ©s des gymnases avec capacitÃ© partielle (GÃ‰NÃ‰RÃ‰E AUTOMATIQUEMENT)',
            'type': 'auto',
            'exemple': {
                'Gymnase': 'PARC DES SPORTS',
                'Semaine': 3,
                'Horaire_Debut': '09:00',
                'Horaire_Fin': '12:00',
                'Capacite_Occupee': 1,
                'Remarques': 'Maintenance'
            },
            'notes': [
                'Capacite_Occupee: Nombre de terrains/crÃ©neaux occupÃ©s (dÃ©faut = capacitÃ© totale si vide)',
                '  â†’ Si gymnase a capacitÃ© 3 et Capacite_Occupee = 1, il reste 2 crÃ©neaux disponibles',
                '  â†’ Si vide ou >= capacitÃ© totale, le gymnase est totalement indisponible',
                'Horaire_Debut/Fin: Plage horaire concernÃ©e (format HH:MM)'
            ]
        },
        'Indispos_Equipes': {
            'colonnes': ['Equipe', 'Semaine', 'Horaire_Debut', 'Horaire_Fin', 'Remarques'],
            'description': 'IndisponibilitÃ©s des Ã©quipes (GÃ‰NÃ‰RÃ‰E AUTOMATIQUEMENT)',
            'type': 'auto',
            'exemple': {
                'Equipe': 'CENTRALE 1 (1)',
                'Semaine': 5,
                'Horaire_Debut': '14:00',
                'Horaire_Fin': '18:00',
                'Remarques': 'CompÃ©tition nationale'
            }
        },
        'Indispos_Institutions': {
            'colonnes': ['Institution', 'Semaine', 'Horaire_Debut', 'Horaire_Fin', 'Remarques'],
            'description': 'IndisponibilitÃ©s par institution - S\'applique Ã  toutes les Ã©quipes de l\'institution',
            'type': 'auto',
            'exemple': {
                'Institution': 'EDP 1',
                'Semaine': 2,
                'Horaire_Debut': '08:00',
                'Horaire_Fin': '10:00',
                'Remarques': 'RÃ©union BDE'
            },
            'notes': [
                'Institution: Nom exact de l\'institution concernÃ©e',
                'Semaine: NumÃ©ro de la semaine (1 Ã  N)',
                'Horaire_Debut: Heure de dÃ©but de l\'indisponibilitÃ© (format HH:MM)',
                'Horaire_Fin: Heure de fin de l\'indisponibilitÃ© (format HH:MM)',
                '  â†’ Si vide ou non renseignÃ©: considÃ©rer la journÃ©e complÃ¨te',
                'Remarques: Informations complÃ©mentaires (optionnel)'
            ]
        },
        'Preferences_Gymnases': {
            'colonnes': ['Institution', 'Gymnase_Pref_1', 'Gymnase_Pref_2', 'Gymnase_Pref_3', 'Gymnase_Pref_4', 'Gymnase_Pref_5'],
            'description': 'PrÃ©fÃ©rences de gymnases par institution - Remplir du plus prÃ©fÃ©rÃ© (1) au moins prÃ©fÃ©rÃ© (5)',
            'type': 'auto',
            'exemple': {
                'Institution': 'EDP 1',
                'Gymnase_Pref_1': 'PARC DES PRINCES',
                'Gymnase_Pref_2': 'GYMNASE CENTRAL',
                'Gymnase_Pref_3': '',
                'Gymnase_Pref_4': '',
                'Gymnase_Pref_5': ''
            },
            'notes': [
                'Institution: Nom exact de l\'institution (automatiquement prÃ©-rempli)',
                'Gymnase_Pref_1 Ã  5: Gymnases prÃ©fÃ©rÃ©s par ordre de prÃ©fÃ©rence',
                '  â†’ Gymnase_Pref_1 = premier choix (bonus maximal)',
                '  â†’ Gymnase_Pref_5 = cinquiÃ¨me choix (bonus minimal)',
                '  â†’ Laisser vide si moins de 5 prÃ©fÃ©rences',
                'Bonus configurables dans le fichier YAML (bonus_preferences_gymnases)'
            ]
        },
        'Obligation_Presence': {
            'colonnes': ['Gymnase', 'Institution_Obligatoire', 'Remarques'],
            'description': 'Obligation de prÃ©sence institutionnelle par gymnase (GÃ‰NÃ‰RÃ‰E AUTOMATIQUEMENT)',
            'type': 'auto',
            'exemple': {
                'Gymnase': 'PARC DES PRINCES',
                'Institution_Obligatoire': 'ENS',
                'Remarques': 'Au moins une Ã©quipe ENS dans chaque match'
            }
        },
        'Groupes_Non_Simultaneite': {
            'colonnes': ['Nom_Groupe', 'Entites', 'Remarques'],
            'description': 'Groupes d\'Ã©quipes/institutions Ã  ne pas faire jouer simultanÃ©ment',
            'type': 'manuel',
            'exemple': {
                'Nom_Groupe': 'Grandes Ã‰coles Lyon',
                'Entites': 'EDP, CENTRALE',
                'Remarques': 'Grandes Ã©coles d\'ingÃ©nieurs qui ne doivent pas jouer simultanÃ©ment'
            }
        },
        'Ententes': {
            'colonnes': ['Institution_1', 'Institution_2', 'Penalite_Non_Planif', 'Remarques'],
            'description': 'Paires d\'institutions avec pÃ©nalitÃ© rÃ©duite si match non planifiÃ©',
            'type': 'manuel',
            'exemple': {
                'Institution_1': 'LYON 1',
                'Institution_2': 'LYON 2',
                'Penalite_Non_Planif': '10000.0',
                'Remarques': 'Entente historique entre les deux universitÃ©s'
            },
            'notes': [
                'Institution_1 et Institution_2: Paire d\'institutions (ordre non important)',
                'Penalite_Non_Planif: PÃ©nalitÃ© si le match entente n\'est pas planifiÃ© (optionnel, dÃ©faut configurÃ© dans YAML)',
                'Les matchs entre ces institutions ont une prioritÃ© plus faible et peuvent plus facilement ne pas Ãªtre planifiÃ©s',
                'Exemple: LYON 1 â†” LYON 2 signifie qu\'un match entre une Ã©quipe LYON 1 et une Ã©quipe LYON 2 est une entente'
            ]
        },
        'Contraintes_Temporelles': {
            'colonnes': ['Equipe_1', 'Equipe_2', 'Genre', 'Type_Contrainte', 'Semaine', 'Horaires_Possibles', 'Remarques'],
            'description': 'Contraintes temporelles sur matchs spÃ©cifiques (ex: CFE aprÃ¨s semaine X)',
            'type': 'manuel',
            'exemple': {
                'Equipe_1': 'LYON 1 (1)',
                'Equipe_2': 'LYON 2 (1)',
                'Genre': 'M',
                'Type_Contrainte': 'Apres',
                'Semaine': '8',
                'Horaires_Possibles': 'Mercredi 18h00, Vendredi 16h00',
                'Remarques': 'Match CFE Ã  planifier aprÃ¨s la semaine 8'
            },
            'notes': [
                'Equipe_1 et Equipe_2: Noms des Ã©quipes sans genre (Institution (numÃ©ro))',
                'Genre: M (masculin) ou F (fÃ©minin) - mÃªme genre pour les deux Ã©quipes',
                'Type_Contrainte: "Avant" ou "Apres" (planifier avant/aprÃ¨s la semaine indiquÃ©e)',
                'Semaine: NumÃ©ro de semaine limite (1-52)',
                'Horaires_Possibles: Liste d\'horaires autorisÃ©s pour ce match (optionnel, sÃ©parÃ©s par virgule)',
                'S\'applique Ã  TOUS les matchs entre ces deux Ã©quipes (ordre bidirectionnel)',
                'Contrainte paramÃ©trable: souple (pÃ©nalitÃ©) ou dure (bloquante) via YAML'
            ]
        },
        'Types_Poules': {
            'colonnes': ['Poule', 'Type', 'Remarques'],
            'description': 'Type de championnat pour chaque poule (Classique ou Aller-Retour)',
            'type': 'auto',
            'exemple': {
                'Poule': 'VBFA1PA',
                'Type': 'Classique',
                'Remarques': 'Championnat simple tour'
            },
            'notes': [
                'Poule: Nom exact de la poule (automatiquement prÃ©-rempli)',
                'Type: "Classique" ou "Aller-Retour"',
                '  â†’ Classique: Chaque paire d\'Ã©quipes joue 1 seul match (nÃ—(n-1)/2 matchs)',
                '  â†’ Aller-Retour: Chaque paire joue 2 matchs (aller ET retour, nÃ—(n-1) matchs)',
                'Aller-Retour implique de jouer le match une fois chez chaque Ã©quipe (inversion ordre)',
                'L\'espacement entre matchs aller et retour est paramÃ©trable via YAML'
            ]
        },
        'Matchs_Fixes': {
            'colonnes': ['Equipe_1', 'Equipe_2', 'Genre', 'Poule', 'Semaine', 'Horaire', 'Gymnase', 'Score', 'Type_Competition', 'Remarques'],
            'description': 'Matchs dÃ©jÃ  jouÃ©s ou planifiÃ©s Ã  intÃ©grer dans le calendrier',
            'type': 'manuel',
            'exemple': {
                'Equipe_1': 'LYON 1 (1)',
                'Equipe_2': 'LYON 2 (1)',
                'Genre': 'F',
                'Poule': 'VBFA1PA',
                'Semaine': '1',
                'Horaire': '18:00',
                'Gymnase': 'PARC DES SPORTS',
                'Score': '3-1',
                'Type_Competition': 'CFE',
                'Remarques': 'Match dÃ©jÃ  jouÃ©'
            },
            'notes': [
                'Equipe_1 et Equipe_2: Noms exacts des Ã©quipes sans le genre (Institution (numÃ©ro) seulement)',
                'Genre: F ou M (obligatoire)',
                'Poule: Code de la poule (doit correspondre aux Ã©quipes)',
                'Semaine: NumÃ©ro de semaine oÃ¹ le match a Ã©tÃ©/sera jouÃ© (1 Ã  nb_semaines)',
                'Horaire: Heure du match (format HH:MM)',
                'Gymnase: Nom exact du gymnase (doit exister dans la feuille Gymnases)',
                'Score: Score du match si dÃ©jÃ  jouÃ© (optionnel, format: "X-Y")',
                'Type_Competition: Nature de la compÃ©tition',
                '  â†’ CFE: Championnat de France des Ã‰coles',
                '  â†’ CFU: Championnat de France Universitaire',
                '  â†’ Acad: Match de championnat rÃ©gulier',
                '  â†’ Autre: Autre type de match',
                'Remarques: Informations complÃ©mentaires (optionnel)',
                '',
                'âš ï¸ IMPORTANT: Ces matchs seront exclus de la planification automatique',
                'Ils apparaÃ®tront dans le calendrier final aux crÃ©neaux indiquÃ©s'
            ]
        },
                'Niveaux_Gymnases': {
            'colonnes': ['Gymnase', 'Niveau', 'Remarque'],
            'description': 'Classification des gymnases par niveau (haut/bas) avec bonus par niveau de match',
            'type': 'manuel',
            'exemple': {
                'Gymnase': 'PARC DES SPORTS',
                'Niveau': 'Haut niveau',
                'Remarque': 'Gymnase principal de la ville'
            },
            'notes': [
                'Gymnase: Nom exact du gymnase (doit exister dans la feuille Gymnases)',
                'Niveau: "Haut niveau" ou "Bas niveau"',
                'Remarque: Explication ou commentaire optionnel',
                'Bonus par niveau de match configurÃ©s dans YAML: bonus_haut_niveau et bonus_bas_niveau',
                'Exemple: bonus_haut_niveau: [10, 8, 5, 1] signifie +10 pour A1, +8 pour A2, etc. dans gymnase haut niveau'
            ]
        }
    }
    
    def __init__(self, fichier_path: str):
        """
        Initialise le gestionnaire de configuration.
        
        Args:
            fichier_path: Chemin vers le fichier Excel de configuration
        """
        self.fichier_path = Path(fichier_path)
        self.fichier_path.parent.mkdir(parents=True, exist_ok=True)
    
    def fichier_existe(self) -> bool:
        """VÃ©rifie si le fichier existe."""
        return self.fichier_path.exists()
    
    def lire_feuille(self, nom_feuille: str) -> Optional[pd.DataFrame]:
        """
        Lit une feuille du fichier Excel.
        
        Args:
            nom_feuille: Nom de la feuille Ã  lire
            
        Returns:
            DataFrame contenant les donnÃ©es, ou None si la feuille n'existe pas
        """
        if not self.fichier_existe():
            logger.warning(f"Le fichier {self.fichier_path} n'existe pas")
            return None
        
        try:
            df = pd.read_excel(self.fichier_path, sheet_name=nom_feuille)
            logger.info(f"Feuille '{nom_feuille}' lue avec succÃ¨s ({len(df)} lignes)")
            return df
        except ValueError:
            logger.warning(f"La feuille '{nom_feuille}' n'existe pas dans le fichier")
            return None
        except Exception as e:
            logger.error(f"Erreur lors de la lecture de '{nom_feuille}': {e}")
            return None
    
    def lire_toutes_feuilles(self) -> Dict[str, pd.DataFrame]:
        """
        Lit toutes les feuilles dÃ©finies dans STRUCTURES.
        
        Returns:
            Dictionnaire {nom_feuille: DataFrame}
        """
        resultat = {}
        for nom_feuille in self.STRUCTURES.keys():
            df = self.lire_feuille(nom_feuille)
            if df is not None:
                resultat[nom_feuille] = df
        return resultat
    
    def valider_structure(self, nom_feuille: str, df: pd.DataFrame) -> Tuple[bool, List[str]]:
        """
        Valide qu'une feuille a la structure correcte.
        
        Args:
            nom_feuille: Nom de la feuille
            df: DataFrame Ã  valider
            
        Returns:
            Tuple (valide, liste des erreurs)
        """
        if nom_feuille not in self.STRUCTURES:
            return False, [f"Feuille '{nom_feuille}' non dÃ©finie dans STRUCTURES"]
        
        colonnes_attendues = set(self.STRUCTURES[nom_feuille]['colonnes'])
        colonnes_presentes = set(df.columns)
        
        erreurs = []
        
        # Colonnes manquantes
        colonnes_manquantes = colonnes_attendues - colonnes_presentes
        if colonnes_manquantes:
            erreurs.append(f"Colonnes manquantes: {', '.join(colonnes_manquantes)}")
        
        # Colonnes en trop (warning, pas erreur bloquante)
        colonnes_extra = colonnes_presentes - colonnes_attendues
        if colonnes_extra:
            logger.warning(f"Colonnes non standard dans '{nom_feuille}': {', '.join(colonnes_extra)}")
        
        return len(erreurs) == 0, erreurs
    
    def _preremplir_preferences_gymnases(self, ws):
        """
        PrÃ©-remplit la feuille Preferences_Gymnases avec toutes les institutions uniques
        trouvÃ©es dans la feuille Equipes.
        
        Args:
            ws: Worksheet Preferences_Gymnases Ã  remplir
        """
        # Lire la feuille Equipes pour extraire les institutions
        try:
            df_equipes = self.lire_feuille('Equipes')
            if df_equipes is None or df_equipes.empty or 'Equipe' not in df_equipes.columns:
                logger.warning("Impossible de prÃ©-remplir Preferences_Gymnases : feuille Equipes vide ou invalide")
                return
            
            # Extraire les institutions depuis les noms d'Ã©quipes
            institutions = set()
            for equipe_nom in df_equipes['Equipe'].dropna():
                equipe_str = str(equipe_nom).strip()
                # Extraire l'institution (tout avant le dernier "(numÃ©ro)")
                if '(' in equipe_str:
                    institution = equipe_str.rsplit('(', 1)[0].strip()
                    if institution:
                        institutions.add(institution)
            
            if not institutions:
                logger.warning("Aucune institution trouvÃ©e dans la feuille Equipes")
                return
            
            # Trier les institutions par ordre alphabÃ©tique
            institutions_triees = sorted(institutions)
            
            # Remplir la colonne Institution (Ã  partir de la ligne 2)
            for idx, institution in enumerate(institutions_triees, start=2):
                ws.cell(row=idx, column=1, value=institution)
            
            logger.info(f"Preferences_Gymnases prÃ©-remplie avec {len(institutions_triees)} institutions")
            
        except Exception as e:
            logger.error(f"Erreur lors du prÃ©-remplissage de Preferences_Gymnases : {e}")
    
    def creer_feuille_vide(self, nom_feuille: str, avec_exemple: bool = False) -> pd.DataFrame:
        """
        CrÃ©e une feuille vide avec la structure correcte.
        
        IMPORTANT: Les exemples ne doivent JAMAIS Ãªtre ajoutÃ©s dans les fichiers de configuration.
        Ils sont dÃ©finis uniquement pour la documentation et la validation.
        
        Args:
            nom_feuille: Nom de la feuille Ã  crÃ©er
            avec_exemple: Si True, ajoute une ligne d'exemple (PAR DÃ‰FAUT: False)
                         âš ï¸ NE JAMAIS UTILISER avec_exemple=True dans les scripts de production
            
        Returns:
            DataFrame vide avec les bonnes colonnes (SANS ligne d'exemple)
        """
        if nom_feuille not in self.STRUCTURES:
            raise ValueError(f"Feuille '{nom_feuille}' non dÃ©finie")
        
        structure = self.STRUCTURES[nom_feuille]
        colonnes = structure['colonnes']
        
        if avec_exemple:
            # âš ï¸ MODE DEBUG UNIQUEMENT - Ne pas utiliser en production
            logger.warning(f"âš ï¸ CrÃ©ation de la feuille '{nom_feuille}' AVEC exemple (mode debug)")
            exemple = structure['exemple']
            df = pd.DataFrame([exemple], columns=colonnes)
        else:
            # Mode normal: crÃ©er DataFrame vide SANS exemple
            df = pd.DataFrame(columns=colonnes)
        
        return df
    
    def generer_feuilles_manquantes(self, conserver_existant: bool = True) -> Dict[str, str]:
        """
        GÃ©nÃ¨re UNIQUEMENT les feuilles manquantes, conserve les existantes intactes.
        
        Args:
            conserver_existant: Si True, conserve les donnÃ©es existantes
            
        Returns:
            Dictionnaire {nom_feuille: 'crÃ©Ã©e' ou 'conservÃ©e'}
        """
        statuts = {}
        
        # Si le fichier n'existe pas, crÃ©er toutes les feuilles
        if not self.fichier_existe():
            logger.info("CrÃ©ation d'un nouveau fichier")
            with pd.ExcelWriter(self.fichier_path, engine='openpyxl') as writer:
                for nom_feuille in self.STRUCTURES.keys():
                    df = self.creer_feuille_vide(nom_feuille, avec_exemple=False)
                    df.to_excel(writer, sheet_name=nom_feuille, index=False)
                    statuts[nom_feuille] = 'crÃ©Ã©e'
                    logger.info(f"Feuille '{nom_feuille}' crÃ©Ã©e")
            
            # Formater le fichier
            self._formater_fichier()
            return statuts
        
        # Fichier existe : n'ajouter QUE les feuilles manquantes
        if conserver_existant:
            import openpyxl
            wb = openpyxl.load_workbook(self.fichier_path)
            feuilles_existantes = set(wb.sheetnames)
            logger.info(f"Fichier existant trouvÃ© avec {len(feuilles_existantes)} feuilles")
            
            feuilles_a_creer = []
            for nom_feuille in self.STRUCTURES.keys():
                if nom_feuille in feuilles_existantes:
                    statuts[nom_feuille] = 'conservÃ©e'
                    logger.info(f"Feuille '{nom_feuille}' conservÃ©e")
                else:
                    feuilles_a_creer.append(nom_feuille)
            
            # CrÃ©er seulement les feuilles manquantes
            if feuilles_a_creer:
                for nom_feuille in feuilles_a_creer:
                    ws = wb.create_sheet(nom_feuille)
                    
                    # Ã‰crire les en-tÃªtes
                    colonnes = self.STRUCTURES[nom_feuille]['colonnes']
                    for idx, col in enumerate(colonnes, 1):
                        ws.cell(1, idx, col)
                    
                    # PrÃ©-remplir les institutions pour Preferences_Gymnases
                    if nom_feuille == 'Preferences_Gymnases':
                        self._preremplir_preferences_gymnases(ws)
                    
                    statuts[nom_feuille] = 'crÃ©Ã©e'
                    logger.info(f"Feuille '{nom_feuille}' crÃ©Ã©e sans exemple")
                
                wb.save(self.fichier_path)
        else:
            # Mode Ã©crasement : recrÃ©er tout
            with pd.ExcelWriter(self.fichier_path, engine='openpyxl') as writer:
                for nom_feuille in self.STRUCTURES.keys():
                    df = self.creer_feuille_vide(nom_feuille, avec_exemple=False)
                    df.to_excel(writer, sheet_name=nom_feuille, index=False)
                    statuts[nom_feuille] = 'crÃ©Ã©e'
                    logger.info(f"Feuille '{nom_feuille}' crÃ©Ã©e")
        
        # Formater le fichier (couleurs, largeurs de colonnes)
        self._formater_fichier()
        
        return statuts
    
    def _ajouter_validation_liste(self, ws, nom_colonne: str, valeurs: List[str], 
                                   ligne_debut: int, ligne_fin: int):
        """
        Ajoute une liste dÃ©roulante Ã  une colonne.
        
        Pour les grandes listes (>100 Ã©lÃ©ments ou >200 caractÃ¨res), utilise une plage
        nommÃ©e dans une feuille cachÃ©e au lieu d'une formule inline.
        """
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.utils import get_column_letter
        
        # Trouver l'index de la colonne
        col_idx = None
        for idx, cell in enumerate(ws[1], 1):
            if hasattr(cell, 'value') and cell.value == nom_colonne:
                col_idx = idx
                break
        
        if col_idx is None or not valeurs:
            return
        
        # Calculer la taille de la formule inline
        formule_inline = ",".join(valeurs)
        use_named_range = len(valeurs) > 50 or len(formule_inline) > 200
        
        if use_named_range:
            # Utiliser une plage nommÃ©e (pour grandes listes)
            nom_plage = f"Liste_{nom_colonne}_{ws.title}"
            nom_plage = nom_plage.replace(" ", "_").replace("-", "_")[:31]  # Limite Excel: 31 caractÃ¨res
            
            # CrÃ©er ou rÃ©cupÃ©rer feuille cachÃ©e pour les listes
            wb = ws.parent
            if '_Listes_Validation' not in wb.sheetnames:
                ws_listes = wb.create_sheet('_Listes_Validation')
                ws_listes.sheet_state = 'hidden'
            else:
                ws_listes = wb['_Listes_Validation']
            
            # Trouver la premiÃ¨re colonne vide dans la feuille cachÃ©e
            col_liste = ws_listes.max_column + 1 if ws_listes.max_column > 0 else 1
            
            # Ã‰crire les valeurs dans la feuille cachÃ©e
            for idx, valeur in enumerate(valeurs, 1):
                ws_listes.cell(row=idx, column=col_liste, value=valeur)
            
            # CrÃ©er la formule de rÃ©fÃ©rence Ã  la plage
            col_letter_liste = get_column_letter(col_liste)
            formule = f"_Listes_Validation!${col_letter_liste}$1:${col_letter_liste}${len(valeurs)}"
            
            logger.debug(f"Validation avec plage nommÃ©e pour {nom_colonne} : {len(valeurs)} valeurs -> {formule}")
        else:
            # Utiliser formule inline (pour petites listes)
            formule = f'"{formule_inline}"'
            logger.debug(f"Validation inline pour {nom_colonne} : {len(valeurs)} valeurs")
        
        # CrÃ©er la validation
        dv = DataValidation(type="list", formula1=formule, allow_blank=True)
        dv.error = 'Valeur invalide'
        dv.errorTitle = 'Erreur de saisie'
        dv.prompt = f'SÃ©lectionnez une valeur dans la liste ({len(valeurs)} choix)'
        dv.promptTitle = 'Liste dÃ©roulante'
        
        # Appliquer la validation Ã  la plage
        col_letter = get_column_letter(col_idx)
        plage = f"{col_letter}{ligne_debut}:{col_letter}{ligne_fin}"
        dv.add(plage)
        ws.add_data_validation(dv)
        
        logger.info(f"âœ“ Validation ajoutÃ©e pour {nom_colonne} : {len(valeurs)} valeurs")
    
    def _extraire_liste_valeurs(self, wb, nom_feuille: str, nom_colonne: str) -> List[str]:
        """
        Extrait les valeurs uniques d'une colonne.
        
        Pour la colonne 'Equipe', gÃ©nÃ¨re automatiquement les variantes avec genre si nÃ©cessaire:
        - Si une Ã©quipe existe en plusieurs genres (M et F): gÃ©nÃ¨re "NOM [M]" et "NOM [F]"
        - Si une Ã©quipe existe en un seul genre ou sans genre: pas de suffixe
        """
        from core.utils import extraire_genre_depuis_poule, formater_nom_avec_genre
        
        if nom_feuille not in wb.sheetnames:
            return []
        
        ws = wb[nom_feuille]
        valeurs = []
        
        # Trouver l'index de la colonne
        col_idx = None
        for idx, cell in enumerate(ws[1], 1):
            if hasattr(cell, 'value') and cell.value == nom_colonne:
                col_idx = idx
                break
        
        if col_idx is None:
            return []
        
        # Cas spÃ©cial pour les Ã©quipes: dÃ©tecter genres multiples
        if nom_colonne == 'Equipe' and nom_feuille == 'Equipes':
            # DÃ©tecter si la feuille a une colonne Poule pour extraire le genre
            poule_col_idx = None
            for idx, cell in enumerate(ws[1], 1):
                if hasattr(cell, 'value') and cell.value == 'Poule':
                    poule_col_idx = idx
                    break
            
            # Mapping: {nom_equipe: {genres}}
            equipes_par_nom = {}
            
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row, col_idx)
                if not (cell.value and str(cell.value).strip()):
                    continue
                
                nom_equipe = str(cell.value).strip()
                
                # Extraire le genre depuis la poule si disponible
                genre = ''
                if poule_col_idx:
                    poule_cell = ws.cell(row, poule_col_idx)
                    if poule_cell.value:
                        genre = extraire_genre_depuis_poule(str(poule_cell.value))
                
                # Stocker dans le mapping
                if nom_equipe not in equipes_par_nom:
                    equipes_par_nom[nom_equipe] = set()
                if genre:
                    equipes_par_nom[nom_equipe].add(genre)
            
            # GÃ©nÃ©rer les variantes
            for nom_equipe, genres in equipes_par_nom.items():
                if len(genres) > 1:
                    # Plusieurs genres: crÃ©er variantes [M] et [F]
                    for genre in sorted(genres):
                        valeurs.append(formater_nom_avec_genre(nom_equipe, genre))
                else:
                    # Un seul genre ou pas de genre: pas de suffixe
                    valeurs.append(nom_equipe)
        else:
            # Extraction standard pour les autres colonnes
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row, col_idx)
                if cell.value and str(cell.value).strip():
                    valeur = str(cell.value).strip()
                    if valeur not in valeurs:
                        valeurs.append(valeur)
        
        return sorted(valeurs)
    
    def _extraire_institutions(self, wb) -> List[str]:
        """Extrait la liste des institutions depuis les Ã©quipes."""
        institutions = set()
        
        if 'Equipes' not in wb.sheetnames:
            return []
        
        ws = wb['Equipes']
        
        # Trouver la colonne Equipe
        col_idx = None
        for idx, cell in enumerate(ws[1], 1):
            if hasattr(cell, 'value') and cell.value == 'Equipe':
                col_idx = idx
                break
        
        if col_idx is None:
            return []
        
        # Extraire les institutions (partie avant le numÃ©ro)
        import re
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row, col_idx)
            if cell.value:
                equipe = str(cell.value).strip()
                # Format: "Institution (numÃ©ro)"
                match = re.match(r'^(.+?)\s*\(\d+\)\s*$', equipe)
                if match:
                    institutions.add(match.group(1).strip())
        
        return sorted(list(institutions))
    
    def _formater_fichier(self):
        """Applique le formatage au fichier Excel (en-tÃªtes en gras, couleurs, listes dÃ©roulantes, centrage, bordures)."""
        try:
            from openpyxl.styles import Border, Side
            
            wb = openpyxl.load_workbook(self.fichier_path)
            
            # Extraire les listes pour les validations
            equipes_list = self._extraire_liste_valeurs(wb, 'Equipes', 'Equipe')
            gymnases_list = self._extraire_liste_valeurs(wb, 'Gymnases', 'Gymnase')
            institutions_list = self._extraire_institutions(wb)
            
            # Styles pour le formatage
            couleur_entete = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
            couleur_ligne_paire = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            couleur_ligne_impaire = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            
            police_entete = Font(bold=True, color="FFFFFF", size=11)
            police_normale = Font(size=10)
            
            alignement_centre = Alignment(horizontal="center", vertical="center", wrap_text=True)
            alignement_gauche = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            # Bordures fines
            bordure_fine = Border(
                left=Side(style='thin', color='D3D3D3'),
                right=Side(style='thin', color='D3D3D3'),
                top=Side(style='thin', color='D3D3D3'),
                bottom=Side(style='thin', color='D3D3D3')
            )
            
            for nom_feuille in self.STRUCTURES.keys():
                if nom_feuille not in wb.sheetnames:
                    continue
                
                ws = wb[nom_feuille]
                
                # DÃ©terminer le nombre de lignes avec donnÃ©es
                max_row = ws.max_row
                max_col = ws.max_column
                
                # Formater la ligne d'en-tÃªte (ligne 1)
                for cell in ws[1]:
                    cell.fill = couleur_entete
                    cell.font = police_entete
                    cell.alignment = alignement_centre
                    cell.border = bordure_fine
                
                # Formater toutes les lignes de donnÃ©es (alternance de couleurs + centrage)
                for row_idx in range(2, max_row + 1):
                    # Alterner les couleurs des lignes
                    couleur_ligne = couleur_ligne_paire if row_idx % 2 == 0 else couleur_ligne_impaire
                    
                    for col_idx in range(1, max_col + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        
                        # Appliquer la couleur de fond
                        cell.fill = couleur_ligne
                        
                        # Appliquer la police
                        cell.font = police_normale
                        
                        # Centrage pour toutes les cellules SAUF les colonnes de remarques/notes
                        col_letter = cell.column_letter
                        header_cell = ws.cell(row=1, column=col_idx)
                        col_name = header_cell.value if header_cell.value else ""
                        
                        # Alignement Ã  gauche pour les colonnes de texte long
                        if any(keyword in str(col_name).lower() for keyword in ['remarque', 'note', 'commentaire', 'description', 'entite']):
                            cell.alignment = alignement_gauche
                        else:
                            cell.alignment = alignement_centre
                        
                        # Appliquer les bordures
                        cell.border = bordure_fine
                
                # Ajuster la largeur des colonnes de maniÃ¨re plus intelligente
                for column in ws.columns:
                    max_length = 0
                    column_letter = None
                    
                    for cell in column:
                        if column_letter is None and hasattr(cell, 'column_letter'):
                            column_letter = cell.column_letter
                        try:
                            if cell.value:
                                cell_length = len(str(cell.value))
                                if cell_length > max_length:
                                    max_length = cell_length
                        except:
                            pass
                    
                    if column_letter:
                        # Largeur minimale de 12, maximale de 50
                        adjusted_width = max(12, min(max_length + 3, 50))
                        ws.column_dimensions[column_letter].width = adjusted_width
                
                # Figer la premiÃ¨re ligne
                ws.freeze_panes = "A2"
                
                # Nettoyer les validations existantes pour Ã©viter les doublons
                ws.data_validations.dataValidation.clear()
                
                # Ajouter les listes dÃ©roulantes pour les feuilles auto-gÃ©nÃ©rÃ©es
                if nom_feuille in ['Indispos_Gymnases', 'Obligation_Presence'] and gymnases_list:
                    self._ajouter_validation_liste(ws, 'Gymnase', gymnases_list, 2, 1000)
                
                if nom_feuille == 'Indispos_Equipes' and equipes_list:
                    self._ajouter_validation_liste(ws, 'Equipe', equipes_list, 2, 1000)
                
                if nom_feuille in ['Indispos_Institutions', 'Preferences_Gymnases', 'Obligation_Presence']:
                    if nom_feuille != 'Obligation_Presence' and institutions_list:
                        self._ajouter_validation_liste(ws, 'Institution', institutions_list, 2, 1000)
                    if nom_feuille == 'Obligation_Presence' and institutions_list:
                        self._ajouter_validation_liste(ws, 'Institution_Obligatoire', institutions_list, 2, 1000)
                
                if nom_feuille == 'Preferences_Gymnases' and gymnases_list:
                    # Ajouter la validation pour toutes les colonnes Gymnase_Pref_*
                    colonnes_pref = [col for col in self.STRUCTURES['Preferences_Gymnases']['colonnes'] 
                                    if col.startswith('Gymnase_Pref_')]
                    for col_name in colonnes_pref:
                        self._ajouter_validation_liste(ws, col_name, gymnases_list, 2, 1000)
                
                # Validation pour Applique_A
                if nom_feuille == 'Indispos_Institutions':
                    self._ajouter_validation_liste(ws, 'Applique_A', 
                                                   ['Toutes', 'Masculin', 'Feminin'], 2, 1000)
                
                # Validation pour Groupes_Non_Simultaneite
                if nom_feuille == 'Groupes_Non_Simultaneite':
                    from openpyxl.comments import Comment
                    # Pour la colonne Entites, on met une note explicative dans la cellule B1
                    # (on ne peut pas faire de vraie validation multi-sÃ©lection dans Excel basique)
                    note_cell = ws['B1']
                    note_cell.comment = Comment(
                        "ASTUCE: Vous pouvez saisir plusieurs entitÃ©s sÃ©parÃ©es par des virgules ou points-virgules.\n\n"
                        "Exemples:\n"
                        "â€¢ ECL, EML, ENS\n"
                        "â€¢ LYON 1; LYON 2; LYON 3\n"
                        "â€¢ LYON 1 (1), LYON 1 (2)\n"
                        "â€¢ INSA\n\n"
                        f"Institutions disponibles: {', '.join(institutions_list[:10])}{'...' if len(institutions_list) > 10 else ''}\n\n"
                        f"Ã‰quipes disponibles: {', '.join(equipes_list[:10])}{'...' if len(equipes_list) > 10 else ''}",
                        "PyCalendar"
                    )
                    # Augmenter la largeur de la colonne Entites
                    ws.column_dimensions['B'].width = 60
                
                # Validation pour Ententes (listes dÃ©roulantes institutions)
                if nom_feuille == 'Ententes' and institutions_list:
                    self._ajouter_validation_liste(ws, 'Institution_1', institutions_list, 2, 1000)
                    self._ajouter_validation_liste(ws, 'Institution_2', institutions_list, 2, 1000)
                
                # Validation pour Contraintes_Temporelles
                if nom_feuille == 'Contraintes_Temporelles':
                    if equipes_list:
                        self._ajouter_validation_liste(ws, 'Equipe_1', equipes_list, 2, 1000)
                        self._ajouter_validation_liste(ws, 'Equipe_2', equipes_list, 2, 1000)
                    self._ajouter_validation_liste(ws, 'Type_Contrainte', ['Avant', 'Apres'], 2, 1000)
            
            wb.save(self.fichier_path)
            logger.info("Formatage du fichier appliquÃ© avec succÃ¨s (avec listes dÃ©roulantes)")
            
        except Exception as e:
            logger.error(f"Erreur lors du formatage: {e}")
    
    def reorganiser_feuilles(self):
        """
        RÃ©organise les feuilles dans l'ordre logique dÃ©fini par STRUCTURES.
        
        Ordre souhaitÃ©:
        1. Equipes (manuel)
        2. Gymnases (manuel)
        3. Indispos_Gymnases (auto)
        4. Indispos_Equipes (auto)
        5. Indispos_Institutions (auto)
        6. Preferences_Gymnases (auto)
        7. Obligation_Presence (auto)
        8. Groupes_Non_Simultaneite (manuel)
        """
        try:
            if not self.fichier_existe():
                logger.warning("Fichier inexistant, impossible de rÃ©organiser")
                return
            
            wb = openpyxl.load_workbook(self.fichier_path)
            
            # Ordre dÃ©sirÃ© (selon STRUCTURES)
            ordre_desire = list(self.STRUCTURES.keys())
            
            # Feuilles existantes
            feuilles_existantes = wb.sheetnames.copy()
            
            # RÃ©organiser uniquement les feuilles qui existent
            ordre_final = [f for f in ordre_desire if f in feuilles_existantes]
            
            # Ajouter les feuilles non dÃ©finies dans STRUCTURES Ã  la fin
            feuilles_extra = [f for f in feuilles_existantes if f not in ordre_desire]
            ordre_final.extend(feuilles_extra)
            
            # RÃ©organiser les feuilles en les dÃ©plaÃ§ant une par une
            for idx_cible, nom_feuille in enumerate(ordre_final):
                idx_actuel = wb.sheetnames.index(nom_feuille)
                if idx_actuel != idx_cible:
                    # DÃ©placer la feuille Ã  la position cible
                    # offset = position_cible - position_actuelle
                    offset = idx_cible - idx_actuel
                    wb.move_sheet(nom_feuille, offset=offset)
            
            wb.save(self.fichier_path)
            logger.info(f"Feuilles rÃ©organisÃ©es dans l'ordre: {', '.join(ordre_final)}")
            
        except Exception as e:
            logger.error(f"Erreur lors de la rÃ©organisation: {e}")
    
    def valider_fichier_complet(self) -> Tuple[bool, Dict[str, List[str]]]:
        """
        Valide que toutes les feuilles du fichier ont la structure correcte.
        
        Returns:
            Tuple (tout_valide, dictionnaire des erreurs par feuille)
        """
        if not self.fichier_existe():
            return False, {'global': ["Le fichier n'existe pas"]}
        
        feuilles = self.lire_toutes_feuilles()
        erreurs_par_feuille = {}
        tout_valide = True
        
        for nom_feuille in self.STRUCTURES.keys():
            if nom_feuille not in feuilles:
                erreurs_par_feuille[nom_feuille] = ["Feuille manquante"]
                tout_valide = False
            else:
                valide, erreurs = self.valider_structure(nom_feuille, feuilles[nom_feuille])
                if not valide:
                    erreurs_par_feuille[nom_feuille] = erreurs
                    tout_valide = False
        
        return tout_valide, erreurs_par_feuille
    
    def afficher_rapport(self):
        """Affiche un rapport sur l'Ã©tat du fichier de configuration."""
        print("\n" + "="*80)
        print(f"ğŸ“‹ RAPPORT DE CONFIGURATION : {self.fichier_path.name}")
        print("="*80 + "\n")
        
        if not self.fichier_existe():
            print("âŒ Le fichier n'existe pas encore.\n")
            print("ğŸ’¡ Utilisez generer_feuilles_manquantes() pour le crÃ©er.\n")
            return
        
        # Lire toutes les feuilles
        feuilles = self.lire_toutes_feuilles()
        
        # Valider
        tout_valide, erreurs = self.valider_fichier_complet()
        
        print(f"âœ… Fichier trouvÃ© : {self.fichier_path}")
        print(f"ğŸ“Š Nombre de feuilles : {len(feuilles)}/{len(self.STRUCTURES)}\n")
        
        # DÃ©tails par feuille
        for nom_feuille, structure in self.STRUCTURES.items():
            print(f"ğŸ“„ {nom_feuille}")
            print(f"   Description : {structure['description']}")
            
            if nom_feuille in feuilles:
                df = feuilles[nom_feuille]
                nb_lignes = len(df)
                nb_remplies = df.dropna(how='all').shape[0] - 1  # -1 pour exclure l'en-tÃªte
                
                print(f"   âœ… PrÃ©sente : {nb_lignes} lignes ({nb_remplies} remplies)")
                
                # Valider la structure
                valide, errs = self.valider_structure(nom_feuille, df)
                if not valide:
                    print(f"   âš ï¸  Erreurs de structure :")
                    for err in errs:
                        print(f"      - {err}")
            else:
                print(f"   âŒ Manquante")
            
            print()
        
        # RÃ©sumÃ© final
        print("="*80)
        if tout_valide:
            print("âœ… Toutes les feuilles sont valides et prÃªtes Ã  l'emploi !")
        else:
            print("âš ï¸  Certaines feuilles nÃ©cessitent des corrections.")
            print("\nğŸ’¡ Utilisez generer_feuilles_manquantes() pour corriger automatiquement.")
        print("="*80 + "\n")


def migrer_depuis_ancien_format(ancien_fichier: str, nouveau_fichier: str):
    """
    Migre les donnÃ©es d'un ancien format vers le nouveau format de configuration central.
    
    Args:
        ancien_fichier: Chemin vers l'ancien fichier (ex: equipes_handball.xlsx)
        nouveau_fichier: Chemin vers le nouveau fichier de configuration
    """
    logger.info(f"Migration de {ancien_fichier} vers {nouveau_fichier}")
    
    # CrÃ©er le gestionnaire pour le nouveau fichier
    config = ConfigManager(nouveau_fichier)
    
    # GÃ©nÃ©rer la structure de base
    config.generer_feuilles_manquantes(conserver_existant=False)
    
    # Lire l'ancien fichier
    try:
        df_ancien = pd.read_excel(ancien_fichier)
        logger.info(f"Ancien fichier lu : {len(df_ancien)} Ã©quipes")
        
        # Mapper les colonnes de l'ancien format vers le nouveau
        colonnes_mapping = {
            'Equipe': 'Equipe',
            'Poule': 'Poule',
            'Horaire_1': 'Horaire_1',
            'Horaire_2': 'Horaire_2',
            'Lieu_1': 'Lieu_1',
            'Lieu_2': 'Lieu_2',
            'Indispo_1': 'Indispo_1',
            'Indispo_2': 'Indispo_2'
        }
        
        # CrÃ©er le DataFrame pour la feuille Equipes
        df_nouveau = pd.DataFrame()
        for col_ancien, col_nouveau in colonnes_mapping.items():
            if col_ancien in df_ancien.columns:
                df_nouveau[col_nouveau] = df_ancien[col_ancien]
        
        # Ã‰crire dans le nouveau fichier
        with pd.ExcelWriter(nouveau_fichier, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            df_nouveau.to_excel(writer, sheet_name='Equipes', index=False)
        
        logger.info(f"Migration rÃ©ussie : {len(df_nouveau)} Ã©quipes migrÃ©es")
        
        # Reformater
        config._formater_fichier()
        
        print(f"âœ… Migration terminÃ©e !")
        print(f"   {len(df_nouveau)} Ã©quipes migrÃ©es vers {nouveau_fichier}")
        
    except Exception as e:
        logger.error(f"Erreur lors de la migration : {e}")
        raise


if __name__ == "__main__":
    # Configuration du logging
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )
    
    # Test avec le fichier exemple
    config = ConfigManager("exemple/config_exemple.xlsx")
    
    print("ğŸ”§ Test du gestionnaire de configuration\n")
    
    # GÃ©nÃ©rer les feuilles
    statuts = config.generer_feuilles_manquantes()
    
    print("\nğŸ“Š Statuts des feuilles :")
    for feuille, statut in statuts.items():
        emoji = "âœ…" if statut == "crÃ©Ã©e" else "ğŸ“‹"
        print(f"  {emoji} {feuille}: {statut}")
    
    # Afficher le rapport
    config.afficher_rapport()
