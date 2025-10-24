#!/usr/bin/env python3
"""
Script d'importation de matchs depuis un fichier Excel partagé en ligne.

Ce script permet d'importer des matchs déjà joués ou planifiés depuis un
tableau Excel partagé (Google Sheets, OneDrive, SharePoint, etc.) vers la feuille
Matchs_Fixes d'une configuration PyCalendar.

Utilisation:
    python importer_matchs_externes.py --config CONFIG --url URL [OPTIONS]

Exemples:
    # Importer tous les matchs avec score de la journée 3
    python importer_matchs_externes.py \\
        --config data_volley/config_volley.yaml \\
        --url "https://docs.google.com/spreadsheets/d/.../export?format=xlsx" \\
        --sport VB \\
        --journee 3 \\
        --avec-score

    # Importer depuis SharePoint avec authentification
    python importer_matchs_externes.py \\
        --config data_volley/config_volley.yaml \\
        --url "https://tenant.sharepoint.com/:x:/g/personal/user_domain_com/...xe=..." \\
        --sport VB

    # Importer depuis un fichier local
    python importer_matchs_externes.py \\
        --config data_volley/config_volley.yaml \\
        --fichier-local "/path/to/file.xlsx" \\
        --sport VB \\
        --tous
"""

import argparse
import sys
import os
import pandas as pd
import requests
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Optional, Tuple
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import xlrd
import tempfile
import urllib.parse
import msal
from requests_oauthlib import OAuth2Session
import json
import re
import base64


class ImporteurMatchsExternes:
    """Classe pour importer des matchs depuis une source externe."""
    
    def __init__(
        self,
        config_path: str,
        url_externe: Optional[str] = None,
        fichier_local: Optional[str] = None,
        sport: str = "VB",
        journee: Optional[int] = None,
        date_limite: Optional[str] = None,
        avec_score: bool = False,
        sans_score: bool = False,
        tous: bool = False,
        dry_run: bool = False,
        ignorer_annules: bool = True
    ):
        """
        Initialise l'importeur.
        
        Args:
            config_path: Chemin vers le fichier de configuration Excel
            url_externe: URL du fichier Excel partagé en ligne (optionnel)
            fichier_local: Chemin vers un fichier Excel local (optionnel)
            sport: Code du sport (VB, HB, etc.)
            journee: Numéro de journée à importer (optionnel)
            date_limite: Date limite pour l'import (format DD/MM/YYYY)
            avec_score: Importer uniquement les matchs avec score
            sans_score: Importer uniquement les matchs sans score
            tous: Importer tous les matchs (défaut si rien spécifié)
            dry_run: Mode simulation (ne modifie pas le fichier)
            ignorer_annules: Ignorer les matchs avec 'annule' dans les remarques
        """
        self.config_yaml_path = Path(config_path)
        self.url_externe = url_externe
        self.fichier_local = Path(fichier_local) if fichier_local else None
        self.sport = sport.upper()
        self.journee = journee
        self.date_limite = self._parse_date(date_limite) if date_limite else None
        self.avec_score = avec_score
        self.sans_score = sans_score
        self.tous = tous or (not avec_score and not sans_score)
        self.dry_run = dry_run
        self.ignorer_annules = ignorer_annules
        
        # Paramètres SharePoint (à configurer via variables d'environnement)
        self.sharepoint_client_id = os.getenv("SHAREPOINT_CLIENT_ID", "")
        self.sharepoint_tenant_id = os.getenv("SHAREPOINT_TENANT_ID", "")
        self.sharepoint_client_secret = os.getenv("SHAREPOINT_CLIENT_SECRET", "")
        
        # Charger le fichier Excel depuis le YAML
        self.config_excel_path = self._charger_chemin_excel()
        
        # Validation
        if not url_externe and not fichier_local:
            raise ValueError("Vous devez spécifier soit --url soit --fichier-local")
        
        # DataFrame du fichier externe
        self.df_externe: Optional[pd.DataFrame] = None
        
        # DataFrame de la configuration
        self.df_matchs_fixes: Optional[pd.DataFrame] = None
        
        # Mapping des gymnases et équipes mixtes (chargé à la demande)
        self._mapping_gymnases: Optional[dict] = None
        self._equipes_mixtes: Optional[set] = None
        
        # Cache pour l'authentification SharePoint
        self._sharepoint_token: Optional[str] = None
    
    def _parse_sharepoint_url(self, url: str) -> Dict[str, Optional[str]]:
        """
        Parse une URL SharePoint pour extraire les informations nécessaires.
        
        Args:
            url: URL SharePoint
            
        Returns:
            Dictionnaire avec les informations extraites
        """
        # URL format: https://tenant.sharepoint.com/sites/site/_layouts/15/Doc.aspx?sourcedoc={item_id}&file=filename.xlsx&action=default&mobileredirect=true
        # or: https://tenant.sharepoint.com/:x:/g/personal/user_domain_com/E...
        
        parsed = urllib.parse.urlparse(url)
        
        if 'sharepoint.com' not in parsed.netloc:
            raise ValueError("URL n'est pas une URL SharePoint valide")
        
        # Extraire le tenant
        tenant = parsed.netloc.split('.')[0]
        
        # Pour les URLs de type /:x:/g/personal/...
        if '/:x:/' in parsed.path:
            # Format: /:x:/g/personal/user_domain_com/encoded_path?e=sharing_param
            path_parts = parsed.path.split('/')
            if len(path_parts) >= 4:
                resource_type = path_parts[2]  # 'g' for group, 'personal' for personal
                # Reconstruct the full container path (everything after resource_type)
                container_path = '/'.join(path_parts[3:])  # 'personal/klucediniz_sport-u_com/EXeOs9uh4dxGkxsKLI0B_tABH0LSZdk2qWOukF8kVmNH_g'
                
                # L'ID de l'item est encodé dans l'URL
                query_params = urllib.parse.parse_qs(parsed.query)
                if 'e' in query_params:
                    sharing_token = query_params['e'][0]
                    
                    return {
                        'tenant': tenant,
                        'resource_type': resource_type,
                        'container': container_path,  # Now includes the full path
                        'sharing_token': sharing_token,
                        'url_type': 'sharing'
                    }
        
        # Pour les URLs classiques avec sourcedoc
        query_params = urllib.parse.parse_qs(parsed.query)
        if 'sourcedoc' in query_params:
            item_id = query_params['sourcedoc'][0]
            
            # Extraire le site depuis l'URL
            path_parts = parsed.path.split('/')
            site_name = None
            for i, part in enumerate(path_parts):
                if part == 'sites' and i + 1 < len(path_parts):
                    site_name = path_parts[i + 1]
                    break
            
            return {
                'tenant': tenant,
                'site_name': site_name,
                'item_id': item_id,
                'url_type': 'classic'
            }
        
        raise ValueError("Format d'URL SharePoint non reconnu")
    
    def _authenticate_sharepoint(self) -> str:
        """
        Authentifie auprès de Microsoft et retourne un token d'accès.
        
        Returns:
            Token d'accès pour Microsoft Graph API
        """
        if not all([self.sharepoint_client_id, self.sharepoint_tenant_id, self.sharepoint_client_secret]):
            raise ValueError(
                "Authentification SharePoint requise. Fournissez:\n"
                "  • Client ID Azure AD\n"
                "  • Tenant ID Azure AD\n"
                "  • Client Secret Azure AD"
            )
        
        return self._authenticate_client_credentials()
    
    def _authenticate_client_credentials(self) -> str:
        """
        Authentification utilisant client credentials (client secret).
        
        Returns:
            Token d'accès
        """
        authority = f"https://login.microsoftonline.com/{self.sharepoint_tenant_id}"
        
        app = msal.ConfidentialClientApplication(
            self.sharepoint_client_id,
            client_credential=self.sharepoint_client_secret,
            authority=authority
        )
        
        result = app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
        
        if "access_token" in result:
            return result["access_token"]
        else:
            error = result.get("error_description", result.get("error", "Erreur inconnue"))
            raise ValueError(f"Échec de l'authentification client credentials: {error}")
    
    def _download_sharepoint_file(self, url: str) -> bytes:
        """
        Télécharge un fichier depuis SharePoint en utilisant Microsoft Graph API.
        
        Args:
            url: URL SharePoint du fichier
            
        Returns:
            Contenu du fichier en bytes
        """
        # Parser l'URL
        sharepoint_info = self._parse_sharepoint_url(url)
        
        # Authentifier
        if not self._sharepoint_token:
            self._sharepoint_token = self._authenticate_sharepoint()
        
        headers = {
            'Authorization': f'Bearer {self._sharepoint_token}',
            'Accept': 'application/octet-stream'
        }
        
        if sharepoint_info['url_type'] == 'sharing':
            # Pour les URLs de partage, utiliser l'API de partage
            sharing_url = f"https://{sharepoint_info['tenant']}.sharepoint.com/:x:/{sharepoint_info['resource_type']}/{sharepoint_info['container']}?e={sharepoint_info['sharing_token']}"
            
            # Encoder l'URL pour l'API Graph
            encoded_url = base64.b64encode(sharing_url.encode()).decode()
            
            graph_url = f"https://graph.microsoft.com/v1.0/shares/u!{encoded_url}/driveItem/content"
            
        else:
            # Pour les URLs classiques, construire l'URL Graph
            graph_url = (
                f"https://graph.microsoft.com/v1.0/sites/{sharepoint_info['tenant']}.sharepoint.com:/sites/{sharepoint_info['site_name']}:/"
                f"drive/items/{sharepoint_info['item_id']}/content"
            )
        
        response = requests.get(graph_url, headers=headers)
        response.raise_for_status()
        
        return response.content
    
    def _charger_chemin_excel(self) -> Path:
        """Charge le chemin du fichier Excel depuis le YAML."""
        import yaml
        
        # Essayer différents encodages
        encodings_to_try = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
        config = None
        
        for encoding in encodings_to_try:
            try:
                with open(self.config_yaml_path, 'r', encoding=encoding) as f:
                    config = yaml.safe_load(f)
                break  # Si ça marche, on sort de la boucle
            except UnicodeDecodeError:
                continue  # Essayer le prochain encodage
        
        if config is None:
            raise ValueError(f"Impossible de lire le fichier YAML {self.config_yaml_path} avec les encodages testés: {encodings_to_try}")
        
        # Essayer plusieurs chemins possibles
        fichier_excel = None
        if 'fichiers' in config and 'donnees' in config['fichiers']:
            fichier_excel = config['fichiers']['donnees']
        elif 'fichier_excel' in config:
            fichier_excel = config['fichier_excel']
        
        if not fichier_excel:
            raise ValueError(f"Aucun fichier Excel trouvé dans {self.config_yaml_path} (cherché: fichiers.donnees ou fichier_excel)")
        
        # Résoudre le chemin relatif par rapport à la racine du projet (où se trouve le YAML, généralement on remonte d'un niveau)
        chemin = Path(fichier_excel)
        if not chemin.is_absolute():
            # Le YAML est dans configs/, les données sont à la racine
            # Donc on prend le parent du YAML (configs/) puis on remonte encore (racine)
            racine_projet = self.config_yaml_path.parent.parent
            chemin = racine_projet / fichier_excel
        
        return chemin
        
    def _parser_score(self, score_raw) -> str:
        """
        Parse et normalise un score en format 'score_E1 - score_E2'.
        
        Args:
            score_raw: Score brut depuis le fichier externe
            
        Returns:
            Score normalisé ou score original si parsing impossible
        """
        if pd.isna(score_raw):
            return ''
        
        score_str = str(score_raw).strip()
        if not score_str:
            return ''
        
        # Essayer différents formats courants
        import re
        
        # Format: "3-1" ou "3 - 1"
        match = re.match(r'^(\d+)\s*-\s*(\d+)$', score_str)
        if match:
            return f"{match.group(1)}-{match.group(2)}"
        
        # Format: "3/1" ou "3 / 1"
        match = re.match(r'^(\d+)\s*/\s*(\d+)$', score_str)
        if match:
            return f"{match.group(1)}-{match.group(2)}"
        
        # Format: "3-1 (25-20, 25-18, 25-15)" - garder tel quel
        if re.match(r'^\d+-\d+\s*\(.+\)$', score_str):
            return score_str
        
        # Format: "3-1, 3-0" (match en 5 sets) - garder tel quel
        if re.match(r'^\d+-\d+,\s*\d+-\d+$', score_str):
            return score_str
        
        # Si on ne peut pas parser, garder tel quel
        return score_str
    
    def _parse_date(self, date_str: str) -> datetime:
        """Parse une date au format DD/MM/YYYY."""
        try:
            return datetime.strptime(date_str, "%d/%m/%Y")
        except ValueError:
            raise ValueError(f"Format de date invalide: {date_str}. Attendu: DD/MM/YYYY")
    
    def mapper_gymnases(self) -> dict:
        """
        Retourne le mapping des gymnases externes vers gymnases de la config.
        
        Returns:
            Dictionnaire {nom_externe: nom_config}
        """
        if self._mapping_gymnases is None:
            self._mapping_gymnases = {
                # Mappings LAURASU → Config
                'HALLE - C.BESSON': 'BESSON',
                'HALLE - C. BESSON': 'BESSON',
                'GYMNASE DESCARTES': 'DESCARTES',
                'GYMNASE ESA': 'ESA',
                'GYMNASE CENTRALE LYON': 'ECL',
                'COMPET C (HAUT) - LEON JOUHAUX': 'L. J. HAUT',
                'HALLE LYON 2': 'LYON 2 HC',
                'HALLE - 3D': 'LAENNEC',
                'LAENNEC': 'LAENNEC',
            }
        return self._mapping_gymnases
    
    def charger_equipes_mixtes(self) -> set:
        """
        Charge la liste des équipes présentes en F et M depuis la config.
        
        Returns:
            Set des noms d'équipes mixtes
        """
        if self._equipes_mixtes is None:
            from collections import defaultdict
            
            # Lire les équipes de la config
            df_equipes = pd.read_excel(self.config_excel_path, sheet_name='Equipes')
            
            # Grouper par équipe et compter les genres
            equipes_genres = defaultdict(set)
            for _, row in df_equipes.iterrows():
                equipe = row['Equipe']
                poule = row['Poule']
                # Le 3ème caractère de la poule indique le genre (F ou M)
                genre = poule[2] if len(poule) > 2 else None
                if genre in ['F', 'M']:
                    equipes_genres[equipe].add(genre)
            
            # Garder uniquement celles avec F et M
            self._equipes_mixtes = {
                equipe for equipe, genres in equipes_genres.items()
                if len(genres) == 2
            }
        
        return self._equipes_mixtes
    
    def ajouter_genre_equipe(self, nom_equipe: str, poule: str) -> str:
        """
        Ajoute le genre entre crochets si l'équipe existe en F et M.
        
        Args:
            nom_equipe: Nom de l'équipe
            poule: Code de la poule (pour extraire le genre)
            
        Returns:
            Nom de l'équipe avec [F] ou [M] si mixte
        """
        equipes_mixtes = self.charger_equipes_mixtes()
        
        if nom_equipe in equipes_mixtes:
            # Extraire le genre de la poule
            genre = poule[2] if len(poule) > 2 else None
            if genre in ['F', 'M']:
                return f"{nom_equipe} [{genre}]"
        
        return nom_equipe
    
    def normaliser_gymnase(self, gymnase_externe: str) -> str:
        """
        Normalise le nom d'un gymnase selon le mapping.
        
        Args:
            gymnase_externe: Nom du gymnase dans le fichier externe
            
        Returns:
            Nom normalisé du gymnase
        """
        if pd.isna(gymnase_externe):
            return ''
        
        mapping = self.mapper_gymnases()
        gymnase_clean = str(gymnase_externe).strip()
        
        # Chercher dans le mapping
        return mapping.get(gymnase_clean, gymnase_clean)
    
    def telecharger_fichier_externe(self) -> pd.DataFrame:
        """
        Télécharge et charge le fichier Excel depuis l'URL ou charge depuis un fichier local.
        
        Returns:
            DataFrame contenant les données du fichier externe
        """
        # Cas 1: Fichier local
        if self.fichier_local:
            print(f"📂 Chargement du fichier local: {self.fichier_local}")
            
            if not self.fichier_local.exists():
                raise FileNotFoundError(f"Fichier introuvable: {self.fichier_local}")
            
            try:
                # Charger sans en-tête d'abord pour trouver la ligne de début
                df_raw = pd.read_excel(self.fichier_local, sheet_name=0, header=None)
                
                # Trouver la ligne avec "Date" et "Sport" (en-têtes)
                header_row = None
                for i, row in df_raw.iterrows():
                    if pd.notna(row[0]) and str(row[0]).strip() == 'Date':
                        if pd.notna(row[1]) and str(row[1]).strip() == 'Sport':
                            header_row = i
                            break
                
                if header_row is None:
                    raise ValueError("Impossible de trouver les en-têtes dans le fichier")
                
                print(f"   → En-têtes trouvés à la ligne {header_row}")
                
                # Recharger avec le bon header
                df = pd.read_excel(self.fichier_local, sheet_name=0, header=header_row)
                
                # Nettoyer les noms de colonnes (enlever espaces)
                df.columns = df.columns.str.strip()
                
                print(f"✓ Fichier chargé: {len(df)} lignes")
                print(f"   Colonnes: {list(df.columns)}")
                return df
            except Exception as e:
                raise RuntimeError(f"Erreur lors du chargement: {e}")
        
        # Cas 2: URL externe
        if self.url_externe:
            print(f"📥 Téléchargement du fichier depuis: {self.url_externe[:50]}...")
            
            try:
                temp_file = None
                # Télécharger le fichier
                response = requests.get(self.url_externe, timeout=30)
                response.raise_for_status()
                
                # Vérifier si la réponse contient du HTML (page d'authentification SharePoint)
                if response.headers.get('content-type', '').startswith('text/html'):
                    print(f"   → Détection d'une page d'authentification SharePoint, tentative d'authentification...")
                    
                    try:
                        # Essayer l'authentification SharePoint
                        file_content = self._download_sharepoint_file(self.url_externe)
                        temp_file = Path("temp_externe.xlsx")
                        temp_file.write_bytes(file_content)
                        
                    except Exception as auth_error:
                        raise ValueError(
                            f"L'URL SharePoint nécessite une authentification. "
                            f"Assurez-vous que les paramètres suivants sont correctement configurés dans le code :\n"
                            f"  • Client ID Azure AD\n"
                            f"  • Tenant ID Azure AD\n"
                            f"  • Client Secret Azure AD\n"
                            f"Erreur d'authentification : {auth_error}"
                        )
                
                else:
                    # Sauvegarder temporairement
                    temp_file = Path("temp_externe.xlsx")
                    temp_file.write_bytes(response.content)
                
                # Charger avec pandas (essayer différents engines Excel)
                df_raw = None
                engine_used = None
                
                # Essayer d'abord openpyxl (format moderne .xlsx)
                try:
                    df_raw = pd.read_excel(temp_file, sheet_name=0, header=None, engine='openpyxl')
                    engine_used = 'openpyxl'
                except Exception:
                    # Si openpyxl échoue, essayer xlrd (format ancien .xls)
                    try:
                        df_raw = pd.read_excel(temp_file, sheet_name=0, header=None, engine='xlrd')
                        engine_used = 'xlrd'
                    except Exception as e:
                        raise ValueError(f"Impossible de lire le fichier Excel avec openpyxl ou xlrd: {e}")
                
                # Trouver la ligne avec "Date" et "Sport"
                header_row = None
                for i, row in df_raw.iterrows():
                    if pd.notna(row[0]) and str(row[0]).strip() == 'Date':
                        if pd.notna(row[1]) and str(row[1]).strip() == 'Sport':
                            header_row = i
                            break
                
                if header_row is None:
                    raise ValueError("Impossible de trouver les en-têtes dans le fichier")
                
                # Recharger avec le bon header
                df = pd.read_excel(temp_file, sheet_name=0, header=header_row, engine=engine_used)
                df.columns = df.columns.str.strip()
                
                # Nettoyer
                temp_file.unlink()
                
                print(f"✓ Fichier téléchargé: {len(df)} lignes")
                return df
                
            except requests.RequestException as e:
                raise RuntimeError(f"Erreur lors du téléchargement: {e}")
            except Exception as e:
                # Nettoyer le fichier temporaire si il existe
                try:
                    if 'temp_file' in locals():
                        temp_file_obj = locals()['temp_file']
                        if temp_file_obj is not None and hasattr(temp_file_obj, 'exists') and temp_file_obj.exists():
                            temp_file_obj.unlink(missing_ok=True)
                except:
                    pass
                
                # Message d'erreur plus informatif
                error_msg = str(e)
                if "Expected BOF record" in error_msg and "<!DOCTYPE" in error_msg:
                    error_msg = (
                        "Le fichier téléchargé n'est pas un Excel valide. Cela indique probablement :\n"
                        "  • L'URL SharePoint nécessite une authentification\n"
                        "  • Le fichier n'est pas accessible publiquement\n"
                        "  • L'URL est une page de connexion plutôt qu'un lien direct vers le fichier\n"
                        "\nSolutions :\n"
                        "  1. Téléchargez manuellement le fichier Excel depuis SharePoint\n"
                        "  2. Utilisez --fichier-local avec le fichier téléchargé\n"
                        "  3. Vérifiez les permissions de partage du fichier SharePoint\n"
                        "  4. Utilisez l'option 'Télécharger une copie' dans SharePoint pour obtenir l'URL directe"
                    )
                elif "Excel file format cannot be determined" in error_msg:
                    error_msg = (
                        "Format de fichier Excel non reconnu. Essayez :\n"
                        "  • De télécharger manuellement le fichier\n"
                        "  • De vérifier que c'est bien un fichier .xlsx ou .xls"
                    )
                
                raise RuntimeError(f"Erreur lors du chargement du fichier: {error_msg}")
        
        raise ValueError("Aucune source de fichier spécifiée")
    
    def explorer_structure(self):
        """
        Explore et affiche la structure du fichier externe.
        Utile pour comprendre le format des données.
        """
        if self.df_externe is None:
            raise ValueError("Fichier externe non chargé. Appelez telecharger_fichier_externe() d'abord.")
        
        print("\n" + "="*70)
        print("📊 EXPLORATION DU FICHIER EXTERNE")
        print("="*70)
        
        print(f"\n📋 Colonnes détectées ({len(self.df_externe.columns)}):")
        for i, col in enumerate(self.df_externe.columns, 1):
            print(f"   {i}. {col}")
        
        print(f"\n📏 Dimensions:")
        print(f"   - Lignes: {len(self.df_externe)}")
        print(f"   - Colonnes: {len(self.df_externe.columns)}")
        
        print(f"\n🔍 Aperçu des 5 premières lignes:")
        print(self.df_externe.head(5).to_string(index=False))
        
        print(f"\n📊 Types de données:")
        for col in self.df_externe.columns:
            dtype = self.df_externe[col].dtype
            non_null = self.df_externe[col].notna().sum()
            print(f"   - {col}: {dtype} ({non_null} valeurs non-nulles)")
        
        print("\n" + "="*70)
    
    def mapper_colonnes(self) -> Dict[str, str]:
        """
        Mappe les colonnes du fichier externe vers le format Matchs_Fixes.
        
        Format fichier externe (LAURASU):
        - Date, Sport, Sexe, Poule, Equipe 1, Equipe 2, Hre Déb, Lieu, Commentaire, Arbitres, Résultats
        
        Format Matchs_Fixes:
        - Equipe_1, Equipe_2, Poule, Semaine, Horaire, Gymnase, Score, Type_Competition, Remarques
        
        Returns:
            Dictionnaire de mapping {colonne_externe: colonne_config}
        """
        # Mapping flexible qui gère différents formats de colonnes
        mapping = {
            # Format LAURASU (avec espaces)
            'Equipe 1': 'Equipe_1',
            'Equipe 2': 'Equipe_2',
            'Poule': 'Poule',
            'Hre Déb': 'Horaire',
            'Lieu': 'Gymnase',
            'Résultats': 'Score',
            'Commentaire': 'Remarques',
            # Format alternatif (underscores)
            'Equipe_1': 'Equipe_1',
            'Equipe_2': 'Equipe_2',
            'Score': 'Score',
            'Remarques': 'Remarques',
            'Gymnase': 'Gymnase',
            'Horaire': 'Horaire',
        }
        return mapping
    
    def filtrer_matchs(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Filtre les matchs selon les critères spécifiés.
        
        Args:
            df: DataFrame source
            
        Returns:
            DataFrame filtré
        """
        df_filtre = df.copy()
        
        print(f"\n🔍 Filtrage des matchs (départ: {len(df_filtre)} matchs)...")
        
        # Filtre par sport
        if 'Sport' in df_filtre.columns:
            df_filtre = df_filtre[df_filtre['Sport'].str.upper() == self.sport]
            print(f"   → Filtre sport '{self.sport}': {len(df_filtre)} matchs")
        
        # Filtre par date (convertir en numéro de semaine)
        if self.date_limite and 'Date' in df_filtre.columns:
            df_filtre['_Date_parsed'] = pd.to_datetime(df_filtre['Date'], errors='coerce')
            df_filtre = df_filtre[df_filtre['_Date_parsed'] <= self.date_limite]
            df_filtre = df_filtre.drop(columns=['_Date_parsed'])
            print(f"   → Filtre date ≤ {self.date_limite.strftime('%d/%m/%Y')}: {len(df_filtre)} matchs")
        
        # Filtre par score (résultats)
        if 'Résultats' in df_filtre.columns:
            if self.avec_score:
                df_filtre = df_filtre[df_filtre['Résultats'].notna() & (df_filtre['Résultats'].astype(str).str.strip() != '')]
                print(f"   → Filtre avec score: {len(df_filtre)} matchs")
            elif self.sans_score:
                df_filtre = df_filtre[df_filtre['Résultats'].isna() | (df_filtre['Résultats'].astype(str).str.strip() == '')]
                print(f"   → Filtre sans score: {len(df_filtre)} matchs")
        
        # Filtre des matchs annulés
        if self.ignorer_annules and 'Remarques' in df_filtre.columns:
            # Utiliser une regex pour détecter différents types d'annulation
            pattern_annule = r'annul|report|forfait|blessure|maladie|erreur'
            mask_annule = df_filtre['Remarques'].astype(str).str.contains(pattern_annule, case=False, na=False, regex=True)
            df_filtre = df_filtre[~mask_annule]
            print(f"   → Filtre annulations: {len(df_filtre)} matchs")
        
        return df_filtre
    
    def convertir_vers_format_config(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Convertit le DataFrame externe vers le format Matchs_Fixes.
        
        Args:
            df: DataFrame source
            
        Returns:
            DataFrame au format Matchs_Fixes
        """
        mapping = self.mapper_colonnes()
        
        # Créer le DataFrame cible
        df_config = pd.DataFrame()
        
        # Mapper les colonnes directes
        for col_externe, col_config in mapping.items():
            if col_externe in df.columns:
                df_config[col_config] = df[col_externe]
        
        # Traitement spécial pour la semaine (calculée depuis la date)
        if 'Date' in df.columns:
            # Convertir en datetime
            df['_Date_parsed'] = pd.to_datetime(df['Date'], errors='coerce')
            
            # Calculer le numéro de semaine
            # La semaine 1 commence le 16 octobre 2025
            date_debut_saison = pd.to_datetime('2025-10-16')  # Début de la semaine 1
            
            def calculer_semaine(date):
                if pd.isna(date):
                    return ''
                delta = (date - date_debut_saison).days
                semaine = (delta // 7) + 1
                return max(1, semaine)  # Au minimum semaine 1
            
            df_config['Semaine'] = df['_Date_parsed'].apply(calculer_semaine)
        else:
            df_config['Semaine'] = ''
        
        # Traitement spécial pour l'horaire (convertir format time en HH:MM)
        if 'Horaire' in df_config.columns:
            def formater_horaire(horaire):
                if pd.isna(horaire):
                    return ''
                # Si c'est déjà un string HH:MM
                if isinstance(horaire, str):
                    return horaire.strip()
                # Si c'est un time object
                if hasattr(horaire, 'hour'):
                    return f"{horaire.hour:02d}:{horaire.minute:02d}"
                # Essayer de parser
                try:
                    horaire_str = str(horaire)
                    if ':' in horaire_str:
                        parts = horaire_str.split(':')
                        return f"{int(parts[0]):02d}:{int(parts[1]):02d}"
                except:
                    pass
                return str(horaire)
            
            df_config['Horaire'] = df_config['Horaire'].apply(formater_horaire)
        
        # Nettoyer les équipes (enlever espaces superflus)
        for col in ['Equipe_1', 'Equipe_2']:
            if col in df_config.columns:
                df_config[col] = df_config[col].astype(str).str.strip()
        
        # Nettoyer les poules et extraire le genre
        if 'Poule' in df_config.columns:
            df_config['Poule'] = df_config['Poule'].astype(str).str.strip()
            
            # Extraire le genre depuis la poule (3ème caractère)
            def extraire_genre(poule):
                if pd.isna(poule) or not str(poule).strip():
                    return ''
                poule_str = str(poule).strip()
                if len(poule_str) >= 3:
                    genre_char = poule_str[2]
                    if genre_char in ['F', 'M']:
                        return genre_char
                return ''
            
            df_config['Genre'] = df_config['Poule'].apply(extraire_genre)
        else:
            df_config['Genre'] = ''
        
        # Enlever le [F] ou [M] des noms d'équipes (déjà fait dans ajouter_genre_equipe, mais on nettoie)
        for col in ['Equipe_1', 'Equipe_2']:
            if col in df_config.columns:
                df_config[col] = df_config[col].str.replace(' [F]', '', regex=False)
                df_config[col] = df_config[col].str.replace(' [M]', '', regex=False)
                df_config[col] = df_config[col].str.strip()
        
        # Normaliser les gymnases
        if 'Gymnase' in df_config.columns:
            df_config['Gymnase'] = df_config['Gymnase'].apply(self.normaliser_gymnase)
        
        # Parser et normaliser les scores
        if 'Score' in df_config.columns:
            df_config['Score'] = df_config['Score'].apply(self._parser_score)
        
        # Ajouter les colonnes manquantes avec valeurs par défaut
        colonnes_requises = [
            'Equipe_1', 'Equipe_2', 'Genre', 'Poule', 'Semaine', 'Horaire',
            'Gymnase', 'Score', 'Type_Competition', 'Remarques'
        ]
        
        for col in colonnes_requises:
            if col not in df_config.columns:
                if col == 'Type_Competition':
                    df_config[col] = 'Acad'
                elif col == 'Remarques':
                    df_config[col] = 'Importé J1'
                elif col == 'Genre':
                    df_config[col] = ''
                else:
                    df_config[col] = ''
        
        # Filtrer les matchs annulés si demandé (après avoir ajouté les remarques par défaut)
        if self.ignorer_annules:
            # Recherche insensible à la casse pour 'annul' ou 'erreur' (annulé, annulé, erreur, etc.)
            remarques_annules = df_config['Remarques'].astype(str).str.lower().str.contains(r'annul|erreur', regex=True)
            nb_annules = remarques_annules.sum()
            if nb_annules > 0:
                print(f"   ⚠️  {nb_annules} match(s) annulé(s) ou avec erreur ignoré(s)")
                df_config = df_config[~remarques_annules]
        
        # Compléter les remarques existantes
        if 'Remarques' in df_config.columns:
            df_config['Remarques'] = df_config['Remarques'].fillna('Importé J1')
            df_config['Remarques'] = df_config['Remarques'].apply(
                lambda x: f"{x} | Importé J1" if x and str(x).strip() and 'Importé' not in str(x) else 'Importé J1'
            )
        
        # Réordonner les colonnes
        df_config = df_config[colonnes_requises]
        
        return df_config
    
    def charger_matchs_fixes_existants(self) -> pd.DataFrame:
        """
        Charge les matchs fixes existants depuis la configuration.
        
        Returns:
            DataFrame des matchs fixes existants
        """
        if not self.config_excel_path.exists():
            raise FileNotFoundError(f"Fichier de configuration introuvable: {self.config_excel_path}")
        
        try:
            df = pd.read_excel(self.config_excel_path, sheet_name='Matchs_Fixes')
            print(f"✓ Matchs fixes existants chargés: {len(df)} matchs")
            return df
        except Exception as e:
            print(f"⚠️  Aucun match fixe existant (feuille vide ou inexistante)")
            # Créer un DataFrame vide avec les bonnes colonnes
            return pd.DataFrame(columns=[
                'Equipe_1', 'Equipe_2', 'Genre', 'Poule', 'Semaine', 'Horaire',
                'Gymnase', 'Score', 'Type_Competition', 'Remarques'
            ])
    
    def fusionner_matchs(self, df_nouveaux: pd.DataFrame, df_existants: pd.DataFrame) -> pd.DataFrame:
        """
        Fusionne les nouveaux matchs avec les existants (évite les doublons).
        Gère la cohérence des scores entre doublons.
        
        Args:
            df_nouveaux: Nouveaux matchs à ajouter
            df_existants: Matchs déjà présents
            
        Returns:
            DataFrame fusionné
        """
        # Créer une clé unique pour détecter les doublons
        # Utiliser les colonnes communes aux deux formats
        def creer_cle(df):
            return (
                df['Equipe_1'].astype(str) + '|' +
                df['Equipe_2'].astype(str) + '|' +
                df['Semaine'].astype(str) if 'Semaine' in df.columns else ''
            )
        
        df_nouveaux['_cle'] = creer_cle(df_nouveaux)
        df_existants['_cle'] = creer_cle(df_existants)
        
        # Identifier les doublons
        doublons = df_nouveaux[df_nouveaux['_cle'].isin(df_existants['_cle'])]
        nouveaux_uniques = df_nouveaux[~df_nouveaux['_cle'].isin(df_existants['_cle'])]
        
        print(f"\n📊 Fusion des matchs:")
        print(f"   - Matchs à ajouter: {len(df_nouveaux)}")
        print(f"   - Doublons détectés: {len(doublons)}")
        print(f"   - Nouveaux matchs uniques: {len(nouveaux_uniques)}")
        
        # Traiter les doublons pour gérer les scores
        df_doublons_traite = pd.DataFrame()
        changements_effectues = 0

        if not doublons.empty:
            print(f"   🔍 Analyse des {len(doublons)} doublon(s)...")

            for cle in doublons['_cle'].unique():
                # Récupérer les versions nouveau et existant
                nouveau = df_nouveaux[df_nouveaux['_cle'] == cle].iloc[0]
                existant = df_existants[df_existants['_cle'] == cle].iloc[0]

                # Comparer les scores
                score_nouveau = str(nouveau.get('Score', '')).strip()
                score_existant = str(existant.get('Score', '')).strip()

                # Considérer comme vide/invalide: '', 'nan', 'NaN', ou seulement des espaces
                def score_est_valide(score):
                    return score and score.lower() not in ['nan', ''] and score.strip() != ''

                score_nouveau_valide = score_est_valide(score_nouveau)
                score_existant_valide = score_est_valide(score_existant)

                match_final = existant.copy()  # Par défaut, garder l'existant

                if score_nouveau_valide and score_existant_valide:
                    # Les deux ont un score valide
                    if score_nouveau == score_existant:
                        # Scores identiques - pas d'affichage
                        pass
                    else:
                        print(f"   ⚠️  Doublon {cle.split('|')[0]} vs {cle.split('|')[1]}: scores différents!")
                        print(f"      Existant: '{score_existant}' | Nouveau: '{score_nouveau}'")
                        print(f"      → Garde le score existant: '{score_existant}'")

                elif score_nouveau_valide and not score_existant_valide:
                    # Nouveau a un score valide, existant n'en a pas ou est invalide → mettre à jour
                    print(f"   📝 Doublon {cle.split('|')[0]} vs {cle.split('|')[1]}: ajout score '{score_nouveau}'")
                    match_final['Score'] = score_nouveau
                    changements_effectues += 1

                elif not score_nouveau_valide and score_existant_valide:
                    # Existant a un score valide, nouveau n'en a pas → garder existant (pas d'affichage)
                    pass

                else:
                    # Aucun score valide (pas d'affichage)
                    pass

                df_doublons_traite = pd.concat([df_doublons_traite, match_final.to_frame().T], ignore_index=True)

        # Afficher un résumé des changements seulement s'il y en a eu
        if changements_effectues > 0:
            print(f"   ✅ {changements_effectues} doublon(s) mis à jour avec de nouveaux scores")
        
        # Fusionner: existants (sans les doublons traités) + doublons traités + nouveaux uniques
        existants_sans_doublons = df_existants[~df_existants['_cle'].isin(doublons['_cle'])]
        df_fusionne = pd.concat([
            existants_sans_doublons.drop(columns=['_cle']),
            df_doublons_traite.drop(columns=['_cle']) if not df_doublons_traite.empty else pd.DataFrame(),
            nouveaux_uniques.drop(columns=['_cle'])
        ], ignore_index=True)
        
        return df_fusionne
    
    def sauvegarder_configuration(self, df_matchs: pd.DataFrame):
        """
        Sauvegarde les matchs dans la feuille Matchs_Fixes de la configuration.
        
        Args:
            df_matchs: DataFrame des matchs à sauvegarder
        """
        if self.dry_run:
            print("\n🔍 MODE SIMULATION - Aucune modification effectuée")
            print(f"   {len(df_matchs)} matchs seraient sauvegardés")
            return
        
        print(f"\n💾 Sauvegarde dans {self.config_excel_path}...")
        
        try:
            # Charger le workbook existant
            wb = openpyxl.load_workbook(self.config_excel_path)
            
            # Supprimer la feuille Matchs_Fixes si elle existe
            if 'Matchs_Fixes' in wb.sheetnames:
                del wb['Matchs_Fixes']
            
            # Créer une nouvelle feuille
            ws = wb.create_sheet('Matchs_Fixes')
            
            # Écrire les données
            for r in dataframe_to_rows(df_matchs, index=False, header=True):
                ws.append(r)
            
            # Sauvegarder
            wb.save(self.config_excel_path)
            
            print(f"✓ Configuration sauvegardée: {len(df_matchs)} matchs dans Matchs_Fixes")
            
        except Exception as e:
            raise RuntimeError(f"Erreur lors de la sauvegarde: {e}")
    
    def executer(self):
        """Exécute le processus complet d'importation."""
        print("="*70)
        print("🔄 IMPORTATION DE MATCHS EXTERNES")
        print("="*70)
        
        print(f"\n⚙️  Configuration:")
        print(f"   - Fichier config: {self.config_yaml_path}")
        print(f"   - Fichier Excel: {self.config_excel_path}")
        print(f"   - Sport: {self.sport}")
        if self.journee:
            print(f"   - Journée: {self.journee}")
        if self.date_limite:
            print(f"   - Date limite: {self.date_limite.strftime('%d/%m/%Y')}")
        print(f"   - Filtre: {'Avec score' if self.avec_score else 'Sans score' if self.sans_score else 'Tous'}")
        if self.dry_run:
            print(f"   - Mode: SIMULATION (dry-run)")
        
        # 1. Télécharger le fichier externe
        self.df_externe = self.telecharger_fichier_externe()
        
        # 2. Explorer la structure (optionnel, pour debug)
        # self.explorer_structure()
        
        # 3. Filtrer les matchs
        print(f"\n🔍 Filtrage des matchs...")
        df_filtre = self.filtrer_matchs(self.df_externe)
        
        if len(df_filtre) == 0:
            print("⚠️  Aucun match ne correspond aux critères. Arrêt.")
            return
        
        # 4. Convertir vers le format config
        print(f"\n🔄 Conversion vers le format Matchs_Fixes...")
        df_convertis = self.convertir_vers_format_config(df_filtre)
        
        # 5. Charger les matchs existants
        print(f"\n📂 Chargement des matchs existants...")
        df_existants = self.charger_matchs_fixes_existants()
        
        # 6. Fusionner
        df_final = self.fusionner_matchs(df_convertis, df_existants)
        
        # 7. Sauvegarder
        self.sauvegarder_configuration(df_final)
        
        print("\n" + "="*70)
        print("✅ IMPORTATION TERMINÉE")
        print("="*70)


def main():
    """Point d'entrée principal."""
    parser = argparse.ArgumentParser(
        description="Importer des matchs depuis un fichier Excel partagé en ligne",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__
    )
    
    # Arguments obligatoires
    parser.add_argument(
        '--config',
        required=True,
        help="Chemin vers le fichier de configuration Excel"
    )
    
    # Source du fichier (l'un des deux requis)
    source_group = parser.add_mutually_exclusive_group(required=True)
    source_group.add_argument(
        '--url',
        help="URL du fichier Excel partagé en ligne"
    )
    source_group.add_argument(
        '--fichier-local',
        help="Chemin vers un fichier Excel local"
    )
    
    # Arguments optionnels
    parser.add_argument(
        '--sport',
        default='VB',
        help="Code du sport (défaut: VB pour Volley-Ball)"
    )
    parser.add_argument(
        '--journee',
        type=int,
        help="Numéro de journée à importer"
    )
    parser.add_argument(
        '--date-limite',
        help="Date limite pour l'import (format DD/MM/YYYY)"
    )
    
    # Filtres
    filtre_group = parser.add_mutually_exclusive_group()
    filtre_group.add_argument(
        '--avec-score',
        action='store_true',
        help="Importer uniquement les matchs avec score"
    )
    filtre_group.add_argument(
        '--sans-score',
        action='store_true',
        help="Importer uniquement les matchs sans score (planifiés)"
    )
    filtre_group.add_argument(
        '--tous',
        action='store_true',
        help="Importer tous les matchs (défaut)"
    )
    
    # Options
    parser.add_argument(
        '--dry-run',
        action='store_true',
        help="Mode simulation (ne modifie pas le fichier)"
    )
    parser.add_argument(
        '--explorer',
        action='store_true',
        help="Explorer la structure du fichier externe sans importer"
    )
    parser.add_argument(
        '--ignorer-annules',
        action='store_true',
        default=True,
        help="Ignorer les matchs avec 'annule' ou 'erreur' dans les remarques (défaut: activé)"
    )
    parser.add_argument(
        '--garder-annules',
        action='store_false',
        dest='ignorer_annules',
        help="Garder les matchs avec 'annule' ou 'erreur' dans les remarques"
    )
    
    args = parser.parse_args()
    
    try:
        importeur = ImporteurMatchsExternes(
            config_path=args.config,
            url_externe=args.url,
            fichier_local=args.fichier_local,
            sport=args.sport,
            journee=args.journee,
            date_limite=args.date_limite,
            avec_score=args.avec_score,
            sans_score=args.sans_score,
            tous=args.tous,
            dry_run=args.dry_run,
            ignorer_annules=args.ignorer_annules
        )
        
        if args.explorer:
            importeur.df_externe = importeur.telecharger_fichier_externe()
            importeur.explorer_structure()
        else:
            importeur.executer()
        
        return 0
        
    except Exception as e:
        print(f"\n❌ Erreur: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())
