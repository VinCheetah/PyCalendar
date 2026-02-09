#!/usr/bin/env python3
"""Inspect Semaine column format in Indispos_Gymnases."""
import sys
sys.path.insert(0, "src")
from pycalendar.core.config import Config

config = Config.from_yaml("configs/config_volleyP2.yaml")
excel_path = config.fichier_donnees
print(f"Excel: {excel_path}")

import openpyxl
wb = openpyxl.load_workbook(excel_path)
ws = wb['Indispos_Gymnases']

print("\nIndispos_Gymnases columns:")
print([ws.cell(1, i).value for i in range(1, 10)])

print("\nFirst 5 rows:")
for i in range(2, 7):
    row = [ws.cell(i, j).value for j in range(1, 7)]
    print(f"  Row {i}: {row}")
