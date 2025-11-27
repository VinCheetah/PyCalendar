#!/usr/bin/env python3
"""
Ajoute les informations de contact dans le fichier config_volley.xlsx

STRATÉGIE DE MATCHING:
1. Normaliser les noms d'institutions (retirer UDL -, ASC, espaces, etc.)
2. Matcher par: Institution + Genre + Niveau + Horaire préféré
3. Pour LYON 1: utiliser les infos détaillées (niveaux/horaires) pour matcher précisément
4. Pour les capitaines LYON 1: compléter après le premier pass
"""

import pandas as pd
import openpyxl
import re
import shutil
from pathlib import Path
from datetime import datetime


def normalize_institution(name: str) -> str:
    """Normalise le nom d'une institution pour le matching."""
    name = name.upper().strip()
    
    # Retirer préfixes UDL, ASC
    prefixes = ['UDL - ', 'ASC ', 'ENTPE - ']
    for prefix in prefixes:
        if name.startswith(prefix):
            name = name[len(prefix):]
    
    # Cas spéciaux
    if 'LYON 1 SCIENCES' in name or 'LYON 1 SANTE' in name:
        return 'LYON 1'
    if name == 'SANTE':
        return 'LYON 1 SANTE'
    
    # Normalisation bidirectionnelle (contact <-> config)
    # Tous les formats sont normalisés vers le format du CONFIG
    mappings = {
        # Format contact -> Format config unifié
        'ISARA LYON': 'ISARA',
        'ESA BRON': 'ESA',
        'LYON 2 IEP': 'LYON 2 (IEP)',
        'INSA LYON': 'INSA',
        'CPE LYON': 'CPE',
        'ECAM LYON': 'ECAM',
        'CATHO LYON': 'CATHO',
        'ESME SUDURIA': 'ESME',
        'ESSCA LYON': 'ESSCA',
        'COB VETO': 'VETO',  # Contact "COB VETO" -> "VETO" (format config)
        'AMOS LYON': 'AMOS',
    }
    
    for contact_format, config_format in mappings.items():
        if name == contact_format:
            return config_format
    
    # Retirer espaces multiples
    name = re.sub(r'\s+', ' ', name).strip()
    
    return name
    return name


def extract_info_from_contact_name(name: str) -> dict:
    """
    Extrait les informations d'un nom d'équipe du fichier de contacts.
    
    Ex: "UDL - LYON 1 SCIENCES (A3 16H, 18H / A4 16H, 18H)"
    -> {'institution': 'LYON 1', 'niveaux': ['A3', 'A4'], 'horaires': ['16H', '18H']}
    """
    # Institution de base (retirer parenthèses, niveaux, horaires)
    base = re.sub(r'\s*\([^)]*\)', '', name)
    base = re.sub(r'\s+\d{1,2}[hH]\b', '', base)
    base = re.sub(r'\s+A[1-4]\b', '', base)
    institution = normalize_institution(base)
    
    # Extraire niveaux (A1, A2, A3, A4)
    niveaux = list(set(re.findall(r'A[1-4]', name)))
    niveaux.sort()
    
    # Extraire horaires (14h, 16h, 18h, 20h - avec H majuscule ou minuscule)
    horaires_brut = re.findall(r'(\d{1,2})[hH]', name)
    horaires = list(set([h.zfill(2) + ':00' for h in horaires_brut]))  # Normaliser en HH:00
    horaires.sort()
    
    return {
        'institution': institution,
        'niveaux': niveaux,
        'horaires': horaires,
        'raw_name': name
    }


def extract_info_from_config_team(equipe: str, poule: str, horaire_pref: str) -> dict:
    """
    Extrait les informations d'une équipe du config.
    
    Ex: equipe="LYON 1 (5)", poule="VBFA2PA", horaire="14:00"
    -> {'institution': 'LYON 1', 'genre': 'F', 'niveau': 'A2', 'horaire': '14:00'}
    """
    # Institution
    match = re.match(r'(.+?)\s*\(', equipe)
    institution = normalize_institution(match.group(1).strip() if match else equipe)
    
    # Genre depuis poule (VBF = F, VBM = M)
    genre = 'F' if 'VBF' in poule else 'M'
    
    # Niveau depuis poule (A1, A2, A3, A4)
    match_niveau = re.search(r'A(\d)', poule)
    niveau = f'A{match_niveau.group(1)}' if match_niveau else None
    
    return {
        'institution': institution,
        'genre': genre,
        'niveau': niveau,
        'horaire': horaire_pref,
        'equipe': equipe,
        'poule': poule
    }


def match_contact_to_team(team_info: dict, contacts_list: list, genre: str) -> dict:
    """
    Trouve le contact correspondant à une équipe du config.
    
    Stratégie:
    1. Match exact: institution + niveau + horaire
    2. Match partiel: institution + niveau
    3. Match large: institution seule
    """
    institution = team_info['institution']
    niveau = team_info['niveau']
    horaire = team_info['horaire']
    
    # Filtrer par genre
    candidates = [c for c in contacts_list if c['genre'] == genre]
    
    # Priorité 1: Institution + Niveau + Horaire
    if niveau and horaire:
        for contact in candidates:
            if contact['info']['institution'] == institution:
                if niveau in contact['info']['niveaux'] or not contact['info']['niveaux']:
                    if horaire in contact['info']['horaires'] or not contact['info']['horaires']:
                        return contact
    
    # Priorité 2: Institution + Niveau
    if niveau:
        for contact in candidates:
            if contact['info']['institution'] == institution:
                if niveau in contact['info']['niveaux'] or not contact['info']['niveaux']:
                    return contact
    
    # Priorité 3: Institution seule
    for contact in candidates:
        if contact['info']['institution'] == institution:
            return contact
    
    return None


def load_contacts_from_file(file_path: Path) -> dict:
    """Charge les contacts depuis CONTACTS RESPO EQUIPES BB_HB_VB.xlsx"""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    contacts = {'F': [], 'M': []}
    
    # Volleyball Féminin (colonnes 1-4)
    for row_idx in range(4, ws.max_row + 1):
        institution = ws.cell(row_idx, 1).value
        nom = ws.cell(row_idx, 2).value
        email = ws.cell(row_idx, 3).value
        tel = ws.cell(row_idx, 4).value
        
        if institution and str(institution).strip():
            institution_str = str(institution).strip()
            info = extract_info_from_contact_name(institution_str)
            
            contacts['F'].append({
                'info': info,
                'nom': str(nom).strip() if nom else None,
                'email': str(email).strip() if email else None,
                'tel': str(tel).strip().replace('.', '').replace(' ', '').replace('.0', '') if tel else None,
                'genre': 'F'
            })
    
    # Volleyball Masculin (colonnes 6-9)
    for row_idx in range(4, ws.max_row + 1):
        institution = ws.cell(row_idx, 6).value
        nom = ws.cell(row_idx, 7).value
        email = ws.cell(row_idx, 8).value
        tel = ws.cell(row_idx, 9).value
        
        if institution and str(institution).strip():
            institution_str = str(institution).strip()
            info = extract_info_from_contact_name(institution_str)
            
            contacts['M'].append({
                'info': info,
                'nom': str(nom).strip() if nom else None,
                'email': str(email).strip() if email else None,
                'tel': str(tel).strip().replace('.', '').replace(' ', '').replace('.0', '') if tel else None,
                'genre': 'M'
            })
    
    return contacts


def load_capitaines_lyon1(file_path: Path) -> dict:
    """Charge les capitaines LYON 1 depuis CAPITAINES 2025 2026.docx"""
    from docx import Document
    
    doc = Document(file_path)
    
    capitaines = {}
    
    # Le fichier a une seule table avec 6 colonnes:
    # N°, NIVEAU, HORAIRE MATCH, ENTRAINEUR, NOM/PRENOM, NUM TEL
    for table in doc.tables:
        for i, row in enumerate(table.rows):
            cells = [cell.text.strip() for cell in row.cells]
            
            # Skip header row
            if i == 0 or not cells[0] or 'N°' in cells[0]:
                continue
            
            try:
                numero = int(cells[0])
                # Déterminer le genre basé sur le numéro (les 13 premiers sont F, le reste M)
                current_genre = 'F' if numero <= 13 else 'M'
                
                nom_prenom = cells[4] if len(cells) > 4 else ''
                tel = cells[5] if len(cells) > 5 else ''
                
                # Nettoyer téléphone
                tel = tel.replace(' ', '').replace('.', '')
                
                # Générer l'email depuis NOM/PRENOM
                if nom_prenom and ' ' in nom_prenom:
                    parts = nom_prenom.split()
                    if len(parts) >= 2:
                        nom = parts[0]
                        prenom = parts[1]
                        email = f"{prenom.lower()}.{nom.lower()}@univ-lyon1.fr"
                        
                        key = (numero, current_genre)
                        capitaines[key] = {
                            'nom': nom_prenom.upper(),
                            'email': email,
                            'tel': tel,
                            'genre': current_genre,
                            'numero': numero
                        }
            except (ValueError, IndexError):
                continue
    
    return capitaines


def main():
    """Fonction principale."""
    base_dir = Path(__file__).parent.parent
    config_file = base_dir / 'examples' / 'volleyball' / 'config_volley.xlsx'
    contacts_file = base_dir / 'config' / 'CONTACTS RESPO EQUIPES BB_HB_VB.xlsx'
    capitaines_file = base_dir / 'config' / 'CAPITAINES 2025 2026.docx'
    
    print("=" * 80)
    print("📥 AJOUT DES CONTACTS DANS CONFIG_VOLLEY.XLSX")
    print("=" * 80)
    print()
    
    # 1. Créer backup
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_file = config_file.parent / f'config_volley_backup_{timestamp}.xlsx'
    shutil.copy2(config_file, backup_file)
    print(f"✅ Backup créé: {backup_file.name}")
    print()
    
    # 2. Charger les contacts
    print("📖 Chargement des contacts depuis CONTACTS RESPO EQUIPES BB_HB_VB.xlsx...")
    contacts = load_contacts_from_file(contacts_file)
    print(f"   Contacts Féminin: {len(contacts['F'])}")
    print(f"   Contacts Masculin: {len(contacts['M'])}")
    print()
    
    # 3. Charger les capitaines LYON 1
    print("📖 Chargement des capitaines LYON 1...")
    capitaines = load_capitaines_lyon1(capitaines_file)
    print(f"   Capitaines chargés: {len(capitaines)}")
    print()
    
    # 4. Charger config_volley
    print("📖 Chargement de config_volley.xlsx...")
    df = pd.read_excel(config_file, sheet_name='Equipes')
    print(f"   Équipes dans config: {len(df)}")
    print()
    
    # 5. Ajouter colonnes si elles n'existent pas
    if 'Responsable_Nom' not in df.columns:
        df['Responsable_Nom'] = None
    if 'Responsable_Email' not in df.columns:
        df['Responsable_Email'] = None
    if 'Responsable_Telephone' not in df.columns:
        df['Responsable_Telephone'] = None
    
    # 6. Matcher et remplir
    print("🔗 Matching des contacts...")
    print()
    
    matched = 0
    not_matched = []
    
    for idx, row in df.iterrows():
        equipe = row['Equipe']
        poule = row['Poule']
        horaire_pref = row['Horaire_Prefere']
        
        # Extraire infos de l'équipe
        team_info = extract_info_from_config_team(equipe, poule, horaire_pref)
        genre = team_info['genre']
        
        # Essayer de matcher avec les capitaines LYON 1 d'abord
        if team_info['institution'] == 'LYON 1':
            # Extraire numéro
            match_num = re.search(r'\((\d+)\)', equipe)
            if match_num:
                numero = int(match_num.group(1))
                key = (numero, genre)
                
                if key in capitaines:
                    cap = capitaines[key]
                    df.at[idx, 'Responsable_Nom'] = cap['nom']
                    df.at[idx, 'Responsable_Email'] = cap['email']
                    df.at[idx, 'Responsable_Telephone'] = cap['tel']
                    matched += 1
                    print(f"✅ {equipe:20} -> Capitaine LYON 1: {cap['nom']}")
                    continue
        
        # Sinon, matcher avec le fichier de contacts
        all_contacts = contacts['F'] + contacts['M']
        contact = match_contact_to_team(team_info, all_contacts, genre)
        
        if contact:
            df.at[idx, 'Responsable_Nom'] = contact['nom']
            df.at[idx, 'Responsable_Email'] = contact['email']
            df.at[idx, 'Responsable_Telephone'] = contact['tel']
            matched += 1
            print(f"✅ {equipe:20} ({genre} {team_info['niveau']}) -> {contact['info']['raw_name'][:40]}")
        else:
            not_matched.append({
                'equipe': equipe,
                'institution': team_info['institution'],
                'genre': genre,
                'niveau': team_info['niveau'],
                'horaire': horaire_pref
            })
    
    print()
    print("=" * 80)
    print("📊 RÉSULTATS")
    print("=" * 80)
    print(f"✅ Équipes avec contact: {matched}/{len(df)}")
    print(f"❌ Équipes sans contact: {len(not_matched)}/{len(df)}")
    print()
    
    if not_matched:
        print("❌ Équipes non matchées:")
        for eq in not_matched[:20]:
            print(f"   {eq['equipe']:20} ({eq['institution']:15} | {eq['genre']} | {eq['niveau']} | {eq['horaire']})")
        if len(not_matched) > 20:
            print(f"   ... et {len(not_matched) - 20} autres")
        print()
    
    # 7. Sauvegarder
    print("💾 Sauvegarde du fichier...")
    
    # Charger le workbook complet pour préserver les autres feuilles
    wb = openpyxl.load_workbook(config_file)
    
    # Supprimer et recréer la feuille Equipes
    if 'Equipes' in wb.sheetnames:
        del wb['Equipes']
    
    ws = wb.create_sheet('Equipes', 0)
    
    # Écrire les données
    for r_idx, row in enumerate([df.columns.tolist()] + df.values.tolist(), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(r_idx, c_idx, value)
    
    wb.save(config_file)
    print(f"✅ Fichier sauvegardé: {config_file}")
    print()
    print("✨ Import terminé avec succès!")


if __name__ == '__main__':
    main()
