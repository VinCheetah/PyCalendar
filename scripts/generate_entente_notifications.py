#!/usr/bin/env python3
"""
Script de génération des notifications d'ententes

Ce script analyse une solution de calendrier et génère les emails de notification
pour chaque équipe ayant des matchs en entente à organiser.

Usage:
    python scripts/generate_entente_notifications.py [options]
    
Options:
    --solution PATH      Chemin vers le fichier solution JSON (défaut: solutions/latest_volley.json)
    --contacts PATH      Chemin vers le fichier contacts Excel (défaut: config/contacts_equipes.xlsx)
    --output PATH        Chemin vers le fichier de sortie (défaut: notifications_ententes.txt)
    --deadline DATE      Date limite pour organiser les ententes (format: JJ/MM/AAAA)
    --test              Mode test : affiche seulement les statistiques sans générer le fichier
"""

import json
import argparse
from pathlib import Path
from datetime import datetime, timedelta
from typing import Dict, List, Set, Tuple, Optional
from collections import defaultdict
import sys

# Ajouter le répertoire parent au path pour importer les modules
sys.path.insert(0, str(Path(__file__).parent.parent))

try:
    import pandas as pd
except ImportError:
    print("❌ pandas n'est pas installé. Installez-le avec: pip install pandas openpyxl")
    sys.exit(1)


class EntenteNotificationGenerator:
    """Générateur de notifications pour les matchs en entente."""
    
    def __init__(self, solution_path: str, config_path: str = None, deadline_date: str = None):
        """
        Initialise le générateur.
        
        Args:
            solution_path: Chemin vers le fichier solution JSON
            config_path: Chemin vers le fichier config Excel (optionnel, déduit de la solution si absent)
            deadline_date: Date limite au format JJ/MM/AAAA (optionnel)
        """
        self.solution_path = Path(solution_path)
        self.config_path = Path(config_path) if config_path else None
        self.deadline_date = deadline_date
        
        # Données chargées
        self.solution_data = None
        self.contacts_df = None
        self.ententes_par_equipe = defaultdict(list)
        self.equipes_sans_contacts = []  # Liste des équipes sans infos de contact
        
    def load_solution(self) -> bool:
        """Charge le fichier solution JSON."""
        if not self.solution_path.exists():
            print(f"❌ Fichier solution introuvable: {self.solution_path}")
            return False
        
        try:
            with open(self.solution_path, 'r', encoding='utf-8') as f:
                self.solution_data = json.load(f)
            print(f"✅ Solution chargée: {self.solution_path}")
            return True
        except Exception as e:
            print(f"❌ Erreur lors du chargement de la solution: {e}")
            return False
    
    def load_contacts(self) -> bool:
        """Charge les contacts depuis la feuille Equipes du fichier config Excel."""
        # Si config_path n'est pas fourni, essayer de le déduire du chemin solution
        if not self.config_path:
            # Chercher un fichier config dans examples/volleyball/
            possible_configs = [
                Path('examples/volleyball/config_volley.xlsx'),
                Path('configs/config_volley.xlsx'),
                Path('config/config_volley.xlsx'),
            ]
            
            for config in possible_configs:
                if config.exists():
                    self.config_path = config
                    break
            
            if not self.config_path:
                print(f"⚠️  Fichier config Excel introuvable")
                print(f"   Essayez de spécifier --config explicitement")
                print(f"   Continuons sans contacts (informations limitées)")
                return False
        
        if not self.config_path.exists():
            print(f"⚠️  Fichier config introuvable: {self.config_path}")
            print(f"   Continuons sans contacts (informations limitées)")
            return False
        
        try:
            # Lire la feuille Equipes du fichier config
            df = pd.read_excel(self.config_path, sheet_name='Equipes')
            
            # Colonnes attendues: Equipe, Responsable_Nom, Responsable_Email, Responsable_Telephone
            # Normaliser les colonnes
            df.columns = [c.strip() for c in df.columns]
            
            # Vérifier les colonnes requises
            required_cols = ['Equipe']
            missing_cols = [col for col in required_cols if col not in df.columns]
            if missing_cols:
                print(f"❌ Fichier config: colonnes manquantes: {missing_cols}")
                return False
            
            # Créer une colonne unifiée pour l'ID équipe (avec et sans genre)
            # Format attendu dans solution: "LYON 1 (1)|M" ou "LYON 1 (1)|F"
            # Format dans config: "LYON 1 (1)" (sans genre)
            self.contacts_df = df
            
            print(f"✅ Contacts chargés depuis config: {len(self.contacts_df)} équipes")
            return True
            
        except Exception as e:
            print(f"❌ Erreur lors du chargement des contacts depuis config: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def extract_ententes(self) -> bool:
        """Extrait les matchs en entente depuis la solution."""
        if not self.solution_data:
            return False
        
        try:
            # Récupérer tous les matchs (v2.0 : dictionnaire avec 'scheduled' et 'unscheduled')
            matches_dict = self.solution_data.get('matches', {})
            all_matches = []
            
            # Combiner les matchs planifiés et non planifiés
            if isinstance(matches_dict, dict):
                all_matches.extend(matches_dict.get('scheduled', []))
                all_matches.extend(matches_dict.get('unscheduled', []))
            
            nb_total = len(all_matches)
            nb_scheduled = len(matches_dict.get('scheduled', []))
            nb_ententes = 0
            
            # Set pour détecter les doublons (même match entre mêmes équipes)
            seen_matches = set()
            
            for match in all_matches:
                # Vérifier si c'est une entente (v2.0 : champ 'is_entente')
                if match.get('is_entente', False):
                    equipe1_id = match['equipe1_id']
                    equipe2_id = match['equipe2_id']
                    
                    # Créer une clé unique pour détecter les doublons
                    # Trier les IDs pour que A-B et B-A soient identiques
                    match_key = tuple(sorted([equipe1_id, equipe2_id]))
                    
                    # Vérifier si ce match a déjà été traité
                    if match_key in seen_matches:
                        # Doublon détecté - ignorer
                        continue
                    
                    seen_matches.add(match_key)
                    nb_ententes += 1
                    
                    equipe1_nom = match['equipe1_nom']
                    equipe2_nom = match['equipe2_nom']
                    
                    # Ajouter aux deux équipes
                    entente_info = {
                        'adversaire_id': equipe2_id,
                        'adversaire_nom': equipe2_nom,
                        'adversaire_institution': match['equipe2_institution'],
                        'poule': match['poule']
                    }
                    self.ententes_par_equipe[equipe1_id].append(entente_info)
                    
                    entente_info_inverse = {
                        'adversaire_id': equipe1_id,
                        'adversaire_nom': equipe1_nom,
                        'adversaire_institution': match['equipe1_institution'],
                        'poule': match['poule']
                    }
                    self.ententes_par_equipe[equipe2_id].append(entente_info_inverse)
            
            print(f"\n📊 Statistiques de la solution:")
            print(f"   Matchs planifiés avec créneau: {nb_scheduled}")
            print(f"   Matchs en entente (is_entente=True): {nb_ententes}")
            print(f"   Total matchs: {nb_total}")
            print(f"   Équipes concernées par des ententes: {len(self.ententes_par_equipe)}")
            
            return True
        except Exception as e:
            print(f"❌ Erreur lors de l'extraction des ententes: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def get_contact_info(self, equipe_id: str) -> Optional[Dict]:
        """
        Récupère les informations de contact d'une équipe.
        
        Args:
            equipe_id: ID de l'équipe (format: "LYON 1 (1)|M" ou "LYON 1 (1)")
            
        Returns:
            Dict avec les infos de contact ou None si non trouvé
        """
        if self.contacts_df is None:
            return None
        
        # Extraire le nom de base sans genre (format config: "LYON 1 (1)")
        # Le format solution peut être "LYON 1 (1)|M" ou "LYON 1 (1)|F"
        base_id = equipe_id.split('|')[0].strip() if '|' in equipe_id else equipe_id
        
        # Chercher l'équipe dans le fichier config
        equipe_row = self.contacts_df[self.contacts_df['Equipe'].astype(str).str.strip() == base_id]
        
        if equipe_row.empty:
            # Ajouter à la liste des équipes sans contacts
            if base_id not in [e['id'] for e in self.equipes_sans_contacts]:
                self.equipes_sans_contacts.append({
                    'id': base_id,
                    'nom': equipe_id
                })
            return None
        
        row = equipe_row.iloc[0]
        
        # Fonction helper pour récupérer valeurs avec fallback
        def safe_get(col, default='Non renseigné'):
            try:
                val = row.get(col, default)
                if pd.isna(val) or str(val).strip() == '' or str(val) == 'nan':
                    return default
                return str(val).strip()
            except Exception:
                return default
        
        return {
            'nom_equipe': safe_get('Equipe', base_id),
            'institution': base_id.split('(')[0].strip() if '(' in base_id else base_id,
            'capitaine_nom': safe_get('Responsable_Nom'),
            'capitaine_email': safe_get('Responsable_Email'),
            'capitaine_telephone': safe_get('Responsable_Telephone'),
            'remarques': safe_get('Remarques', '')
        }
    
    def format_email(self, equipe_id: str, ententes: List[Dict]) -> str:
        """
        Formate l'email de notification pour une équipe.
        
        Args:
            equipe_id: ID de l'équipe
            ententes: Liste des ententes pour cette équipe
            
        Returns:
            Texte de l'email formaté
        """
        # Récupérer les infos de l'équipe
        equipe_contact = self.get_contact_info(equipe_id)
        if not equipe_contact:
            equipe_nom = equipe_id
            email_dest = "[EMAIL NON RENSEIGNÉ]"
            capitaine_nom = ""
        else:
            equipe_nom = equipe_contact['nom_equipe']
            email_dest = equipe_contact['capitaine_email']
            capitaine_nom = equipe_contact.get('capitaine_nom', '')
        
        # Nombre d'ententes
        nb_ententes = len(ententes)
        
        # Adapter le texte au nombre d'ententes
        if nb_ententes == 1:
            intro_match = "un match à organiser en entente"
            votre_match = "votre match"
            ce_match = "ce match"
        else:
            intro_match = f"{nb_ententes} matchs à organiser en entente"
            votre_match = "vos matchs"
            ce_match = "ces matchs"
        
        # En-tête de l'email
        email_lines = []
        email_lines.append("=" * 80)
        email_lines.append(f"DESTINATAIRE: {email_dest}")
        if capitaine_nom:
            email_lines.append(f"CAPITAINE: {capitaine_nom}")
        email_lines.append(f"ÉQUIPE: {equipe_nom}")
        email_lines.append("")
        if nb_ententes == 1:
            email_lines.append(f"OBJET: Match en entente à organiser - {equipe_nom}")
        else:
            email_lines.append(f"OBJET: {nb_ententes} matchs en entente à organiser - {equipe_nom}")
        email_lines.append("=" * 80)
        email_lines.append("")
        
        # Corps du message - Introduction personnalisée
        if capitaine_nom:
            email_lines.append(f"Bonjour {capitaine_nom},")
        else:
            email_lines.append(f"Bonjour,")
        email_lines.append("")
        
        # Message adapté au nombre d'ententes
        if nb_ententes == 1:
            email_lines.append(f"Votre équipe « {equipe_nom} » a un match à organiser en entente.")
            email_lines.append("")
            email_lines.append("Ce match n'a pas pu être placé sur le calendrier officiel faute de créneaux")
            email_lines.append("disponibles. Il vous revient donc de vous organiser directement avec l'équipe")
            email_lines.append("adverse pour trouver une date, un horaire et un lieu qui conviennent aux deux")
            email_lines.append("équipes.")
        else:
            email_lines.append(f"Votre équipe « {equipe_nom} » a {nb_ententes} matchs à organiser en entente.")
            email_lines.append("")
            email_lines.append("Ces matchs n'ont pas pu être placés sur le calendrier officiel faute de créneaux")
            email_lines.append("disponibles. Il vous revient donc de vous organiser directement avec les équipes")
            email_lines.append("adverses pour trouver des dates, horaires et lieux qui conviennent à tous.")
        
        email_lines.append("")
        
        # Date limite si spécifiée
        if self.deadline_date:
            email_lines.append(f"Date limite : {self.deadline_date}")
            email_lines.append(f"Merci d'organiser {ce_match} avant cette échéance.")
            email_lines.append("")
        
        # Liste des ententes
        if nb_ententes == 1:
            email_lines.append("VOTRE MATCH EN ENTENTE :")
        else:
            email_lines.append(f"VOS {nb_ententes} MATCHS EN ENTENTE :")
        email_lines.append("")
        
        for idx, entente in enumerate(ententes, 1):
            if nb_ententes > 1:
                email_lines.append(f"  ── Match {idx}/{nb_ententes} ──")
            else:
                email_lines.append(f"  ────────────")
            email_lines.append(f"  Poule: {entente['poule']}")
            email_lines.append(f"  Adversaire: {entente['adversaire_nom']} ({entente['adversaire_institution']})")
            email_lines.append(f"  ────────────")
            email_lines.append("")
            
            # Informations de contact de l'adversaire
            adv_contact = self.get_contact_info(entente['adversaire_id'])
            if adv_contact:
                email_lines.append(f"  Coordonnées du capitaine à contacter :")
                cap_nom = adv_contact['capitaine_nom']
                cap_email = adv_contact['capitaine_email']
                cap_tel = adv_contact['capitaine_telephone']
                
                # Formatter proprement les coordonnées
                if str(cap_nom) != 'nan' and cap_nom:
                    email_lines.append(f"    - {cap_nom}")
                if str(cap_email) != 'nan' and cap_email:
                    email_lines.append(f"    - {cap_email}")
                if str(cap_tel) != 'nan' and cap_tel:
                    # Formater le téléphone
                    tel_str = str(cap_tel).replace('.0', '')
                    if len(tel_str) == 10 and tel_str.startswith('0'):
                        tel_formatted = f"{tel_str[0:2]} {tel_str[2:4]} {tel_str[4:6]} {tel_str[6:8]} {tel_str[8:10]}"
                        email_lines.append(f"    - {tel_formatted}")
                    else:
                        email_lines.append(f"    - {tel_str}")
                
                if adv_contact.get('remarques') and str(adv_contact['remarques']) != 'nan':
                    email_lines.append(f"    - Remarque : {adv_contact['remarques']}")
            else:
                email_lines.append(f"  Coordonnées non disponibles pour cette équipe.")
                email_lines.append(f"  Contactez l'organisation pour obtenir les informations.")
            
            email_lines.append("")
        
        # Instructions finales adaptées
        email_lines.append("─" * 80)
        email_lines.append("")
        email_lines.append("Comment procéder :")
        email_lines.append("")
        
        if nb_ententes == 1:
            email_lines.append("  1. Contactez le capitaine adverse dès que possible")
            email_lines.append("  2. Convenez ensemble d'une date et d'un horaire")
            email_lines.append("  3. Trouvez un gymnase disponible (le vôtre, le leur, ou un autre)")
            email_lines.append("  4. Confirmez la rencontre auprès de l'organisation")
        else:
            email_lines.append("  1. Contactez les capitaines adverses dès que possible")
            email_lines.append("  2. Convenez ensemble des dates et horaires pour chaque match")
            email_lines.append("  3. Trouvez des gymnases disponibles")
            email_lines.append("  4. Confirmez les rencontres auprès de l'organisation")
        
        email_lines.append("")
        
        if self.deadline_date:
            email_lines.append(f"Rappel : échéance au {self.deadline_date}")
            email_lines.append("")
        
        email_lines.append("En cas de difficulté (capitaine injoignable, désaccord sur la date, etc.),")
        email_lines.append("contactez-nous rapidement pour que nous puissions vous aider.")
        email_lines.append("")
        email_lines.append("Bon courage pour l'organisation et bonne saison sportive !")
        email_lines.append("")
        email_lines.append("Cordialement,")
        email_lines.append("L'équipe d'organisation du championnat")
        email_lines.append("")
        email_lines.append("=" * 80)
        email_lines.append("")
        email_lines.append("")
        
        return "\n".join(email_lines)
    
    def generate_notifications(self, output_path: str) -> bool:
        """
        Génère le fichier de notifications.
        
        Args:
            output_path: Chemin vers le fichier de sortie
            
        Returns:
            True si succès, False sinon
        """
        if not self.ententes_par_equipe:
            print("ℹ️  Aucune entente trouvée dans la solution.")
            return True
        
        try:
            output_file = Path(output_path)
            
            with open(output_file, 'w', encoding='utf-8') as f:
                # En-tête du fichier
                f.write("╔" + "═" * 78 + "╗\n")
                f.write("║" + " " * 78 + "║\n")
                f.write("║" + "  NOTIFICATIONS DES MATCHS EN ENTENTE".center(78) + "║\n")
                f.write("║" + " " * 78 + "║\n")
                f.write("║" + f"  Généré le: {datetime.now().strftime('%d/%m/%Y à %H:%M')}".ljust(78) + "║\n")
                f.write("║" + f"  Solution: {self.solution_path.name}".ljust(78) + "║\n")
                f.write("║" + f"  Nombre d'équipes concernées: {len(self.ententes_par_equipe)}".ljust(78) + "║\n")
                f.write("║" + " " * 78 + "║\n")
                f.write("╚" + "═" * 78 + "╝\n")
                f.write("\n\n\n")
                
                # Générer un email pour chaque équipe
                equipes_triees = sorted(self.ententes_par_equipe.keys())
                
                for idx, equipe_id in enumerate(equipes_triees, 1):
                    ententes = self.ententes_par_equipe[equipe_id]
                    
                    # Écrire l'email
                    email_text = self.format_email(equipe_id, ententes)
                    f.write(email_text)
                    
                    # Séparateur entre emails (sauf pour le dernier)
                    if idx < len(equipes_triees):
                        f.write("\n\n\n")
                
                # Footer du fichier
                f.write("\n\n")
                f.write("╔" + "═" * 78 + "╗\n")
                f.write("║" + " " * 78 + "║\n")
                f.write("║" + "  FIN DU FICHIER DE NOTIFICATIONS".center(78) + "║\n")
                f.write("║" + f"  Total: {len(equipes_triees)} équipes notifiées".center(78) + "║\n")
                f.write("║" + " " * 78 + "║\n")
                f.write("╚" + "═" * 78 + "╝\n")
            
            print(f"\n✅ Fichier de notifications généré: {output_file}")
            print(f"   {len(equipes_triees)} emails créés")
            
            return True
            
        except Exception as e:
            print(f"❌ Erreur lors de la génération du fichier: {e}")
            return False
    
    def generate_excel(self, output_path: str) -> bool:
        """
        Génère un fichier Excel avec la liste des ententes à organiser.
        Format identique au match_sheet_generator.
        
        Args:
            output_path: Chemin vers le fichier Excel de sortie
            
        Returns:
            True si succès, False sinon
        """
        if not self.ententes_par_equipe:
            print("ℹ️  Aucune entente trouvée - pas de fichier Excel à générer.")
            return True
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            
            # Créer le workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Matchs"
            
            # Définir les styles (identiques au match_sheet_generator)
            header_font = Font(bold=True, size=11)
            header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            cell_alignment = Alignment(horizontal="left", vertical="center")
            center_alignment = Alignment(horizontal="center", vertical="center")
            
            border_style = Side(style="thin", color="000000")
            border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
            
            # En-têtes (format match_sheet_generator)
            headers = ['Date', 'Sport', 'Sexe', 'Poule', 'Equipe 1', 'Equipe 2', 'Hre Déb', 'Lieu']
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Collecter toutes les ententes uniques
            ententes_uniques = set()
            ententes_data = []
            
            for equipe_id, ententes in self.ententes_par_equipe.items():
                for entente in ententes:
                    # Créer une clé unique pour éviter les doublons
                    match_key = tuple(sorted([equipe_id, entente['adversaire_id']]))
                    if match_key not in ententes_uniques:
                        ententes_uniques.add(match_key)
                        
                        # Extraire le genre depuis le code de poule (ex: VBFA1PA -> F, VBMA1PA -> M)
                        genre = 'M'  # Par défaut
                        code_poule = entente.get('poule', '')
                        if len(code_poule) >= 3:
                            genre = code_poule[2] if code_poule[2] in ['F', 'M'] else 'M'
                        
                        ententes_data.append({
                            'poule': entente['poule'],
                            'genre': genre,
                            'equipe1_id': equipe_id,
                            'equipe1_nom': equipe_id.split('|')[0] if '|' in equipe_id else equipe_id,
                            'equipe2_id': entente['adversaire_id'],
                            'equipe2_nom': entente['adversaire_nom']
                        })
            
            # Trier par genre (F avant M) puis par poule puis équipe1
            genre_order = {'F': 0, 'M': 1}
            ententes_data.sort(key=lambda x: (
                genre_order.get(x['genre'], 2),
                x['poule'],
                x['equipe1_nom']
            ))
            
            # Remplir les données
            row_num = 2
            for entente in ententes_data:
                # Date - vide
                ws.cell(row=row_num, column=1, value='').alignment = center_alignment
                ws.cell(row=row_num, column=1).border = border
                
                # Sport - VB
                ws.cell(row=row_num, column=2, value='VB').alignment = center_alignment
                ws.cell(row=row_num, column=2).border = border
                
                # Sexe
                ws.cell(row=row_num, column=3, value=entente['genre']).alignment = center_alignment
                ws.cell(row=row_num, column=3).border = border
                
                # Poule
                ws.cell(row=row_num, column=4, value=entente['poule']).alignment = cell_alignment
                ws.cell(row=row_num, column=4).border = border
                
                # Équipe 1
                ws.cell(row=row_num, column=5, value=entente['equipe1_nom']).alignment = cell_alignment
                ws.cell(row=row_num, column=5).border = border
                
                # Équipe 2
                ws.cell(row=row_num, column=6, value=entente['equipe2_nom']).alignment = cell_alignment
                ws.cell(row=row_num, column=6).border = border
                
                # Hre Déb - vide
                ws.cell(row=row_num, column=7, value='').alignment = center_alignment
                ws.cell(row=row_num, column=7).border = border
                
                # Lieu - ENTENTE
                ws.cell(row=row_num, column=8, value='ENTENTE').alignment = cell_alignment
                ws.cell(row=row_num, column=8).border = border
                
                row_num += 1
            
            # Ajuster les largeurs de colonnes (identiques au match_sheet_generator)
            ws.column_dimensions['A'].width = 12   # Date
            ws.column_dimensions['B'].width = 8    # Sport
            ws.column_dimensions['C'].width = 6    # Sexe
            ws.column_dimensions['D'].width = 12   # Poule
            ws.column_dimensions['E'].width = 15   # Equipe 1
            ws.column_dimensions['F'].width = 15   # Equipe 2
            ws.column_dimensions['G'].width = 10   # Hre Déb
            ws.column_dimensions['H'].width = 35   # Lieu
            
            # Sauvegarder
            wb.save(output_path)
            print(f"\n✅ Fichier Excel généré: {output_path}")
            print(f"   {len(ententes_data)} ententes listées")
            
            # Afficher les équipes sans contacts
            if self.equipes_sans_contacts:
                print(f"\n⚠️  {len(self.equipes_sans_contacts)} équipes sans contacts:")
                for equipe in self.equipes_sans_contacts[:10]:  # Limiter à 10
                    print(f"   - {equipe['nom']}")
                if len(self.equipes_sans_contacts) > 10:
                    print(f"   ... et {len(self.equipes_sans_contacts) - 10} autres")
            
            return True
            
        except Exception as e:
            print(f"❌ Erreur lors de la génération du fichier Excel: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def run(self, output_path: str, excel_path: str = None, test_mode: bool = False) -> bool:
        """
        Execute le processus complet de génération.
        
        Args:
            output_path: Chemin vers le fichier de sortie texte
            excel_path: Chemin vers le fichier Excel de sortie (optionnel)
            test_mode: Si True, n'écrit pas les fichiers
            
        Returns:
            True si succès, False sinon
        """
        print("\n🚀 Génération des notifications d'ententes\n")
        
        # Charger la solution
        if not self.load_solution():
            return False
        
        # Charger les contacts
        if not self.load_contacts():
            print("\n⚠️  Continuons sans le fichier config (infos limitées)")
        
        # Extraire les ententes
        if not self.extract_ententes():
            return False
        
        if not self.ententes_par_equipe:
            print("\n✅ Aucune entente trouvée - pas de notification à générer")
            return True
        
        # Mode test
        if test_mode:
            print("\n🧪 Mode test activé - pas de génération de fichiers")
            print("\nAperçu d'un email:")
            print("-" * 80)
            premier_equipe = list(self.ententes_par_equipe.keys())[0]
            print(self.format_email(premier_equipe, self.ententes_par_equipe[premier_equipe]))
            
            # Afficher les équipes sans contacts
            if self.equipes_sans_contacts:
                print(f"\n⚠️  {len(self.equipes_sans_contacts)} équipes sans contacts détectées")
            
            return True
        
        # Générer le fichier Excel si demandé
        if excel_path:
            if not self.generate_excel(excel_path):
                print("⚠️  Échec de génération du fichier Excel, continuons...")
        
        # Générer les notifications texte
        return self.generate_notifications(output_path)


def main():
    """Point d'entrée principal du script."""
    parser = argparse.ArgumentParser(
        description="Génère les notifications d'ententes depuis une solution de calendrier",
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    
    parser.add_argument(
        '--solution',
        default='solutions/latest_volley.json',
        help='Chemin vers le fichier solution JSON (défaut: solutions/latest_volley.json)'
    )
    
    parser.add_argument(
        '--config',
        help='Chemin vers le fichier config Excel (défaut: auto-détection depuis examples/volleyball/config_volley.xlsx)'
    )
    
    parser.add_argument(
        '--output',
        default='notifications_ententes.txt',
        help='Chemin vers le fichier de sortie texte (défaut: notifications_ententes.txt)'
    )
    
    parser.add_argument(
        '--excel',
        default='ententes.xlsx',
        help='Chemin vers le fichier Excel de sortie (défaut: ententes.xlsx)'
    )
    
    parser.add_argument(
        '--deadline',
        help='Date limite pour organiser les ententes (format: JJ/MM/AAAA)'
    )
    
    parser.add_argument(
        '--test',
        action='store_true',
        help='Mode test : affiche seulement les statistiques sans générer les fichiers'
    )
    
    args = parser.parse_args()
    
    # Créer le générateur
    generator = EntenteNotificationGenerator(
        solution_path=args.solution,
        config_path=args.config,
        deadline_date=args.deadline
    )
    
    # Exécuter
    success = generator.run(args.output, excel_path=args.excel, test_mode=args.test)
    
    if success:
        print("\n✨ Génération terminée avec succès!\n")
        sys.exit(0)
    else:
        print("\n❌ Erreur lors de la génération\n")
        sys.exit(1)


if __name__ == '__main__':
    main()
