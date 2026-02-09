#!/usr/bin/env python3
"""
Générateur de présentation des poules au format Excel.

Ce script génère un fichier Excel présentant les poules de manière structurée,
avec le même format que PoulesVB.xlsx:
- Organisation par genre (Féminin puis Masculin)
- Colonnes par niveau (A1, A2, A3, A4)
- Équipes avec numéro et horaire préféré
- Type de championnat (Classique ou Aller-Retour)
- Surlignage bleu clair des équipes avec horaires aménagés

Usage:
    python scripts/generate_pools_presentation.py --sport volley
    python scripts/generate_pools_presentation.py --config configs/config_volley.yaml
    python scripts/generate_pools_presentation.py --solution solutions/latest_volley.json

Sources de données:
    - Fichier JSON de solution exporté (entités avec poules et équipes)
    - Fichier de configuration Excel (feuilles Equipes, Dispos_Gymnases_Equipes)
"""

import sys
import json
import argparse
from pathlib import Path
from typing import Dict, List, Optional, Any, Tuple
from dataclasses import dataclass, field
from datetime import datetime
from collections import defaultdict

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Setup path pour imports
SCRIPTS_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPTS_DIR.parent
SRC_DIR = PROJECT_ROOT / "src"

for _path in [str(PROJECT_ROOT), str(SRC_DIR)]:
    if _path not in sys.path:
        sys.path.insert(0, _path)

from scripts.script_base import ScriptContext, Sport, create_base_parser, SOLUTIONS_DIR, DATA_DIR


# ============================================================================
# CONSTANTES DE STYLE
# ============================================================================

# Couleurs
BLUE_HEADER = "FF366092"          # Bleu foncé pour titre et niveaux
RED_FEMININE = "FFFF0000"         # Rouge pour section Féminin
BLUE_MASCULINE = "FF1E90FF"       # Bleu pour section Masculin
GRAY_POOL = "FF808080"            # Gris pour noms des poules
WHITE_TEXT = "FFFFFFFF"           # Texte blanc
BLACK_TEXT = "FF000000"           # Texte noir
LIGHT_BLUE_HIGHLIGHT = "FFADD8E6" # Bleu clair pour horaires aménagés

# Bordures
THIN_BORDER = Side(style='thin', color='FF000000')
MEDIUM_BORDER = Side(style='medium', color='FF000000')
THICK_BORDER = Side(style='thick', color='FF000000')

# Dimensions
COL_NUM_WIDTH = 4.0               # Colonne numéro
COL_TEAM_WIDTH = 18.0             # Colonne équipe
COL_HOUR_WIDTH = 12.0             # Colonne horaire


# ============================================================================
# CLASSES DE DONNÉES
# ============================================================================

@dataclass
class TeamInfo:
    """Information sur une équipe."""
    nom: str
    nom_complet: str
    genre: str
    poule: str
    horaire_prefere: str
    institution: str = ""
    has_amenaged_schedule: bool = False  # Horaire aménagé selon gymnase


@dataclass
class PoolInfo:
    """Information sur une poule."""
    nom: str
    genre: str
    niveau: str
    type_championnat: str  # "Classique" ou "Aller-Retour"
    equipes: List[TeamInfo] = field(default_factory=list)


@dataclass 
class PoolsData:
    """Données complètes des poules."""
    sport_name: str
    sport_emoji: str
    phase: str
    date_generation: str
    poules_by_genre_niveau: Dict[Tuple[str, str], List[PoolInfo]] = field(default_factory=dict)
    equipes_amenagees: set = field(default_factory=set)  # Équipes avec horaires aménagés


# ============================================================================
# CHARGEMENT DES DONNÉES
# ============================================================================

def load_from_json(solution_path: Path) -> PoolsData:
    """Charge les données depuis un fichier JSON de solution."""
    with open(solution_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # Extraire les infos du sport
    sport_info = data.get('sport', {})
    sport_name = sport_info.get('name', 'Sport')
    sport_emoji = sport_info.get('emoji', '🏅')
    
    # Extraire la date
    generated_at = data.get('generated_at', datetime.now().isoformat())
    if 'T' in generated_at:
        date_str = generated_at.split('T')[0]
        try:
            date_obj = datetime.strptime(date_str, '%Y-%m-%d')
            date_generation = date_obj.strftime('%d/%m/%y')
        except:
            date_generation = date_str
    else:
        date_generation = generated_at
    
    # Déterminer la phase
    config = data.get('config', {})
    phase = config.get('phase', 'Phase 1')
    if not phase:
        phase = "Phase 1"
    
    # Charger les équipes
    entities = data.get('entities', {})
    equipes_data = entities.get('equipes', [])
    poules_data = entities.get('poules', [])
    
    # Créer un dictionnaire équipe_id -> TeamInfo
    equipes_dict: Dict[str, TeamInfo] = {}
    for eq in equipes_data:
        eq_id = eq.get('id', '')
        nom = eq.get('nom', eq_id)
        nom_complet = eq.get('nom_complet', nom)
        genre = eq.get('genre', 'M')
        poule = eq.get('poule', '')
        
        horaires = eq.get('horaires_preferes', [])
        horaire_prefere = horaires[0] if horaires else '14:00'
        # Formater l'horaire (14:00 -> 14H)
        if ':' in horaire_prefere:
            h, m = horaire_prefere.split(':')
            horaire_prefere = f"{h}H"
        
        institution = eq.get('institution', '')
        
        equipes_dict[eq_id] = TeamInfo(
            nom=nom,
            nom_complet=nom_complet,
            genre=genre,
            poule=poule,
            horaire_prefere=horaire_prefere,
            institution=institution,
            has_amenaged_schedule=False  # Pas de données d'aménagement dans JSON
        )
    
    # Construire les poules
    poules_by_genre_niveau: Dict[Tuple[str, str], List[PoolInfo]] = defaultdict(list)
    
    for p in poules_data:
        nom = p.get('nom', p.get('id', ''))
        genre = p.get('genre', 'M')
        niveau = p.get('niveau', 'A1')
        type_champ = p.get('type', 'Classique')
        equipes_ids = p.get('equipes_ids', [])
        
        # Récupérer les infos des équipes
        equipes: List[TeamInfo] = []
        for eq_id in equipes_ids:
            if eq_id in equipes_dict:
                equipes.append(equipes_dict[eq_id])
        
        pool = PoolInfo(
            nom=nom,
            genre=genre,
            niveau=niveau,
            type_championnat=type_champ,
            equipes=equipes
        )
        
        poules_by_genre_niveau[(genre, niveau)].append(pool)
    
    # Trier les poules par nom (ordre alphabétique)
    for key in poules_by_genre_niveau:
        poules_by_genre_niveau[key].sort(key=lambda p: p.nom)
    
    return PoolsData(
        sport_name=sport_name,
        sport_emoji=sport_emoji,
        phase=phase,
        date_generation=date_generation,
        poules_by_genre_niveau=dict(poules_by_genre_niveau),
        equipes_amenagees=set()
    )


def load_from_excel(config_path: Path, excel_path: Path) -> PoolsData:
    """Charge les données depuis un fichier de configuration Excel."""
    import yaml
    
    # Charger la config YAML pour les infos du sport
    with open(config_path, 'r', encoding='utf-8') as f:
        yaml_config = yaml.safe_load(f)
    
    sport_type = yaml_config.get('sport_type', 'volleyball')
    sport = Sport.from_type(sport_type)
    
    # Lire les feuilles Excel
    df_equipes = pd.read_excel(excel_path, sheet_name='Equipes')
    
    # Charger les équipes avec horaires aménagés
    equipes_amenagees = set()
    try:
        df_dispos = pd.read_excel(excel_path, sheet_name='Dispos_Gymnases_Equipes')
        if not df_dispos.empty:
            for _, row in df_dispos.iterrows():
                equipe = str(row.get('Equipe', '')).strip()
                genre = str(row.get('Genre', '')).strip()
                if equipe and genre:
                    # Clé unique: nom|genre
                    equipes_amenagees.add(f"{equipe}|{genre}")
    except Exception as e:
        print(f"Note: Feuille Dispos_Gymnases_Equipes non trouvée ou vide: {e}")
    
    # Construire les poules
    poules_by_genre_niveau: Dict[Tuple[str, str], List[PoolInfo]] = defaultdict(list)
    equipes_by_poule: Dict[str, List[TeamInfo]] = defaultdict(list)
    
    for _, row in df_equipes.iterrows():
        nom = str(row.get('Equipe', '')).strip()
        if not nom:
            continue
        
        niveau = str(row.get('Niveau_Equipe', 'A1')).strip()
        genre = str(row.get('Genre_Equipe', 'M')).strip()
        poule = str(row.get('Poule', '')).strip()
        
        horaire = row.get('Horaire_Prefere', '14:00')
        if pd.isna(horaire):
            horaire = '14:00'
        horaire = str(horaire)
        
        # Formater l'horaire (14:00 -> 14H)
        if ':' in horaire:
            h, m = horaire.split(':')
            horaire = f"{h}H"
        elif 'H' not in horaire.upper():
            horaire = f"{horaire}H"
        
        # Vérifier si l'équipe a un horaire aménagé
        cle_equipe = f"{nom}|{genre}"
        has_amenaged = cle_equipe in equipes_amenagees
        
        team = TeamInfo(
            nom=nom,
            nom_complet=nom,
            genre=genre,
            poule=poule,
            horaire_prefere=horaire,
            has_amenaged_schedule=has_amenaged
        )
        
        equipes_by_poule[poule].append(team)
    
    # Charger les types de poules
    types_poules: Dict[str, str] = {}
    try:
        df_types = pd.read_excel(excel_path, sheet_name='Types_Poules')
        for _, row in df_types.iterrows():
            poule_nom = str(row.get('Poule', '')).strip()
            type_champ = str(row.get('Type_Championnat', 'Classique')).strip()
            if poule_nom:
                types_poules[poule_nom] = type_champ
    except Exception:
        pass
    
    # Créer les objets PoolInfo
    for poule_nom, equipes in equipes_by_poule.items():
        if not equipes:
            continue
        
        genre = equipes[0].genre
        niveau = equipes[0].poule[2:4] if len(equipes[0].poule) >= 4 else 'A1'
        # Extraire le niveau depuis le nom de poule (ex: VBFA1PA -> A1)
        if len(poule_nom) >= 4:
            niveau = poule_nom[3:5] if poule_nom[3].isalpha() else poule_nom[2:4]
        
        type_champ = types_poules.get(poule_nom, 'Classique')
        
        pool = PoolInfo(
            nom=poule_nom,
            genre=genre,
            niveau=niveau,
            type_championnat=type_champ,
            equipes=equipes
        )
        
        poules_by_genre_niveau[(genre, niveau)].append(pool)
    
    # Trier les poules
    for key in poules_by_genre_niveau:
        poules_by_genre_niveau[key].sort(key=lambda p: p.nom)
    
    date_generation = datetime.now().strftime('%d/%m/%y')
    
    return PoolsData(
        sport_name=sport.name,
        sport_emoji=sport.emoji,
        phase="Phase 1",
        date_generation=date_generation,
        poules_by_genre_niveau=dict(poules_by_genre_niveau),
        equipes_amenagees=equipes_amenagees
    )


# ============================================================================
# GÉNÉRATION DU FICHIER EXCEL
# ============================================================================

class PoolsExcelGenerator:
    """Générateur de fichier Excel pour la présentation des poules."""
    
    def __init__(self, data: PoolsData):
        self.data = data
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        
        # Définir les niveaux à afficher
        self.niveaux = ['A1', 'A2', 'A3', 'A4']
        
        # Position courante d'écriture
        self.current_row = 1
    
    def generate(self, output_path: Path):
        """Génère le fichier Excel complet."""
        # Configurer la feuille
        sheet_name = f"Poules {self.data.sport_name.upper()[:2]} {self.data.phase}"
        self.ws.title = sheet_name[:31]  # Limite Excel sur le nom
        
        # Définir les largeurs de colonnes
        self._setup_column_widths()
        
        # Écrire le titre principal
        self._write_main_title()
        
        # Ligne vide
        self.current_row += 1
        
        # Écrire les en-têtes de niveaux
        self._write_level_headers()
        
        # Ligne vide
        self.current_row += 1
        
        # Écrire la section Féminin
        self._write_gender_section('F', 'FÉMININ', RED_FEMININE)
        
        # Lignes vides entre sections
        self.current_row += 2
        
        # Écrire la section Masculin
        self._write_gender_section('M', 'MASCULIN', BLUE_MASCULINE)
        
        # Ajouter le total des équipes
        self._write_total()
        
        # Sauvegarder
        self.wb.save(output_path)
        print(f"✅ Fichier généré: {output_path}")
    
    def _setup_column_widths(self):
        """Configure les largeurs des colonnes."""
        # Structure: [num, team, hour] x 4 niveaux
        cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']
        widths = [COL_NUM_WIDTH, COL_TEAM_WIDTH, COL_HOUR_WIDTH,
                  COL_NUM_WIDTH + 1.5, COL_TEAM_WIDTH, COL_HOUR_WIDTH,
                  COL_NUM_WIDTH + 2, COL_TEAM_WIDTH, COL_HOUR_WIDTH,
                  COL_NUM_WIDTH + 2, COL_TEAM_WIDTH, COL_HOUR_WIDTH]
        
        for col, width in zip(cols, widths):
            self.ws.column_dimensions[col].width = width
    
    def _write_main_title(self):
        """Écrit le titre principal du document."""
        title = f"{self.data.sport_name.upper()} - {self.data.phase.upper()} - POULES 2025-2026 ({self.data.date_generation})"
        
        cell = self.ws.cell(row=self.current_row, column=1, value=title)
        cell.font = Font(name='Arial', size=16, bold=True, color=WHITE_TEXT)
        cell.fill = PatternFill(start_color=BLUE_HEADER, end_color=BLUE_HEADER, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Fusionner les cellules A1:L1
        self.ws.merge_cells(f'A{self.current_row}:L{self.current_row}')
        self.ws.row_dimensions[self.current_row].height = 34.5
        
        self.current_row += 1
    
    def _write_level_headers(self):
        """Écrit les en-têtes des niveaux (A1, A2, A3, A4)."""
        # Position des colonnes pour chaque niveau
        level_cols = [('B', 'C'), ('E', 'F'), ('H', 'I'), ('K', 'L')]
        
        for (col_start, col_end), niveau in zip(level_cols, self.niveaux):
            cell = self.ws.cell(row=self.current_row, column=ord(col_start) - ord('A') + 1, 
                               value=f'NIVEAU {niveau}')
            cell.font = Font(name='Arial', size=14, bold=True, color=WHITE_TEXT)
            cell.fill = PatternFill(start_color=BLUE_HEADER, end_color=BLUE_HEADER, fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=MEDIUM_BORDER)
            
            self.ws.merge_cells(f'{col_start}{self.current_row}:{col_end}{self.current_row}')
        
        self.ws.row_dimensions[self.current_row].height = 27.75
        self.current_row += 1
    
    def _write_gender_section(self, genre_code: str, genre_label: str, color: str):
        """Écrit une section de genre (Féminin ou Masculin)."""
        # Titre de la section
        sport_prefix = self.data.sport_name.upper()[:6]
        cell = self.ws.cell(row=self.current_row, column=1, value=f'{sport_prefix} {genre_label}')
        cell.font = Font(name='Arial', size=18, bold=True, color=WHITE_TEXT)
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        self.ws.merge_cells(f'A{self.current_row}:L{self.current_row}')
        self.ws.row_dimensions[self.current_row].height = 30.0
        
        self.current_row += 2  # Ligne vide après le titre
        
        # Récupérer les poules par niveau
        poules_par_niveau: Dict[str, List[PoolInfo]] = {}
        for niveau in self.niveaux:
            key = (genre_code, niveau)
            poules_par_niveau[niveau] = self.data.poules_by_genre_niveau.get(key, [])
        
        # Écrire les poules de manière indépendante par colonne
        # On garde un tracker de la position actuelle pour chaque niveau
        self._write_pools_independently(poules_par_niveau)
        
        # Ajouter le compte des équipes
        total_equipes = sum(
            len(p.equipes) 
            for pools in poules_par_niveau.values() 
            for p in pools
        )
        
        cell = self.ws.cell(row=self.current_row, column=2, 
                           value=f'{total_equipes} Equipes J{"F" if genre_code == "F" else "G"}')
        cell.font = Font(name='Calibri', size=11, bold=True)
        self.current_row += 1
    
    
    def _write_pools_independently(self, poules_par_niveau: Dict[str, List[PoolInfo]]):
        """Écrit les poules de manière indépendante pour chaque niveau."""
        # Colonnes pour chaque niveau
        level_cols = {
            'A1': (1, 2, 3),    # A, B, C
            'A2': (4, 5, 6),    # D, E, F
            'A3': (7, 8, 9),    # G, H, I
            'A4': (10, 11, 12)  # J, K, L
        }
        
        # Position de départ pour chaque niveau (ligne courante)
        positions = {niveau: self.current_row for niveau in self.niveaux}
        
        # Écrire chaque poule pour chaque niveau
        for niveau in self.niveaux:
            pools = poules_par_niveau.get(niveau, [])
            col_num, col_team, col_hour = level_cols[niveau]
            current_pos = positions[niveau]
            
            for pool in pools:
                # Écrire le nom de la poule
                cell = self.ws.cell(row=current_pos, column=col_team, value=pool.nom)
                cell.font = Font(name='Arial', size=12, bold=True, color=WHITE_TEXT)
                cell.fill = PatternFill(start_color=GRAY_POOL, end_color=GRAY_POOL, fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=MEDIUM_BORDER,
                    right=MEDIUM_BORDER,
                    top=MEDIUM_BORDER,
                    bottom=MEDIUM_BORDER
                )
                self.ws.merge_cells(
                    start_row=current_pos, start_column=col_team,
                    end_row=current_pos, end_column=col_hour
                )
                self.ws.row_dimensions[current_pos].height = 15.75
                current_pos += 1
                
                # Écrire les équipes
                nb_equipes = len(pool.equipes)
                for eq_idx, equipe in enumerate(pool.equipes):
                    is_first = (eq_idx == 0)
                    is_last = (eq_idx == nb_equipes - 1)
                    
                    # Numéro
                    cell_num = self.ws.cell(row=current_pos, column=col_num, value=eq_idx + 1)
                    cell_num.font = Font(name='Calibri', size=10, bold=True)
                    
                    # Nom de l'équipe
                    cell_team = self.ws.cell(row=current_pos, column=col_team, value=equipe.nom)
                    cell_team.font = Font(name='Arial', size=11, bold=True, color=BLACK_TEXT)
                    cell_team.alignment = Alignment(horizontal='center', vertical='center')
                    cell_team.border = Border(
                        left=THIN_BORDER,
                        right=THIN_BORDER,
                        top=None if is_first else THIN_BORDER,
                        bottom=None if is_last else THIN_BORDER
                    )
                    
                    # Horaire
                    cell_hour = self.ws.cell(row=current_pos, column=col_hour, value=equipe.horaire_prefere)
                    cell_hour.font = Font(name='Arial', size=11, bold=True, color=BLACK_TEXT)
                    cell_hour.alignment = Alignment(horizontal='center', vertical='center')
                    cell_hour.border = Border(
                        left=THIN_BORDER,
                        right=THIN_BORDER,
                        top=None if is_first else THIN_BORDER,
                        bottom=None if is_last else THIN_BORDER
                    )
                    
                    # Surligner en bleu clair si horaire aménagé
                    if equipe.has_amenaged_schedule:
                        light_blue_fill = PatternFill(start_color=LIGHT_BLUE_HIGHLIGHT, 
                                                       end_color=LIGHT_BLUE_HIGHLIGHT, 
                                                       fill_type='solid')
                        cell_team.fill = light_blue_fill
                        cell_hour.fill = light_blue_fill
                    
                    self.ws.row_dimensions[current_pos].height = 15.75
                    current_pos += 1
                
                # Ligne CHAMPIONNAT
                label = "CHAMPIONNAT AR" if pool.type_championnat.strip().lower() in ["aller-retour", "ar"] else "CHAMPIONNAT"
                cell = self.ws.cell(row=current_pos, column=col_team, value=label)
                cell.font = Font(name='Calibri', size=11, bold=True)
                cell.fill = PatternFill(start_color='EEECE1', end_color='EEECE1', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
                cell.border = Border(
                    left=THIN_BORDER,
                    right=THIN_BORDER,
                    top=THIN_BORDER,
                    bottom=THIN_BORDER
                )
                self.ws.merge_cells(
                    start_row=current_pos, start_column=col_team,
                    end_row=current_pos, end_column=col_hour
                )
                self.ws.row_dimensions[current_pos].height = 15.75
                current_pos += 1
                
                # Ligne vide entre deux poules
                current_pos += 1
            
            # Mettre à jour la position finale pour ce niveau
            positions[niveau] = current_pos
        
        # Mettre à jour current_row à la position maximale parmi tous les niveaux
        self.current_row = max(positions.values())
    
    def _write_total(self):
        """Écrit le total des équipes."""
        total = sum(
            len(p.equipes)
            for pools in self.data.poules_by_genre_niveau.values()
            for p in pools
        )
        
        cell = self.ws.cell(row=self.current_row, column=2, value=f'Total équipes : {total}')
        cell.font = Font(name='Calibri', size=11, bold=True)
        self.current_row += 1


# ============================================================================
# POINT D'ENTRÉE
# ============================================================================

def create_parser():
    """Crée le parser d'arguments."""
    parser = argparse.ArgumentParser(
        description="Génère un fichier Excel de présentation des poules.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Exemples:
  %(prog)s --sport volley
  %(prog)s --config configs/config_volley.yaml
  %(prog)s --solution solutions/latest_volley.json
  %(prog)s --sport volley --output poules_volley.xlsx
        """
    )
    
    parser.add_argument(
        '--sport', '-s',
        type=str,
        help="Type de sport (volley, hand, basket)"
    )
    
    parser.add_argument(
        '--config', '-c',
        type=str,
        help="Chemin vers le fichier de configuration YAML"
    )
    
    parser.add_argument(
        '--solution', '-j',
        type=str,
        help="Chemin vers le fichier JSON de solution"
    )
    
    parser.add_argument(
        '--output', '-o',
        type=str,
        help="Chemin du fichier Excel de sortie"
    )
    
    parser.add_argument(
        '--verbose', '-v',
        action='store_true',
        help="Affiche plus de détails"
    )
    
    return parser


def main():
    """Point d'entrée principal."""
    parser = create_parser()
    args = parser.parse_args()
    
    # Déterminer la source des données
    solution_path = None
    config_path = None
    excel_path = None
    
    if args.solution:
        solution_path = Path(args.solution)
        if not solution_path.exists():
            print(f"❌ Fichier solution introuvable: {solution_path}")
            sys.exit(1)
    
    elif args.config:
        config_path = Path(args.config)
        if not config_path.exists():
            print(f"❌ Fichier config introuvable: {config_path}")
            sys.exit(1)
        
        # Charger le YAML pour trouver le fichier Excel
        import yaml
        with open(config_path, 'r', encoding='utf-8') as f:
            yaml_config = yaml.safe_load(f)
        
        # Chercher dans fichiers.donnees ou fichier_donnees
        excel_file = None
        if 'fichiers' in yaml_config and 'donnees' in yaml_config['fichiers']:
            excel_file = yaml_config['fichiers']['donnees']
        elif 'fichier_donnees' in yaml_config:
            excel_file = yaml_config.get('fichier_donnees')
        
        if excel_file:
            excel_path = Path(excel_file)
            if not excel_path.is_absolute():
                # Essayer depuis le dossier parent du projet
                excel_path = PROJECT_ROOT / excel_file
                if not excel_path.exists():
                    excel_path = config_path.parent / excel_file
        
        if not excel_path or not excel_path.exists():
            print(f"❌ Fichier Excel introuvable: {excel_path}")
            sys.exit(1)
    
    elif args.sport:
        # Chercher le fichier latest_{sport}.json
        sport_pattern = args.sport.lower()
        if sport_pattern in ['volley', 'volleyball']:
            sport_pattern = 'volley'
        elif sport_pattern in ['hand', 'handball']:
            sport_pattern = 'hand'
        elif sport_pattern in ['basket', 'basketball']:
            sport_pattern = 'basket'
        
        solution_path = SOLUTIONS_DIR / f"latest_{sport_pattern}.json"
        if not solution_path.exists():
            print(f"❌ Fichier solution introuvable: {solution_path}")
            print(f"   Essayez: python scripts/generate_pools_presentation.py --solution <chemin>")
            sys.exit(1)
    
    else:
        # Chercher la solution la plus récente
        json_files = list(SOLUTIONS_DIR.glob("latest_*.json"))
        if json_files:
            solution_path = sorted(json_files, key=lambda f: f.stat().st_mtime, reverse=True)[0]
            print(f"📁 Utilisation de: {solution_path.name}")
        else:
            print("❌ Aucune solution trouvée. Spécifiez --sport, --config ou --solution")
            sys.exit(1)
    
    # Charger les données
    if solution_path:
        print(f"📖 Chargement depuis JSON: {solution_path}")
        data = load_from_json(solution_path)
    else:
        print(f"📖 Chargement depuis Excel: {excel_path}")
        data = load_from_excel(config_path, excel_path)
    
    # Afficher les stats
    if args.verbose:
        print(f"\n📊 Statistiques:")
        print(f"   Sport: {data.sport_name} {data.sport_emoji}")
        print(f"   Phase: {data.phase}")
        print(f"   Date: {data.date_generation}")
        
        total_poules = sum(len(pools) for pools in data.poules_by_genre_niveau.values())
        total_equipes = sum(len(p.equipes) for pools in data.poules_by_genre_niveau.values() for p in pools)
        print(f"   Poules: {total_poules}")
        print(f"   Équipes: {total_equipes}")
        
        if data.equipes_amenagees:
            print(f"   Équipes avec horaires aménagés: {len(data.equipes_amenagees)}")
    
    # Déterminer le fichier de sortie
    if args.output:
        output_path = Path(args.output)
    else:
        sport_short = data.sport_name[:2].upper()
        output_path = PROJECT_ROOT / "exports" / f"Poules{sport_short}.xlsx"
    
    # Créer le dossier exports si nécessaire
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Générer le fichier
    print(f"\n📝 Génération du fichier Excel...")
    generator = PoolsExcelGenerator(data)
    generator.generate(output_path)
    
    print(f"\n🎉 Terminé! Fichier créé: {output_path}")


if __name__ == "__main__":
    main()
