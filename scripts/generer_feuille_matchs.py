#!/usr/bin/env python3
"""
Génère une feuille Excel de matchs formatée pour une semaine donnée.

Ce script extrait les matchs d'une semaine depuis la solution JSON et génère
un fichier Excel formaté pour impression/partage avec le SIUAPS.

Usage:
    python scripts/generer_feuille_matchs.py --semaine 1
    python scripts/generer_feuille_matchs.py -w 2 --date-depart "23/10/25"
    python scripts/generer_feuille_matchs.py -w 3 --config configs/config_hand.yaml

Exemples:
    # Semaine 1 avec date auto-calculée
    python scripts/generer_feuille_matchs.py -w 1 --auto-date

    # Semaine 2 pour handball
    python scripts/generer_feuille_matchs.py -w 2 --config configs/config_hand.yaml

    # Fichier de sortie personnalisé
    python scripts/generer_feuille_matchs.py -w 3 -o matchs_s3.xlsx
"""

import sys
from pathlib import Path
from datetime import timedelta
from collections import defaultdict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Setup paths
PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))
sys.path.insert(0, str(PROJECT_ROOT / 'src'))

from scripts.script_base import (
    ScriptContext,
    create_base_parser,
    print_header,
    print_success,
    print_error,
    print_info,
)
from scripts.sport_utils import extraire_sport_code, extraire_genre_niveau
from pycalendar.core.constants import format_user_date, parse_user_date, DATE_USER_FORMAT_LABEL


# ==============================================================================
# Configuration
# ==============================================================================

# Mapping des codes gymnases vers noms complets SIUAPS
GYMNASES_SIUAPS = {
    "DESCARTES": "ENS DESCARTES",
    "ECL": "CENTRALE",
    "ESA": "GYMNASE ESA",
    "LYON 2 HC": "HALLE LYON 2",
    "LAENNEC": "HALLE - 3D",
    "BESSON": "HALLE - C.BESSON",
    "L. J. HAUT": "COMPET C (HAUT) - LEON JOUHAUX",
    "ENTPE": "ENTPE",
    "GRENOBLE": "GRENOBLE",
}

# Ordre de tri pour niveaux
NIVEAU_ORDER = {'A1': 0, 'A2': 1, 'A3': 2, 'A4': 3, 'EXT': 4}
GENRE_ORDER = {'F': 0, 'M': 1}


# ==============================================================================
# Fonctions de traitement des données
# ==============================================================================

def calculer_date_semaine(date_depart: str, numero_semaine: int) -> str:
    """Calcule la date correspondant à une semaine donnée."""
    date_obj = parse_user_date(date_depart)
    if not date_obj:
        raise ValueError(f"Date invalide '{date_depart}', attendu: {DATE_USER_FORMAT_LABEL}")
    date_cible = date_obj + timedelta(weeks=(numero_semaine - 1))
    return format_user_date(date_cible)


def compter_equipes_par_institution(solution: dict) -> dict:
    """Compte le nombre d'équipes par institution et genre."""
    equipes = solution['entities']['equipes']
    counts = defaultdict(lambda: defaultdict(int))
    for e in equipes:
        counts[e['institution']][e['genre']] += 1
    return counts


def simplifier_nom_equipe(nom_complet: str, institution: str, genre: str, counts: dict) -> str:
    """Simplifie le nom d'une équipe si c'est la seule de l'institution/genre."""
    if counts[institution][genre] == 1:
        return institution
    return nom_complet


def mapper_gymnase(code: str) -> str:
    """Convertit le code gymnase en nom complet SIUAPS."""
    return GYMNASES_SIUAPS.get(code, code)


def filtrer_matchs_par_semaine(solution: dict, numero_semaine: int, date_str: str) -> pd.DataFrame:
    """
    Filtre les matchs planifiés pour une semaine donnée et les formate.
    
    Returns:
        DataFrame avec les matchs filtrés et formatés, ou None si aucun match
    """
    matchs_planifies = solution['matches']['scheduled']
    matchs_semaine = [m for m in matchs_planifies if m.get('semaine') == numero_semaine]
    
    if not matchs_semaine:
        return None
    
    counts = compter_equipes_par_institution(solution)
    data = []
    
    for match in matchs_semaine:
        poule = match.get('poule', 'HORS_CHAMPIONNAT')
        championship_type = match.get('championship_type', 'Acad')
        
        # Déterminer si c'est un match CFE/CFU
        is_cfe_cfu = (
            championship_type in ['CFE', 'CFU'] or 
            (poule == '' and match.get('is_fixed', False))
        )
        
        # Détection CFE pour anciennes solutions
        if championship_type == 'Acad' and poule == '' and match.get('is_fixed', False):
            championship_type = 'CFE'
        
        # Genre et niveau
        genre = match.get('genre', match.get('equipe1_genre', 'M'))
        
        if poule and poule != 'HORS_CHAMPIONNAT' and not is_cfe_cfu:
            _, niveau = extraire_genre_niveau(poule)
            sport = extraire_sport_code(poule)
        else:
            niveau = 'EXT'
            sport = 'VB'
        
        # Gestion des équipes externes
        equipe1_is_externe = match['equipe1_nom_complet'] == 'EXTERNE'
        equipe2_is_externe = match['equipe2_nom_complet'] == 'EXTERNE'
        
        equipe1_nom = match['equipe1_nom'] if equipe1_is_externe else match['equipe1_nom_complet']
        equipe2_nom = match['equipe2_nom'] if equipe2_is_externe else match['equipe2_nom_complet']
        
        # Simplification des noms pour CFE/CFU
        if is_cfe_cfu and not (equipe1_is_externe or equipe2_is_externe):
            equipe1_nom = simplifier_nom_equipe(
                match['equipe1_nom_complet'], 
                match['equipe1_institution'], 
                match['equipe1_genre'],
                counts
            )
            equipe2_nom = simplifier_nom_equipe(
                match['equipe2_nom_complet'],
                match['equipe2_institution'],
                match['equipe2_genre'],
                counts
            )
            poule_affichee = championship_type
        elif is_cfe_cfu:
            poule_affichee = championship_type
        else:
            poule_affichee = poule if poule else 'HORS_CHAMPIONNAT'
        
        data.append({
            'Date': date_str,
            'Sport': sport,
            'Sexe': genre,
            'Poule': poule_affichee,
            'Equipe 1': equipe1_nom,
            'Equipe 2': equipe2_nom,
            'Hre Déb': match['horaire'],
            'Lieu': mapper_gymnase(match['gymnase']),
            '_niveau': niveau,
            '_is_cfe_cfu': is_cfe_cfu,
            '_priorite': 0 if is_cfe_cfu else 1,
        })
    
    df = pd.DataFrame(data)
    
    # Tri: genre (F avant M), priorité (CFE/CFU d'abord), niveau, horaire
    df['_genre_order'] = df['Sexe'].map(GENRE_ORDER).fillna(2)
    df['_niveau_order'] = df['_niveau'].map(NIVEAU_ORDER).fillna(5)
    
    df = df.sort_values(
        by=['_genre_order', '_priorite', '_niveau_order', 'Hre Déb'],
        ascending=[True, True, True, True]
    )
    
    # Nettoyage colonnes temporaires
    df = df.drop(columns=['_genre_order', '_niveau_order', '_niveau', '_priorite'])
    df = df.reset_index(drop=True)
    
    return df


# ==============================================================================
# Génération Excel
# ==============================================================================

def appliquer_mise_en_forme(worksheet, df: pd.DataFrame):
    """Applique la mise en forme au fichier Excel."""
    # Styles
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    cell_alignment = Alignment(horizontal="left", vertical="center")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000")
    )
    
    # En-têtes
    for col_num, column_title in enumerate(df.columns, 1):
        if column_title == '_is_cfe_cfu':
            continue
        cell = worksheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Données
    for row_num in range(2, len(df) + 2):
        is_cfe_cfu = df.iloc[row_num - 2].get('_is_cfe_cfu', False)
        
        for col_num in range(1, len(df.columns) + 1):
            col_name = df.columns[col_num - 1]
            if col_name == '_is_cfe_cfu':
                continue
            
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.border = border
            
            if is_cfe_cfu:
                cell.fill = yellow_fill
            
            # Colonnes centrées: Date, Sport, Sexe, Hre Déb
            if col_num in [1, 2, 3, 7]:
                cell.alignment = center_alignment
            else:
                cell.alignment = cell_alignment
    
    # Largeurs colonnes
    widths = {'A': 12, 'B': 8, 'C': 6, 'D': 12, 'E': 15, 'F': 15, 'G': 10, 'H': 35}
    for col, width in widths.items():
        worksheet.column_dimensions[col].width = width


def generer_feuille_excel(df: pd.DataFrame, fichier_sortie: Path):
    """Génère un fichier Excel avec mise en forme."""
    df_export = df.drop(columns=['_is_cfe_cfu'], errors='ignore')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Matchs"
    
    # En-têtes
    for col_num, column_title in enumerate(df_export.columns, 1):
        ws.cell(row=1, column=col_num, value=column_title)
    
    # Données
    for row_num, row_data in enumerate(df_export.itertuples(index=False), 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    appliquer_mise_en_forme(ws, df)
    wb.save(fichier_sortie)
    
    return len(df)


# ==============================================================================
# Main
# ==============================================================================

def main():
    parser = create_base_parser(
        description="Génère une feuille Excel de matchs formatée pour une semaine donnée"
    )
    
    parser.add_argument(
        '-w', '--semaine',
        type=int,
        required=True,
        help='Numéro de la semaine (1, 2, 3, ...)'
    )
    
    parser.add_argument(
        '-d', '--date-depart',
        type=str,
        default="22/01/26",
        help='Date de départ pour la semaine 1 (format: DD/MM/YY). Par défaut: 22/01/26'
    )
    
    parser.add_argument(
        '--auto-date',
        action='store_true',
        help='Calcule automatiquement la date en fonction de la semaine'
    )
    
    parser.add_argument(
        '-o', '--output',
        type=str,
        help='Fichier de sortie (défaut: Matchs_Semaine_X_DATE.xlsx)'
    )
    
    parser.epilog = """
Exemples:
    python scripts/generer_feuille_matchs.py -w 1
    python scripts/generer_feuille_matchs.py -w 2 --auto-date
    python scripts/generer_feuille_matchs.py -w 3 --config configs/config_hand.yaml
    python scripts/generer_feuille_matchs.py -w 1 -o matchs_semaine_1.xlsx
    """
    
    args = parser.parse_args()
    
    # Créer le contexte
    try:
        ctx = ScriptContext.from_args(args)
    except FileNotFoundError as e:
        print_error(str(e))
        return 1
    
    # Afficher le header
    print_header(f"Génération feuille de matchs {ctx.sport.name}", ctx.sport.emoji)
    ctx.print_status()
    
    # Vérifier qu'on a une solution
    if not ctx.solution_path or not ctx.solution_path.exists():
        print_error("Aucune solution trouvée")
        print_info("Spécifiez une solution avec --solution ou --config")
        return 1
    
    # Calculer la date
    try:
        if args.auto_date:
            date_str = calculer_date_semaine(args.date_depart, args.semaine)
        else:
            date_str = args.date_depart
    except ValueError as e:
        print_error(str(e))
        return 1
    
    print()
    print_info(f"Semaine: {args.semaine}")
    print_info(f"Date: {date_str}")
    
    # Charger la solution et filtrer les matchs
    try:
        solution = ctx.solution_data
        df_matchs = filtrer_matchs_par_semaine(solution, args.semaine, date_str)
    except Exception as e:
        print_error(f"Erreur lors du chargement: {e}")
        if args.verbose:
            import traceback
            traceback.print_exc()
        return 1
    
    if df_matchs is None or df_matchs.empty:
        print_error(f"Aucun match trouvé pour la semaine {args.semaine}")
        return 1
    
    # Fichier de sortie
    if args.output:
        fichier_sortie = Path(args.output)
        if not fichier_sortie.is_absolute():
            fichier_sortie = PROJECT_ROOT / args.output
    else:
        date_clean = date_str.replace('/', '-')
        fichier_sortie = PROJECT_ROOT / f"Matchs_Semaine_{args.semaine}_{date_clean}.xlsx"
    
    fichier_sortie.parent.mkdir(parents=True, exist_ok=True)
    
    # Générer le fichier Excel
    nb_matchs = generer_feuille_excel(df_matchs, fichier_sortie)
    
    print()
    print_success(f"Fichier généré: {fichier_sortie.name}")
    print()
    print(f"📊 Résumé:")
    print(f"   Matchs féminins : {len(df_matchs[df_matchs['Sexe'] == 'F'])}")
    print(f"   Matchs masculins : {len(df_matchs[df_matchs['Sexe'] == 'M'])}")
    print(f"   Lieux uniques : {df_matchs['Lieu'].nunique()}")
    print(f"   Total: {nb_matchs} matchs")
    
    return 0


if __name__ == "__main__":
    sys.exit(main())
