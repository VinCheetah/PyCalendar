#!/usr/bin/env python3
"""
Import des matchs depuis une solution calculée vers la feuille Matchs_Fixes.

Ce script permet d'importer tout ou une partie des matchs d'une solution 
calculée par PyCalendar vers la feuille Matchs_Fixes du fichier Excel de 
configuration. Il gère intelligemment les doublons et les types de poules.

Usage:
    # Importer tous les matchs de la solution par défaut
    python scripts/import_solution_to_fixed.py --config configs/config_volley.yaml
    
    # Importer uniquement les matchs des semaines 1 à 5
    python scripts/import_solution_to_fixed.py -c configs/config_volley.yaml --semaines 1-5
    
    # Importer une solution spécifique
    python scripts/import_solution_to_fixed.py -s solutions/solution_volley_2026-01-27.json
    
    # Importer uniquement certaines poules
    python scripts/import_solution_to_fixed.py -c configs/config_volley.yaml --poules VBFA1PA,VBMA1PA
    
    # Importer uniquement les matchs féminins
    python scripts/import_solution_to_fixed.py -c configs/config_volley.yaml --genre F
    
    # Simulation (dry-run)
    python scripts/import_solution_to_fixed.py -c configs/config_volley.yaml --dry-run

Exemples avancés:
    # Remplacer les matchs existants (au lieu de fusionner)
    python scripts/import_solution_to_fixed.py -c configs/config_volley.yaml --remplacer
    
    # Exclure les matchs CFE/CFU (importer uniquement Acad)
    python scripts/import_solution_to_fixed.py -c configs/config_volley.yaml --type Acad
    
    # Inclure les matchs en entente
    python scripts/import_solution_to_fixed.py -c configs/config_volley.yaml --inclure-ententes
"""

import sys
from pathlib import Path
import argparse
import json
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Set, Tuple

import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

# Setup paths
PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))
sys.path.insert(0, str(PROJECT_ROOT / 'src'))

from scripts.script_base import (
    ScriptContext,
    create_base_parser,
    print_header,
    print_success,
    print_error,
    print_warning,
    print_info,
    SOLUTIONS_DIR,
    CONFIGS_DIR,
    PROJECT_ROOT,
)


# ==============================================================================
# Constantes
# ==============================================================================

COLONNES_MATCHS_FIXES = [
    'Equipe_1', 'Equipe_2', 'Genre', 'Poule', 'Semaine', 'Date',
    'Horaire', 'Gymnase', 'Score', 'Type_Competition', 'Remarques', 'Arbitres'
]

IMPORT_TAG = "[Import Solution]"


# ==============================================================================
# Fonctions utilitaires
# ==============================================================================

def parse_semaines_arg(arg: str) -> List[int]:
    """
    Parse un argument de semaines (ex: "1-5", "3,5,7", "1-3,6,8-10").
    
    Args:
        arg: Chaîne représentant les semaines
        
    Returns:
        Liste des numéros de semaines
    """
    semaines = set()
    for part in arg.split(','):
        part = part.strip()
        if '-' in part:
            debut, fin = part.split('-')
            semaines.update(range(int(debut), int(fin) + 1))
        else:
            semaines.add(int(part))
    return sorted(semaines)


def creer_cle_match(equipe1: str, equipe2: str, genre: str, poule: str = '', 
                    semaine: Optional[int] = None, is_aller_retour: bool = False) -> str:
    """
    Crée une clé unique pour identifier un match.
    
    Pour les poules Classiques, la clé est symétrique (A vs B = B vs A).
    Pour les poules Aller-Retour, la clé inclut la semaine pour distinguer aller/retour.
    
    Args:
        equipe1: Nom de la première équipe
        equipe2: Nom de la deuxième équipe
        genre: Genre du match (M/F)
        poule: Code de la poule
        semaine: Numéro de semaine
        is_aller_retour: True si la poule est en aller-retour
        
    Returns:
        Clé unique pour le match
    """
    # Trier les équipes pour avoir une clé symétrique
    equipes_triees = tuple(sorted([equipe1.strip(), equipe2.strip()]))
    
    if is_aller_retour and semaine is not None:
        # Pour aller-retour, inclure la semaine pour distinguer les matchs
        return f"{equipes_triees[0]}|{equipes_triees[1]}|{genre}|{poule}|S{semaine}"
    else:
        # Pour classique, juste les équipes et le genre
        return f"{equipes_triees[0]}|{equipes_triees[1]}|{genre}|{poule}"


def calculer_date_depuis_semaine(
    semaine: int, 
    date_debut: Optional[datetime], 
    jour_match: str = 'jeudi'
) -> Optional[str]:
    """
    Calcule la date d'un match à partir du numéro de semaine.
    
    Args:
        semaine: Numéro de la semaine
        date_debut: Date de début de saison
        jour_match: Jour officiel des matchs
        
    Returns:
        Date au format DD/MM/YYYY ou None
    """
    if not date_debut:
        return None
    
    # Mapping jour -> index (lundi = 0)
    jours_mapping = {
        'lundi': 0, 'mardi': 1, 'mercredi': 2, 'jeudi': 3,
        'vendredi': 4, 'samedi': 5, 'dimanche': 6
    }
    
    jour_index = jours_mapping.get(jour_match.lower(), 3)  # jeudi par défaut
    
    # Calculer la date: début + (semaine - 1) * 7 jours
    date_match = date_debut + timedelta(weeks=semaine - 1)
    
    # Ajuster au jour de la semaine si nécessaire
    current_day = date_match.weekday()
    if current_day != jour_index:
        # Avancer jusqu'au bon jour
        delta = (jour_index - current_day) % 7
        date_match = date_match + timedelta(days=delta)
    
    return date_match.strftime('%d/%m/%Y')


def formater_score(score_dict: Optional[Dict]) -> str:
    """
    Formate un score depuis le format solution vers le format Excel.
    
    Args:
        score_dict: Dictionnaire avec equipe1, equipe2, has_score
        
    Returns:
        Score formaté (ex: "3-1") ou chaîne vide
    """
    if not score_dict or not score_dict.get('has_score'):
        return ''
    
    s1 = score_dict.get('equipe1')
    s2 = score_dict.get('equipe2')
    
    if s1 is not None and s2 is not None:
        return f"{s1}-{s2}"
    return ''


# ==============================================================================
# Classe principale
# ==============================================================================

class ImporteurSolutionVersFixed:
    """
    Importeur de matchs depuis une solution JSON vers la feuille Matchs_Fixes.
    """
    
    def __init__(
        self,
        ctx: ScriptContext,
        semaines: Optional[List[int]] = None,
        poules: Optional[List[str]] = None,
        genre: Optional[str] = None,
        types_competition: Optional[List[str]] = None,
        inclure_ententes: bool = False,
        remplacer: bool = False,
        dry_run: bool = False,
        verbose: bool = False,
    ):
        """
        Initialise l'importeur.
        
        Args:
            ctx: Contexte du script
            semaines: Liste des semaines à importer (None = toutes)
            poules: Liste des poules à importer (None = toutes)
            genre: Genre à filtrer (None = tous)
            types_competition: Types de compétition à inclure (None = tous)
            inclure_ententes: Inclure les matchs en entente
            remplacer: Remplacer les matchs existants au lieu de fusionner
            dry_run: Mode simulation
            verbose: Mode verbeux
        """
        self.ctx = ctx
        self.semaines = semaines
        self.poules = poules
        self.genre = genre
        self.types_competition = types_competition
        self.inclure_ententes = inclure_ententes
        self.remplacer = remplacer
        self.dry_run = dry_run
        self.verbose = verbose
        
        # Données chargées
        self.types_poules: Dict[str, str] = {}  # {poule: "Classique" ou "Aller-Retour"}
        self.excel_path: Optional[Path] = None
        
    def charger_types_poules(self) -> Dict[str, str]:
        """
        Charge les types de poules depuis la solution.
        
        Returns:
            Dictionnaire {poule_id: type}
        """
        if not self.ctx.solution_data:
            return {}
        
        poules = self.ctx.solution_data.get('entities', {}).get('poules', [])
        types = {}
        
        for poule in poules:
            poule_id = poule.get('id', poule.get('nom', ''))
            poule_type = poule.get('type', 'Classique')
            types[poule_id] = poule_type
        
        return types
    
    def est_aller_retour(self, poule: str) -> bool:
        """Vérifie si une poule est en aller-retour."""
        return self.types_poules.get(poule, 'Classique').lower() in ['aller-retour', 'ar']
    
    def charger_matchs_solution(self) -> List[Dict]:
        """
        Charge et filtre les matchs depuis la solution.
        
        Returns:
            Liste des matchs filtrés
        """
        scheduled = self.ctx.get_scheduled_matches()
        
        if not scheduled:
            print_warning("Aucun match planifié trouvé dans la solution")
            return []
        
        matchs_filtres = []
        stats = {
            'total': len(scheduled),
            'semaines_exclues': 0,
            'poules_exclues': 0,
            'genre_exclu': 0,
            'type_exclu': 0,
            'ententes_exclues': 0,
        }
        
        for match in scheduled:
            # Filtrer par semaine
            semaine = match.get('semaine')
            if self.semaines and semaine not in self.semaines:
                stats['semaines_exclues'] += 1
                continue
            
            # Filtrer par poule
            poule = match.get('poule', '')
            if self.poules and poule not in self.poules:
                stats['poules_exclues'] += 1
                continue
            
            # Filtrer par genre
            match_genre = match.get('genre', '')
            if self.genre and match_genre != self.genre:
                stats['genre_exclu'] += 1
                continue
            
            # Filtrer par type de compétition
            type_comp = match.get('championship_type', 'Acad')
            if self.types_competition and type_comp not in self.types_competition:
                stats['type_exclu'] += 1
                continue
            
            # Filtrer les ententes si non demandées
            is_entente = match.get('is_entente', False)
            gymnase = match.get('gymnase', '')
            if gymnase.upper() == 'ENTENTE':
                is_entente = True
            
            if is_entente and not self.inclure_ententes:
                stats['ententes_exclues'] += 1
                continue
            
            matchs_filtres.append(match)
        
        # Afficher les stats de filtrage
        print(f"\n📊 Filtrage des matchs:")
        print(f"   Total dans la solution: {stats['total']}")
        if stats['semaines_exclues']:
            print(f"   Exclus (semaines): {stats['semaines_exclues']}")
        if stats['poules_exclues']:
            print(f"   Exclus (poules): {stats['poules_exclues']}")
        if stats['genre_exclu']:
            print(f"   Exclus (genre): {stats['genre_exclu']}")
        if stats['type_exclu']:
            print(f"   Exclus (type compétition): {stats['type_exclu']}")
        if stats['ententes_exclues']:
            print(f"   Exclus (ententes): {stats['ententes_exclues']}")
        print(f"   → Matchs retenus: {len(matchs_filtres)}")
        
        return matchs_filtres
    
    def convertir_en_dataframe(self, matchs: List[Dict]) -> pd.DataFrame:
        """
        Convertit les matchs de la solution en DataFrame compatible Matchs_Fixes.
        
        Args:
            matchs: Liste des matchs depuis la solution
            
        Returns:
            DataFrame au format Matchs_Fixes
        """
        # Récupérer les infos du calendrier pour calculer les dates
        date_debut = None
        jour_match = 'jeudi'
        
        if self.ctx.config_data:
            calendrier = self.ctx.config_data.get('calendrier', {})
            date_str = calendrier.get('date_debut')
            if date_str:
                try:
                    date_debut = datetime.fromisoformat(date_str)
                except:
                    pass
            jour_match = calendrier.get('jour_match', 'jeudi')
        
        rows = []
        for match in matchs:
            semaine = match.get('semaine')
            
            # Calculer la date si possible
            date_str = calculer_date_depuis_semaine(semaine, date_debut, jour_match)
            
            # Formater le score
            score = formater_score(match.get('score'))
            
            # Créer la ligne
            row = {
                'Equipe_1': match.get('equipe1_nom', ''),
                'Equipe_2': match.get('equipe2_nom', ''),
                'Genre': match.get('genre', ''),
                'Poule': match.get('poule', ''),
                'Semaine': semaine,
                'Date': date_str or '',
                'Horaire': match.get('horaire', ''),
                'Gymnase': match.get('gymnase', ''),
                'Score': score,
                'Type_Competition': match.get('championship_type', 'Acad'),
                'Remarques': IMPORT_TAG,
                'Arbitres': '',
            }
            rows.append(row)
        
        return pd.DataFrame(rows, columns=COLONNES_MATCHS_FIXES)
    
    def charger_matchs_existants(self) -> pd.DataFrame:
        """
        Charge les matchs fixes existants depuis le fichier Excel.
        
        Returns:
            DataFrame des matchs existants
        """
        if not self.excel_path or not self.excel_path.exists():
            return pd.DataFrame(columns=COLONNES_MATCHS_FIXES)
        
        try:
            df = pd.read_excel(self.excel_path, sheet_name='Matchs_Fixes')
            
            # S'assurer que toutes les colonnes existent
            for col in COLONNES_MATCHS_FIXES:
                if col not in df.columns:
                    df[col] = ''
            
            print(f"✓ Matchs fixes existants chargés: {len(df)} matchs")
            return df[COLONNES_MATCHS_FIXES]
        except Exception as e:
            if self.verbose:
                print(f"   ℹ️  Feuille Matchs_Fixes non trouvée ({e})")
            return pd.DataFrame(columns=COLONNES_MATCHS_FIXES)
    
    def fusionner_matchs(
        self, 
        df_nouveaux: pd.DataFrame, 
        df_existants: pd.DataFrame
    ) -> Tuple[pd.DataFrame, Dict[str, int]]:
        """
        Fusionne les nouveaux matchs avec les existants en évitant les doublons.
        
        Pour les poules Classiques: évite les matchs identiques (même paire d'équipes)
        Pour les poules Aller-Retour: permet 2 matchs par paire (un par sens)
        
        Args:
            df_nouveaux: Nouveaux matchs à ajouter
            df_existants: Matchs déjà présents
            
        Returns:
            Tuple (DataFrame fusionné, statistiques)
        """
        stats = {
            'total_nouveaux': len(df_nouveaux),
            'doublons_ignores': 0,
            'matchs_ajoutes': 0,
        }
        
        if df_existants.empty:
            stats['matchs_ajoutes'] = len(df_nouveaux)
            return df_nouveaux, stats
        
        # Créer les clés pour les matchs existants
        cles_existantes: Set[str] = set()
        
        for _, row in df_existants.iterrows():
            poule = str(row.get('Poule', '')).strip()
            is_ar = self.est_aller_retour(poule)
            semaine = row.get('Semaine') if is_ar else None
            
            cle = creer_cle_match(
                str(row.get('Equipe_1', '')),
                str(row.get('Equipe_2', '')),
                str(row.get('Genre', '')),
                poule,
                semaine,
                is_ar
            )
            cles_existantes.add(cle)
        
        # Filtrer les nouveaux matchs
        lignes_a_ajouter = []
        
        for _, row in df_nouveaux.iterrows():
            poule = str(row.get('Poule', '')).strip()
            is_ar = self.est_aller_retour(poule)
            semaine = row.get('Semaine') if is_ar else None
            
            cle = creer_cle_match(
                str(row.get('Equipe_1', '')),
                str(row.get('Equipe_2', '')),
                str(row.get('Genre', '')),
                poule,
                semaine,
                is_ar
            )
            
            if cle in cles_existantes:
                stats['doublons_ignores'] += 1
                if self.verbose:
                    print(f"   ⏭️  Doublon ignoré: {row.get('Equipe_1')} vs {row.get('Equipe_2')} (S{row.get('Semaine')})")
            else:
                lignes_a_ajouter.append(row)
                cles_existantes.add(cle)  # Éviter les doublons dans les nouveaux aussi
        
        stats['matchs_ajoutes'] = len(lignes_a_ajouter)
        
        # Concaténer
        if lignes_a_ajouter:
            df_a_ajouter = pd.DataFrame(lignes_a_ajouter)
            df_fusionne = pd.concat([df_existants, df_a_ajouter], ignore_index=True)
        else:
            df_fusionne = df_existants
        
        return df_fusionne, stats
    
    def sauvegarder(self, df_matchs: pd.DataFrame):
        """
        Sauvegarde les matchs dans la feuille Matchs_Fixes.
        
        Args:
            df_matchs: DataFrame à sauvegarder
        """
        if self.dry_run:
            print("\n🔍 MODE SIMULATION - Aucune modification effectuée")
            print(f"   {len(df_matchs)} matchs seraient sauvegardés dans Matchs_Fixes")
            return
        
        if not self.excel_path:
            raise ValueError("Chemin Excel non configuré")
        
        print(f"\n💾 Sauvegarde dans {self.excel_path.name}...")
        
        try:
            # Charger le workbook existant
            wb = openpyxl.load_workbook(self.excel_path)
            
            # Supprimer la feuille Matchs_Fixes si elle existe
            if 'Matchs_Fixes' in wb.sheetnames:
                del wb['Matchs_Fixes']
            
            # Créer une nouvelle feuille
            ws = wb.create_sheet('Matchs_Fixes')
            
            # Écrire les données
            for r in dataframe_to_rows(df_matchs, index=False, header=True):
                ws.append(r)
            
            # Sauvegarder
            wb.save(self.excel_path)
            
            print_success(f"{len(df_matchs)} matchs sauvegardés dans Matchs_Fixes")
            
        except Exception as e:
            raise RuntimeError(f"Erreur lors de la sauvegarde: {e}")
    
    def executer(self) -> bool:
        """
        Exécute l'importation complète.
        
        Returns:
            True si succès, False sinon
        """
        # Vérifier la solution
        if not self.ctx.solution_data:
            print_error("Aucune solution chargée")
            return False
        
        # Récupérer le chemin Excel
        self.excel_path = self.ctx.excel_path
        if not self.excel_path:
            print_error("Chemin Excel de configuration non trouvé")
            return False
        
        if not self.excel_path.exists():
            print_error(f"Fichier Excel introuvable: {self.excel_path}")
            return False
        
        print(f"📂 Configuration Excel: {self.excel_path.name}")
        print(f"📄 Solution: {self.ctx.solution_path.name if self.ctx.solution_path else 'N/A'}")
        
        # Charger les types de poules
        self.types_poules = self.charger_types_poules()
        nb_ar = sum(1 for t in self.types_poules.values() if t.lower() in ['aller-retour', 'ar'])
        print(f"🏐 Types de poules: {len(self.types_poules)} ({nb_ar} aller-retour)")
        
        # Charger et filtrer les matchs de la solution
        matchs = self.charger_matchs_solution()
        
        if not matchs:
            print_warning("Aucun match à importer après filtrage")
            return True
        
        # Convertir en DataFrame
        df_nouveaux = self.convertir_en_dataframe(matchs)
        
        if self.remplacer:
            # Mode remplacement: ignorer les matchs existants
            print("\n⚠️  Mode REMPLACEMENT: les matchs existants seront remplacés")
            df_final = df_nouveaux
            stats = {'matchs_ajoutes': len(df_nouveaux), 'doublons_ignores': 0}
        else:
            # Mode fusion: charger et fusionner avec les existants
            df_existants = self.charger_matchs_existants()
            df_final, stats = self.fusionner_matchs(df_nouveaux, df_existants)
        
        # Afficher le résumé
        print(f"\n📊 Résumé de l'import:")
        print(f"   Matchs de la solution: {stats.get('total_nouveaux', len(df_nouveaux))}")
        print(f"   Doublons ignorés: {stats['doublons_ignores']}")
        print(f"   Matchs ajoutés: {stats['matchs_ajoutes']}")
        print(f"   Total final: {len(df_final)}")
        
        # Sauvegarder
        self.sauvegarder(df_final)
        
        return True


# ==============================================================================
# Main
# ==============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Import des matchs depuis une solution vers Matchs_Fixes",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Exemples:
    # Importer tous les matchs
    python scripts/import_solution_to_fixed.py -c configs/config_volley.yaml
    
    # Importer les semaines 1 à 5
    python scripts/import_solution_to_fixed.py -c configs/config_volley.yaml --semaines 1-5
    
    # Importer une solution spécifique
    python scripts/import_solution_to_fixed.py -s solutions/solution_volley_2026-01-27.json
    
    # Simulation (dry-run)
    python scripts/import_solution_to_fixed.py -c configs/config_volley.yaml --dry-run
"""
    )
    
    # Arguments de base
    parser.add_argument(
        '--config', '-c',
        type=str,
        default=None,
        help='Fichier de configuration YAML (ex: configs/config_volley.yaml)'
    )
    
    parser.add_argument(
        '--solution', '-s',
        type=str,
        default=None,
        help='Fichier solution JSON (ex: solutions/latest_volley.json)'
    )
    
    parser.add_argument(
        '--sport',
        type=str,
        default=None,
        help='Code sport (ex: volley, hand). Auto-détecté si non fourni.'
    )
    
    # Filtres
    parser.add_argument(
        '--semaines',
        type=str,
        default=None,
        help='Semaines à importer (ex: "1-5", "3,5,7", "1-3,6,8-10")'
    )
    
    parser.add_argument(
        '--poules',
        type=str,
        default=None,
        help='Poules à importer, séparées par virgule (ex: "VBFA1PA,VBMA1PA")'
    )
    
    parser.add_argument(
        '--genre', '-g',
        type=str,
        choices=['M', 'F'],
        default=None,
        help='Genre à filtrer (M ou F)'
    )
    
    parser.add_argument(
        '--type',
        type=str,
        default=None,
        help='Types de compétition à inclure, séparés par virgule (ex: "Acad,CFE")'
    )
    
    # Options
    parser.add_argument(
        '--inclure-ententes',
        action='store_true',
        help='Inclure les matchs en entente (exclus par défaut)'
    )
    
    parser.add_argument(
        '--remplacer',
        action='store_true',
        help='Remplacer les matchs existants au lieu de fusionner'
    )
    
    parser.add_argument(
        '--dry-run',
        action='store_true',
        help='Simulation: ne modifie aucun fichier'
    )
    
    parser.add_argument(
        '--verbose', '-v',
        action='store_true',
        help='Mode verbeux'
    )
    
    args = parser.parse_args()
    
    # Créer le contexte
    try:
        ctx = ScriptContext.from_args(args)
    except Exception as e:
        print_error(f"Erreur lors du chargement du contexte: {e}")
        return 1
    
    # Afficher l'en-tête
    print_header(f"Import Solution → Matchs Fixes {ctx.sport.name}", ctx.sport.emoji)
    ctx.print_status()
    
    # Parser les arguments de filtre
    semaines = parse_semaines_arg(args.semaines) if args.semaines else None
    poules = [p.strip() for p in args.poules.split(',')] if args.poules else None
    types_competition = [t.strip() for t in args.type.split(',')] if args.type else None
    
    # Afficher les filtres actifs
    if semaines or poules or args.genre or types_competition:
        print("\n🔍 Filtres actifs:")
        if semaines:
            print(f"   Semaines: {semaines}")
        if poules:
            print(f"   Poules: {poules}")
        if args.genre:
            print(f"   Genre: {args.genre}")
        if types_competition:
            print(f"   Types: {types_competition}")
    
    # Créer et exécuter l'importeur
    importeur = ImporteurSolutionVersFixed(
        ctx=ctx,
        semaines=semaines,
        poules=poules,
        genre=args.genre,
        types_competition=types_competition,
        inclure_ententes=args.inclure_ententes,
        remplacer=args.remplacer,
        dry_run=args.dry_run,
        verbose=args.verbose,
    )
    
    try:
        success = importeur.executer()
        
        if success:
            print("\n" + "=" * 70)
            print_success("Import terminé avec succès!")
            return 0
        else:
            return 1
            
    except Exception as e:
        print_error(f"Erreur lors de l'import: {e}")
        if args.verbose:
            import traceback
            traceback.print_exc()
        return 1


if __name__ == '__main__':
    sys.exit(main())
