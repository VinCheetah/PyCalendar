#!/usr/bin/env python3
from __future__ import annotations
"""
Script de synchronisation MySportU <-> Configuration Excel.

Utilise le module ``pycalendar.mysportu`` pour la connexion et la récupération
des données depuis l'API MySportU, puis synchronise avec le fichier Excel de
configuration du calendrier.

Opérations disponibles :

1. MAPPING (--mapping) :
   - Met à jour la colonne 'MySportU' dans la feuille 'Equipes'
   - Met à jour la colonne 'MySportU' dans la feuille 'Gymnases'
   - Établit la correspondance entre les noms MySportU et les noms config

2. SYNCHRONISATION (--sync) :
   - Récupère **tous** les matchs depuis l'API MySportU (avec cache)
   - Importe les nouveaux matchs vers 'Matchs_Fixes'
   - Met à jour les scores des matchs terminés
   - Supprime les matchs Acad reportés/annulés (sauf --no-clean)
   - Trie les matchs par semaine/type/date/genre/niveau/poule (sauf --no-sort)

Principes :
   - Les matchs CFU/CFE ne sont PAS dans MySportU et sont préservés
   - Le genre est TOUJOURS pris en compte (clé = (équipe, genre))
   - Les matchs hors jour normal (ententes) sont conservés et marqués
   - Les matchs sans date/gymnase connu sont conservés avec avertissement
   - Le format Semaine est uniformisé : "N (dd/mm)"

Usage :
    python sync_mysportu.py --sync                        # Sync depuis l'API
    python sync_mysportu.py --sync --force                # Ignorer le cache
    python sync_mysportu.py --sync --dry-run              # Aperçu sans modif
    python sync_mysportu.py --mapping                     # Mapping équipes/gyms
    python sync_mysportu.py --all                         # Tout faire
    python sync_mysportu.py --sync --fix-scores           # Corriger scores
    python sync_mysportu.py --sync --excel                # Depuis un fichier Excel
"""

import re
import sys
import argparse
from datetime import datetime
from pathlib import Path
from typing import Optional, Tuple, Dict, List, Set, Any
from dataclasses import dataclass, field
from enum import Enum

import pandas as pd
import yaml
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

from rich.console import Console
from rich.table import Table
from rich.panel import Panel
from rich.text import Text

# ── Paths ───────────────────────────────────────────────────────────────────

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))
sys.path.insert(0, str(PROJECT_ROOT / "src"))

# ── Module MySportU propre ──────────────────────────────────────────────────

from pycalendar.mysportu import MySportU
from pycalendar.mysportu.models import MatchInfo, MatchState, Score

# ── Console Rich ────────────────────────────────────────────────────────────

console = Console()


# =============================================================================
# CONSTANTES ET MAPPINGS
# =============================================================================

# Couleurs Excel
COLOR_ERROR = PatternFill(start_color='FFCCCB', end_color='FFCCCB', fill_type='solid')
COLOR_WARNING = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
COLOR_SUCCESS = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
COLOR_INFO = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
COLOR_ENTENTE = PatternFill(start_color='E8DAEF', end_color='E8DAEF', fill_type='solid')

# Mapping code institution MySportU -> nom court config
INSTITUTION_CODE_MAP = {
    '069069001': 'ENTPE',
    '069069006': 'LYON 2',
    '069069007': 'LYON 3',
    '069069008': 'ESA',
    '069069009': 'ISARA',
    '069069012': 'CATHO',
    '069069014': 'CPE',
    '069069015': 'ECAM',
    '069069016': 'ECL',
    '069069018': 'EML',
    '069069019': 'ENS',
    '069069020': 'ENTPE',
    '069069021': 'ESME',
    '069069022': 'ESSCA',
    '069069025': 'INSA',
    '069069026': 'ISOSTEO',
    '069069028': 'VETO',
    '069069029': 'CESI',
    '069069031': 'LYON 1',
}

# Mapping gymnases MySportU -> gymnases config
GYMNASE_MAP = {
    'CENTRALE': 'ECL',
    'COMPET C (HAUT) - LEON JOUHAUX': 'L. J. HAUT',
    'COMPET. LAPLANCHE- R. LISSMANN': None,
    'ENS DESCARTES': 'DESCARTES',
    'ESA BRON': 'ESA',
    'HALLE - 3D': 'LAENNEC',
    'HALLE - C.BESSON': 'BESSON',
    'HALLE C LYON 2': 'LYON 2 HC',
    'COMPET PIBAROT - R.VANEL': None,
    'N-A': None,
}

# Surcharges de noms d'équipes API
API_TEAM_NAME_MAP: dict[str, Optional[str]] = {
    'AS UD Lyon': None,   # Nom générique, inutilisable
    'ENS LYON': 'ENS',
}


# =============================================================================
# TYPES
# =============================================================================

class IssueType(Enum):
    """Types d'incohérences."""
    TEAM_NOT_FOUND = "Équipe non trouvée"
    TEAM_NO_MAPPING = "Équipe sans mapping"
    GENRE_MISMATCH = "Genre incohérent"
    POULE_MISMATCH = "Poules différentes"
    GYM_NOT_FOUND = "Gymnase non trouvé"
    SCORE_CONFLICT = "Conflit de score"
    DATE_MISSING = "Date manquante"
    DATA_MISMATCH = "Données incohérentes MySportU/Config"
    ENTENTE_RESOLVED = "Entente résolue"


@dataclass
class Issue:
    """Incohérence détectée."""
    type: IssueType
    message: str
    mysportu_data: Optional[str] = None
    config_data: Optional[str] = None
    severity: str = "warning"   # error, warning, info


@dataclass
class SyncResult:
    """Résultat de la synchronisation."""
    matches_added: int = 0
    scores_added: int = 0
    scores_corrected: int = 0
    duplicates_found: int = 0
    matches_removed: int = 0
    matches_reordered: int = 0
    matches_kept_no_date: int = 0
    matches_kept_no_gym: int = 0
    matches_updated: int = 0
    ententes_resolved: int = 0
    issues: List[Issue] = field(default_factory=list)


# =============================================================================
# NORMALISATION DES NOMS D'ÉQUIPES
# =============================================================================

def normalize_team_name(libelle_court: str, club_code: str, genre: str) -> Optional[str]:
    """
    Normalise un nom d'équipe MySportU vers le format config.

    Utilise le code club pour identifier l'institution, puis le libellé court
    pour extraire le numéro d'équipe.

    Args:
        libelle_court: Nom court API (ex: "INSA (2)", "SANTE (3)")
        club_code: Code club MySportU (ex: "069069025")
        genre: Genre du match (F/M)

    Returns:
        Nom config (ex: "INSA (2)") ou None
    """
    if not libelle_court or not club_code:
        return None

    # Surcharge explicite
    if libelle_court in API_TEAM_NAME_MAP:
        return API_TEAM_NAME_MAP[libelle_court]

    institution = INSTITUTION_CODE_MAP.get(club_code)
    if not institution:
        return None

    # Extraire le numéro d'équipe
    num_match = re.search(r'\((\d+)\)$', libelle_court)
    if num_match:
        numero = num_match.group(1)
    else:
        num_match = re.search(r'(\d+)$', libelle_court)
        numero = num_match.group(1) if num_match else '1'

    # Cas spéciaux
    clean = libelle_court.upper()
    if 'IEP' in clean:
        return 'LYON 2 (IEP) (4)'
    if 'SANT' in clean or 'SANTÉ' in clean or 'SANTE' in clean:
        return f"SANTE ({numero})"

    return f"{institution} ({numero})"


def normalize_excel_team(name: Any, genre: Any) -> Tuple[Optional[str], Optional[str]]:
    """
    Normalise un nom d'équipe depuis un fichier Excel MySportU.

    Format attendu : "069069XXX - NOM EQUIPE".

    Returns:
        Tuple (config_name, genre)
    """
    if pd.isna(name) or not genre:
        return None, None

    match = re.match(r'^(\d+)\s*-\s*(.+)$', str(name))
    if not match:
        return None, None

    code = match.group(1)
    rest = match.group(2).strip()

    institution = INSTITUTION_CODE_MAP.get(code)
    if not institution:
        return None, None

    numero = _extract_team_number(institution, rest)

    if numero:
        config_name = f"{institution} ({numero})"
    else:
        config_name = f"{institution} (1)"

    # Cas spéciaux
    if institution == 'LYON 2' and 'IEP' in rest.upper():
        config_name = 'LYON 2 (IEP) (4)'
    elif institution == 'LYON 1' and ('SANTÉ' in rest.upper() or 'SANTE' in rest.upper()):
        num_match = re.search(r'\((\d+)\)', rest)
        n = num_match.group(1) if num_match else '1'
        config_name = f'SANTE ({n})'

    return config_name, genre


def _extract_team_number(institution: str, rest: str) -> Optional[str]:
    """Extrait le numéro d'équipe selon les règles par institution."""
    rest_upper = rest.upper()

    if institution == 'LYON 2':
        m = re.search(r'\((\d+)\)|VB\s*[MF]\s*(\d+)', rest)
        return (m.group(1) or m.group(2)) if m else None
    elif institution == 'LYON 1':
        m = re.search(r'\((\d+)\)', rest)
        return m.group(1) if m else None
    elif institution == 'INSA':
        m = re.search(r'(\d+)\s*$', rest)
        return m.group(1) if m else None
    elif institution == 'ECL':
        m = re.search(r'[MF]\s*(\d+)|([MF])(\d+)', rest, re.IGNORECASE)
        if m:
            return m.group(1) or m.group(3)
        return None
    elif institution == 'EML':
        m = re.search(r'EML\s*(\d+)', rest)
        return m.group(1) if m else None
    elif institution == 'ENS':
        m = re.search(r'[FG](\d+)', rest)
        return m.group(1) if m else None
    elif institution == 'ENTPE':
        if 'FILLES' in rest_upper:
            return '1'
        m = re.search(r'\((\d+)\)|ENTPE\s*(\d+)', rest)
        return (m.group(1) or m.group(2)) if m else None
    elif institution in ('ESA', 'ESME'):
        m = re.search(r'(\d+)\s*$', rest)
        return m.group(1) if m else None
    elif institution in ('ESSCA', 'CPE', 'CATHO'):
        m = re.search(r'\((\d+)\)|n°(\d+)', rest)
        return (m.group(1) or m.group(2)) if m else '1'
    else:
        m = re.search(r'\((\d+)\)|(\d+)\s*$', rest)
        return (m.group(1) or m.group(2)) if m else None


def normalize_gymnase(lieu_libelle: Optional[str]) -> Optional[str]:
    """Normalise un lieu MySportU vers un gymnase config."""
    if not lieu_libelle:
        return None
    return GYMNASE_MAP.get(lieu_libelle.strip())


def get_genre_from_championnat(champ: str) -> Optional[str]:
    """Déduit le genre (F/M) depuis le libellé championnat."""
    if pd.isna(champ):
        return None
    upper = str(champ).upper()
    if 'VBF' in upper or 'HBF' in upper or 'BBF' in upper:
        return 'F'
    if 'VBM' in upper or 'HBM' in upper or 'BBM' in upper:
        return 'M'
    return None


# =============================================================================
# DATES ET SEMAINES
# =============================================================================

def calculate_week_number(date_str: Optional[str], start_date: datetime) -> Optional[int]:
    """Calcule le numéro de semaine depuis la date de début de saison."""
    if not date_str or date_str == 'N-A':
        return None
    try:
        if isinstance(date_str, str):
            date = datetime.strptime(date_str.split(' ')[0], '%d/%m/%Y')
        else:
            date = pd.to_datetime(date_str)
        delta = (date - start_date).days
        return delta // 7 + 1
    except (ValueError, TypeError):
        return None


def format_week_display(week_num: int, start_date: datetime) -> str:
    """Formate "N (dd/mm)" depuis le numéro de semaine."""
    match_date = start_date + pd.Timedelta(days=(week_num - 1) * 7)
    return f"{week_num} ({match_date.strftime('%d/%m')})"


def parse_week_number(semaine_str) -> Optional[int]:
    """Extrait le numéro de semaine depuis 'N' ou 'N (dd/mm)'."""
    if pd.isna(semaine_str):
        return None
    m = re.match(r'^(\d+)', str(semaine_str).strip())
    return int(m.group(1)) if m else None


def format_date_ddmmyy(date_str: Optional[str]) -> Optional[str]:
    """Formate une date en dd/mm/yy."""
    if not date_str or date_str == 'N-A':
        return None
    try:
        dt = datetime.strptime(date_str.split(' ')[0], '%d/%m/%Y')
        return dt.strftime('%d/%m/%y')
    except (ValueError, TypeError):
        return None


def is_entente_match(date_str: Optional[str], jour_match: str = 'jeudi') -> bool:
    """Vérifie si un match est joué un autre jour que le jour normal."""
    if not date_str:
        return False
    jour_mapping = {
        'lundi': 0, 'mardi': 1, 'mercredi': 2, 'jeudi': 3,
        'vendredi': 4, 'samedi': 5, 'dimanche': 6,
    }
    jour_normal = jour_mapping.get(jour_match.lower(), 3)
    try:
        dt = datetime.strptime(date_str.split(' ')[0], '%d/%m/%Y')
        return dt.weekday() != jour_normal
    except (ValueError, IndexError):
        return False


# =============================================================================
# CHARGEMENT DES DONNÉES EXCEL CONFIG
# =============================================================================

def load_config(config_path: Path) -> dict:
    """Charge la configuration YAML."""
    with open(config_path, 'r', encoding='utf-8') as f:
        return yaml.safe_load(f)


def load_config_teams(config_excel_path: Path) -> pd.DataFrame:
    """Charge les équipes de la configuration Excel."""
    return pd.read_excel(config_excel_path, sheet_name='Equipes')


def load_existing_fixed_matches(config_excel_path: Path) -> pd.DataFrame:
    """Charge les matchs fixes existants."""
    try:
        return pd.read_excel(config_excel_path, sheet_name='Matchs_Fixes')
    except Exception:
        return pd.DataFrame()


def get_gymnases(config_excel_path: Path) -> Set[str]:
    """Récupère la liste des gymnases valides."""
    df = pd.read_excel(config_excel_path, sheet_name='Gymnases')
    return set(df['Gymnase'].dropna().tolist())


def get_team_poules(config_excel_path: Path) -> Dict[Tuple[str, str], str]:
    """Récupère la poule de chaque équipe -> {(equipe, genre): poule}."""
    df = pd.read_excel(config_excel_path, sheet_name='Equipes')
    poules = {}
    for _, row in df.iterrows():
        key = (row['Equipe'], row['Genre_Equipe'])
        poules[key] = row.get('Poule')
    return poules


def load_mysportu_excel(file_path: Path, championship_pattern: str = 'PH2') -> pd.DataFrame:
    """Charge les données depuis un fichier Excel MySportU exporté."""
    df = pd.read_excel(file_path)
    if championship_pattern:
        mask = df['Championnat'].str.contains(championship_pattern, na=False, case=False)
        df = df[mask].copy()
    return df


# =============================================================================
# CONVERSION MatchInfo -> DONNÉES SYNC
# =============================================================================

@dataclass
class SyncMatch:
    """
    Match normalisé prêt pour la synchronisation avec la config Excel.

    Créé soit depuis un MatchInfo (API), soit depuis un DataFrame (Excel).
    """
    eq_a_config: str            # Nom config du receveur
    eq_b_config: str            # Nom config du visiteur
    genre: str                  # M ou F
    poule_msu: str              # Code poule MySportU
    date_str: Optional[str]     # dd/mm/yyyy
    heure: Optional[str]        # HH:MM
    lieu_libelle: Optional[str] # Nom du lieu MySportU
    score: Optional[str]        # "X-Y"
    state: MatchState           # État du match
    is_entente: bool            # Match hors jour normal
    api_id: Optional[int]       # ID MySportU
    gym_resolved: Optional[str] = None  # Gymnase réel résolu (même pour ententes)

    # Noms bruts pour les messages
    display_a: str = ""
    display_b: str = ""


def matches_to_sync(
    matches: list[MatchInfo],
    valid_teams: Set[Tuple[str, str]],
    jour_match: str = 'jeudi',
) -> Tuple[list[SyncMatch], list[Issue]]:
    """
    Convertit des MatchInfo en SyncMatch avec normalisation des noms.

    Conserve TOUS les matchs, y compris ceux sans date ou hors jour normal.

    Returns:
        Tuple (sync_matches, issues)
    """
    sync_matches: list[SyncMatch] = []
    issues: list[Issue] = []

    for m in matches:
        # Normaliser les équipes
        eq_a = normalize_team_name(m.receveur.libelle_court, m.receveur.club_code, m.genre)
        eq_b = normalize_team_name(m.visiteur.libelle_court, m.visiteur.club_code, m.genre)

        if not eq_a or not eq_b:
            issues.append(Issue(
                type=IssueType.TEAM_NOT_FOUND,
                message="Impossible de résoudre les équipes",
                mysportu_data=f"{m.receveur.libelle_court} vs {m.visiteur.libelle_court}",
                severity="error",
            ))
            continue

        # Vérifier existence dans la config
        missing = False
        for eq, display in [(eq_a, m.receveur.libelle_court), (eq_b, m.visiteur.libelle_court)]:
            if (eq, m.genre) not in valid_teams:
                issues.append(Issue(
                    type=IssueType.TEAM_NOT_FOUND,
                    message="Équipe non trouvée dans la config",
                    mysportu_data=display,
                    config_data=f"{eq} ({m.genre})",
                    severity="error",
                ))
                missing = True
        if missing:
            continue

        # Score
        score_str = str(m.score) if m.score else None

        # Entente ?
        entente = is_entente_match(m.date, jour_match) if m.date else False

        # Résoudre le gymnase réel même pour les ententes
        lieu_lib = m.lieu.libelle if m.lieu else None
        gym_resolved = normalize_gymnase(lieu_lib) if lieu_lib else None

        sync_matches.append(SyncMatch(
            eq_a_config=eq_a,
            eq_b_config=eq_b,
            genre=m.genre,
            poule_msu=m.poule,
            date_str=m.date,
            heure=m.heure if m.heure else None,
            lieu_libelle=lieu_lib,
            score=score_str,
            state=m.state,
            is_entente=entente,
            api_id=m.id,
            gym_resolved=gym_resolved,
            display_a=m.receveur.libelle_court,
            display_b=m.visiteur.libelle_court,
        ))

    return sync_matches, issues


def excel_to_sync(
    df: pd.DataFrame,
    valid_teams: Set[Tuple[str, str]],
    jour_match: str = 'jeudi',
) -> Tuple[list[SyncMatch], list[Issue]]:
    """Convertit un DataFrame Excel MySportU en SyncMatch."""
    sync_matches: list[SyncMatch] = []
    issues: list[Issue] = []

    for _, row in df.iterrows():
        genre = get_genre_from_championnat(row.get('Championnat', ''))
        if not genre:
            continue

        eq_a, _ = normalize_excel_team(row.get('Equipe A'), genre)
        eq_b, _ = normalize_excel_team(row.get('Equipe B'), genre)

        if not eq_a or not eq_b:
            issues.append(Issue(
                type=IssueType.TEAM_NOT_FOUND,
                message="Impossible de parser les équipes",
                mysportu_data=f"{row.get('Equipe A')} vs {row.get('Equipe B')}",
                severity="error",
            ))
            continue

        missing = False
        for eq, display in [(eq_a, row.get('Equipe A')), (eq_b, row.get('Equipe B'))]:
            if (eq, genre) not in valid_teams:
                issues.append(Issue(
                    type=IssueType.TEAM_NOT_FOUND,
                    message="Équipe non trouvée dans la config",
                    mysportu_data=str(display),
                    config_data=f"{eq} ({genre})",
                    severity="error",
                ))
                missing = True
        if missing:
            continue

        # État
        etat_raw = row.get('etat', '')
        etat_str = str(etat_raw).lower().strip() if not pd.isna(etat_raw) else ''
        if 'termin' in etat_str:
            state = MatchState.TERMINE
        elif 'report' in etat_str:
            state = MatchState.REPORTE
        elif 'annul' in etat_str:
            state = MatchState.ANNULE
        else:
            state = MatchState.NON_JOUE

        # Score
        score_raw = row.get('Score')
        score = None
        if not pd.isna(score_raw) and str(score_raw).strip() not in ('', 'N-A', 'nan'):
            score = str(score_raw).strip()

        date_str = row.get('Date')
        if pd.isna(date_str):
            date_str = None
        else:
            date_str = str(date_str).strip()

        entente = is_entente_match(date_str, jour_match) if date_str else False

        lieu_lib = str(row.get('Lieu')) if not pd.isna(row.get('Lieu')) else None
        gym_resolved = normalize_gymnase(lieu_lib) if lieu_lib else None

        sync_matches.append(SyncMatch(
            eq_a_config=eq_a,
            eq_b_config=eq_b,
            genre=genre,
            poule_msu=row.get('Poule', '') if not pd.isna(row.get('Poule', '')) else '',
            date_str=date_str,
            heure=str(row.get('Heure')) if not pd.isna(row.get('Heure')) else None,
            lieu_libelle=lieu_lib,
            score=score if state == MatchState.TERMINE else None,
            state=state,
            is_entente=entente,
            api_id=row.get('_api_id') if '_api_id' in row.index else None,
            gym_resolved=gym_resolved,
            display_a=str(row.get('Equipe A', '')),
            display_b=str(row.get('Equipe B', '')),
        ))

    return sync_matches, issues


# =============================================================================
# CLÉS DE MATCH & RECHERCHE
# =============================================================================

def create_match_key(eq1: str, eq2: str, genre: str) -> frozenset:
    """Clé unique pour un match (indépendante de l'ordre des équipes)."""
    return frozenset([eq1, eq2, f"GENRE:{genre}"])


def find_match_in_existing(
    eq1: str, eq2: str, genre: str,
    existing: pd.DataFrame,
) -> Tuple[Optional[int], bool]:
    """
    Trouve un match existant dans le DataFrame.

    Returns:
        (index, is_reversed) — index pandas ou None
    """
    for idx, row in existing.iterrows():
        if pd.isna(row.get('Equipe_1')) or pd.isna(row.get('Equipe_2')):
            continue
        if row.get('Genre') != genre:
            continue
        if row['Equipe_1'] == eq1 and row['Equipe_2'] == eq2:
            return int(idx), False  # type: ignore[arg-type]
        if row['Equipe_1'] == eq2 and row['Equipe_2'] == eq1:
            return int(idx), True  # type: ignore[arg-type]
    return None, False


def reverse_score(score: str) -> str:
    """Inverse un score ("3-1" -> "1-3")."""
    if not score or '-' not in str(score):
        return score
    parts = str(score).split('-')
    return f"{parts[1]}-{parts[0]}" if len(parts) == 2 else score


# =============================================================================
# PARTIE 1 : MAPPING
# =============================================================================

def build_team_mapping_from_api(
    matches: list[MatchInfo],
    config_teams: Set[Tuple[str, str]],
    verbose: bool = False,
) -> Tuple[Dict[Tuple[str, str], List[str]], List[Issue]]:
    """
    Construit le mapping équipes API -> équipes config.

    Returns:
        ({(config_name, genre): [api_libellés_courts]}, issues)
    """
    issues: list[Issue] = []
    mapping: dict[Tuple[str, str], list[str]] = {}
    seen_api: set[Tuple[str, str]] = set()

    for m in matches:
        for eq in (m.receveur, m.visiteur):
            key_api = (eq.libelle_court, m.genre)
            if key_api in seen_api:
                continue
            seen_api.add(key_api)

            config_name = normalize_team_name(eq.libelle_court, eq.club_code, m.genre)
            if not config_name:
                issues.append(Issue(
                    type=IssueType.TEAM_NOT_FOUND,
                    message="Impossible de résoudre l'équipe API",
                    mysportu_data=f"{eq.libelle_court} (code={eq.club_code})",
                    severity="error",
                ))
                continue

            cfg_key = (config_name, m.genre)
            if cfg_key not in config_teams:
                issues.append(Issue(
                    type=IssueType.TEAM_NOT_FOUND,
                    message="Équipe non trouvée dans la config",
                    mysportu_data=eq.libelle_court,
                    config_data=f"{config_name} ({m.genre})",
                    severity="error",
                ))
                continue

            mapping.setdefault(cfg_key, [])
            if eq.libelle_court not in mapping[cfg_key]:
                mapping[cfg_key].append(eq.libelle_court)

    # Équipes config sans mapping
    for team, genre in sorted(config_teams - set(mapping.keys())):
        issues.append(Issue(
            type=IssueType.TEAM_NO_MAPPING,
            message="Aucune correspondance MySportU",
            config_data=f"{team} ({genre})",
            severity="warning",
        ))

    if verbose:
        console.print(f"  Équipes API: {len(seen_api)}, config: {len(config_teams)}, "
                       f"mappings: {len(mapping)}")

    return mapping, issues


def update_teams_excel_column(
    config_excel_path: Path,
    mapping: Dict[Tuple[str, str], List[str]],
    dry_run: bool = False,
    verbose: bool = False,
) -> int:
    """Met à jour la colonne 'MySportU' dans la feuille 'Equipes'."""
    wb = load_workbook(config_excel_path)
    ws = wb['Equipes']

    headers: Dict[Any, int] = {cell.value: cell.column for cell in ws[1]}  # type: ignore[misc]
    equipe_col = headers.get('Equipe', 0)
    genre_col = headers.get('Genre_Equipe', 0)

    # Créer la colonne MySportU si absente
    if 'MySportU' not in headers:
        mysportu_col = max(headers.values()) + 1
        ws.cell(row=1, column=mysportu_col, value='MySportU')
        ws.cell(row=1, column=mysportu_col).font = Font(bold=True)
    else:
        mysportu_col = headers['MySportU']

    updated = 0
    for row in range(2, ws.max_row + 1):
        equipe = ws.cell(row=row, column=equipe_col).value
        genre = ws.cell(row=row, column=genre_col).value
        if not equipe or not genre:
            continue

        key = (equipe, genre)
        if key in mapping:
            value = ' | '.join(sorted(set(mapping[key])))
            current = ws.cell(row=row, column=mysportu_col).value
            if current != value:
                if not dry_run:
                    ws.cell(row=row, column=mysportu_col, value=value)
                    ws.cell(row=row, column=mysportu_col).fill = COLOR_SUCCESS
                updated += 1
                if verbose:
                    console.print(f"    {equipe} ({genre}): {value}")

    if not dry_run:
        wb.save(config_excel_path)

    return updated


def update_gymnases_excel_column(
    config_excel_path: Path,
    dry_run: bool = False,
    verbose: bool = False,
) -> int:
    """Met à jour la colonne 'MySportU' dans la feuille 'Gymnases'."""
    wb = load_workbook(config_excel_path)
    ws = wb['Gymnases']

    headers: Dict[Any, int] = {cell.value: cell.column for cell in ws[1]}  # type: ignore[misc]
    gymnase_col = headers.get('Gymnase', 0)
    if not gymnase_col:
        return 0

    if 'MySportU' not in headers:
        mysportu_col = max(headers.values()) + 1
        ws.cell(row=1, column=mysportu_col, value='MySportU')
        ws.cell(row=1, column=mysportu_col).font = Font(bold=True)
    else:
        mysportu_col = headers['MySportU']

    # Inverser le mapping : config -> [mysportu]
    reverse_map: dict[str, list[str]] = {}
    for msu_name, cfg_name in GYMNASE_MAP.items():
        if cfg_name:
            reverse_map.setdefault(cfg_name, []).append(msu_name)

    updated = 0
    for row in range(2, ws.max_row + 1):
        gymnase = ws.cell(row=row, column=gymnase_col).value
        if not gymnase or gymnase not in reverse_map:
            continue

        value = ' | '.join(sorted(reverse_map[gymnase]))
        current = ws.cell(row=row, column=mysportu_col).value
        if current != value:
            if not dry_run:
                ws.cell(row=row, column=mysportu_col, value=value)
                ws.cell(row=row, column=mysportu_col).fill = COLOR_SUCCESS
            updated += 1
            if verbose:
                console.print(f"    {gymnase}: {value}")

    if not dry_run:
        wb.save(config_excel_path)

    return updated


# =============================================================================
# PARTIE 2 : SYNCHRONISATION DES MATCHS
# =============================================================================

def sync_matches(
    config_excel_path: Path,
    sync_data: list[SyncMatch],
    start_date: datetime,
    valid_gymnases: Set[str],
    team_poules: Dict[Tuple[str, str], str],
    competition_type: str = 'Acad',
    fix_scores: bool = False,
    dry_run: bool = False,
    verbose: bool = False,
) -> SyncResult:
    """
    Synchronise les matchs avec la configuration Excel.

    - Ajoute les nouveaux matchs (non annulés)
    - Met à jour les scores des matchs existants
    - Gère les ententes et matchs sans date/gymnase
    - Ne filtre PAS par jour : tous les matchs sont conservés
    """
    result = SyncResult()
    existing = load_existing_fixed_matches(config_excel_path)

    matches_to_add: list[dict] = []
    matches_to_update: list[Tuple[int, dict]] = []

    for sm in sync_data:
        # ── Poule ──
        poule_a = team_poules.get((sm.eq_a_config, sm.genre))
        poule_b = team_poules.get((sm.eq_b_config, sm.genre))
        match_poule = None
        if poule_a and poule_b and poule_a == poule_b:
            match_poule = poule_a
        elif poule_a and poule_b and poule_a != poule_b:
            result.issues.append(Issue(
                type=IssueType.POULE_MISMATCH,
                message="Poules différentes",
                mysportu_data=f"{sm.eq_a_config} vs {sm.eq_b_config}",
                config_data=f"{poule_a} vs {poule_b}",
                severity="info",
            ))

        # ── Semaine ──
        semaine_num = calculate_week_number(sm.date_str, start_date)
        semaine_display = format_week_display(semaine_num, start_date) if semaine_num else None

        # ── Gymnase ──
        # Pour les ententes : utiliser le gymnase réel MySportU s'il existe,
        # sinon marquer 'ENTENTE'
        if sm.is_entente:
            if sm.gym_resolved and sm.gym_resolved in valid_gymnases:
                # Match entente mais avec un vrai gymnase connu → utiliser le gym réel
                gymnase = sm.gym_resolved
                if verbose:
                    console.print(f"    🤝 Entente résolue: {sm.eq_a_config} vs {sm.eq_b_config} "
                                   f"({sm.genre}) - {sm.date_str} → {gymnase}")
            else:
                # Pas de gymnase connu → ENTENTE
                gymnase = 'ENTENTE'
                if verbose:
                    console.print(f"    🤝 Entente (sans gym): {sm.eq_a_config} vs {sm.eq_b_config} "
                                   f"({sm.genre}) - {sm.date_str}")
        else:
            gymnase = normalize_gymnase(sm.lieu_libelle)
            if gymnase and gymnase not in valid_gymnases:
                result.issues.append(Issue(
                    type=IssueType.GYM_NOT_FOUND,
                    message="Gymnase non trouvé dans la config",
                    mysportu_data=sm.lieu_libelle,
                    config_data=gymnase,
                    severity="warning",
                ))

        # ── Avertissements pour matchs sans date/gymnase ──
        if semaine_num is None and sm.date_str:
            result.matches_kept_no_date += 1
            if not sm.state.is_cancelled:
                result.issues.append(Issue(
                    type=IssueType.DATE_MISSING,
                    message="Semaine hors calendrier",
                    mysportu_data=f"{sm.eq_a_config} vs {sm.eq_b_config} ({sm.genre}) [{sm.date_str}]",
                    severity="info",
                ))
        elif not sm.date_str and not sm.state.is_cancelled:
            result.matches_kept_no_date += 1

        if gymnase is None and not sm.is_entente and sm.lieu_libelle:
            result.matches_kept_no_gym += 1
            result.issues.append(Issue(
                type=IssueType.GYM_NOT_FOUND,
                message="Gymnase inconnu",
                mysportu_data=sm.lieu_libelle,
                config_data=f"{sm.eq_a_config} vs {sm.eq_b_config}",
                severity="info",
            ))

        # ── Score (seulement pour matchs terminés) ──
        score = sm.score if sm.state == MatchState.TERMINE else None

        # ── Date formatée ──
        date_formatted = format_date_ddmmyy(sm.date_str)

        # ── Chercher le match existant ──
        existing_idx, is_reversed = find_match_in_existing(
            sm.eq_a_config, sm.eq_b_config, sm.genre, existing,
        )

        if existing_idx is not None:
            result.duplicates_found += 1
            existing_row = existing.loc[existing_idx]
            existing_score = existing_row.get('Score')
            existing_gymnase = str(existing_row.get('Gymnase', '')).strip()
            existing_date = str(existing_row.get('Date', '')).strip() if not pd.isna(existing_row.get('Date')) else ''
            existing_horaire = str(existing_row.get('Horaire', '')).strip() if not pd.isna(existing_row.get('Horaire')) else ''
            existing_semaine = str(existing_row.get('Semaine', '')).strip() if not pd.isna(existing_row.get('Semaine')) else ''

            # ── Mise à jour des matchs ENTENTE existants ──
            # Si le match existant a gymnase='ENTENTE' et que MySportU fournit
            # maintenant un vrai gymnase + date, mettre à jour
            updates_for_existing: dict = {}

            if existing_gymnase.upper() == 'ENTENTE' and gymnase and gymnase != 'ENTENTE':
                # L'entente a été résolue avec un vrai gymnase
                updates_for_existing['Gymnase'] = gymnase
                result.ententes_resolved += 1
                result.issues.append(Issue(
                    type=IssueType.ENTENTE_RESOLVED,
                    message=f"Entente résolue: {sm.eq_a_config} vs {sm.eq_b_config} ({sm.genre})",
                    mysportu_data=f"Gym: {gymnase}, Date: {date_formatted or '?'}",
                    config_data=f"Ancien gym: ENTENTE",
                    severity="info",
                ))
                if verbose:
                    console.print(f"    ✅ Entente résolue: {sm.eq_a_config} vs {sm.eq_b_config} "
                                   f"({sm.genre}) → {gymnase}")

            # ── Mise à jour date/horaire/semaine si MySportU a des données plus récentes ──
            if date_formatted and existing_date and date_formatted != existing_date:
                updates_for_existing['Date'] = date_formatted
                if semaine_display:
                    updates_for_existing['Semaine'] = semaine_display
                result.issues.append(Issue(
                    type=IssueType.DATA_MISMATCH,
                    message=f"Date différente: {sm.eq_a_config} vs {sm.eq_b_config} ({sm.genre})",
                    mysportu_data=f"{date_formatted}",
                    config_data=f"{existing_date}",
                    severity="warning",
                ))
            elif date_formatted and not existing_date:
                updates_for_existing['Date'] = date_formatted
                if semaine_display:
                    updates_for_existing['Semaine'] = semaine_display

            if sm.heure and existing_horaire and sm.heure != existing_horaire:
                updates_for_existing['Horaire'] = sm.heure
                result.issues.append(Issue(
                    type=IssueType.DATA_MISMATCH,
                    message=f"Horaire différent: {sm.eq_a_config} vs {sm.eq_b_config} ({sm.genre})",
                    mysportu_data=f"{sm.heure}",
                    config_data=f"{existing_horaire}",
                    severity="warning",
                ))
            elif sm.heure and not existing_horaire:
                updates_for_existing['Horaire'] = sm.heure

            # Gymnase non-ENTENTE : vérifier la cohérence
            if (gymnase and gymnase != 'ENTENTE'
                    and existing_gymnase and existing_gymnase.upper() != 'ENTENTE'
                    and gymnase != existing_gymnase):
                result.issues.append(Issue(
                    type=IssueType.DATA_MISMATCH,
                    message=f"Gymnase différent: {sm.eq_a_config} vs {sm.eq_b_config} ({sm.genre})",
                    mysportu_data=f"{gymnase}",
                    config_data=f"{existing_gymnase}",
                    severity="warning",
                ))

            # Semaine : vérifier cohérence
            if semaine_display and existing_semaine:
                existing_wn = parse_week_number(existing_semaine)
                new_wn = parse_week_number(semaine_display)
                if existing_wn and new_wn and existing_wn != new_wn:
                    updates_for_existing['Semaine'] = semaine_display
                    result.issues.append(Issue(
                        type=IssueType.DATA_MISMATCH,
                        message=f"Semaine différente: {sm.eq_a_config} vs {sm.eq_b_config} ({sm.genre})",
                        mysportu_data=f"{semaine_display}",
                        config_data=f"{existing_semaine}",
                        severity="warning",
                    ))

            if updates_for_existing:
                # Ajouter une remarque
                remarks = []
                if 'Gymnase' in updates_for_existing:
                    remarks.append(f"Gym: {existing_gymnase}→{updates_for_existing['Gymnase']}")
                if 'Date' in updates_for_existing:
                    remarks.append(f"Date: {existing_date or '?'}→{updates_for_existing['Date']}")
                if 'Horaire' in updates_for_existing:
                    remarks.append(f"Heure: {existing_horaire or '?'}→{updates_for_existing['Horaire']}")
                if 'Semaine' in updates_for_existing:
                    remarks.append(f"Sem: {existing_semaine or '?'}→{updates_for_existing['Semaine']}")
                updates_for_existing['Remarques'] = 'MAJ MySportU: ' + ', '.join(remarks)
                matches_to_update.append((existing_idx, updates_for_existing))
                result.matches_updated += 1
                if verbose:
                    console.print(f"    🔄 MAJ: {sm.eq_a_config} vs {sm.eq_b_config} ({sm.genre}) "
                                   f"— {', '.join(remarks)}")

            # ── Mise à jour du score ──
            if score:
                score_to_use = reverse_score(score) if is_reversed else score

                if pd.isna(existing_score) or str(existing_score).strip() == '':
                    matches_to_update.append((existing_idx, {
                        'Score': score_to_use,
                        'Remarques': 'Score MySportU',
                    }))
                    result.scores_added += 1
                    if verbose:
                        inv = " (inversé)" if is_reversed else ""
                        console.print(f"    📝 Score: {existing_row['Equipe_1']} vs "
                                       f"{existing_row['Equipe_2']} = {score_to_use}{inv}")
                elif str(existing_score).strip() != score_to_use:
                    if fix_scores:
                        matches_to_update.append((existing_idx, {
                            'Score': score_to_use,
                            'Remarques': f"Score corrigé (était {existing_score})",
                        }))
                        result.scores_corrected += 1
                        if verbose:
                            console.print(f"    ✏️  Score corrigé: "
                                           f"{existing_row['Equipe_1']} vs {existing_row['Equipe_2']}: "
                                           f"{existing_score} → {score_to_use}")
                    else:
                        result.issues.append(Issue(
                            type=IssueType.SCORE_CONFLICT,
                            message=f"{existing_row['Equipe_1']} vs {existing_row['Equipe_2']} ({sm.genre})",
                            mysportu_data=score_to_use,
                            config_data=str(existing_score),
                            severity="warning",
                        ))
        else:
            # Nouveau match — n'ajouter que si pas annulé
            if sm.state.is_cancelled:
                continue

            # Un nouveau match sans semaine ET sans gymnase n'est pas importable
            if semaine_num is None and gymnase is None:
                continue

            # Remarque : indiquer si c'est un match entente
            remarque = 'Import MySportU'
            if sm.is_entente and gymnase != 'ENTENTE':
                remarque = f'Import MySportU (entente, gym résolu: {gymnase})'
            elif sm.is_entente:
                remarque = 'Import MySportU (entente)'

            match_data = {
                'Equipe_1': sm.eq_a_config,
                'Equipe_2': sm.eq_b_config,
                'Genre': sm.genre,
                'Poule': match_poule,
                'Semaine': semaine_display,
                'Date': date_formatted,
                'Horaire': sm.heure,
                'Gymnase': gymnase,
                'Score': score,
                'Type_Competition': competition_type,
                'Remarques': remarque,
                'Arbitres': None,
                'Ignorer': None,
            }
            matches_to_add.append(match_data)

    # ── Appliquer les modifications ──
    if not dry_run and (matches_to_update or matches_to_add):
        wb = load_workbook(config_excel_path)
        ws = wb['Matchs_Fixes']
        headers: Dict[Any, int] = {cell.value: cell.column for cell in ws[1]}  # type: ignore[misc]

        # Scores existants
        for idx, updates in matches_to_update:
            excel_row = idx + 2
            for col_name, value in updates.items():
                if col_name in headers:
                    col = headers[col_name]
                    ws.cell(row=excel_row, column=col, value=value)
                    ws.cell(row=excel_row, column=col).fill = COLOR_INFO

        # Nouveaux matchs
        start_row = ws.max_row + 1
        for i, match in enumerate(matches_to_add):
            row_num = start_row + i
            for col_name, value in match.items():
                if col_name in headers and value is not None:
                    col_idx = headers[col_name]
                    ws.cell(row=row_num, column=col_idx, value=value)
                    ws.cell(row=row_num, column=col_idx).fill = COLOR_SUCCESS

        wb.save(config_excel_path)

    result.matches_added = len(matches_to_add)
    return result


def remove_cancelled_matches(
    config_excel_path: Path,
    sync_data: list[SyncMatch],
    competition_type: str = 'Acad',
    dry_run: bool = False,
    verbose: bool = False,
) -> Tuple[int, list[dict]]:
    """
    Supprime les matchs config dont le correspondant MySportU est annulé/reporté.

    Ne touche QUE les matchs du type ``competition_type`` (Acad par défaut).
    """
    cancelled_keys: set[frozenset] = set()
    for sm in sync_data:
        if sm.state.is_cancelled:
            cancelled_keys.add(create_match_key(sm.eq_a_config, sm.eq_b_config, sm.genre))

    if not cancelled_keys:
        return 0, []

    df = load_existing_fixed_matches(config_excel_path)
    rows_to_remove: list[int] = []
    removed: list[dict] = []

    for idx, row in df.iterrows():
        type_comp = str(row.get('Type_Competition', '')).strip().lower()
        if type_comp != competition_type.lower():
            continue

        eq1 = row.get('Equipe_1')
        eq2 = row.get('Equipe_2')
        genre = row.get('Genre')

        if pd.isna(eq1) or pd.isna(eq2) or pd.isna(genre):
            continue

        key = create_match_key(eq1, eq2, genre)
        if key in cancelled_keys:
            rows_to_remove.append(int(idx))  # type: ignore[arg-type]
            removed.append({
                'Equipe_1': eq1, 'Equipe_2': eq2,
                'Genre': genre, 'Semaine': row.get('Semaine'),
            })
            if verbose:
                console.print(f"    🗑️  {eq1} vs {eq2} ({genre})")

    if not dry_run and rows_to_remove:
        wb = load_workbook(config_excel_path)
        ws = wb['Matchs_Fixes']
        for idx in sorted(rows_to_remove, reverse=True):
            ws.delete_rows(idx + 2)
        wb.save(config_excel_path)

    return len(rows_to_remove), removed


# =============================================================================
# TRI ET NORMALISATION
# =============================================================================

def parse_poule_for_sorting(poule) -> Tuple[int, int]:
    """Parse une poule pour le tri (niveau, lettre_poule)."""
    if not poule or pd.isna(poule):
        return (99, 99)
    m = re.match(r'VB[FM]A(\d)P([A-F])', str(poule))
    if m:
        return (int(m.group(1)), ord(m.group(2)) - ord('A'))
    return (99, 99)


def match_sort_key(row: pd.Series, start_date: datetime) -> Tuple:
    """
    Clé de tri pour un match.

    Ordre : Semaine → Non-Acad en premier → Date → Entente en dernier
            → Genre (F avant M) → Niveau → Pool
    """
    week = parse_week_number(row.get('Semaine')) or 999

    type_comp = str(row.get('Type_Competition', '')).strip().lower()
    is_acad = 1 if type_comp == 'acad' else 0

    date_str = row.get('Date')
    if pd.isna(date_str):
        date_parsed = datetime(2099, 12, 31)
    else:
        date_parsed = datetime(2099, 12, 31)
        for fmt in ('%d/%m/%y', '%d/%m/%Y'):
            try:
                date_parsed = datetime.strptime(str(date_str), fmt)
                break
            except ValueError:
                continue

    gymnase = str(row.get('Gymnase', '')).strip().upper()
    is_entente = 1 if gymnase == 'ENTENTE' else 0
    genre_order = 0 if row.get('Genre') == 'F' else 1
    niveau, pool = parse_poule_for_sorting(row.get('Poule'))

    return (week, is_acad, date_parsed, is_entente, genre_order, niveau, pool)


def sort_and_normalize_weeks(
    config_excel_path: Path,
    start_date: datetime,
    dry_run: bool = False,
    verbose: bool = False,
) -> int:
    """
    Trie les matchs et uniformise le format des semaines.

    Returns:
        Nombre de matchs réordonnés
    """
    df = pd.read_excel(config_excel_path, sheet_name='Matchs_Fixes')
    if df.empty:
        return 0

    # Normaliser les semaines
    old_semaines = df['Semaine'].astype(str).tolist()

    def normalize_week(val):
        wn = parse_week_number(val)
        return format_week_display(wn, start_date) if wn else val

    df['Semaine'] = df['Semaine'].apply(normalize_week)
    new_semaines = df['Semaine'].astype(str).tolist()
    weeks_changed = sum(1 for o, n in zip(old_semaines, new_semaines) if o != n)

    # Trier
    sort_keys = df.apply(lambda r: match_sort_key(r, start_date), axis=1)
    sorted_indices = sorted(range(len(df)), key=lambda i: sort_keys.iloc[i])
    df_sorted = df.iloc[sorted_indices].reset_index(drop=True)

    order_changes = sum(1 for i, orig_idx in enumerate(sorted_indices) if i != orig_idx)
    has_changes = weeks_changed > 0 or order_changes > 0

    if verbose:
        if weeks_changed > 0:
            console.print(f"    {weeks_changed} semaines normalisées")
        console.print(f"    {order_changes} matchs réordonnés sur {len(df)}")

    if not dry_run and has_changes:
        wb = load_workbook(config_excel_path)
        ws = wb['Matchs_Fixes']
        headers = [str(cell.value) for cell in ws[1] if cell.value]

        # Effacer les données
        for row in range(2, ws.max_row + 1):
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).value = None
                ws.cell(row=row, column=col).fill = PatternFill()

        # Réécrire triées
        for i, (_, row_data) in enumerate(df_sorted.iterrows()):
            excel_row = i + 2
            for j, col_name in enumerate(headers):
                if col_name in row_data.index:
                    value = row_data[col_name]
                    if pd.notna(value):
                        ws.cell(row=excel_row, column=j + 1, value=value)

        wb.save(config_excel_path)

    return order_changes


# =============================================================================
# AFFICHAGE RICH
# =============================================================================

def print_state_summary(sync_data: list[SyncMatch]) -> None:
    """Affiche un résumé des états des matchs dans un tableau Rich."""
    counts: dict[MatchState, int] = {}
    for sm in sync_data:
        counts[sm.state] = counts.get(sm.state, 0) + 1

    table = Table(title="États des matchs MySportU", show_lines=False, padding=(0, 1))
    table.add_column("État", style="bold")
    table.add_column("Icône")
    table.add_column("Nombre", justify="right", style="cyan")

    for state in MatchState:
        if state in counts:
            table.add_row(state.label, state.icon, str(counts[state]))

    total = sum(counts.values())
    table.add_section()
    table.add_row("Total", "📊", str(total), style="bold")

    console.print(table)


def print_sync_result(result: SyncResult, dry_run: bool = False) -> None:
    """Affiche le résultat de synchronisation dans un panel Rich."""
    lines: list[str] = []
    if result.matches_added:
        lines.append(f"  ➕ {result.matches_added} matchs ajoutés{' (simulation)' if dry_run else ''}")
    if result.matches_updated:
        lines.append(f"  🔄 {result.matches_updated} matchs mis à jour (date/gym/horaire)")
    if result.ententes_resolved:
        lines.append(f"  🤝 {result.ententes_resolved} ententes résolues (vrai gymnase)")
    if result.scores_added:
        lines.append(f"  📝 {result.scores_added} scores ajoutés")
    if result.scores_corrected:
        lines.append(f"  ✏️  {result.scores_corrected} scores corrigés")
    if result.duplicates_found:
        lines.append(f"  🔄 {result.duplicates_found} matchs existants")
    if result.matches_removed:
        lines.append(f"  🗑️  {result.matches_removed} matchs supprimés")
    if result.matches_reordered:
        lines.append(f"  📊 {result.matches_reordered} matchs réordonnés")
    if result.matches_kept_no_date:
        lines.append(f"  ⏳ {result.matches_kept_no_date} matchs sans date")
    if result.matches_kept_no_gym:
        lines.append(f"  🏠 {result.matches_kept_no_gym} matchs gymnase inconnu")

    if not lines:
        lines.append("  Aucune modification")

    content = "\n".join(lines)
    title = "Résultat (simulation)" if dry_run else "Résultat"
    style = "yellow" if dry_run else "green"
    console.print(Panel(content, title=title, border_style=style))


def print_issues(issues: list[Issue], max_per_type: int = 10) -> None:
    """Affiche les incohérences dans des tableaux Rich par type."""
    if not issues:
        console.print("\n[green]✅ Aucune incohérence détectée[/green]")
        return

    by_type: dict[IssueType, list[Issue]] = {}
    for issue in issues:
        by_type.setdefault(issue.type, []).append(issue)

    severity_styles = {"error": "red", "warning": "yellow", "info": "blue"}
    severity_icons = {"error": "❌", "warning": "⚠️", "info": "ℹ️"}

    for issue_type, type_issues in by_type.items():
        sev = type_issues[0].severity
        style = severity_styles.get(sev, "white")
        icon = severity_icons.get(sev, "•")

        table = Table(
            title=f"{icon} {issue_type.value} ({len(type_issues)})",
            show_lines=False, padding=(0, 1),
            title_style=style,
        )
        table.add_column("Message")
        table.add_column("MySportU", style="dim")
        table.add_column("Config", style="dim")

        for issue in type_issues[:max_per_type]:
            table.add_row(
                issue.message,
                issue.mysportu_data or "",
                issue.config_data or "",
            )

        if len(type_issues) > max_per_type:
            table.add_row(f"... et {len(type_issues) - max_per_type} autres", "", "")

        console.print(table)
        console.print()


def print_entente_summary(sync_data: list[SyncMatch], valid_gymnases: Set[str] = None) -> None:
    """Affiche les matchs en entente, distinguant résolues et non résolues."""
    ententes = [sm for sm in sync_data if sm.is_entente]
    if not ententes:
        return

    table = Table(title=f"🤝 Matchs en entente ({len(ententes)})", show_lines=False)
    table.add_column("Match")
    table.add_column("Genre")
    table.add_column("Date")
    table.add_column("Gymnase MySportU")
    table.add_column("Statut")
    table.add_column("État")

    for sm in ententes:
        gym_display = sm.gym_resolved or sm.lieu_libelle or "?"
        if sm.gym_resolved and valid_gymnases and sm.gym_resolved in valid_gymnases:
            statut = "[green]Résolu[/green]"
        elif sm.gym_resolved:
            statut = "[yellow]Gym inconnu config[/yellow]"
        else:
            statut = "[red]Non résolu[/red]"

        table.add_row(
            f"{sm.eq_a_config} vs {sm.eq_b_config}",
            sm.genre,
            sm.date_str or "?",
            gym_display,
            statut,
            sm.state.icon,
        )

    console.print(table)


# =============================================================================
# FONCTION PRINCIPALE
# =============================================================================

def main() -> int:
    parser = argparse.ArgumentParser(
        description='Synchronisation MySportU ↔ Configuration Excel.',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )

    parser.add_argument(
        '--config', '-c', type=Path,
        default=Path('configs/config_volleyP2.yaml'),
        help='Fichier de configuration YAML',
    )

    # Source
    source = parser.add_argument_group('Source des données')
    source.add_argument('--excel', action='store_true',
                        help="Utiliser un fichier Excel au lieu de l'API")
    source.add_argument('--mysportu', '-m', type=Path, default=None,
                        help='Fichier Excel MySportU (avec --excel)')
    source.add_argument('--sport', type=str, default='VB',
                        help='Code sport pour le filtre (default: VB)')

    # Modes
    mode = parser.add_argument_group("Modes d'exécution")
    mode.add_argument('--mapping', action='store_true',
                      help='Mettre à jour les colonnes MySportU')
    mode.add_argument('--sync', action='store_true',
                      help='Synchroniser les matchs')
    mode.add_argument('--all', action='store_true',
                      help='Mapping + Sync')

    # Options sync
    opts = parser.add_argument_group('Options de synchronisation')
    opts.add_argument('--no-clean', action='store_true',
                      help='Ne pas supprimer les matchs reportés')
    opts.add_argument('--no-sort', action='store_true',
                      help='Ne pas trier les matchs')
    opts.add_argument('--fix-scores', action='store_true',
                      help='Corriger les scores en conflit')
    opts.add_argument('--force', action='store_true',
                      help='Ignorer le cache MySportU')

    # Général
    parser.add_argument('--championship', '-p', type=str, default='PH2',
                        help='Pattern championnat (default: PH2)')
    parser.add_argument('--type', '-t', type=str, default='Acad',
                        help='Type de compétition (default: Acad)')
    parser.add_argument('--dry-run', '-n', action='store_true',
                        help='Aperçu sans modification')
    parser.add_argument('--verbose', '-v', action='store_true',
                        help='Mode verbeux')
    parser.add_argument('--username', type=str, default=None,
                        help='Identifiant MySportU (sinon: env ou configs/default.yaml)')
    parser.add_argument('--password', type=str, default=None,
                        help='Mot de passe MySportU (sinon: env ou configs/default.yaml)')

    args = parser.parse_args()

    # Validation
    if not any([args.mapping, args.sync, args.all]):
        parser.print_help()
        console.print("\n[red]❌ Spécifiez un mode: --mapping, --sync, ou --all[/red]")
        return 1

    # ── Charger la configuration YAML ──
    console.print(Panel.fit(
        f"[bold]Synchronisation MySportU[/bold]\n"
        f"Config: {args.config}\n"
        f"Mode: {'Simulation' if args.dry_run else 'Production'}",
        border_style="blue",
    ))

    config = load_config(args.config)
    data_file = Path(config['fichiers']['donnees'])
    if not data_file.is_absolute():
        data_file = args.config.parent.parent / data_file

    cal_config = config.get('calendrier', {})
    date_debut_str = cal_config.get('date_debut', '2026-01-22')
    start_date = datetime.strptime(date_debut_str, '%Y-%m-%d')
    jour_match = cal_config.get('jour_match', 'jeudi')

    console.print(f"  📂 Excel config: [bold]{data_file.name}[/bold]")
    console.print(f"  📅 Début saison: [bold]{start_date.strftime('%d/%m/%Y')}[/bold] "
                   f"(jour: {jour_match})")

    if args.dry_run:
        console.print("\n  [yellow]🔍 MODE SIMULATION — aucune modification[/yellow]\n")

    # ── Charger les données config ──
    console.print("\n[bold]📥 Chargement des données config...[/bold]")
    df_config_teams = load_config_teams(data_file)
    valid_gymnases = get_gymnases(data_file)
    team_poules = get_team_poules(data_file)
    valid_teams: Set[Tuple[str, str]] = {
        (row['Equipe'], row['Genre_Equipe']) for _, row in df_config_teams.iterrows()
    }

    console.print(f"  • {len(df_config_teams)} équipes config")
    console.print(f"  • {len(valid_gymnases)} gymnases")

    # ── Récupérer les matchs ──
    all_issues: list[Issue] = []
    sync_data: list[SyncMatch] = []
    all_matches: list[MatchInfo] = []

    if args.excel:
        # Mode Excel
        if args.mysportu is None:
            mysportu_file = data_file.parent / 'mysportu.xlsx'
            if not mysportu_file.exists():
                candidates = list(data_file.parent.glob('*mysportu*.xlsx'))
                if not candidates:
                    console.print(f"[red]❌ Aucun fichier MySportU trouvé dans {data_file.parent}[/red]")
                    return 1
                mysportu_file = candidates[0]
        else:
            mysportu_file = args.mysportu

        console.print(f"\n[bold]📄 Source: Excel ({mysportu_file.name})[/bold]")
        df_mysportu = load_mysportu_excel(mysportu_file, args.championship)
        sync_data, conv_issues = excel_to_sync(df_mysportu, valid_teams, jour_match)
        all_issues.extend(conv_issues)
        console.print(f"  • {len(df_mysportu)} matchs Excel → {len(sync_data)} matchs résolus")
    else:
        # Mode API (défaut)
        console.print(f"\n[bold]🌐 Source: API MySportU[/bold] "
                       f"(sport={args.sport}, championnat={args.championship})")

        try:
            msu = MySportU(
                config_path=args.config,
                username=args.username,
                password=args.password,
                verbose=args.verbose,
                cache_enabled=not args.force,
            )
            with msu:
                all_matches = msu.get_matches(
                    sport=args.sport,
                    championship=args.championship,
                    force_refresh=args.force,
                )
        except Exception as e:
            console.print(f"[red]❌ Erreur de connexion API: {e}[/red]")
            console.print("[dim]💡 Utilisez --excel pour un fichier Excel MySportU[/dim]")
            return 1

        sync_data, conv_issues = matches_to_sync(all_matches, valid_teams, jour_match)
        all_issues.extend(conv_issues)
        console.print(f"  • {len(all_matches)} matchs API → {len(sync_data)} matchs résolus")

    if not sync_data:
        console.print("[yellow]⚠️  Aucun match à traiter[/yellow]")
        if all_issues:
            print_issues(all_issues)
        return 1

    # Résumé des états
    print_state_summary(sync_data)

    # Ententes
    print_entente_summary(sync_data)

    # =========================================================================
    # MODE: MAPPING
    # =========================================================================
    if args.mapping or args.all:
        console.rule("[bold blue]MAPPING")

        if args.excel:
            console.print("[yellow]⚠️  Mapping depuis Excel non supporté — utilisez l'API[/yellow]")
        else:
            console.print("\n🔄 Construction du mapping des équipes...")
            mapping, mapping_issues = build_team_mapping_from_api(
                all_matches, valid_teams, verbose=args.verbose,
            )
            all_issues.extend(mapping_issues)

            console.print("✏️  Mise à jour colonne MySportU (Equipes)...")
            updated_teams = update_teams_excel_column(
                data_file, mapping, dry_run=args.dry_run, verbose=args.verbose,
            )
            act = "seraient mises à jour" if args.dry_run else "mises à jour ✓"
            console.print(f"  • {updated_teams} équipes {act}")

            console.print("✏️  Mise à jour colonne MySportU (Gymnases)...")
            updated_gyms = update_gymnases_excel_column(
                data_file, dry_run=args.dry_run, verbose=args.verbose,
            )
            act = "seraient mis à jour" if args.dry_run else "mis à jour ✓"
            console.print(f"  • {updated_gyms} gymnases {act}")

    # =========================================================================
    # MODE: SYNCHRONISATION
    # =========================================================================
    if args.sync or args.all:
        console.rule("[bold blue]SYNCHRONISATION")

        # 1. Sync des matchs
        console.print("\n🔄 Synchronisation des matchs...")
        if args.fix_scores:
            console.print("  [yellow]⚠️  Correction des scores activée[/yellow]")

        sync_result = sync_matches(
            data_file, sync_data, start_date,
            valid_gymnases, team_poules,
            competition_type=args.type,
            fix_scores=args.fix_scores,
            dry_run=args.dry_run,
            verbose=args.verbose,
        )
        all_issues.extend(sync_result.issues)

        # 2. Supprimer les matchs reportés
        if not args.no_clean:
            console.print("\n🗑️  Suppression des matchs reportés/annulés...")
            removed_count, _ = remove_cancelled_matches(
                data_file, sync_data,
                competition_type=args.type,
                dry_run=args.dry_run,
                verbose=args.verbose,
            )
            sync_result.matches_removed = removed_count
            act = "seraient supprimés" if args.dry_run else "supprimés ✓"
            console.print(f"  • {removed_count} matchs {act}")

        # 3. Trier et normaliser
        if not args.no_sort:
            console.print("\n📊 Tri et normalisation des semaines...")
            reordered = sort_and_normalize_weeks(
                data_file, start_date,
                dry_run=args.dry_run,
                verbose=args.verbose,
            )
            sync_result.matches_reordered = reordered
            act = "seraient réordonnés" if args.dry_run else "réordonnés ✓"
            console.print(f"  • {reordered} matchs {act}")

        # Résultat
        console.print()
        print_sync_result(sync_result, dry_run=args.dry_run)

    # ── Incohérences ──
    print_issues(all_issues)

    # ── Fin ──
    console.print()
    if args.dry_run:
        console.print("[yellow]✅ Vérification terminée (aucune modification)[/yellow]")
    else:
        console.print("[green]✅ Synchronisation terminée ![/green]")

    return 0


if __name__ == '__main__':
    exit(main())
