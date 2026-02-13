#!/usr/bin/env python3
from __future__ import annotations
"""
Script de synchronisation MySportU <-> Configuration Excel.

Ce script gère deux opérations distinctes :

1. MAPPING (--mapping) :
   - Met à jour la colonne 'MySportU' dans la feuille 'Equipes'
   - Met à jour la colonne 'MySportU' dans la feuille 'Gymnases'
   - Établit la correspondance entre les noms MySportU et les noms de la config

2. SYNCHRONISATION (--sync) :
   - Récupère les matchs depuis l'API MySportU (par défaut) ou un fichier Excel (--excel)
   - Importe les nouveaux matchs vers 'Matchs_Fixes'
   - Met à jour les scores des matchs existants
   - Supprime les matchs Acad reportés/annulés (par défaut, --no-clean pour désactiver)
   - Trie les matchs par semaine/type/date/genre/niveau/poule (par défaut, --no-sort pour désactiver)

IMPORTANT :
- Les matchs CFU/CFE ne sont PAS dans MySportU, ils sont préservés et repositionnés
- Le genre est TOUJOURS pris en compte (une équipe peut avoir le même nom en M et F)
- Le format de la colonne Semaine est uniformisé : "N (dd/mm)"
- Les matchs hors jeudi (ententes) sont inclus et marqués comme ENTENTE

Usage:
    # Vérifier le mapping sans modifier
    python sync_mysportu.py --mapping --dry-run

    # Mettre à jour le mapping
    python sync_mysportu.py --mapping

    # Synchroniser les matchs depuis l'API (par défaut)
    python sync_mysportu.py --sync

    # Synchroniser depuis un fichier Excel exporté
    python sync_mysportu.py --sync --excel
    python sync_mysportu.py --sync --excel --mysportu data/volleyball/mysportu.xlsx

    # Synchroniser sans supprimer les reportés
    python sync_mysportu.py --sync --no-clean
    
    # Tout faire
    python sync_mysportu.py --all
"""

import re
import sys
import time
import argparse
from datetime import datetime
from pathlib import Path
from typing import Optional, Tuple, Dict, List, Set, Any
from dataclasses import dataclass, field
from enum import Enum

import requests
from bs4 import BeautifulSoup
import pandas as pd
import yaml
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font


# =============================================================================
# CONSTANTES ET CONFIGURATION
# =============================================================================

# API MySportU
BASE_URL = "https://gestion.mysportu.com"
LOGIN_URL = f"{BASE_URL}/auth/login"
DEFAULT_USERNAME = "0838827"
DEFAULT_PASSWORD = "CheeGliFFSU2!"

# Couleurs Excel pour le feedback visuel
COLOR_ERROR = PatternFill(start_color='FFCCCB', end_color='FFCCCB', fill_type='solid')
COLOR_WARNING = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
COLOR_SUCCESS = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
COLOR_INFO = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
COLOR_ENTENTE = PatternFill(start_color='E8DAEF', end_color='E8DAEF', fill_type='solid')

# Mapping code institution MySportU -> nom court config
INSTITUTION_CODE_MAP = {
    '069069001': 'ENTPE',
    '069069006': 'LYON 2',
    '069069007': 'LYON 3',
    '069069008': 'ESA',
    '069069009': 'ISARA',
    '069069012': 'CATHO',
    '069069014': 'CPE',
    '069069015': 'ECAM',
    '069069016': 'ECL',
    '069069018': 'EML',
    '069069019': 'ENS',
    '069069020': 'ENTPE',
    '069069021': 'ESME',
    '069069022': 'ESSCA',
    '069069025': 'INSA',
    '069069026': 'ISOSTEO',
    '069069028': 'VETO',
    '069069029': 'CESI',
    '069069031': 'LYON 1',
}

# Mapping gymnases MySportU -> gymnases config
GYMNASE_MAP = {
    'CENTRALE': 'ECL',
    'COMPET C (HAUT) - LEON JOUHAUX': 'L. J. HAUT',
    'COMPET. LAPLANCHE- R. LISSMANN': None,
    'ENS DESCARTES': 'DESCARTES',
    'ESA BRON': 'ESA',
    'HALLE - 3D': 'LAENNEC',
    'HALLE - C.BESSON': 'BESSON',
    'HALLE C LYON 2': 'LYON 2 HC',
    'COMPET PIBAROT - R.VANEL': None,
    'COMPET. LAPLANCHE- R. LISSMANN': None,
    'N-A': None,
}

# Mapping des noms courts MySportU API -> noms config
# L'API retourne des libellés courts (ex: "ENTPE (1)", "INSA (2)")
# qui sont souvent directement utilisables mais parfois différents
API_TEAM_NAME_MAP = {
    'AS UD Lyon': None,  # Nom générique, pas utilisable
    'ENS LYON': 'ENS',
}


# =============================================================================
# CLIENT API MYSPORTU
# =============================================================================

class MySportUClient:
    """Client pour l'API MySportU (connexion + récupération de matchs)."""
    
    def __init__(self, username: str = DEFAULT_USERNAME, password: str = DEFAULT_PASSWORD):
        self.username = username
        self.password = password
        self.session = requests.Session()
        self.logged_in = False
    
    def login(self) -> bool:
        """Connexion à MySportU via session avec CSRF."""
        try:
            resp = self.session.get(LOGIN_URL)
            if resp.status_code != 200:
                return False
            
            soup = BeautifulSoup(resp.text, 'html.parser')
            csrf_input = soup.find('input', {'name': '_token'})
            if not csrf_input:
                return False
            
            csrf_token = csrf_input.get('value')
            login_data = {
                '_token': csrf_token,
                'username': self.username,
                'password': self.password
            }
            self.session.post(LOGIN_URL, data=login_data, allow_redirects=True)
            self.logged_in = 'ffsu_session' in self.session.cookies
            return self.logged_in
        except Exception as e:
            print(f"  ❌ Erreur de connexion: {e}")
            return False
    
    def _api_get(self, url: str) -> Optional[Dict]:
        """Requête GET JSON à l'API."""
        if not self.logged_in:
            return None
        try:
            resp = self.session.get(url, headers={
                'Accept': 'application/json',
                'X-Requested-With': 'XMLHttpRequest'
            })
            if resp.status_code == 200:
                return resp.json()
        except Exception:
            pass
        return None
    
    def get_all_matches(self) -> List[Dict]:
        """Récupère tous les matchs (paginés)."""
        all_matches = []
        page = 1
        last_page = None
        
        while True:
            url = f"{BASE_URL}/feuille-de-match/rencontres?page={page}"
            data = self._api_get(url)
            
            if not data or 'data' not in data:
                break
            
            matches = data['data']
            if not matches:
                break
            
            meta = data.get('meta', {})
            if last_page is None:
                last_page = meta.get('last_page', 1)
            
            all_matches.extend(matches)
            print(f"\r  📥 Récupération: {len(all_matches)} matchs (page {page}/{last_page})...", end='', flush=True)
            
            if page >= last_page:
                break
            
            page += 1
            time.sleep(0.05)
        
        print()
        return all_matches


def normalize_api_team_name(libelle_court: str, club_code: str, genre: str) -> Optional[str]:
    """
    Normalise un nom d'équipe depuis l'API MySportU vers le format config.
    
    L'API fournit directement un libellé court (ex: "ENTPE (1)", "INSA (2)")
    et le code club qui identifie l'institution.
    
    Args:
        libelle_court: Nom court de l'équipe API (ex: "ENTPE (1)", "SANTE (2)")
        club_code: Code club MySportU (ex: "069069025")
        genre: Genre du match (F/M)
    
    Returns:
        Nom config (ex: "ENTPE (1)", "SANTE (2)") ou None si non résolu
    """
    if not libelle_court or not club_code:
        return None
    
    # Vérifier le mapping explicite
    if libelle_court in API_TEAM_NAME_MAP:
        if API_TEAM_NAME_MAP[libelle_court] is None:
            return None
        return API_TEAM_NAME_MAP[libelle_court]
    
    institution = INSTITUTION_CODE_MAP.get(club_code)
    if not institution:
        return None
    
    # Extraire le numéro d'équipe du libellé court
    # Format typique : "INST (N)" ou "INST N" ou "INST (IEP) (N)"
    num_match = re.search(r'\((\d+)\)$', libelle_court)
    if num_match:
        numero = num_match.group(1)
    else:
        # Parfois juste "INST N" à la fin
        num_match = re.search(r'(\d+)$', libelle_court)
        numero = num_match.group(1) if num_match else '1'
    
    # Cas spéciaux
    clean_name = libelle_court.upper()
    if 'IEP' in clean_name:
        return 'LYON 2 (IEP) (4)'
    
    if 'SANT' in clean_name or 'SANTÉ' in clean_name or 'SANTE' in clean_name:
        return f"SANTE ({numero})"
    
    return f"{institution} ({numero})"


def extract_score_from_api(match_data: Dict) -> Optional[str]:
    """
    Extrait le score au format "X-Y" (receveur-visiteur) depuis les données API.
    
    L'API fournit le score comme une liste de dicts avec equipe_id et score.
    """
    scores = match_data.get('score', [])
    if not scores or len(scores) < 2:
        return None
    
    receveur_id = match_data.get('receveur', {}).get('id')
    visiteur_id = match_data.get('visiteur', {}).get('id')
    
    score_rec = None
    score_vis = None
    
    for s in scores:
        if s.get('equipe_id') == receveur_id:
            score_rec = s.get('score')
        elif s.get('equipe_id') == visiteur_id:
            score_vis = s.get('score')
    
    if score_rec is not None and score_vis is not None:
        return f"{score_rec}-{score_vis}"
    return None


def get_api_match_etat(match_data: Dict) -> Optional[str]:
    """
    Détermine l'état d'un match depuis l'API.
    
    Returns:
        'T' (terminé), 'R' (reporté), 'N' (annulé/forfait), None (non joué)
    """
    etat = match_data.get('etat')
    forfait = match_data.get('forfait')
    
    if forfait is not None:
        return 'N'  # Forfait = annulé
    if etat == 'T':
        return 'T'  # Terminé
    if etat == 'R':
        return 'R'  # Reporté
    if etat == 'N':
        return 'N'  # Annulé
    return None  # Non joué


def is_entente_match(date_str: str, jour_match: str = 'jeudi') -> bool:
    """
    Détermine si un match est un match en entente (joué hors du jour normal).
    
    Args:
        date_str: Date au format "dd/mm/yyyy" ou "dd/mm/yyyy HH:MM"
        jour_match: Jour normal des matchs (default: "jeudi")
    
    Returns:
        True si le match est joué un autre jour que le jour normal
    """
    if not date_str:
        return False
    
    jour_mapping = {
        'lundi': 0, 'mardi': 1, 'mercredi': 2, 'jeudi': 3,
        'vendredi': 4, 'samedi': 5, 'dimanche': 6
    }
    jour_normal = jour_mapping.get(jour_match.lower(), 3)
    
    try:
        date_part = date_str.split(' ')[0]
        dt = datetime.strptime(date_part, '%d/%m/%Y')
        return dt.weekday() != jour_normal
    except (ValueError, IndexError):
        return False


def fetch_matches_from_api(
    championship_pattern: str = 'PH2',
    sport_pattern: str = 'VB',
    verbose: bool = False
) -> Tuple[pd.DataFrame, list]:
    """
    Récupère les matchs depuis l'API MySportU et les convertit en DataFrame.
    
    Le format du DataFrame est compatible avec l'ancien format Excel MySportU
    pour permettre la réutilisation de toute la logique de sync.
    
    Args:
        championship_pattern: Pattern pour filtrer les championnats (ex: 'PH2')
        sport_pattern: Pattern sport (ex: 'VB' pour volleyball)
        verbose: Afficher les détails
    
    Returns:
        Tuple (DataFrame des matchs, liste des issues)
    """
    issues = []
    
    # Connexion
    print("  🔐 Connexion à MySportU...")
    client = MySportUClient()
    if not client.login():
        issues.append(Issue(
            type=IssueType.TEAM_NOT_FOUND,
            message="Impossible de se connecter à MySportU",
            severity="error"
        ))
        return pd.DataFrame(), issues
    print("  ✅ Connecté")
    
    # Récupérer tous les matchs
    all_matches = client.get_all_matches()
    print(f"  📊 {len(all_matches)} matchs récupérés au total")
    
    # Filtrer par championnat et sport
    filtered = []
    for m in all_matches:
        infos = m.get('infosRencontre', {})
        comp = infos.get('competition_libelle', '') or ''
        
        # Vérifier le sport et le championnat
        if sport_pattern.upper() in comp.upper() and championship_pattern.upper() in comp.upper():
            filtered.append(m)
    
    print(f"  🎯 {len(filtered)} matchs {sport_pattern} {championship_pattern}")
    
    if not filtered:
        issues.append(Issue(
            type=IssueType.TEAM_NOT_FOUND,
            message=f"Aucun match trouvé pour {sport_pattern} {championship_pattern}",
            severity="error"
        ))
        return pd.DataFrame(), issues
    
    # Convertir en DataFrame compatible avec l'ancien format
    rows = []
    for m in filtered:
        infos = m.get('infosRencontre', {})
        rec = m.get('receveur', {})
        vis = m.get('visiteur', {})
        
        # Date et heure
        date_rencontre = infos.get('date_rencontre', '')
        date_parts = date_rencontre.split(' ') if date_rencontre else []
        date_str = date_parts[0] if len(date_parts) >= 1 else None
        heure_str = date_parts[1] if len(date_parts) >= 2 else None
        
        # Lieu
        lieu_data = infos.get('lieu_pratique') or {}
        lieu_str = lieu_data.get('libelle') if isinstance(lieu_data, dict) else None
        
        # Score (seulement pour les matchs terminés)
        etat = get_api_match_etat(m)
        score = extract_score_from_api(m) if etat == 'T' else None
        
        # État
        etat_label = {
            'T': 'Terminé', 'R': 'Reporté', 'N': 'Annulé'
        }.get(etat, 'Non joué')
        
        # Poule MySportU
        poule_data = m.get('poule', {})
        poule_msu = poule_data.get('libelle') if isinstance(poule_data, dict) else None
        
        # Construction du nom d'équipe au format Excel MySportU
        # pour rester compatible avec l'ancien format
        rec_code = rec.get('club', {}).get('code', '')
        vis_code = vis.get('club', {}).get('code', '')
        rec_nom = rec.get('club', {}).get('nom', '')
        vis_nom = vis.get('club', {}).get('nom', '')
        rec_court = rec.get('libelle_court', '')
        vis_court = vis.get('libelle_court', '')
        
        # Reconstituer le nom au format "CODE - NOM" pour l'ancien pipeline
        equipe_a = f"{rec_code} - {rec_court}" if rec_code else rec_court
        equipe_b = f"{vis_code} - {vis_court}" if vis_code else vis_court
        
        row = {
            'Discipline': 'Volley 6X6',
            'Championnat': infos.get('competition_libelle', ''),
            'Phase': infos.get('phase_libelle', ''),
            'Journée/Tour': m.get('tour'),
            'Date': date_str,
            'Heure': heure_str,
            'Lieu': lieu_str,
            'Equipe A': equipe_a,
            'Equipe B': equipe_b,
            'Score': score,
            'etat': etat_label,
            # Champs supplémentaires pour la sync API
            '_api_id': m.get('id'),
            '_receveur_id': rec.get('id'),
            '_visiteur_id': vis.get('id'),
            '_receveur_code': rec_code,
            '_visiteur_code': vis_code,
            '_receveur_court': rec_court,
            '_visiteur_court': vis_court,
            '_poule_msu': poule_msu,
            '_etat_code': etat,
        }
        rows.append(row)
    
    df = pd.DataFrame(rows)
    
    # Stats
    etats = df['etat'].value_counts()
    if verbose:
        print(f"\n  📈 Répartition des états:")
        for etat_val, count in etats.items():
            print(f"      {etat_val}: {count}")
    
    # Matchs hors jeudi (ententes)
    matchs_entente = 0
    for _, row in df.iterrows():
        if is_entente_match(row.get('Date', ''), 'jeudi'):
            matchs_entente += 1
    if matchs_entente > 0:
        print(f"  🤝 {matchs_entente} matchs en entente (hors jeudi)")
    
    return df, issues

class IssueType(Enum):
    """Types d'incohérences détectées."""
    TEAM_NOT_FOUND = "Équipe non trouvée"
    TEAM_NO_MAPPING = "Équipe config sans mapping"
    GENRE_MISMATCH = "Genre incohérent"
    POULE_MISMATCH = "Poules différentes"
    GYM_NOT_FOUND = "Gymnase non trouvé"
    SCORE_CONFLICT = "Conflit de score"


@dataclass
class Issue:
    """Représente une incohérence détectée."""
    type: IssueType
    message: str
    mysportu_data: Optional[str] = None
    config_data: Optional[str] = None
    severity: str = "warning"


@dataclass
class SyncResult:
    """Résultat de la synchronisation des matchs."""
    matches_added: int = 0
    scores_added: int = 0
    scores_corrected: int = 0
    duplicates_found: int = 0
    matches_removed: int = 0
    matches_reordered: int = 0
    issues: List[Issue] = field(default_factory=list)


# =============================================================================
# FONCTIONS UTILITAIRES - PARSING
# =============================================================================

def get_genre_from_championnat(champ: str) -> Optional[str]:
    """Détermine le genre (F/M) à partir du nom du championnat MySportU."""
    if pd.isna(champ):
        return None
    champ_upper = str(champ).upper()
    if 'VBF' in champ_upper:
        return 'F'
    elif 'VBM' in champ_upper:
        return 'M'
    return None


def normalize_mysportu_team(name: str, genre_from_championnat: str) -> Tuple[Optional[str], Optional[str]]:
    """
    Normalise un nom d'équipe MySportU vers le format config.
    
    Le genre du championnat est PRIORITAIRE sur le genre déduit du nom.
    
    Args:
        name: Nom brut MySportU (ex: "069069025 - AS INSA LYON VOLLEY-BALL MASCULIN 3")
        genre_from_championnat: Genre déduit du championnat (F/M) - OBLIGATOIRE
    
    Returns:
        Tuple (config_name, genre) - genre est celui du championnat
    """
    if pd.isna(name) or not genre_from_championnat:
        return None, None
    
    # Format: "069069XXX - NOM EQUIPE"
    match = re.match(r'^(\d+)\s*-\s*(.+)$', str(name))
    if not match:
        return None, None
    
    code = match.group(1)
    rest = match.group(2).strip()
    
    institution = INSTITUTION_CODE_MAP.get(code)
    if not institution:
        return None, None
    
    # Extraire le numéro d'équipe selon l'institution
    numero = _extract_team_number(institution, rest)
    
    # Construire le nom config
    if numero:
        config_name = f"{institution} ({numero})"
    else:
        config_name = f"{institution} (1)"
    
    # Cas spéciaux
    if institution == 'LYON 2' and 'IEP' in rest.upper():
        config_name = 'LYON 2 (IEP) (4)'
    elif institution == 'LYON 1' and ('SANTÉ' in rest.upper() or 'SANTE' in rest.upper()):
        num_match = re.search(r'\((\d+)\)', rest)
        numero = num_match.group(1) if num_match else '1'
        config_name = f'SANTE ({numero})'
    
    # Le genre est TOUJOURS celui du championnat
    return config_name, genre_from_championnat


def _extract_team_number(institution: str, rest: str) -> Optional[str]:
    """Extrait le numéro d'équipe selon les règles de chaque institution."""
    rest_upper = rest.upper()
    
    if institution == 'LYON 2':
        num_match = re.search(r'\((\d+)\)|VB\s*[MF]\s*(\d+)', rest)
        if num_match:
            return num_match.group(1) or num_match.group(2)
    
    elif institution == 'LYON 1':
        num_match = re.search(r'\((\d+)\)', rest)
        if num_match:
            return num_match.group(1)
    
    elif institution == 'INSA':
        num_match = re.search(r'(\d+)\s*$', rest)
        if num_match:
            return num_match.group(1)
    
    elif institution == 'ECL':
        num_match = re.search(r'[MF]\s*(\d+)|([MF])(\d+)', rest, re.IGNORECASE)
        if num_match:
            return num_match.group(1) or num_match.group(3)
    
    elif institution == 'EML':
        num_match = re.search(r'EML\s*(\d+)', rest)
        if num_match:
            return num_match.group(1)
    
    elif institution == 'ENS':
        num_match = re.search(r'[FG](\d+)', rest)
        if num_match:
            return num_match.group(1)
    
    elif institution == 'ENTPE':
        if 'FILLES' in rest_upper:
            return '1'
        num_match = re.search(r'\((\d+)\)|ENTPE\s*(\d+)', rest)
        if num_match:
            return num_match.group(1) or num_match.group(2)
    
    elif institution in ['ESA', 'ESME']:
        num_match = re.search(r'(\d+)\s*$', rest)
        if num_match:
            return num_match.group(1)
    
    elif institution in ['ESSCA', 'CPE', 'CATHO']:
        num_match = re.search(r'\((\d+)\)|n°(\d+)', rest)
        if num_match:
            return num_match.group(1) or num_match.group(2)
        return '1'
    
    else:
        num_match = re.search(r'\((\d+)\)|(\d+)\s*$', rest)
        if num_match:
            return num_match.group(1) or num_match.group(2)
    
    return None


def normalize_gymnase(lieu: str) -> Optional[str]:
    """Normalise un lieu MySportU vers un gymnase de config."""
    if pd.isna(lieu):
        return None
    return GYMNASE_MAP.get(str(lieu).strip(), None)


# =============================================================================
# FONCTIONS UTILITAIRES - DATES ET SEMAINES
# =============================================================================

def calculate_week_number(date_value, start_date: datetime) -> Optional[int]:
    """Calcule le numéro de semaine à partir de la date de début de saison."""
    if pd.isna(date_value) or date_value == 'N-A':
        return None
    try:
        if isinstance(date_value, str):
            date = datetime.strptime(date_value, '%d/%m/%Y')
        else:
            date = pd.to_datetime(date_value)
        delta = (date - start_date).days
        return delta // 7 + 1
    except (ValueError, TypeError):
        return None


def format_week_display(week_num: int, start_date: datetime) -> str:
    """
    Formate le numéro de semaine pour affichage.
    
    Format uniforme: "N (dd/mm)" où N est le numéro de semaine
    et dd/mm est le premier jour de cette semaine.
    """
    if week_num is None:
        return None
    match_date = start_date + pd.Timedelta(days=(week_num - 1) * 7)
    return f"{week_num} ({match_date.strftime('%d/%m')})"


def parse_week_number(semaine_str) -> Optional[int]:
    """Extrait le numéro de semaine depuis le format 'N' ou 'N (dd/mm)'."""
    if pd.isna(semaine_str):
        return None
    match = re.match(r'^(\d+)', str(semaine_str).strip())
    if match:
        return int(match.group(1))
    return None


def format_date_ddmmyy(date_value) -> Optional[str]:
    """Formate une date en dd/mm/yy."""
    if pd.isna(date_value) or date_value == 'N-A':
        return None
    try:
        if isinstance(date_value, str):
            dt = datetime.strptime(date_value, '%d/%m/%Y')
        else:
            dt = pd.to_datetime(date_value)
        return dt.strftime('%d/%m/%y')
    except (ValueError, TypeError):
        return None


# =============================================================================
# FONCTIONS DE CHARGEMENT
# =============================================================================

def load_config(config_path: Path) -> dict:
    """Charge la configuration YAML."""
    with open(config_path, 'r', encoding='utf-8') as f:
        return yaml.safe_load(f)


def load_mysportu_data(file_path: Path, championship_pattern: str = 'PH2') -> pd.DataFrame:
    """
    Charge les données MySportU en filtrant par championnat.
    
    Note: Les matchs reportés ne sont PAS exclus ici, ils le seront lors de la sync.
    Cela permet de détecter les matchs config qui doivent être supprimés.
    """
    df = pd.read_excel(file_path)
    
    if championship_pattern:
        mask = df['Championnat'].str.contains(championship_pattern, na=False, case=False)
        df = df[mask].copy()
    
    return df


def load_config_teams(config_excel_path: Path) -> pd.DataFrame:
    """Charge les équipes de la configuration."""
    return pd.read_excel(config_excel_path, sheet_name='Equipes')


def load_existing_fixed_matches(config_excel_path: Path) -> pd.DataFrame:
    """Charge les matchs fixes existants."""
    try:
        return pd.read_excel(config_excel_path, sheet_name='Matchs_Fixes')
    except Exception:
        return pd.DataFrame()


def get_gymnases(config_excel_path: Path) -> Set[str]:
    """Récupère la liste des gymnases valides."""
    df = pd.read_excel(config_excel_path, sheet_name='Gymnases')
    return set(df['Gymnase'].dropna().tolist())


def get_team_poules(config_excel_path: Path) -> Dict[Tuple[str, str], str]:
    """
    Récupère la poule de chaque équipe.
    
    Returns:
        Dict {(equipe, genre): poule}
    """
    df = pd.read_excel(config_excel_path, sheet_name='Equipes')
    poules = {}
    for _, row in df.iterrows():
        key = (row['Equipe'], row['Genre_Equipe'])
        poules[key] = row.get('Poule')
    return poules


# =============================================================================
# PARTIE 1 : MAPPING (Colonnes MySportU dans Equipes et Gymnases)
# =============================================================================

def build_team_mapping(
    df_mysportu: pd.DataFrame,
    df_config_teams: pd.DataFrame,
    verbose: bool = False
) -> Tuple[Dict[Tuple[str, str], List[str]], List[Issue]]:
    """
    Construit le mapping équipes MySportU -> équipes config.
    
    IMPORTANT: Le mapping utilise la clé (nom_equipe, genre) pour gérer
    les équipes qui ont le même nom dans les deux genres.
    
    Returns:
        Tuple (mapping_dict, issues)
        - mapping_dict: {(config_name, genre): [liste_noms_mysportu]}
        - issues: Liste des problèmes détectés
    """
    issues = []
    mapping = {}  # (config_name, genre) -> [mysportu_names]
    
    # Ensemble des équipes config existantes (avec genre)
    config_teams_set = {(row['Equipe'], row['Genre_Equipe']) 
                        for _, row in df_config_teams.iterrows()}
    
    # Récupérer toutes les équipes MySportU uniques
    all_mysportu_teams = set(df_mysportu['Equipe A'].dropna()) | set(df_mysportu['Equipe B'].dropna())
    
    # Pour chaque équipe MySportU
    for msu_team in sorted(all_mysportu_teams):
        # Trouver tous les matchs avec cette équipe
        mask = (df_mysportu['Equipe A'] == msu_team) | (df_mysportu['Equipe B'] == msu_team)
        matches = df_mysportu[mask]
        
        # Pour chaque championnat unique (une équipe peut jouer en M et F)
        for champ in matches['Championnat'].unique():
            genre = get_genre_from_championnat(champ)
            if not genre:
                continue
            
            config_name, _ = normalize_mysportu_team(msu_team, genre)
            if not config_name:
                issues.append(Issue(
                    type=IssueType.TEAM_NOT_FOUND,
                    message="Impossible de parser l'équipe MySportU",
                    mysportu_data=msu_team,
                    severity="error"
                ))
                continue
            
            key = (config_name, genre)
            
            # Vérifier que l'équipe existe dans la config
            if key not in config_teams_set:
                issues.append(Issue(
                    type=IssueType.TEAM_NOT_FOUND,
                    message=f"Équipe non trouvée dans la config",
                    mysportu_data=msu_team,
                    config_data=f"{config_name} ({genre})",
                    severity="error"
                ))
                continue
            
            # Ajouter au mapping
            if key not in mapping:
                mapping[key] = []
            if msu_team not in mapping[key]:
                mapping[key].append(msu_team)
    
    # Détecter les équipes config SANS mapping MySportU
    for team_name, genre in sorted(config_teams_set - set(mapping.keys())):
        issues.append(Issue(
            type=IssueType.TEAM_NO_MAPPING,
            message=f"Équipe config sans correspondance MySportU",
            config_data=f"{team_name} ({genre})",
            severity="warning"
        ))
    
    if verbose:
        print(f"  • Équipes MySportU: {len(all_mysportu_teams)}")
        print(f"  • Équipes config: {len(config_teams_set)}")
        print(f"  • Mappings créés: {len(mapping)}")
    
    return mapping, issues


def update_teams_excel_column(
    config_excel_path: Path,
    mapping: Dict[Tuple[str, str], List[str]],
    dry_run: bool = False,
    verbose: bool = False
) -> int:
    """
    Met à jour la colonne 'MySportU' dans la feuille 'Equipes'.
    
    Utilise le couple (Equipe, Genre_Equipe) comme clé.
    """
    wb = load_workbook(config_excel_path)
    ws = wb['Equipes']
    
    headers = {cell.value: cell.column for cell in ws[1]}
    equipe_col = headers.get('Equipe')
    genre_col = headers.get('Genre_Equipe')
    
    # Créer la colonne MySportU si elle n'existe pas
    if 'MySportU' not in headers:
        mysportu_col = max(headers.values()) + 1
        ws.cell(row=1, column=mysportu_col, value='MySportU')
        ws.cell(row=1, column=mysportu_col).font = Font(bold=True)
    else:
        mysportu_col = headers['MySportU']
    
    updated_count = 0
    
    for row in range(2, ws.max_row + 1):
        equipe = ws.cell(row=row, column=equipe_col).value
        genre = ws.cell(row=row, column=genre_col).value
        
        if not equipe or not genre:
            continue
        
        key = (equipe, genre)
        if key in mapping:
            mysportu_value = ' | '.join(sorted(set(mapping[key])))
            current_value = ws.cell(row=row, column=mysportu_col).value
            
            if current_value != mysportu_value:
                if not dry_run:
                    ws.cell(row=row, column=mysportu_col, value=mysportu_value)
                    ws.cell(row=row, column=mysportu_col).fill = COLOR_SUCCESS
                updated_count += 1
                if verbose:
                    print(f"    {equipe} ({genre}): {mysportu_value[:50]}...")
    
    if not dry_run:
        wb.save(config_excel_path)
    
    return updated_count


def update_gymnases_excel_column(
    config_excel_path: Path,
    dry_run: bool = False,
    verbose: bool = False
) -> int:
    """Met à jour la colonne 'MySportU' dans la feuille 'Gymnases'."""
    wb = load_workbook(config_excel_path)
    ws = wb['Gymnases']
    
    headers = {cell.value: cell.column for cell in ws[1]}
    gymnase_col = headers.get('Gymnase')
    
    if not gymnase_col:
        return 0
    
    # Créer la colonne MySportU si elle n'existe pas
    if 'MySportU' not in headers:
        mysportu_col = max(headers.values()) + 1
        ws.cell(row=1, column=mysportu_col, value='MySportU')
        ws.cell(row=1, column=mysportu_col).font = Font(bold=True)
    else:
        mysportu_col = headers['MySportU']
    
    # Inverser le mapping : config -> [mysportu]
    reverse_map = {}
    for msu_name, cfg_name in GYMNASE_MAP.items():
        if cfg_name:
            reverse_map.setdefault(cfg_name, []).append(msu_name)
    
    updated_count = 0
    
    for row in range(2, ws.max_row + 1):
        gymnase = ws.cell(row=row, column=gymnase_col).value
        if not gymnase:
            continue
        
        if gymnase in reverse_map:
            mysportu_value = ' | '.join(sorted(reverse_map[gymnase]))
            current_value = ws.cell(row=row, column=mysportu_col).value
            
            if current_value != mysportu_value:
                if not dry_run:
                    ws.cell(row=row, column=mysportu_col, value=mysportu_value)
                    ws.cell(row=row, column=mysportu_col).fill = COLOR_SUCCESS
                updated_count += 1
                if verbose:
                    print(f"    {gymnase}: {mysportu_value}")
    
    if not dry_run:
        wb.save(config_excel_path)
    
    return updated_count


# =============================================================================
# PARTIE 2 : SYNCHRONISATION DES MATCHS
# =============================================================================

def create_match_key(eq1: str, eq2: str, genre: str) -> frozenset:
    """
    Crée une clé unique pour un match (indépendante de l'ordre des équipes).
    
    Inclut le genre pour différencier les matchs M et F.
    """
    return frozenset([eq1, eq2, f"GENRE:{genre}"])


def find_match_in_existing(
    eq1: str, eq2: str, genre: str,
    existing_matches: pd.DataFrame
) -> Tuple[Optional[int], bool]:
    """
    Trouve l'index d'un match existant.
    
    Returns:
        Tuple (index, is_reversed)
        - index: Index de la ligne ou None
        - is_reversed: True si les équipes sont inversées
    """
    for idx, row in existing_matches.iterrows():
        if pd.isna(row.get('Equipe_1')) or pd.isna(row.get('Equipe_2')):
            continue
        if row.get('Genre') != genre:
            continue
        
        if row['Equipe_1'] == eq1 and row['Equipe_2'] == eq2:
            return idx, False
        elif row['Equipe_1'] == eq2 and row['Equipe_2'] == eq1:
            return idx, True
    
    return None, False


def reverse_score(score: str) -> str:
    """Inverse un score (ex: '3-1' -> '1-3')."""
    if not score or '-' not in str(score):
        return score
    parts = str(score).split('-')
    if len(parts) == 2:
        return f"{parts[1]}-{parts[0]}"
    return score


def is_match_cancelled(etat) -> bool:
    """Vérifie si un match est reporté ou annulé."""
    if pd.isna(etat):
        return False
    return str(etat).lower().strip() in ['reporté', 'reporte', 'annulé', 'annule']


def sync_matches(
    config_excel_path: Path,
    df_mysportu: pd.DataFrame,
    start_date: datetime,
    valid_teams: Set[Tuple[str, str]],
    valid_gymnases: Set[str],
    team_poules: Dict[Tuple[str, str], str],
    competition_type: str = 'Acad',
    fix_scores: bool = False,
    dry_run: bool = False,
    verbose: bool = False,
    jour_match: str = 'jeudi'
) -> SyncResult:
    """
    Synchronise les matchs MySportU avec la configuration.
    
    Supporte les données provenant de l'API (colonnes _receveur_court etc.)
    ou de l'import Excel classique.
    
    - Ajoute les nouveaux matchs
    - Met à jour les scores des matchs existants
    - Gère les matchs en entente (hors jour normal)
    - Gère correctement le genre et l'ordre des équipes
    """
    result = SyncResult()
    existing_matches = load_existing_fixed_matches(config_excel_path)
    
    # Détecter si les données proviennent de l'API
    is_api_data = '_receveur_court' in df_mysportu.columns
    
    matches_to_add = []
    matches_to_update = []  # (row_index, updates_dict)
    
    # Filtrer les matchs non-reportés
    if 'etat' in df_mysportu.columns:
        df_active = df_mysportu[~df_mysportu['etat'].apply(is_match_cancelled)]
    else:
        df_active = df_mysportu
    
    for _, row in df_active.iterrows():
        genre = get_genre_from_championnat(row.get('Championnat', ''))
        if not genre:
            continue
        
        # Normaliser les équipes selon la source de données
        if is_api_data:
            eq_a_config = normalize_api_team_name(
                row.get('_receveur_court', ''),
                row.get('_receveur_code', ''),
                genre
            )
            eq_b_config = normalize_api_team_name(
                row.get('_visiteur_court', ''),
                row.get('_visiteur_code', ''),
                genre
            )
            display_a = row.get('_receveur_court', row.get('Equipe A'))
            display_b = row.get('_visiteur_court', row.get('Equipe B'))
        else:
            eq_a_config, _ = normalize_mysportu_team(row.get('Equipe A'), genre)
            eq_b_config, _ = normalize_mysportu_team(row.get('Equipe B'), genre)
            display_a = row.get('Equipe A')
            display_b = row.get('Equipe B')
        
        if not eq_a_config or not eq_b_config:
            result.issues.append(Issue(
                type=IssueType.TEAM_NOT_FOUND,
                message="Impossible de parser les équipes",
                mysportu_data=f"{display_a} vs {display_b}",
                severity="error"
            ))
            continue
        
        # Vérifier que les équipes existent (avec le bon genre)
        if (eq_a_config, genre) not in valid_teams:
            result.issues.append(Issue(
                type=IssueType.TEAM_NOT_FOUND,
                message=f"Équipe A non trouvée",
                mysportu_data=str(display_a),
                config_data=f"{eq_a_config} ({genre})",
                severity="error"
            ))
            continue
        
        if (eq_b_config, genre) not in valid_teams:
            result.issues.append(Issue(
                type=IssueType.TEAM_NOT_FOUND,
                message=f"Équipe B non trouvée",
                mysportu_data=str(display_b),
                config_data=f"{eq_b_config} ({genre})",
                severity="error"
            ))
            continue
        
        # Déterminer la poule du match
        poule_a = team_poules.get((eq_a_config, genre))
        poule_b = team_poules.get((eq_b_config, genre))
        
        match_poule = None
        if poule_a and poule_b and poule_a == poule_b:
            match_poule = poule_a
        elif poule_a and poule_b and poule_a != poule_b:
            result.issues.append(Issue(
                type=IssueType.POULE_MISMATCH,
                message=f"Équipes de poules différentes",
                mysportu_data=f"{eq_a_config} vs {eq_b_config}",
                config_data=f"{poule_a} vs {poule_b}",
                severity="info"
            ))
        
        # Détecter les matchs en entente (hors jour normal)
        entente = is_entente_match(row.get('Date', ''), jour_match)
        
        Calculer le numéro de semaine
        semaine_num = calculate_week_number(row.get('Date'), start_date)
        
        # Normaliser le gymnase
        if entente:
            gymnase = 'ENTENTE'
            if verbose:
                print(f"    🤝 Entente: {eq_a_config} vs {eq_b_config} ({genre}) - {row.get('Date')}")
        else:
            gymnase = normalize_gymnase(row.get('Lieu'))
            if gymnase and gymnase not in valid_gymnases:
                result.issues.append(Issue(
                    type=IssueType.GYM_NOT_FOUND,
                    message=f"Gymnase non trouvé",
                    mysportu_data=row.get('Lieu'),
                    config_data=gymnase,
                    severity="warning"
                ))
                config_data=gymnase,
                    severity="warning"
                ))
        semaine_display = format_week_display(semaine_num, start_date) if semaine_num else None
        
        # Ignorer matchs non programmés
        if semaine_num is None or gymnase is None:
            continue
        
        # Score
        score_raw = row.get('Score')
        score = None
        if not pd.isna(score_raw) and str(score_raw).strip() not in ['', 'N-A', 'nan']:
            score = str(score_raw).strip()
        
        # Date au format dd/mm/yy
        date_formatted = format_date_ddmmyy(row.get('Date'))
        
        # Chercher si le match existe déjà
        existing_idx, is_reversed = find_match_in_existing(
            eq_a_config, eq_b_config, genre, existing_matches
        )
        
        if existing_idx is not None:
            result.duplicates_found += 1
            existing_row = existing_matches.loc[existing_idx]
            existing_score = existing_row.get('Score')
            
            if score:
                score_to_compare = reverse_score(score) if is_reversed else score
                
                if pd.isna(existing_score) or str(existing_score).strip() == '':
                    matches_to_update.append((existing_idx, {
                        'Score': score_to_compare,
                        'Remarques': 'Score MySportU'
                    }))
                    result.scores_added += 1
                    if verbose:
                        order_info = " (inversé)" if is_reversed else ""
                        print(f"    Score à ajouter: {existing_row['Equipe_1']} vs {existing_row['Equipe_2']} = {score_to_compare}{order_info}")
                elif str(existing_score).strip() != score_to_compare:
                    if fix_scores:
                        matches_to_update.append((existing_idx, {
                            'Score': score_to_compare,
                            'Remarques': f"Score corrigé MySportU (était {existing_score})"
                        }))
                        result.scores_corrected += 1
                        if verbose:
                            print(f"    Score corrigé: {existing_row['Equipe_1']} vs {existing_row['Equipe_2']}: {existing_score} → {score_to_compare}")
                    else:
                        result.issues.append(Issue(
                            type=IssueType.SCORE_CONFLICT,
                            message=f"{existing_row['Equipe_1']} vs {existing_row['Equipe_2']} ({genre})",
                            mysportu_data=score_to_compare,
                            config_data=str(existing_score),
                            severity="warning"
                        ))
        else:
            # Nouveau match
            match_data = {
                'Equipe_1': eq_a_config,
                'Equipe_2': eq_b_config,
                'Genre': genre,
                'Poule': match_poule,
                'Semaine': semaine_display,
                'Date': date_formatted,
                'Horaire': row.get('Heure') if row.get('Heure') != 'N-A' else None,
                'Gymnase': gymnase,
                'Score': score,
                'Type_Competition': competition_type,
                'Remarques': 'Import MySportU',
                'Arbitres': None,
                'Ignorer': None,  # Toujours vide lors de l'import pour que le match soit fixé
            }
            matches_to_add.append(match_data)
    
    # Appliquer les modifications
    if not dry_run:
        wb = load_workbook(config_excel_path)
        ws = wb['Matchs_Fixes']
        headers = {cell.value: cell.column for cell in ws[1]}
        
        # Mettre à jour les scores
        for existing_idx, updates in matches_to_update:
            excel_row = existing_idx + 2  # +2 : header + 0-based
            for col_name, value in updates.items():
                if col_name in headers:
                    col = headers[col_name]
                    ws.cell(row=excel_row, column=col, value=value)
                    ws.cell(row=excel_row, column=col).fill = COLOR_INFO
        
        # Ajouter les nouveaux matchs
        start_row = ws.max_row + 1
        for i, match in enumerate(matches_to_add):
            row_num = start_row + i
            for col_name, value in match.items():
                if col_name in headers and value is not None:
                    col_idx = headers[col_name]
                    ws.cell(row=row_num, column=col_idx, value=value)
                    ws.cell(row=row_num, column=col_idx).fill = COLOR_SUCCESS
        
        wb.save(config_excel_path)
    
    result.matches_added = len(matches_to_add)
    
    return result



def remove_cancelled_acad_matches(
    config_excel_path: Path,
    df_mysportu: pd.DataFrame,
    start_date: datetime,
    competition_type: str = 'Acad',
    dry_run: bool = False,
    verbose: bool = False
) -> Tuple[int, List[dict]]:
    """
    Supprime les matchs Acad qui sont reportés/annulés dans MySportU.
    
    IMPORTANT: Ne supprime QUE les matchs de type Acad.
    Les matchs CFU/CFE sont préservés.
    """
    # Construire l'ensemble des matchs reportés dans MySportU
    cancelled_keys = set()
    
    # Détecter si les données proviennent de l'API
    is_api_data = '_receveur_court' in df_mysportu.columns
    
    if 'etat' in df_mysportu.columns:
        df_cancelled = df_mysportu[df_mysportu['etat'].apply(is_match_cancelled)]
        
        for _, row in df_cancelled.iterrows():
            genre = get_genre_from_championnat(row.get('Championnat', ''))
            if not genre:
                continue
            
            if is_api_data:
                eq_a = normalize_api_team_name(
                    row.get('_receveur_court', ''),
                    row.get('_receveur_code', ''),
                    genre
                )
                eq_b = normalize_api_team_name(
                    row.get('_visiteur_court', ''),
                    row.get('_visiteur_code', ''),
                    genre
                )
            else:
                eq_a, _ = normalize_mysportu_team(row.get('Equipe A'), genre)
                eq_b, _ = normalize_mysportu_team(row.get('Equipe B'), genre)
            
            if eq_a and eq_b:
                cancelled_keys.add(create_match_key(eq_a, eq_b, genre))
    
    if not cancelled_keys:
        return 0, []
    
    # Charger les matchs config
    df_config = load_existing_fixed_matches(config_excel_path)
    
    rows_to_remove = []
    removed_matches = []
    
    for idx, row in df_config.iterrows():
        # Ne supprimer QUE les matchs Acad
        type_comp = str(row.get('Type_Competition', '')).strip()
        if type_comp.lower() != competition_type.lower():
            continue
        
        eq1 = row.get('Equipe_1')
        eq2 = row.get('Equipe_2')
        genre = row.get('Genre')
        
        if pd.isna(eq1) or pd.isna(eq2) or pd.isna(genre):
            continue
        
        match_key = create_match_key(eq1, eq2, genre)
        
        if match_key in cancelled_keys:
            rows_to_remove.append(idx)
            removed_matches.append({
                'Equipe_1': eq1,
                'Equipe_2': eq2,
                'Genre': genre,
                'Semaine': row.get('Semaine')
            })
            if verbose:
                print(f"    \U0001f5d1\ufe0f  {eq1} vs {eq2} ({genre})")
    
    if not dry_run and rows_to_remove:
        wb = load_workbook(config_excel_path)
        ws = wb['Matchs_Fixes']
        
        # Supprimer de bas en haut
        for idx in sorted(rows_to_remove, reverse=True):
            ws.delete_rows(idx + 2)  # +2 : header + 0-based
        
        wb.save(config_excel_path)
    
    return len(rows_to_remove), removed_matches



def parse_poule_for_sorting(poule) -> Tuple[int, int]:
    """Parse une poule pour le tri (niveau, numéro de pool)."""
    if not poule or pd.isna(poule):
        return (99, 99)
    
    match = re.match(r'VB[FM]A(\d)P([A-F])', str(poule))
    if match:
        niveau = int(match.group(1))
        pool_num = ord(match.group(2)) - ord('A')
        return (niveau, pool_num)
    
    return (99, 99)


def match_sort_key(row: pd.Series, start_date: datetime) -> Tuple:
    """
    Clé de tri pour un match.
    
    Ordre: Semaine, Non-Acad en premier, Date, Entente en dernier, Genre (F avant M), Niveau, Pool
    
    Les matchs avec gymnase 'ENTENTE' sont placés en fin de journée car ils sont
    généralement joués sur des créneaux spéciaux hors du gymnase principal.
    """
    week = parse_week_number(row.get('Semaine')) or 999
    
    # Type de compétition: non-Acad = 0, Acad = 1
    type_comp = str(row.get('Type_Competition', '')).strip().lower()
    is_acad = 1 if type_comp == 'acad' else 0
    
    # Date
    date_str = row.get('Date')
    if pd.isna(date_str):
        date_parsed = datetime(2099, 12, 31)
    else:
        try:
            date_parsed = datetime.strptime(str(date_str), '%d/%m/%y')
        except ValueError:
            try:
                date_parsed = datetime.strptime(str(date_str), '%d/%m/%Y')
            except ValueError:
                date_parsed = datetime(2099, 12, 31)
    
    # Gymnase ENTENTE en fin de journée: 0 = normal, 1 = entente
    gymnase = str(row.get('Gymnase', '')).strip().upper()
    is_entente = 1 if gymnase == 'ENTENTE' else 0
    
    # Genre: F = 0, M = 1
    genre_order = 0 if row.get('Genre') == 'F' else 1
    
    niveau, pool = parse_poule_for_sorting(row.get('Poule'))
    
    return (week, is_acad, date_parsed, is_entente, genre_order, niveau, pool)


def sort_and_normalize_weeks(
    config_excel_path: Path,
    start_date: datetime,
    dry_run: bool = False,
    verbose: bool = False
) -> int:
    """
    Trie les matchs et uniformise le format des semaines.
    
    - Tri: Semaine > Non-Acad en premier > Date > Genre > Niveau > Pool
    - Format semaine uniformisé: "N (dd/mm)"
    """
    df = pd.read_excel(config_excel_path, sheet_name='Matchs_Fixes')
    
    if df.empty:
        return 0
    
    # Sauvegarder les anciennes semaines pour détecter les changements
    old_semaines = df['Semaine'].astype(str).tolist()
    
    # Normaliser les semaines
    def normalize_week(semaine_val):
        week_num = parse_week_number(semaine_val)
        if week_num:
            return format_week_display(week_num, start_date)
        return semaine_val
    
    df['Semaine'] = df['Semaine'].apply(normalize_week)
    new_semaines = df['Semaine'].astype(str).tolist()
    
    # Compter les semaines modifiées
    weeks_changed = sum(1 for old, new in zip(old_semaines, new_semaines) if old != new)
    
    # Créer les clés de tri
    sort_keys = df.apply(lambda r: match_sort_key(r, start_date), axis=1)
    sorted_indices = sorted(range(len(df)), key=lambda i: sort_keys.iloc[i])
    df_sorted = df.iloc[sorted_indices].reset_index(drop=True)
    
    # Compter les changements d'ordre
    order_changes = sum(1 for i, orig_idx in enumerate(sorted_indices) if i != orig_idx)
    
    # Total des changements
    has_changes = weeks_changed > 0 or order_changes > 0
    
    if verbose:
        if weeks_changed > 0:
            print(f"    {weeks_changed} semaines à normaliser")
        print(f"    {order_changes} matchs à réordonner sur {len(df)}")
    
    if not dry_run and has_changes:
        wb = load_workbook(config_excel_path)
        ws = wb['Matchs_Fixes']
        
        headers = [cell.value for cell in ws[1]]
        
        # Effacer les données
        for row in range(2, ws.max_row + 1):
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).value = None
                ws.cell(row=row, column=col).fill = PatternFill()
        
        # Réécrire les données triées
        for i, (_, row_data) in enumerate(df_sorted.iterrows()):
            excel_row = i + 2
            for j, col_name in enumerate(headers):
                if col_name and col_name in row_data:
                    value = row_data[col_name]
                    if pd.notna(value):
                        ws.cell(row=excel_row, column=j+1, value=value)
    return order_changes
        
        wb.save(config_excel_path)
    
    return order_changes
    


# =============================================================================
# AFFICHAGE
# =============================================================================

def print_issues(issues: List[Issue], max_per_type: int = 10):
    """Affiche les incoh\u00e9rences de mani\u00e8re structur\u00e9e."""
    if not issues:
        print("\n\u2705 Aucune incoh\u00e9rence d\u00e9tect\u00e9e!")
        return
    
    by_type = {}
    for issue in issues:
        by_type.setdefault(issue.type, []).append(issue)
    
    print("\n" + "=" * 60)
    print("INCOH\u00c9RENCES D\u00c9TECT\u00c9ES")
    print("=" * 60)
    
    severity_icons = {"error": "\u274c", "warning": "\u26a0\ufe0f", "info": "\u2139\ufe0f"}
    
    for issue_type, type_issues in by_type.items():
        icon = severity_icons.get(type_issues[0].severity, '\u2022')
        print(f"\n{icon} {issue_type.value} ({len(type_issues)})")
        print("-" * 40)
        
        for issue in type_issues[:max_per_type]:
            print(f"  \u2022 {issue.message}")
            if issue.mysportu_data:
                print(f"    MySportU: {issue.mysportu_data}")
            if issue.config_data:
                print(f"    Config:   {issue.config_data}")
        
        if len(type_issues) > max_per_type:
            print(f"  ... et {len(type_issues) - max_per_type} autres")


# =============================================================================
# FONCTION PRINCIPALE
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description='Synchronisation MySportU <-> Configuration Excel.',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__
    )
    
    parser.add_argument(
        '--config', '-c',
        type=Path,
        default=Path('configs/config_volleyP2.yaml'),
        help='Fichier de configuration YAML'
    )
    
    # Source des donn\u00e9es
    source_group = parser.add_argument_group('Source des donn\u00e9es')
    source_group.add_argument(
        '--excel',
        action='store_true',
        help="Utiliser un fichier Excel MySportU au lieu de l'API web (fallback)"
    )
    source_group.add_argument(
        '--mysportu', '-m',
        type=Path,
        default=None,
        help='Fichier MySportU Excel (utilis\u00e9 avec --excel, auto-d\u00e9tect\u00e9 si non sp\u00e9cifi\u00e9)'
    )
    source_group.add_argument(
        '--sport',
        type=str,
        default='VB',
        help='Pattern sport pour le filtre API (default: VB)'
    )
    
    # Modes d'ex\u00e9cution
    mode_group = parser.add_argument_group("Modes d'ex\u00e9cution")
    mode_group.add_argument(
        '--mapping',
        action='store_true',
        help='Mettre \u00e0 jour les colonnes MySportU (Equipes + Gymnases) - n\u00e9cessite --excel'
    )
    mode_group.add_argument(
        '--sync',
        action='store_true',
        help='Synchroniser les matchs (inclut clean + sort par d\u00e9faut)'
    )
    mode_group.add_argument(
        '--all',
        action='store_true',
        help='Ex\u00e9cuter mapping + sync (mapping n\u00e9cessite --excel)'
    )
    
    # Options de synchronisation
    sync_group = parser.add_argument_group('Options de synchronisation')
    sync_group.add_argument(
        '--no-clean',
        action='store_true',
        help='Ne pas supprimer les matchs Acad report\u00e9s'
    )
    sync_group.add_argument(
        '--no-sort',
        action='store_true',
        help='Ne pas trier les matchs'
    )
    sync_group.add_argument(
        '--fix-scores',
        action='store_true',
        help='Corriger les scores en conflit avec MySportU'
    )
    
    # Options g\u00e9n\u00e9rales
    parser.add_argument(
        '--championship', '-p',
        type=str,
        default='PH2',
        help='Pattern championnat (default: PH2)'
    )
    parser.add_argument(
        '--type', '-t',
        type=str,
        default='Acad',
        help='Type de comp\u00e9tition pour les nouveaux matchs (default: Acad)'
    )
    parser.add_argument(
        '--dry-run', '-n',
        action='store_true',
        help='Afficher les changements sans modifier'
    )
    parser.add_argument(
        '--verbose', '-v',
        action='store_true',
        help='Afficher des informations d\u00e9taill\u00e9es'
    )
    
    args = parser.parse_args()
    
    # V\u00e9rifier qu'au moins un mode est s\u00e9lectionn\u00e9
    if not any([args.mapping, args.sync, args.all]):
        parser.print_help()
        print("\n\u274c Veuillez sp\u00e9cifier au moins un mode: --mapping, --sync, ou --all")
        return 1
    
    # Le mapping n\u00e9cessite un fichier Excel
    if (args.mapping or args.all) and not args.excel:
        print("\u26a0\ufe0f  Le mode --mapping n\u00e9cessite --excel. Activation automatique.")
        args.excel = True
    
    # Charger la configuration
    print(f"\U0001f4c1 Configuration: {args.config}")
    config = load_config(args.config)
    
    # D\u00e9terminer les chemins
    data_file = Path(config['fichiers']['donnees'])
    if not data_file.is_absolute():
        data_file = args.config.parent.parent / data_file
    
    print(f"\U0001f4ca Config Excel: {data_file}")
    
    # Date de d\u00e9but et jour de match
    cal_config = config.get('calendrier', {})
    date_debut_str = cal_config.get('date_debut', '2026-01-22')
    start_date = datetime.strptime(date_debut_str, '%Y-%m-%d')
    jour_match = cal_config.get('jour_match', 'jeudi')
    print(f"\U0001f4c5 D\u00e9but saison: {start_date.strftime('%d/%m/%Y')} (jour: {jour_match})")
    
    if args.dry_run:
        print("\n\U0001f50d MODE DRY-RUN: Aucune modification\n")
    
    # =============================================
    # CHARGEMENT DES DONN\u00c9ES
    # =============================================
    print("\n\U0001f4e5 Chargement des donn\u00e9es...")
    
    df_mysportu = None
    api_issues = []
    
    if args.excel:
        # Mode Excel classique
        if args.mysportu is None:
            mysportu_file = data_file.parent / 'mysportu.xlsx'
            if not mysportu_file.exists():
                mysportu_files = list(data_file.parent.glob('*mysportu*.xlsx'))
                if not mysportu_files:
                    print(f"\u274c Erreur: Aucun fichier MySportU trouv\u00e9 dans {data_file.parent}")
                    return 1
                mysportu_file = mysportu_files[0]
        else:
            mysportu_file = args.mysportu
        
        print(f"  \U0001f4c4 Source: Excel ({mysportu_file})")
        df_mysportu = load_mysportu_data(mysportu_file, args.championship)
    else:
        # Mode API (par d\u00e9faut)
        print(f"  \U0001f310 Source: API MySportU (sport={args.sport}, championnat={args.championship})")
        df_mysportu, api_issues = fetch_matches_from_api(
            championship_pattern=args.championship,
            sport_pattern=args.sport,
            verbose=args.verbose
        )
        
        if df_mysportu.empty:
            print("\u274c Erreur: Aucun match r\u00e9cup\u00e9r\u00e9 depuis l'API")
            if api_issues:
                for issue in api_issues:
                    print(f"  \u2022 {issue.message}")
            print("\n\U0001f4a1 Essayez avec --excel pour utiliser un fichier Excel MySportU")
            return 1
    
    # Charger les donn\u00e9es config
    df_config_teams = load_config_teams(data_file)
    valid_gymnases = get_gymnases(data_file)
    team_poules = get_team_poules(data_file)
    
    valid_teams = {(row['Equipe'], row['Genre_Equipe']) 
                   for _, row in df_config_teams.iterrows()}
    
    source_label = "Excel" if args.excel else "API"
    print(f"  \u2022 {len(df_mysportu)} matchs MySportU {source_label} ({args.championship})")
    print(f"  \u2022 {len(df_config_teams)} \u00e9quipes config")
    print(f"  \u2022 {len(valid_gymnases)} gymnases")
    
    all_issues = list(api_issues)
    
    # =========================
    # MODE: MAPPING
    # =========================
    if args.mapping or args.all:
        if not args.excel:
            print("\n\u26a0\ufe0f  Mapping ignor\u00e9: n\u00e9cessite --excel")
        else:
            print(f"\n{'='*60}")
            print("MAPPING : Mise \u00e0 jour des colonnes MySportU")
            print('='*60)
            
            print("\n\U0001f504 Construction du mapping des \u00e9quipes...")
            mapping, mapping_issues = build_team_mapping(
                df_mysportu, df_config_teams, verbose=args.verbose
            )
            all_issues.extend(mapping_issues)
            
            print("\n\u270f\ufe0f  Mise \u00e0 jour colonne MySportU (Equipes)...")
            updated_teams = update_teams_excel_column(
                data_file, mapping, dry_run=args.dry_run, verbose=args.verbose
            )
            action = "seraient mises \u00e0 jour" if args.dry_run else "mises \u00e0 jour \u2713"
            print(f"  \u2022 {updated_teams} \u00e9quipes {action}")
            
            print("\n\u270f\ufe0f  Mise \u00e0 jour colonne MySportU (Gymnases)...")
            updated_gyms = update_gymnases_excel_column(
                data_file, dry_run=args.dry_run, verbose=args.verbose
            )
            action = "seraient mis \u00e0 jour" if args.dry_run else "mis \u00e0 jour \u2713"
            print(f"  \u2022 {updated_gyms} gymnases {action}")
    
    # =========================
    # MODE: SYNCHRONISATION
    # =========================
    if args.sync or args.all:
        print(f"\n{'='*60}")
        print("SYNCHRONISATION : Import et mise \u00e0 jour des matchs")
        print('='*60)
        
        # 1. Synchroniser les matchs
        print("\n\U0001f504 Synchronisation des matchs...")
        if args.fix_scores:
            print("  \u26a0\ufe0f  Mode correction des scores activ\u00e9")
        
        sync_result = sync_matches(
            data_file,
            df_mysportu,
            start_date,
            valid_teams,
            valid_gymnases,
            team_poules,
            competition_type=args.type,
            fix_scores=args.fix_scores,
            dry_run=args.dry_run,
            verbose=args.verbose,
            jour_match=jour_match
        )
        all_issues.extend(sync_result.issues)
        
        action_add = "seraient ajout\u00e9s" if args.dry_run else "ajout\u00e9s \u2713"
        action_score = "seraient ajout\u00e9s" if args.dry_run else "ajout\u00e9s \u2713"
        print(f"  \u2022 {sync_result.matches_added} matchs {action_add}")
        print(f"  \u2022 {sync_result.scores_added} scores {action_score}")
        if sync_result.scores_corrected > 0:
            action_corr = "seraient corrig\u00e9s" if args.dry_run else "corrig\u00e9s \u2713"
            print(f"  \u2022 {sync_result.scores_corrected} scores {action_corr}")
        print(f"  \u2022 {sync_result.duplicates_found} doublons d\u00e9tect\u00e9s")
        
        # 2. Supprimer les matchs report\u00e9s (par d\u00e9faut)
        if not args.no_clean:
            print("\n\U0001f5d1\ufe0f  Suppression des matchs Acad report\u00e9s...")
            removed_count, _ = remove_cancelled_acad_matches(
                data_file,
                df_mysportu,
                start_date,
                competition_type=args.type,
                dry_run=args.dry_run,
                verbose=args.verbose
            )
            action = "seraient supprim\u00e9s" if args.dry_run else "supprim\u00e9s \u2713"
            print(f"  \u2022 {removed_count} matchs {action}")
        
        # 3. Trier et uniformiser (par d\u00e9faut)
        if not args.no_sort:
            print("\n\U0001f4ca Tri et uniformisation des semaines...")
            reordered = sort_and_normalize_weeks(
                data_file,
                start_date,
                dry_run=args.dry_run,
                verbose=args.verbose
            )
            action = "seraient r\u00e9ordonn\u00e9s" if args.dry_run else "r\u00e9ordonn\u00e9s \u2713"
            print(f"  \u2022 {reordered} matchs {action}")
    
    # Afficher les incoh\u00e9rences
    print_issues(all_issues)
    
    # R\u00e9sum\u00e9
    print("\n" + "=" * 60)
    if args.dry_run:
        print("\u2705 V\u00e9rification termin\u00e9e (aucune modification)")
    else:
        print("\u2705 Synchronisation termin\u00e9e!")
    print("=" * 60)
    
    return 0


if __name__ == '__main__':
    exit(main())
