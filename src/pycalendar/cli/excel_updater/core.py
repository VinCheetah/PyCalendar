"""
Core actualization logic for Excel configuration files.

This module contains the main ConfigActualisateurV2 class that orchestrates
validation, correction, and formatting of Excel configuration files.
"""

import sys
import re
import difflib
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font
import yaml

# Add project root to path
current_file = Path(__file__).resolve()
project_root = current_file.parent.parent.parent.parent.parent
if str(project_root / 'src') not in sys.path:
    sys.path.insert(0, str(project_root / 'src'))

from pycalendar.core.config_manager import ConfigManager
from pycalendar.core.coach_groups import (
    COACH_GROUP_SLOT_COLUMNS,
    CoachSlotSpec,
    CoachSlotParseError,
    parse_coach_slot,
)
from pycalendar.core.utils import (
    extraire_genre_depuis_poule,
    formater_nom_avec_genre,
    parser_nom_avec_genre,
)
from pycalendar.core.calendar_manager import CalendarManager, CalendarConfig
from pycalendar.core.constants import DATE_DISPLAY_FORMAT, DATE_USER_FORMAT_LABEL

from .modes import UpdateMode, UpdateOptions, prompt_user_correction
from .reports import RapportFeuille, RapportGlobal, ValidationResult, Severity, CellIssue
from .validators import ColumnValidator
from .formatters import ExcelFormatter, format_workbook, Colors
from .dropdowns import DropdownManager, setup_all_dropdowns
from .match_duplicates import MatchDuplicateDetector, detect_match_duplicates
from .display import (
    print_header, print_section, print_step, print_step_result,
    print_success, print_error, print_warning, print_info,
    print_global_report, print_mode_info, print_correction_summary
)


# Default match day
OFFICIAL_MATCH_WEEKDAY = 3  # Thursday
OFFICIAL_MATCH_DAY_LABEL = "jeudi"


class ConfigActualisateurV2:
    """
    Actualise et valide un fichier de configuration Excel.
    
    Fonctionnalités:
    - Validation complète de la structure et du contenu
    - Correction automatique ou interactive des erreurs
    - Formatage visuel amélioré
    - Génération des listes déroulantes avec dates de semaines
    - Rapport détaillé des problèmes
    """
    
    def __init__(self, fichier_path: str, options: Optional[UpdateOptions] = None):
        self.fichier_path = Path(fichier_path)
        self.options = options or UpdateOptions()
        self.config = ConfigManager(str(fichier_path))
        self.rapport = RapportGlobal(fichier=str(fichier_path))
        self.validator = ColumnValidator()
        
        # Reference data
        self.equipes_ref: Set[str] = set()
        self.equipes_par_nom: Dict[str, Set[str]] = {}
        self.equipes_toutes_variantes: Set[str] = set()
        self.equipes_sans_genre: Set[str] = set()
        self.gymnases_ref: Set[str] = set()
        self.institutions_ref: Set[str] = set()
        self.poules_ref: Set[str] = set()
        self.equipes_par_institution: Dict[str, Dict[str, Set[str]]] = {}
        self._institution_lookup: Dict[str, str] = {}
        self._team_lookup: Dict[str, str] = {}
        
        # Coach groups
        self.coach_slot_columns = list(COACH_GROUP_SLOT_COLUMNS)
        
        # Calendar manager
        self.calendar_manager: Optional[CalendarManager] = None
        self.week_dates: Dict[int, datetime] = {}  # Toutes les semaines du calendrier (1 à nb_semaines)
        self.nb_semaines: int = 13  # Nombre total de semaines
        self.semaine_min: int = 1  # Première semaine planifiable (pour nouveaux matchs)
        self.semaines_banalisees: Set[int] = set()  # Semaines banisées (vacances, etc.)
        self._official_weekday_index = OFFICIAL_MATCH_WEEKDAY
        self._official_weekday_label = OFFICIAL_MATCH_DAY_LABEL
        
        # Interactive mode state
        self._accept_all_corrections = False
        self._ignore_all_corrections = False
        
        # Initialize calendar
        self._init_calendar()
    
    def _init_calendar(self):
        """Initialize calendar manager from YAML config."""
        data = self._load_yaml_config()
        if not data:
            return
        
        calendrier = data.get('calendrier', {}) or {}
        date_debut = calendrier.get('date_debut')
        if not date_debut:
            return
        
        jour_match = calendrier.get('jour_match', OFFICIAL_MATCH_DAY_LABEL)
        semaines_banalisees = calendrier.get('semaines_banalisees', []) or []
        
        # Récupérer le nombre de semaines depuis la config de planification
        planification = data.get('planification', {}) or {}
        nb_semaines = planification.get('nb_semaines', 13)
        semaine_min = planification.get('semaine_min', 1)
        
        # Stocker pour utilisation ultérieure (validation, dropdowns)
        self.nb_semaines = nb_semaines
        self.semaine_min = semaine_min
        self.semaines_banalisees = set(semaines_banalisees)
        
        try:
            config = CalendarConfig(
                date_debut=date_debut,
                jour_match=jour_match,
                semaines_banalisees=semaines_banalisees,
            )
            self.calendar_manager = CalendarManager(config)
            
            # Build week -> date mapping pour TOUTES les semaines du calendrier (1 à nb_semaines)
            # Note: week_dates couvre tout le calendrier, pas seulement les semaines planifiables
            # semaine_min indique à partir de quelle semaine on peut PLANIFIER de nouveaux matchs
            for semaine in range(1, nb_semaines + 1):
                if not self.calendar_manager.est_semaine_banalisee(semaine):
                    date = self.calendar_manager.semaine_to_date(semaine)
                    if date:
                        self.week_dates[semaine] = date
            
            # Set official match day
            jour_english = self.calendar_manager.jour_match
            self._official_weekday_index = self._weekday_name_to_index(jour_english)
            self._official_weekday_label = jour_match.lower() if jour_match else OFFICIAL_MATCH_DAY_LABEL
            
        except Exception as e:
            print_warning(f"Impossible d'initialiser le calendrier: {e}")
    
    def _load_yaml_config(self) -> Optional[Dict[str, Any]]:
        """Load associated YAML configuration.
        
        Priority:
        1. Explicit path provided in options.yaml_config_path
        2. Search in configs/ directory for matching Excel file
        3. Fallback to default.yaml
        """
        # Priority 1: Explicit path from options
        if self.options.yaml_config_path:
            explicit_path = Path(self.options.yaml_config_path)
            if not explicit_path.is_absolute():
                explicit_path = project_root / explicit_path
            if explicit_path.exists():
                try:
                    with open(explicit_path, 'r', encoding='utf-8') as f:
                        data = yaml.safe_load(f) or {}
                    return data
                except Exception as e:
                    print_warning(f"Impossible de charger {explicit_path}: {e}")
        
        # Priority 2: Try to find associated YAML by matching Excel path
        configs_dir = project_root / 'configs'
        if not configs_dir.exists():
            return None
        
        target = self.fichier_path.resolve()
        
        for yaml_path in sorted(configs_dir.glob('*.yaml')):
            try:
                with open(yaml_path, 'r', encoding='utf-8') as f:
                    data = yaml.safe_load(f) or {}
                
                fichiers = data.get('fichiers', {}) or {}
                donnees = fichiers.get('donnees')
                if not donnees:
                    continue
                
                donnees_path = Path(donnees)
                if not donnees_path.is_absolute():
                    donnees_path = (project_root / donnees_path).resolve()
                
                if donnees_path == target:
                    return data
            except Exception:
                continue
        
        # Priority 3: Fallback to default.yaml
        default_yaml = configs_dir / 'default.yaml'
        if default_yaml.exists():
            try:
                with open(default_yaml, 'r', encoding='utf-8') as f:
                    return yaml.safe_load(f) or {}
            except Exception:
                pass
        
        return None
    
    @staticmethod
    def _weekday_name_to_index(name: str) -> int:
        mapping = {
            'monday': 0, 'tuesday': 1, 'wednesday': 2, 'thursday': 3,
            'friday': 4, 'saturday': 5, 'sunday': 6,
        }
        return mapping.get((name or '').strip().lower(), OFFICIAL_MATCH_WEEKDAY)
    
    def actualiser(self) -> bool:
        """
        Actualise complètement le fichier de configuration.
        
        Returns:
            True si l'actualisation a réussi (pas d'erreurs bloquantes)
        """
        # Header
        print_header("ACTUALISATEUR DE CONFIGURATION EXCEL", "🔄")
        
        if not self.fichier_path.exists():
            print_error(f"Le fichier {self.fichier_path} n'existe pas")
            return False
        
        print(f"  📂 Fichier : {self.fichier_path.name}")
        print(f"  📍 Emplacement : {self.fichier_path.parent}")
        
        # Show mode
        mode_names = {
            UpdateMode.VALIDATE: ("Validation seule", "Analyse sans modification"),
            UpdateMode.AUTO_CORRECT: ("Correction automatique", "Applique toutes les corrections possibles"),
            UpdateMode.INTERACTIVE: ("Mode interactif", "Demande confirmation pour chaque correction"),
        }
        mode_name, mode_desc = mode_names.get(self.options.mode, ("Standard", ""))
        print_mode_info(mode_name, mode_desc)
        
        # Backup if needed
        if self.options.mode != UpdateMode.VALIDATE and self.options.backup:
            self._create_backup()
        
        print_section("ANALYSE", "🚀")
        
        # Step 1: Load references
        print_step(1, 6, "Chargement des données de référence")
        self._charger_references()
        print_step_result("✓", len(self.equipes_ref))
        print(f"        → {len(self.equipes_ref)} équipe(s), {len(self.gymnases_ref)} gymnase(s), {len(self.poules_ref)} poule(s)")
        
        # Step 2: Validate sheets
        print_step(2, 6, "Validation des feuilles")
        nb_problemes = self._valider_toutes_feuilles()
        print_step_result("✓")
        
        # Step 3: Generate missing sheets
        print_step(3, 6, "Génération des feuilles manquantes")
        nb_generees = self._generer_feuilles_manquantes()
        print_step_result("✓", nb_generees if nb_generees > 0 else None)
        
        # Step 4: Clean examples
        print_step(4, 6, "Nettoyage des exemples")
        nb_suppressions = self._supprimer_exemples()
        print_step_result("✓", nb_suppressions if nb_suppressions > 0 else None)
        
        # Step 5: Pre-fill Types_Poules
        print_step(5, 6, "Configuration des types de poules")
        nb_poules = self._preremplir_types_poules()
        print_step_result("✓", nb_poules if nb_poules > 0 else None)
        
        # Step 6: Apply corrections and formatting
        print_step(6, 6, "Finalisation et formatage")
        nb_corrections = 0
        
        if self.options.mode != UpdateMode.VALIDATE:
            nb_corrections = self._appliquer_corrections()
        
        if self.options.format_output:
            self._appliquer_formatage()
        
        if self.options.add_dropdowns:
            self._configurer_dropdowns()
        
        print_step_result("✓", nb_corrections if nb_corrections > 0 else None)
        
        # Display report
        print_global_report(self.rapport, self.options.verbose)
        
        return self.rapport.est_valide
    
    def _create_backup(self):
        """Create a backup of the file before modifications."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = self.fichier_path.with_suffix(f'.backup_{timestamp}.xlsx')
        shutil.copy2(self.fichier_path, backup_path)
        print_info(f"Sauvegarde créée: {backup_path.name}")
    
    def _charger_references(self):
        """Load reference data from main sheets."""
        self.equipes_ref.clear()
        self.equipes_par_nom.clear()
        self.equipes_par_institution.clear()
        self.equipes_toutes_variantes.clear()
        self.equipes_sans_genre.clear()
        self.gymnases_ref.clear()
        self.institutions_ref.clear()
        self.poules_ref.clear()
        self._institution_lookup.clear()
        self._team_lookup.clear()
        
        # Load teams
        df_equipes = self.config.lire_feuille('Equipes')
        if df_equipes is not None and 'Equipe' in df_equipes.columns:
            for _, row in df_equipes.iterrows():
                equipe = row.get('Equipe')
                if pd.isna(equipe):
                    continue
                
                equipe_str = str(equipe).strip()
                if not equipe_str:
                    continue
                
                self.equipes_ref.add(equipe_str)
                self._team_lookup[equipe_str.lower()] = equipe_str
                
                # Extract gender
                nom_sans_genre, genre_depuis_nom = parser_nom_avec_genre(equipe_str)
                genre = genre_depuis_nom
                
                if not genre and 'Genre' in df_equipes.columns:
                    genre_col = row.get('Genre')
                    if pd.notna(genre_col):
                        genre_candidate = str(genre_col).strip().upper()
                        if genre_candidate in ['M', 'F']:
                            genre = genre_candidate
                
                if not genre and 'Poule' in df_equipes.columns:
                    poule = row.get('Poule')
                    if pd.notna(poule):
                        genre = extraire_genre_depuis_poule(str(poule))
                        self.poules_ref.add(str(poule).strip())
                
                # Index team
                self._indexer_equipe(nom_sans_genre, genre, equipe_str)
        
        # Load teams from Equipes_Hors_Championnat
        df_hors = self.config.lire_feuille('Equipes_Hors_Championnat')
        if df_hors is not None and not df_hors.empty:
            for _, row in df_hors.iterrows():
                equipe = row.get('Equipe')
                if pd.isna(equipe):
                    continue
                
                equipe_str = str(equipe).strip()
                if equipe_str:
                    self.equipes_ref.add(equipe_str)
                    nom_sans_genre, genre_depuis_nom = parser_nom_avec_genre(equipe_str)
                    genre = genre_depuis_nom or str(row.get('Genre', '')).strip().upper()
                    if genre not in ['M', 'F']:
                        genre = None
                    self._indexer_equipe(nom_sans_genre, genre, equipe_str)
        
        # Load gymnases
        df_gymnases = self.config.lire_feuille('Gymnases')
        if df_gymnases is not None and 'Gymnase' in df_gymnases.columns:
            for gymnase in df_gymnases['Gymnase'].dropna():
                gymnase_str = str(gymnase).strip()
                if gymnase_str:
                    self.gymnases_ref.add(gymnase_str)
        
        # Generate team variants
        self.equipes_toutes_variantes = set(self._generer_variantes_equipes())
        for variante in self.equipes_toutes_variantes:
            self._team_lookup[variante.lower()] = variante
        self.equipes_sans_genre = set(self.equipes_par_nom.keys())
    
    def _indexer_equipe(self, nom_sans_genre: str, genre: Optional[str], etiquette: str,
                        institution_hint: Optional[str] = None):
        """Index a team for lookup."""
        if not nom_sans_genre:
            return
        
        genres = self.equipes_par_nom.setdefault(nom_sans_genre, set())
        if genre in ['M', 'F']:
            genres.add(genre)
        
        institution = institution_hint or self._extraire_institution(nom_sans_genre)
        if not institution:
            return
        
        canonical = self._enregistrer_institution(institution)
        if not canonical:
            return
        
        buckets = self.equipes_par_institution.setdefault(
            canonical, {'ALL': set(), 'M': set(), 'F': set()}
        )
        buckets['ALL'].add(etiquette)
        if genre in ['M', 'F']:
            buckets[genre].add(formater_nom_avec_genre(nom_sans_genre, genre))
    
    def _extraire_institution(self, nom: str) -> Optional[str]:
        """Extract institution from team name."""
        if not nom:
            return None
        match = re.match(r'^(.+?)\s*\(\d+\)\s*$', nom)
        return match.group(1).strip() if match else None
    
    def _enregistrer_institution(self, institution: str) -> Optional[str]:
        """Register an institution."""
        if not institution:
            return None
        propre = institution.strip()
        if not propre:
            return None
        cle = propre.lower()
        if cle in self._institution_lookup:
            return self._institution_lookup[cle]
        self._institution_lookup[cle] = propre
        self.institutions_ref.add(propre)
        return propre
    
    def _generer_variantes_equipes(self) -> List[str]:
        """Generate team variants with gender suffixes.
        
        Génère toutes les variantes possibles:
        - Le nom de base (ex: "LYON 1 (1)")
        - Les variantes avec genre si le genre est connu (ex: "LYON 1 (1) [M]", "LYON 1 (1) [F]")
        """
        variantes = []
        for nom_equipe, genres in self.equipes_par_nom.items():
            # Toujours ajouter le nom de base
            variantes.append(nom_equipe)
            # Ajouter les variantes avec genre
            for genre in genres:
                variantes.append(formater_nom_avec_genre(nom_equipe, genre))
        return sorted(set(variantes))
    
    def _valider_toutes_feuilles(self) -> int:
        """Validate all sheets and detect extra/unused sheets."""
        feuilles_attendues = set(self.config.STRUCTURES.keys())
        nb_problemes = 0
        
        # Load workbook to get actual sheets
        import openpyxl
        wb = openpyxl.load_workbook(self.fichier_path, read_only=True)
        feuilles_presentes = set(wb.sheetnames)
        wb.close()
        
        # Feuilles système/internes et feuilles de données de base à ignorer
        feuilles_systeme = {
            '_Listes_Validation', 
            'Listes_Deroulantes',
        }
        
        # Detect extra sheets (present but not expected)
        feuilles_extra = feuilles_presentes - feuilles_attendues - feuilles_systeme - {'Equipes'}
        if feuilles_extra:
            # Add info to global report
            self.rapport.feuilles_extra = list(feuilles_extra)
        
        # Detect missing sheets
        feuilles_manquantes = feuilles_attendues - feuilles_presentes
        if feuilles_manquantes:
            self.rapport.feuilles_manquantes = list(feuilles_manquantes)
        
        # Validate expected sheets
        for nom_feuille in feuilles_attendues:
            rapport = self._valider_feuille(nom_feuille)
            self.rapport.rapports_feuilles[nom_feuille] = rapport
            
            if rapport.has_problems:
                nb_problemes += 1
        
        # Validate Equipes sheet (special handling)
        if 'Equipes' in feuilles_presentes:
            rapport_equipes = self._valider_feuille_equipes()
            self.rapport.rapports_feuilles['Equipes'] = rapport_equipes
            if rapport_equipes.has_problems:
                nb_problemes += 1
        
        return nb_problemes
    
    def _valider_feuille(self, nom_feuille: str) -> RapportFeuille:
        """Validate a single sheet."""
        rapport = RapportFeuille(nom=nom_feuille)
        
        df = self.config.lire_feuille(nom_feuille)
        if df is None:
            return rapport
        
        structure = self.config.STRUCTURES.get(nom_feuille, {})
        colonnes_attendues = structure.get('colonnes', [])
        
        # Check structure
        self._verifier_structure_colonnes(df, colonnes_attendues, rapport)
        
        if df.empty:
            return rapport
        
        # Validate content
        rapport.nb_lignes_total = len(df)
        self._valider_contenu(df, nom_feuille, rapport)
        
        # Sheet-specific validations
        if nom_feuille == 'Matchs_Fixes':
            self._valider_matchs_fixes(df, rapport)
        elif nom_feuille == 'Contraintes_Temporelles':
            self._valider_contraintes_temporelles(df, rapport)
        elif nom_feuille == 'Ententes':
            self._valider_ententes(df, rapport)
        elif nom_feuille == 'Dispos_Gymnases_Equipes':
            self._valider_dispos_gymnases_equipes(df, rapport)
        elif nom_feuille == 'Coach_Groups':
            self._valider_coach_groups(df, rapport)
        
        return rapport
    
    def _verifier_structure_colonnes(self, df: pd.DataFrame, 
                                      colonnes_attendues: List[str],
                                      rapport: RapportFeuille):
        """Verify column structure."""
        colonnes_presentes = set(df.columns)
        colonnes_attendues_set = set(colonnes_attendues)
        
        rapport.colonnes_manquantes = list(colonnes_attendues_set - colonnes_presentes)
        
        colonnes_extra = colonnes_presentes - colonnes_attendues_set
        
        # Try fuzzy matching for misspelled columns
        for col_extra in colonnes_extra:
            matches = difflib.get_close_matches(col_extra, colonnes_attendues, n=1, cutoff=0.7)
            if matches and matches[0] not in colonnes_presentes:
                rapport.colonnes_renommees.append((col_extra, matches[0]))
            else:
                rapport.colonnes_extra.append(col_extra)
        
        if rapport.colonnes_manquantes or rapport.colonnes_renommees or rapport.colonnes_extra:
            rapport.structure_modifiee = True
    
    def _valider_contenu(self, df: pd.DataFrame, nom_feuille: str, rapport: RapportFeuille):
        """Validate content of each cell."""
        lignes_valides = 0
        
        for idx, row in df.iterrows():
            ligne_num = int(str(idx)) + 2  # type: ignore
            ligne_valide = True
            
            if row.isna().all():
                continue
            
            for colonne in df.columns:
                if str(colonne).startswith('EXTRA_'):
                    continue
                
                valeur = row[colonne]
                result = self._valider_cellule(colonne, valeur, nom_feuille, row)
                
                if not result.valide or result.message:
                    if not result.valide:
                        ligne_valide = False
                    
                    rapport.add_issue(
                        ligne=ligne_num,
                        colonne=colonne,
                        message=result.message or "Valeur invalide",
                        severite=result.severite,
                        valeur_actuelle=valeur,
                        valeur_suggeree=result.valeur_corrigee,
                        auto_correctable=result.auto_correctable
                    )
                    
                    # Handle correction
                    if result.valeur_corrigee is not None and result.auto_correctable:
                        if self._should_apply_correction(ligne_num, colonne, valeur, result.valeur_corrigee):
                            self._planifier_correction(rapport, idx, colonne, result.valeur_corrigee)
                            rapport.add_correction(ligne_num, colonne, valeur, result.valeur_corrigee)
            
            if ligne_valide:
                lignes_valides += 1
        
        rapport.nb_lignes_valides = lignes_valides
    
    def _valider_cellule(self, colonne: str, valeur: Any, nom_feuille: str,
                         row: Optional[pd.Series] = None) -> ValidationResult:
        """Validate a single cell."""
        # Column-specific validation
        if colonne == 'Semaine':
            semaines_valides = set(self.week_dates.keys()) if self.week_dates else None
            
            # Pour Matchs_Fixes: validation allégée car ces matchs peuvent être:
            # 1. Des matchs déjà joués (semaines passées)
            # 2. Des ententes (dates hors jour officiel)
            # 3. Des matchs à venir dans la plage de planification
            # On ne valide pas strictement les semaines des matchs fixes
            if nom_feuille == 'Matchs_Fixes':
                # Accepter toute semaine entre 1 et 52 pour les matchs fixes
                return self.validator.valider_semaine(
                    valeur, 
                    nb_semaines_max=52,
                    semaines_valides=None,  # Pas de restriction
                    week_dates=self.week_dates
                )
            
            return self.validator.valider_semaine(
                valeur, 
                semaines_valides=semaines_valides,
                week_dates=self.week_dates,
                semaines_banalisees=self.semaines_banalisees
            )
        
        elif colonne in ['Heure_Debut', 'Heure_Fin', 'Horaire', 'Horaire_Dispo']:
            return self.validator.valider_horaire(valeur)
        
        elif colonne == 'Date':
            return self.validator.valider_date(valeur)
        
        elif colonne == 'Genre':
            obligatoire = nom_feuille in ['Matchs_Fixes', 'Contraintes_Temporelles', 
                                          'Dispos_Gymnases_Equipes', 'Equipes_Hors_Championnat']
            return self.validator.valider_genre(valeur, obligatoire=obligatoire)
        
        elif colonne == 'Genre_Prioritaire':
            return self.validator.valider_genre(valeur, obligatoire=False)
        
        elif colonne == 'Niveau':
            return self.validator.valider_niveau(valeur)
        
        elif colonne == 'Type_Contrainte':
            return self.validator.valider_type_contrainte(valeur)
        
        elif colonne == 'Type' and nom_feuille == 'Types_Poules':
            return self.validator.valider_type_poule(valeur)
        
        elif colonne == 'Type_Competition' or colonne == 'Type_Championnat':
            return self.validator.valider_type_championnat(valeur)
        
        elif colonne == 'Score':
            return self.validator.valider_score(valeur)
        
        elif colonne == 'Gymnase' or colonne.startswith('Gymnase_'):
            if pd.isna(valeur) or str(valeur).strip() == '':
                if colonne.startswith('Gymnase_Pref_') or colonne.startswith('Gymnase_'):
                    return ValidationResult.ok()  # Optional
                return ValidationResult.error("Gymnase manquant")
            return self.validator.valider_gymnase(valeur, self.gymnases_ref)
        
        elif colonne == 'Institution' or colonne.startswith('Institution_'):
            if nom_feuille == 'Equipes_Hors_Championnat':
                if pd.isna(valeur) or str(valeur).strip() == '':
                    return ValidationResult.error("Institution manquante")
                return ValidationResult.ok()  # External institutions allowed
            return self.validator.valider_institution(valeur, self.institutions_ref)
        
        elif colonne == 'Equipe':
            if nom_feuille == 'Equipes_Hors_Championnat':
                if pd.isna(valeur) or str(valeur).strip() == '':
                    return ValidationResult.error("Équipe manquante")
                return ValidationResult.ok()  # External teams allowed
            # Dispos_Gymnases_Equipes: équipes sans genre [M]/[F]
            if nom_feuille == 'Dispos_Gymnases_Equipes':
                return self.validator.valider_equipe(valeur, self.equipes_sans_genre)
            # Indispos_Equipes: peut contenir des équipes avec ou sans [M]/[F]
            if nom_feuille == 'Indispos_Equipes':
                return self.validator.valider_equipe(valeur, self.equipes_toutes_variantes)
            return self.validator.valider_equipe(valeur, self.equipes_toutes_variantes)
        
        elif colonne in ['Equipe_1', 'Equipe_2']:
            if pd.isna(valeur) or str(valeur).strip() == '':
                if nom_feuille == 'Matchs_Fixes':
                    return ValidationResult.ok()  # One team can be empty
                return ValidationResult.error(f"{colonne} manquante")
            # For Matchs_Fixes and Contraintes: teams are without [F]/[M]
            return self.validator.valider_equipe(valeur, self.equipes_sans_genre)
        
        elif colonne == 'Capacite_Occupee':
            return self.validator.valider_capacite(valeur)
        
        elif colonne in ['Remarques', 'Remarque', 'notes', 'Motif', 'Arbitres']:
            return self.validator.valider_texte_libre(valeur)
        
        # Default: accept
        return ValidationResult.ok()
    
    def _should_apply_correction(self, ligne: int, colonne: str, 
                                  old_value: Any, new_value: Any) -> bool:
        """Determine if a correction should be applied based on mode."""
        if self.options.mode == UpdateMode.VALIDATE:
            return False
        
        if self.options.mode == UpdateMode.AUTO_CORRECT:
            return True
        
        if self.options.mode == UpdateMode.INTERACTIVE:
            if self._accept_all_corrections:
                return True
            if self._ignore_all_corrections:
                return False
            
            result = prompt_user_correction(
                f"Ligne {ligne}, colonne '{colonne}'",
                str(old_value),
                str(new_value)
            )
            
            if result == '__ACCEPT_ALL__':
                self._accept_all_corrections = True
                return True
            elif result == '__IGNORE_ALL__':
                self._ignore_all_corrections = True
                return False
            elif result is None:
                return False
            else:
                return True
        
        return False
    
    def _planifier_correction(self, rapport: RapportFeuille, row_idx: Any, 
                               colonne: str, valeur: Any):
        """Schedule a correction to be applied."""
        mises = rapport.mises_a_jour_cellules.setdefault(row_idx, {})
        mises[colonne] = valeur
    
    def _valider_matchs_fixes(self, df: pd.DataFrame, rapport: RapportFeuille):
        """Validate Matchs_Fixes sheet."""
        # Build team -> poule mapping
        equipe_vers_poules = {}
        df_equipes = self.config.lire_feuille('Equipes')
        if df_equipes is not None:
            for _, row in df_equipes.iterrows():
                equipe = str(row.get('Equipe', '')).strip()
                poule = str(row.get('Poule', '')).strip()
                if equipe and poule:
                    equipe_vers_poules.setdefault(equipe, []).append(poule)
        
        for idx, row in df.iterrows():
            ligne_num = int(str(idx)) + 2  # type: ignore
            
            eq1 = row.get('Equipe_1')
            eq2 = row.get('Equipe_2')
            genre = row.get('Genre')
            poule = row.get('Poule')
            semaine = row.get('Semaine')
            date = row.get('Date')
            type_competition = row.get('Type_Competition') or row.get('Type_Championnat')
            
            # Skip empty rows
            if pd.isna(eq1) and pd.isna(eq2):
                continue
            
            # Validate at least one team
            eq1_str = str(eq1).strip() if pd.notna(eq1) else None
            eq2_str = str(eq2).strip() if pd.notna(eq2) else None
            
            if not eq1_str and not eq2_str:
                rapport.add_issue(ligne_num, 'Equipe_1', 
                                  "Au moins une équipe requise", Severity.ERROR)
                continue
            
            # Déterminer si c'est une compétition académique (Acad)
            # Les auto-détections de poule ne s'appliquent qu'aux matchs Acad
            is_acad = True  # Par défaut
            if pd.notna(type_competition):
                type_comp_str = str(type_competition).strip().upper()
                # Si c'est CFE, CFU ou autre, ce n'est pas Acad
                if type_comp_str in ['CFE', 'CFU', 'AUTRE', 'AMICAL', 'TOURNOI']:
                    is_acad = False
            
            # Vérifier la cohérence de la poule renseignée
            poule_str = str(poule).strip() if pd.notna(poule) else ''
            
            if poule_str:
                # La poule est renseignée : vérifier que les deux équipes en font partie
                poules_eq1 = set(equipe_vers_poules.get(eq1_str, []))
                poules_eq2 = set(equipe_vers_poules.get(eq2_str, []))
                
                erreurs_poule = []
                if eq1_str and poule_str not in poules_eq1:
                    erreurs_poule.append(f"Équipe 1 ({eq1_str}) n'est pas dans la poule {poule_str}")
                if eq2_str and poule_str not in poules_eq2:
                    erreurs_poule.append(f"Équipe 2 ({eq2_str}) n'est pas dans la poule {poule_str}")
                
                if erreurs_poule:
                    rapport.add_issue(
                        ligne_num, 'Poule',
                        f"Incohérence de poule: {'; '.join(erreurs_poule)}",
                        Severity.ERROR
                    )
            
            # Auto-detect poule if missing - SEULEMENT pour les matchs Acad
            if pd.isna(poule) or not str(poule).strip():
                if is_acad and eq1_str and eq2_str:
                    poules_eq1 = set(equipe_vers_poules.get(eq1_str, []))
                    poules_eq2 = set(equipe_vers_poules.get(eq2_str, []))
                    poules_communes = poules_eq1 & poules_eq2
                    
                    if len(poules_communes) == 1:
                        poule_detectee = list(poules_communes)[0]
                        rapport.add_issue(
                            ligne_num, 'Poule',
                            f"Poule auto-détectée: '{poule_detectee}'",
                            Severity.INFO,
                            valeur_suggeree=poule_detectee,
                            auto_correctable=True
                        )
                        self._planifier_correction(rapport, idx, 'Poule', poule_detectee)
                    elif len(poules_communes) > 1:
                        rapport.add_issue(
                            ligne_num, 'Poule',
                            f"Poule ambiguë: {', '.join(poules_communes)}",
                            Severity.ERROR
                        )
                elif not is_acad:
                    # Pour les matchs non-Acad, la poule n'est pas obligatoire
                    pass
            
            # Auto-fill week from date or vice versa
            if self.calendar_manager:
                if pd.notna(date) and (pd.isna(semaine) or not str(semaine).strip()):
                    try:
                        date_parsed = pd.to_datetime(date, dayfirst=True).to_pydatetime()
                        semaine_inferee = self.calendar_manager.infer_semaine_from_date(date_parsed)
                        if semaine_inferee:
                            rapport.add_issue(
                                ligne_num, 'Semaine',
                                f"Semaine déduite de la date: {semaine_inferee}",
                                Severity.INFO,
                                valeur_suggeree=semaine_inferee,
                                auto_correctable=True
                            )
                            self._planifier_correction(rapport, idx, 'Semaine', semaine_inferee)
                    except Exception:
                        pass
                
                elif pd.notna(semaine) and (pd.isna(date) or not str(date).strip()):
                    try:
                        semaine_int = int(float(semaine))
                        date_calculee = self.calendar_manager.semaine_to_date(semaine_int)
                        if date_calculee:
                            date_str = date_calculee.strftime(DATE_DISPLAY_FORMAT)
                            rapport.add_issue(
                                ligne_num, 'Date',
                                f"Date déduite de la semaine: {date_str}",
                                Severity.INFO,
                                valeur_suggeree=date_str,
                                auto_correctable=True
                            )
                            self._planifier_correction(rapport, idx, 'Date', date_str)
                    except Exception:
                        pass
        
        # === Détection des doublons de matchs ===
        self._detecter_doublons_matchs(df, rapport)
    
    def _detecter_doublons_matchs(self, df: pd.DataFrame, rapport: RapportFeuille):
        """
        Détecte les matchs en doublon dans Matchs_Fixes.
        
        Règles:
        - Un même match (mêmes équipes, même genre) ne peut apparaître qu'une fois
          dans une poule Classique
        - Dans une poule Aller-Retour, deux matchs sont autorisés
        - Les matchs marqués "compte double" sont exclus de la détection
        - L'ordre des équipes n'a pas d'importance (A vs B == B vs A)
        """
        # Charger les types de poules pour identifier les Aller-Retour
        df_types_poules = self.config.lire_feuille('Types_Poules')
        
        # Détecter les doublons avec le module dédié
        issues = detect_match_duplicates(df, df_types_poules)
        
        # Ajouter les issues au rapport
        for issue in issues:
            rapport.issues.append(issue)
            
            # Legacy compatibility
            if issue.severite == Severity.ERROR:
                rapport.erreurs_contenu.append(
                    f"L{issue.ligne}, '{issue.colonne}': {issue.message}"
                )
            elif issue.severite == Severity.WARNING:
                rapport.warnings_contenu.append(
                    f"L{issue.ligne}, '{issue.colonne}': {issue.message}"
                )
    
    def _valider_contraintes_temporelles(self, df: pd.DataFrame, rapport: RapportFeuille):
        """Validate Contraintes_Temporelles sheet."""
        seen = {}
        
        for idx, row in df.iterrows():
            ligne_num = int(str(idx)) + 2  # type: ignore
            
            eq1 = row.get('Equipe_1')
            eq2 = row.get('Equipe_2')
            genre = row.get('Genre')
            type_c = row.get('Type_Contrainte')
            semaine = row.get('Semaine')
            
            # Check duplicates
            if pd.notna(eq1) and pd.notna(eq2) and pd.notna(genre):
                key = tuple(sorted([str(eq1).strip(), str(eq2).strip()])) + (str(genre).strip(),)
                if key in seen:
                    rapport.add_issue(
                        ligne_num, 'Equipe_1',
                        f"Contrainte en doublon avec ligne {seen[key]}",
                        Severity.WARNING
                    )
                else:
                    seen[key] = ligne_num
    
    def _valider_ententes(self, df: pd.DataFrame, rapport: RapportFeuille):
        """Validate Ententes sheet."""
        seen = {}
        
        for idx, row in df.iterrows():
            ligne_num = int(str(idx)) + 2  # type: ignore
            
            inst1 = row.get('Institution_1')
            inst2 = row.get('Institution_2')
            
            if pd.isna(inst1) or pd.isna(inst2):
                continue
            
            key = tuple(sorted([str(inst1).strip(), str(inst2).strip()]))
            if key in seen:
                rapport.add_issue(
                    ligne_num, 'Institution_1',
                    f"Entente en doublon avec ligne {seen[key]}",
                    Severity.ERROR
                )
            else:
                seen[key] = ligne_num
    
    def _valider_dispos_gymnases_equipes(self, df: pd.DataFrame, rapport: RapportFeuille):
        """Validate Dispos_Gymnases_Equipes sheet."""
        seen = {}
        
        for idx, row in df.iterrows():
            ligne_num = int(str(idx)) + 2  # type: ignore
            
            equipe = row.get('Equipe')
            genre = row.get('Genre')
            
            if pd.isna(equipe) or pd.isna(genre):
                continue
            
            key = (str(equipe).strip(), str(genre).strip().upper())
            if key in seen:
                rapport.add_issue(
                    ligne_num, 'Equipe',
                    f"Disponibilité en doublon avec ligne {seen[key]}",
                    Severity.ERROR
                )
            else:
                seen[key] = ligne_num
    
    def _valider_coach_groups(self, df: pd.DataFrame, rapport: RapportFeuille):
        """Validate Coach_Groups sheet."""
        team_owners = {}
        
        for idx, row in df.iterrows():
            ligne_num = int(str(idx)) + 2  # type: ignore
            
            coach_name = str(row.get('coach_name', '')).strip()
            if not coach_name:
                rapport.add_issue(
                    ligne_num, 'coach_name',
                    "Nom du coach obligatoire",
                    Severity.ERROR
                )
                continue
    
    def _valider_feuille_equipes(self) -> RapportFeuille:
        """
        Valide la feuille Equipes avec gestion des colonnes Niveau_Equipe et Genre_Equipe.
        
        Logique pour chaque attribut (Niveau_Equipe et Genre_Equipe):
        1. Si colonne vide et Poule est définie -> auto-remplir depuis Poule
        2. Si colonne définie et Poule est définie -> vérifier cohérence
        3. Si colonne définie et Poule est vide -> OK (équipe avec attribut mais pas assignée)
        4. Si colonne vide et Poule est vide -> OK (équipe sans attribut ni poule)
        """
        from pycalendar.core.utils import extraire_niveau_depuis_poule, extraire_genre_depuis_poule
        
        rapport = RapportFeuille(nom='Equipes')
        
        df = self.config.lire_feuille('Equipes')
        if df is None or df.empty:
            return rapport
        
        rapport.nb_lignes_total = len(df)
        lignes_valides = 0
        
        # Vérifier si les colonnes existent
        a_colonne_niveau = 'Niveau_Equipe' in df.columns
        a_colonne_genre = 'Genre_Equipe' in df.columns
        
        for idx, row in df.iterrows():
            ligne_num = int(str(idx)) + 2  # type: ignore
            ligne_valide = True
            
            equipe = row.get('Equipe')
            if pd.isna(equipe) or not str(equipe).strip():
                continue
            
            equipe_str = str(equipe).strip()
            poule = row.get('Poule')
            poule_str = str(poule).strip() if pd.notna(poule) else ''
            
            # ===== Validation Niveau_Equipe =====
            niveau_equipe = row.get('Niveau_Equipe') if a_colonne_niveau else None
            niveau_str = str(niveau_equipe).strip().upper() if pd.notna(niveau_equipe) and str(niveau_equipe).strip() else ''
            
            # Extraire le niveau depuis la poule (si poule définie)
            niveau_depuis_poule = extraire_niveau_depuis_poule(poule_str) if poule_str else ''
            
            # Cas 1: Niveau_Equipe vide et Poule définie -> auto-remplir
            if not niveau_str and niveau_depuis_poule:
                rapport.add_issue(
                    ligne_num, 'Niveau_Equipe',
                    f"Niveau auto-détecté depuis poule: '{niveau_depuis_poule}'",
                    Severity.INFO,
                    valeur_suggeree=niveau_depuis_poule,
                    auto_correctable=True
                )
                if self._should_apply_correction(ligne_num, 'Niveau_Equipe', niveau_equipe, niveau_depuis_poule):
                    self._planifier_correction_equipes(rapport, idx, 'Niveau_Equipe', niveau_depuis_poule)
                    rapport.add_correction(ligne_num, 'Niveau_Equipe', niveau_equipe, niveau_depuis_poule)
            
            # Cas 2: Niveau_Equipe défini et Poule définie -> vérifier cohérence
            elif niveau_str and niveau_depuis_poule:
                if niveau_str != niveau_depuis_poule:
                    rapport.add_issue(
                        ligne_num, 'Niveau_Equipe',
                        f"Incohérence: niveau '{niveau_str}' ≠ poule '{poule_str}' (niveau attendu: '{niveau_depuis_poule}')",
                        Severity.ERROR
                    )
                    ligne_valide = False
            
            # Valider le format du niveau si défini
            if niveau_str:
                result = self.validator.valider_niveau_equipe(niveau_str)
                if not result.valide:
                    rapport.add_issue(
                        ligne_num, 'Niveau_Equipe',
                        result.message or "Niveau invalide",
                        result.severite
                    )
                    ligne_valide = False
                elif result.message and result.valeur_corrigee:
                    rapport.add_issue(
                        ligne_num, 'Niveau_Equipe',
                        result.message,
                        result.severite,
                        valeur_suggeree=result.valeur_corrigee,
                        auto_correctable=result.auto_correctable
                    )
                    if result.auto_correctable and self._should_apply_correction(
                        ligne_num, 'Niveau_Equipe', niveau_equipe, result.valeur_corrigee
                    ):
                        self._planifier_correction_equipes(rapport, idx, 'Niveau_Equipe', result.valeur_corrigee)
            
            # ===== Validation Genre_Equipe =====
            genre_equipe = row.get('Genre_Equipe') if a_colonne_genre else None
            genre_str = str(genre_equipe).strip().upper() if pd.notna(genre_equipe) and str(genre_equipe).strip() else ''
            
            # Extraire le genre depuis la poule (si poule définie)
            genre_depuis_poule = extraire_genre_depuis_poule(poule_str) if poule_str else ''
            
            # Cas 1: Genre_Equipe vide et Poule définie -> auto-remplir
            if not genre_str and genre_depuis_poule:
                rapport.add_issue(
                    ligne_num, 'Genre_Equipe',
                    f"Genre auto-détecté depuis poule: '{genre_depuis_poule}'",
                    Severity.INFO,
                    valeur_suggeree=genre_depuis_poule,
                    auto_correctable=True
                )
                if self._should_apply_correction(ligne_num, 'Genre_Equipe', genre_equipe, genre_depuis_poule):
                    self._planifier_correction_equipes(rapport, idx, 'Genre_Equipe', genre_depuis_poule)
                    rapport.add_correction(ligne_num, 'Genre_Equipe', genre_equipe, genre_depuis_poule)
            
            # Cas 2: Genre_Equipe défini et Poule définie -> vérifier cohérence
            elif genre_str and genre_depuis_poule:
                if genre_str != genre_depuis_poule:
                    rapport.add_issue(
                        ligne_num, 'Genre_Equipe',
                        f"Incohérence: genre '{genre_str}' ≠ poule '{poule_str}' (genre attendu: '{genre_depuis_poule}')",
                        Severity.ERROR
                    )
                    ligne_valide = False
            
            # Valider le format du genre si défini
            if genre_str:
                result = self.validator.valider_genre_equipe(genre_str)
                if not result.valide:
                    rapport.add_issue(
                        ligne_num, 'Genre_Equipe',
                        result.message or "Genre invalide",
                        result.severite
                    )
                    ligne_valide = False
                elif result.message and result.valeur_corrigee:
                    rapport.add_issue(
                        ligne_num, 'Genre_Equipe',
                        result.message,
                        result.severite,
                        valeur_suggeree=result.valeur_corrigee,
                        auto_correctable=result.auto_correctable
                    )
                    if result.auto_correctable and self._should_apply_correction(
                        ligne_num, 'Genre_Equipe', genre_equipe, result.valeur_corrigee
                    ):
                        self._planifier_correction_equipes(rapport, idx, 'Genre_Equipe', result.valeur_corrigee)
            
            if ligne_valide:
                lignes_valides += 1
        
        rapport.nb_lignes_valides = lignes_valides
        return rapport
    
    def _planifier_correction_equipes(self, rapport: RapportFeuille, row_idx: Any,
                                       colonne: str, valeur: Any):
        """Schedule a correction for the Equipes sheet."""
        mises = rapport.mises_a_jour_cellules.setdefault(row_idx, {})
        mises[colonne] = valeur

    def _generer_feuilles_manquantes(self) -> int:
        """Generate missing sheets."""
        statuts = self.config.generer_feuilles_manquantes(conserver_existant=True)
        return len([s for s in statuts.values() if s == 'créée'])
    
    def _supprimer_exemples(self) -> int:
        """Remove example rows from all sheets."""
        wb = openpyxl.load_workbook(self.fichier_path)
        nb_suppressions = 0
        
        for nom_feuille in self.config.STRUCTURES.keys():
            if nom_feuille not in wb.sheetnames:
                continue
            
            structure = self.config.STRUCTURES.get(nom_feuille, {})
            exemple = structure.get('exemple', {})
            if not exemple:
                continue
            
            df = self.config.lire_feuille(nom_feuille)
            if df is None or len(df) == 0:
                continue
            
            # Check if first row matches example
            premiere_ligne = df.iloc[0]
            est_exemple = True
            
            for col in df.columns:
                if col in exemple:
                    if str(premiere_ligne[col]).strip() != str(exemple[col]).strip():
                        est_exemple = False
                        break
            
            if est_exemple:
                # Remove the example row
                ws = wb[nom_feuille]
                ws.delete_rows(2)  # Delete row 2 (first data row)
                nb_suppressions += 1
        
        if nb_suppressions > 0:
            wb.save(self.fichier_path)
        
        return nb_suppressions
    
    def _preremplir_types_poules(self) -> int:
        """Pre-fill Types_Poules sheet with all pools."""
        df_equipes = self.config.lire_feuille('Equipes')
        if df_equipes is None or 'Poule' not in df_equipes.columns:
            return 0
        
        # Get unique pools
        poules_uniques = set()
        for poule in df_equipes['Poule'].dropna():
            poule_str = str(poule).strip()
            if poule_str:
                poules_uniques.add(poule_str)
        
        if not poules_uniques:
            return 0
        
        # Get existing pools in Types_Poules
        df_types = self.config.lire_feuille('Types_Poules')
        poules_existantes = set()
        if df_types is not None and 'Poule' in df_types.columns:
            for poule in df_types['Poule'].dropna():
                poules_existantes.add(str(poule).strip())
        
        # Find pools to add
        poules_a_ajouter = [p for p in sorted(poules_uniques) if p not in poules_existantes]
        
        if not poules_a_ajouter:
            return 0
        
        # Add new pools
        wb = openpyxl.load_workbook(self.fichier_path)
        if 'Types_Poules' not in wb.sheetnames:
            return 0
        
        ws = wb['Types_Poules']
        
        # Find last row
        last_row: int = 1
        for row in ws.iter_rows(min_row=2):
            if row[0].value and row[0].row is not None:
                last_row = row[0].row
        
        # Add pools
        for poule in poules_a_ajouter:
            last_row += 1
            ws.cell(last_row, 1, poule)
            ws.cell(last_row, 2, 'Classique')
        
        wb.save(self.fichier_path)
        return len(poules_a_ajouter)
    
    def _appliquer_corrections(self) -> int:
        """Apply all scheduled corrections."""
        if self.options.mode == UpdateMode.VALIDATE:
            return 0
        
        corrections_appliquees = 0
        wb = openpyxl.load_workbook(self.fichier_path)
        
        for nom_feuille, rapport in self.rapport.rapports_feuilles.items():
            if not rapport.mises_a_jour_cellules and not rapport.structure_modifiee:
                continue
            
            if nom_feuille not in wb.sheetnames:
                continue
            
            df = self.config.lire_feuille(nom_feuille)
            if df is None:
                continue
            
            # Handle Equipes sheet specially
            if nom_feuille == 'Equipes':
                corrections_appliquees += self._appliquer_corrections_equipes(wb, df, rapport)
                continue
            
            structure = self.config.STRUCTURES.get(nom_feuille, {})
            colonnes_attendues = structure.get('colonnes', [])
            
            # Rename columns
            for ancien, nouveau in rapport.colonnes_renommees:
                if ancien in df.columns:
                    df.rename(columns={ancien: nouveau}, inplace=True)
                    corrections_appliquees += 1
            
            # Add missing columns
            for col in rapport.colonnes_manquantes:
                df[col] = None
                corrections_appliquees += 1
            
            # Rename extra columns
            for col in rapport.colonnes_extra:
                if col in df.columns:
                    df.rename(columns={col: f"EXTRA_{col}"}, inplace=True)
                    corrections_appliquees += 1
            
            # Apply cell updates
            for row_idx, updates in rapport.mises_a_jour_cellules.items():
                if row_idx not in df.index:
                    continue
                for col, val in updates.items():
                    if col in df.columns:
                        df.at[row_idx, col] = val
                        corrections_appliquees += 1
            
            # Reorder columns
            colonnes_ordonnees = [c for c in colonnes_attendues if c in df.columns]
            colonnes_ordonnees += [c for c in df.columns if c.startswith('EXTRA_')]
            df = df[[c for c in colonnes_ordonnees if c in df.columns]]
            
            # Rewrite sheet
            ws_index = wb.sheetnames.index(nom_feuille)
            del wb[nom_feuille]
            ws = wb.create_sheet(nom_feuille, ws_index)
            
            # Write headers
            for c_idx, col in enumerate(df.columns, 1):
                ws.cell(1, c_idx, col)
            
            # Write data
            for r_idx, row in enumerate(df.itertuples(index=False), 2):
                for c_idx, value in enumerate(row, 1):
                    if not pd.isna(value):
                        ws.cell(r_idx, c_idx, value)
        
        wb.save(self.fichier_path)
        return corrections_appliquees
    
    def _appliquer_corrections_equipes(self, wb: openpyxl.Workbook, df: pd.DataFrame,
                                        rapport: RapportFeuille) -> int:
        """
        Applique les corrections à la feuille Equipes.
        
        Gère spécialement les colonnes Niveau_Equipe et Genre_Equipe:
        - Création des colonnes si elles n'existent pas
        - Insertion avant la colonne Poule (Genre_Equipe puis Niveau_Equipe)
        - Application des valeurs auto-détectées
        """
        corrections = 0
        ws = wb['Equipes']
        
        # Trouver les indices des colonnes
        colonnes_ws = {}
        for idx, cell in enumerate(ws[1], 1):
            if cell.value:
                colonnes_ws[str(cell.value)] = idx
        
        # Trouver la position de Poule pour les insertions
        poule_idx = colonnes_ws.get('Poule')
        
        # Vérifier si Genre_Equipe existe, sinon la créer avant Poule
        if 'Genre_Equipe' not in colonnes_ws:
            if poule_idx:
                insert_idx = poule_idx
            else:
                insert_idx = ws.max_column + 1
            
            # Insérer une colonne
            ws.insert_cols(insert_idx)
            ws.cell(1, insert_idx, 'Genre_Equipe')
            colonnes_ws['Genre_Equipe'] = insert_idx
            
            # Mettre à jour les indices des colonnes après l'insertion
            for col_name in list(colonnes_ws.keys()):
                if colonnes_ws[col_name] >= insert_idx and col_name != 'Genre_Equipe':
                    colonnes_ws[col_name] = colonnes_ws[col_name] + 1
            
            # Mettre à jour poule_idx si elle a été décalée
            if poule_idx:
                poule_idx = colonnes_ws.get('Poule')
            
            corrections += 1
            print_info("Colonne 'Genre_Equipe' créée avant 'Poule'")
        
        # Vérifier si Niveau_Equipe existe, sinon la créer avant Poule (après Genre_Equipe)
        if 'Niveau_Equipe' not in colonnes_ws:
            if poule_idx:
                insert_idx = poule_idx
            else:
                insert_idx = ws.max_column + 1
            
            # Insérer une colonne
            ws.insert_cols(insert_idx)
            ws.cell(1, insert_idx, 'Niveau_Equipe')
            colonnes_ws['Niveau_Equipe'] = insert_idx
            
            # Mettre à jour les indices des colonnes après l'insertion
            for col_name in list(colonnes_ws.keys()):
                if colonnes_ws[col_name] >= insert_idx and col_name != 'Niveau_Equipe':
                    colonnes_ws[col_name] = colonnes_ws[col_name] + 1
            
            corrections += 1
            print_info("Colonne 'Niveau_Equipe' créée avant 'Poule'")
        
        # Appliquer les mises à jour de cellules
        niveau_col_idx = colonnes_ws.get('Niveau_Equipe')
        genre_col_idx = colonnes_ws.get('Genre_Equipe')
        
        if rapport.mises_a_jour_cellules:
            for row_idx, updates in rapport.mises_a_jour_cellules.items():
                excel_row = int(str(row_idx)) + 2  # type: ignore
                for col, val in updates.items():
                    if col == 'Niveau_Equipe' and niveau_col_idx:
                        ws.cell(excel_row, niveau_col_idx, val)
                        corrections += 1
                    elif col == 'Genre_Equipe' and genre_col_idx:
                        ws.cell(excel_row, genre_col_idx, val)
                        corrections += 1
        
        return corrections
    
    def _appliquer_formatage(self):
        """Apply visual formatting to the workbook."""
        try:
            format_workbook(str(self.fichier_path))
            # Appliquer le surlignage des erreurs après le formatage de base
            self._appliquer_surlignage_erreurs()
        except Exception as e:
            print_warning(f"Erreur de formatage: {e}")
    
    def _appliquer_surlignage_erreurs(self):
        """
        Surligne les cellules contenant des erreurs ou avertissements dans Excel.
        
        Couleurs utilisées:
        - Rouge clair pour les erreurs
        - Jaune/ambre pour les avertissements
        """
        try:
            wb = openpyxl.load_workbook(self.fichier_path)
            
            # Fills pour les différents niveaux de sévérité
            error_fill = PatternFill(
                start_color=Colors.ERROR_BG,
                end_color=Colors.ERROR_BG,
                fill_type="solid"
            )
            warning_fill = PatternFill(
                start_color=Colors.WARNING_BG,
                end_color=Colors.WARNING_BG,
                fill_type="solid"
            )
            
            # Parcourir tous les rapports de feuilles
            for nom_feuille, rapport in self.rapport.rapports_feuilles.items():
                if nom_feuille not in wb.sheetnames:
                    continue
                
                if not rapport.issues:
                    continue
                
                ws = wb[nom_feuille]
                
                # Construire le mapping colonne -> index
                col_name_to_idx: Dict[str, int] = {}
                for col_idx, cell in enumerate(ws[1], 1):
                    if cell.value:
                        col_name_to_idx[str(cell.value)] = col_idx
                
                # Appliquer le surlignage pour chaque issue
                for issue in rapport.issues:
                    # Ne surligner que les erreurs et warnings
                    if issue.severite not in (Severity.ERROR, Severity.CRITICAL, Severity.WARNING):
                        continue
                    
                    col_idx = col_name_to_idx.get(issue.colonne)
                    if col_idx is None:
                        continue
                    
                    row_idx = issue.ligne
                    if row_idx < 2:  # Skip header row
                        continue
                    
                    cell = ws.cell(row_idx, col_idx)
                    
                    if issue.severite in (Severity.ERROR, Severity.CRITICAL):
                        cell.fill = error_fill
                    elif issue.severite == Severity.WARNING:
                        cell.fill = warning_fill
            
            wb.save(self.fichier_path)
            
        except Exception as e:
            print_warning(f"Erreur lors du surlignage des erreurs: {e}")
    
    def _configurer_dropdowns(self):
        """Configure dropdown lists for all sheets."""
        try:
            setup_all_dropdowns(
                filepath=str(self.fichier_path),
                equipes=list(self.equipes_toutes_variantes),
                gymnases=list(self.gymnases_ref),
                institutions=list(self.institutions_ref),
                poules=list(self.poules_ref),
                week_dates=self.week_dates,
                nb_semaines=len(self.week_dates) if self.week_dates else 13
            )
        except Exception as e:
            print_warning(f"Erreur lors de la configuration des dropdowns: {e}")


def actualiser_fichier_v2(fichier: str, options: Optional[UpdateOptions] = None) -> bool:
    """
    Actualise un fichier de configuration avec validation avancée.
    
    Args:
        fichier: Chemin vers le fichier de configuration
        options: Options d'actualisation (optionnel)
        
    Returns:
        True si l'actualisation a réussi (pas d'erreurs bloquantes)
    """
    actualisateur = ConfigActualisateurV2(fichier, options)
    return actualisateur.actualiser()


# For backwards compatibility with the original module
def main():
    """Entry point for command-line usage."""
    import argparse
    
    parser = argparse.ArgumentParser(
        description="Actualise et valide un fichier de configuration Excel"
    )
    
    parser.add_argument(
        '--fichier', '-f',
        help="Fichier de configuration à actualiser"
    )
    parser.add_argument(
        '--mode', '-m',
        choices=['validate', 'auto', 'interactive'],
        default='validate',
        help="Mode: validate (défaut), auto (correction automatique), interactive"
    )
    parser.add_argument(
        '--verbose', '-v',
        action='store_true',
        help="Mode verbeux"
    )
    parser.add_argument(
        '--no-format',
        action='store_true',
        help="Ne pas appliquer le formatage visuel"
    )
    parser.add_argument(
        '--no-dropdowns',
        action='store_true',
        help="Ne pas ajouter les listes déroulantes"
    )
    parser.add_argument(
        '--no-backup',
        action='store_true',
        help="Ne pas créer de sauvegarde"
    )
    
    args = parser.parse_args()
    
    if not args.fichier:
        print("Erreur: --fichier requis")
        return 1
    
    mode_map = {
        'validate': UpdateMode.VALIDATE,
        'auto': UpdateMode.AUTO_CORRECT,
        'interactive': UpdateMode.INTERACTIVE,
    }
    
    options = UpdateOptions(
        mode=mode_map[args.mode],
        verbose=args.verbose,
        format_output=not args.no_format,
        add_dropdowns=not args.no_dropdowns,
        backup=not args.no_backup,
    )
    
    success = actualiser_fichier_v2(args.fichier, options)
    return 0 if success else 1


if __name__ == "__main__":
    import sys
    sys.exit(main())
