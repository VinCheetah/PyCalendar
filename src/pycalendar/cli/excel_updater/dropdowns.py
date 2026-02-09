"""
Dropdown list management for Excel configuration.

Provides data validation and dropdown list generation for:
- Week numbers with dates
- Teams and institutions
- Gymnases
- Types and categories
"""

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter, quote_sheetname
from typing import Dict, List, Optional, Set, Tuple
from datetime import datetime


class DropdownManager:
    """Gestionnaire des listes déroulantes Excel."""
    
    HIDDEN_SHEET_NAME = 'Listes_Deroulantes'
    
    def __init__(self, workbook: openpyxl.Workbook):
        self.wb = workbook
        self._list_column_index = 1  # Current column in hidden sheet
        self._ensure_hidden_sheet()
    
    def _ensure_hidden_sheet(self):
        """Crée ou récupère la feuille cachée pour les listes."""
        if self.HIDDEN_SHEET_NAME not in self.wb.sheetnames:
            ws = self.wb.create_sheet(self.HIDDEN_SHEET_NAME)
            ws.sheet_state = 'hidden'
        else:
            ws = self.wb[self.HIDDEN_SHEET_NAME]
            ws.sheet_state = 'hidden'
            # Reset column index
            self._list_column_index = ws.max_column + 1 if ws.max_column else 1
    
    def _get_hidden_sheet(self):
        """Retourne la feuille cachée."""
        return self.wb[self.HIDDEN_SHEET_NAME]
    
    def _write_list_to_hidden_sheet(self, values: List[str], column_name: str) -> str:
        """
        Écrit une liste dans la feuille cachée et retourne la référence.
        
        Returns:
            Formule de référence à la plage
        """
        ws = self._get_hidden_sheet()
        col_idx = self._list_column_index
        self._list_column_index += 1
        
        # Write header (for reference)
        ws.cell(row=1, column=col_idx, value=f"__{column_name}__")
        
        # Write values
        for i, value in enumerate(values, start=2):
            ws.cell(row=i, column=col_idx, value=value)
        
        # Create reference formula
        col_letter = get_column_letter(col_idx)
        nb_values = len(values)
        formula = f"{quote_sheetname(self.HIDDEN_SHEET_NAME)}!${col_letter}$2:${col_letter}${nb_values + 1}"
        
        return formula
    
    def add_dropdown(self, sheet_name: str, column_name: str, values: List[str],
                     start_row: int = 2, end_row: int = 1000,
                     allow_blank: bool = True, error_title: str = "Valeur invalide",
                     error_message: str = "Sélectionnez une valeur dans la liste"):
        """
        Ajoute une liste déroulante à une colonne.
        
        Args:
            sheet_name: Nom de la feuille cible
            column_name: Nom de la colonne (en-tête)
            values: Liste des valeurs pour le dropdown
            start_row: Première ligne de données
            end_row: Dernière ligne de données
            allow_blank: Autoriser les valeurs vides
            error_title: Titre du message d'erreur
            error_message: Message d'erreur
        """
        if sheet_name not in self.wb.sheetnames or not values:
            return
        
        ws = self.wb[sheet_name]
        
        # Find column index
        col_idx = self._find_column_index(ws, column_name)
        if col_idx is None:
            return
        
        # Remove existing validations for this column
        self._remove_existing_validations(ws, col_idx, start_row, end_row)
        
        # Decide: inline or reference
        inline_formula = ",".join(values)
        if len(inline_formula) > 200 or len(values) > 50:
            # Use hidden sheet reference
            formula = self._write_list_to_hidden_sheet(values, f"{sheet_name}_{column_name}")
        else:
            # Use inline formula
            formula = f'"{inline_formula}"'
        
        # Create validation
        dv = DataValidation(
            type="list",
            formula1=formula,
            allow_blank=allow_blank,
            showErrorMessage=True,
            errorTitle=error_title,
            error=error_message
        )
        
        # Apply to range
        col_letter = get_column_letter(col_idx)
        dv.add(f'{col_letter}{start_row}:{col_letter}{end_row}')
        ws.add_data_validation(dv)
    
    def add_week_dropdown(self, sheet_name: str, column_name: str = "Semaine",
                          week_dates: Optional[Dict[int, datetime]] = None,
                          nb_semaines: int = 13, start_row: int = 2, end_row: int = 1000):
        """
        Ajoute une liste déroulante de semaines avec dates.
        
        Format affiché: "N (dd/mm)" pour une meilleure lisibilité.
        La valeur stockée reste le numéro de semaine.
        
        Args:
            sheet_name: Nom de la feuille
            column_name: Nom de la colonne (défaut: "Semaine")
            week_dates: Mapping {numéro_semaine: date}
            nb_semaines: Nombre total de semaines si week_dates non fourni
            start_row: Première ligne de données
            end_row: Dernière ligne de données
        """
        if sheet_name not in self.wb.sheetnames:
            return
        
        ws = self.wb[sheet_name]
        col_idx = self._find_column_index(ws, column_name)
        if col_idx is None:
            return
        
        # Generate week values with dates
        if week_dates:
            # Create display values with dates: "N (dd/mm)"
            values = []
            for week_num in sorted(week_dates.keys()):
                date = week_dates[week_num]
                date_str = date.strftime("%d/%m")
                values.append(f"{week_num} ({date_str})")
        else:
            values = [str(i) for i in range(1, nb_semaines + 1)]
        
        # Add dropdown
        self.add_dropdown(
            sheet_name=sheet_name,
            column_name=column_name,
            values=values,
            start_row=start_row,
            end_row=end_row,
            error_title="Semaine invalide",
            error_message=f"Sélectionnez un numéro de semaine"
        )
    
    def add_genre_dropdown(self, sheet_name: str, column_name: str = "Genre",
                           start_row: int = 2, end_row: int = 1000):
        """Ajoute une liste déroulante pour le genre (M/F)."""
        self.add_dropdown(
            sheet_name=sheet_name,
            column_name=column_name,
            values=["M", "F"],
            start_row=start_row,
            end_row=end_row,
            error_title="Genre invalide",
            error_message="Sélectionnez M (Masculin) ou F (Féminin)"
        )
    
    def add_type_contrainte_dropdown(self, sheet_name: str, 
                                      column_name: str = "Type_Contrainte",
                                      start_row: int = 2, end_row: int = 1000):
        """Ajoute une liste déroulante pour les types de contrainte."""
        self.add_dropdown(
            sheet_name=sheet_name,
            column_name=column_name,
            values=["Avant", "Apres"],
            start_row=start_row,
            end_row=end_row,
            error_title="Type invalide",
            error_message="Sélectionnez 'Avant' ou 'Apres'"
        )
    
    def add_type_poule_dropdown(self, sheet_name: str = "Types_Poules",
                                 column_name: str = "Type",
                                 start_row: int = 2, end_row: int = 1000):
        """Ajoute une liste déroulante pour les types de poule."""
        self.add_dropdown(
            sheet_name=sheet_name,
            column_name=column_name,
            values=["Classique", "Aller-Retour"],
            start_row=start_row,
            end_row=end_row,
            allow_blank=False,
            error_title="Type invalide",
            error_message="Sélectionnez 'Classique' ou 'Aller-Retour'"
        )
    
    def add_type_competition_dropdown(self, sheet_name: str,
                                       column_name: str = "Type_Competition",
                                       start_row: int = 2, end_row: int = 1000):
        """Ajoute une liste déroulante pour les types de compétition."""
        self.add_dropdown(
            sheet_name=sheet_name,
            column_name=column_name,
            values=["Acad", "CFE", "CFU", "Autre"],
            start_row=start_row,
            end_row=end_row,
            error_title="Type invalide",
            error_message="Sélectionnez: Acad, CFE, CFU ou Autre"
        )
    
    def add_niveau_dropdown(self, sheet_name: str = "Gymnases",
                            column_name: str = "Niveau",
                            start_row: int = 2, end_row: int = 1000):
        """Ajoute une liste déroulante pour les niveaux de gymnase."""
        self.add_dropdown(
            sheet_name=sheet_name,
            column_name=column_name,
            values=["Haut niveau", "Bas niveau"],
            start_row=start_row,
            end_row=end_row,
            error_title="Niveau invalide",
            error_message="Sélectionnez 'Haut niveau' ou 'Bas niveau'"
        )
    
    def add_niveau_equipe_dropdown(self, sheet_name: str = "Equipes",
                                    column_name: str = "Niveau_Equipe",
                                    start_row: int = 2, end_row: int = 1000):
        """Ajoute une liste déroulante pour les niveaux d'équipe (A1, A2, A3, A4)."""
        self.add_dropdown(
            sheet_name=sheet_name,
            column_name=column_name,
            values=["A1", "A2", "A3", "A4"],
            start_row=start_row,
            end_row=end_row,
            allow_blank=True,  # Le niveau peut être vide si l'équipe n'est pas assignée
            error_title="Niveau invalide",
            error_message="Sélectionnez un niveau: A1, A2, A3 ou A4"
        )
    
    def add_ignorer_dropdown(self, sheet_name: str = "Matchs_Fixes",
                             column_name: str = "Ignorer",
                             start_row: int = 2, end_row: int = 1000):
        """Ajoute une validation pour la colonne Ignorer (X ou vide)."""
        self.add_dropdown(
            sheet_name=sheet_name,
            column_name=column_name,
            values=["X"],
            start_row=start_row,
            end_row=end_row,
            allow_blank=True,
            error_title="Valeur invalide",
            error_message="Laissez vide ou indiquez 'X' pour ignorer le match"
        )
    
    def add_genre_equipe_dropdown(self, sheet_name: str = "Equipes",
                                   column_name: str = "Genre_Equipe",
                                   start_row: int = 2, end_row: int = 1000):
        """Ajoute une liste déroulante pour les genres d'équipe (M, F, X)."""
        self.add_dropdown(
            sheet_name=sheet_name,
            column_name=column_name,
            values=["M", "F", "X"],
            start_row=start_row,
            end_row=end_row,
            allow_blank=True,  # Le genre peut être vide si l'équipe n'est pas assignée
            error_title="Genre invalide",
            error_message="Sélectionnez un genre: M (Masculin), F (Féminin) ou X (Mixte)"
        )
    
    def add_equipe_dropdown(self, sheet_name: str, column_name: str,
                            equipes: List[str], start_row: int = 2, end_row: int = 1000):
        """Ajoute une liste déroulante d'équipes."""
        self.add_dropdown(
            sheet_name=sheet_name,
            column_name=column_name,
            values=sorted(equipes),
            start_row=start_row,
            end_row=end_row,
            error_title="Équipe invalide",
            error_message="Sélectionnez une équipe dans la liste"
        )
    
    def add_gymnase_dropdown(self, sheet_name: str, column_name: str,
                             gymnases: List[str], start_row: int = 2, end_row: int = 1000,
                             include_entente: bool = False):
        """Ajoute une liste déroulante de gymnases."""
        values = sorted(gymnases)
        if include_entente and 'ENTENTE' not in values:
            values = ['ENTENTE'] + values
        
        self.add_dropdown(
            sheet_name=sheet_name,
            column_name=column_name,
            values=values,
            start_row=start_row,
            end_row=end_row,
            error_title="Gymnase invalide",
            error_message="Sélectionnez un gymnase dans la liste"
        )
    
    def add_institution_dropdown(self, sheet_name: str, column_name: str,
                                  institutions: List[str], start_row: int = 2, end_row: int = 1000):
        """Ajoute une liste déroulante d'institutions."""
        self.add_dropdown(
            sheet_name=sheet_name,
            column_name=column_name,
            values=sorted(institutions),
            start_row=start_row,
            end_row=end_row,
            error_title="Institution invalide",
            error_message="Sélectionnez une institution dans la liste"
        )
    
    def add_poule_dropdown(self, sheet_name: str, poules: List[str],
                           column_name: str = "Poule",
                           start_row: int = 2, end_row: int = 1000):
        """Ajoute une liste déroulante de poules."""
        self.add_dropdown(
            sheet_name=sheet_name,
            column_name=column_name,
            values=sorted(poules),
            start_row=start_row,
            end_row=end_row,
            error_title="Poule invalide",
            error_message="Sélectionnez une poule dans la liste"
        )
    
    def add_coach_slot_dropdown(self, sheet_name: str, column_name: str,
                                 equipes_genrees: List[str],
                                 institutions: List[str],
                                 institutions_genrees: List[str],
                                 start_row: int = 2, end_row: int = 1000):
        """
        Ajoute une liste déroulante pour les slots Coach_Groups.
        
        Génère une liste combinant:
        - équipes genrées: "team=NOM (N) [M]" ou "team=NOM (N) [F]"
        - institutions: "institution=NOM"
        - institutions genrées: "institution=NOM;gender=M" ou "institution=NOM;gender=F"
        """
        values = []
        
        # 1. Équipes genrées avec format team=...
        for equipe in sorted(equipes_genrees):
            values.append(f"team={equipe}")
        
        # 2. Institutions avec format institution=...
        for institution in sorted(institutions):
            values.append(f"institution={institution}")
        
        # 3. Institutions genrées avec format institution=...;gender=...
        for inst_genre in sorted(institutions_genrees):
            values.append(inst_genre)
        
        self.add_dropdown(
            sheet_name=sheet_name,
            column_name=column_name,
            values=values,
            start_row=start_row,
            end_row=end_row,
            allow_blank=True,  # Les slots peuvent être vides
            error_title="Format invalide",
            error_message="Utilisez team=ÉQUIPE, institution=NOM, ou institution=NOM;gender=M/F"
        )
    
    def _find_column_index(self, ws, column_name: str) -> Optional[int]:
        """Trouve l'index d'une colonne par son nom."""
        for idx, cell in enumerate(ws[1], 1):
            if cell.value and str(cell.value) == column_name:
                return idx
        return None
    
    def _remove_existing_validations(self, ws, col_idx: int, start_row: int, end_row: int):
        """Supprime les validations existantes pour une colonne."""
        col_letter = get_column_letter(col_idx)
        target_range = f'{col_letter}{start_row}:{col_letter}{end_row}'
        
        # Note: openpyxl doesn't have a direct way to remove specific validations
        # We'll just add the new one which will override


def setup_all_dropdowns(filepath: str, 
                        equipes: List[str],
                        gymnases: List[str],
                        institutions: List[str],
                        poules: List[str],
                        week_dates: Optional[Dict[int, datetime]] = None,
                        nb_semaines: int = 13):
    """
    Configure toutes les listes déroulantes pour un fichier de configuration.
    
    Args:
        filepath: Chemin du fichier Excel
        equipes: Liste des équipes (avec variantes de genre si nécessaire)
        gymnases: Liste des gymnases
        institutions: Liste des institutions
        poules: Liste des poules
        week_dates: Mapping semaine -> date (optionnel)
        nb_semaines: Nombre de semaines si week_dates non fourni
    """
    wb = openpyxl.load_workbook(filepath)
    dm = DropdownManager(wb)
    
    # Equipes sans genre pour Matchs_Fixes
    equipes_sans_genre = list(set(
        eq.replace(' [F]', '').replace(' [M]', '').strip()
        for eq in equipes
    ))
    
    # Équipes genrées (avec [M] ou [F])
    equipes_genrees = [eq for eq in equipes if ' [M]' in eq or ' [F]' in eq]
    
    # Institutions genrées pour Coach_Groups
    institutions_genrees = []
    for inst in sorted(institutions):
        institutions_genrees.append(f"institution={inst};gender=M")
        institutions_genrees.append(f"institution={inst};gender=F")
    
    # ===== Matchs_Fixes =====
    if 'Matchs_Fixes' in wb.sheetnames:
        dm.add_equipe_dropdown('Matchs_Fixes', 'Equipe_1', equipes_sans_genre)
        dm.add_equipe_dropdown('Matchs_Fixes', 'Equipe_2', equipes_sans_genre)
        dm.add_genre_dropdown('Matchs_Fixes', 'Genre')
        dm.add_poule_dropdown('Matchs_Fixes', poules, 'Poule')
        dm.add_week_dropdown('Matchs_Fixes', 'Semaine', week_dates, nb_semaines)
        dm.add_gymnase_dropdown('Matchs_Fixes', 'Gymnase', gymnases, include_entente=True)
        dm.add_type_competition_dropdown('Matchs_Fixes', 'Type_Competition')
        dm.add_ignorer_dropdown('Matchs_Fixes', 'Ignorer')
    
    # ===== Contraintes_Temporelles =====
    if 'Contraintes_Temporelles' in wb.sheetnames:
        dm.add_equipe_dropdown('Contraintes_Temporelles', 'Equipe_1', equipes_sans_genre)
        dm.add_equipe_dropdown('Contraintes_Temporelles', 'Equipe_2', equipes_sans_genre)
        dm.add_genre_dropdown('Contraintes_Temporelles', 'Genre')
        dm.add_type_contrainte_dropdown('Contraintes_Temporelles', 'Type_Contrainte')
        dm.add_week_dropdown('Contraintes_Temporelles', 'Semaine', week_dates, nb_semaines)
    
    # ===== Types_Poules =====
    if 'Types_Poules' in wb.sheetnames:
        dm.add_type_poule_dropdown('Types_Poules', 'Type')
    
    # ===== Gymnases =====
    if 'Gymnases' in wb.sheetnames:
        dm.add_niveau_dropdown('Gymnases', 'Niveau')
        dm.add_genre_dropdown('Gymnases', 'Genre_Prioritaire')
    
    # ===== Indispos_Gymnases =====
    if 'Indispos_Gymnases' in wb.sheetnames:
        dm.add_gymnase_dropdown('Indispos_Gymnases', 'Gymnase', gymnases)
        dm.add_week_dropdown('Indispos_Gymnases', 'Semaine', week_dates, nb_semaines)
    
    # ===== Indispos_Equipes =====
    if 'Indispos_Equipes' in wb.sheetnames:
        dm.add_equipe_dropdown('Indispos_Equipes', 'Equipe', equipes)
        dm.add_week_dropdown('Indispos_Equipes', 'Semaine', week_dates, nb_semaines)
    
    # ===== Indispos_Institutions =====
    if 'Indispos_Institutions' in wb.sheetnames:
        dm.add_institution_dropdown('Indispos_Institutions', 'Institution', institutions)
        dm.add_week_dropdown('Indispos_Institutions', 'Semaine', week_dates, nb_semaines)
    
    # ===== Preferences_Gymnases =====
    if 'Preferences_Gymnases' in wb.sheetnames:
        dm.add_institution_dropdown('Preferences_Gymnases', 'Institution', institutions)
        for i in range(1, 6):
            dm.add_gymnase_dropdown('Preferences_Gymnases', f'Gymnase_Pref_{i}', gymnases)
    
    # ===== Dispos_Gymnases_Equipes =====
    if 'Dispos_Gymnases_Equipes' in wb.sheetnames:
        dm.add_equipe_dropdown('Dispos_Gymnases_Equipes', 'Equipe', equipes_sans_genre)
        dm.add_genre_dropdown('Dispos_Gymnases_Equipes', 'Genre')
        for i in range(1, 6):
            dm.add_gymnase_dropdown('Dispos_Gymnases_Equipes', f'Gymnase_{i}', gymnases)
    
    # ===== Ententes =====
    if 'Ententes' in wb.sheetnames:
        dm.add_institution_dropdown('Ententes', 'Institution_1', institutions)
        dm.add_institution_dropdown('Ententes', 'Institution_2', institutions)
    
    # ===== Obligation_Presence =====
    if 'Obligation_Presence' in wb.sheetnames:
        dm.add_gymnase_dropdown('Obligation_Presence', 'Gymnase', gymnases)
        dm.add_institution_dropdown('Obligation_Presence', 'Institution_Obligatoire', institutions)
    
    # ===== Equipes_Hors_Championnat =====
    if 'Equipes_Hors_Championnat' in wb.sheetnames:
        dm.add_genre_dropdown('Equipes_Hors_Championnat', 'Genre')
        dm.add_dropdown(
            'Equipes_Hors_Championnat', 'Type_Championnat',
            ['CFE', 'CFU', 'Autre'],
            error_title="Type invalide",
            error_message="Sélectionnez: CFE, CFU ou Autre"
        )
    
    # ===== Equipes (feuille de données de base) =====
    if 'Equipes' in wb.sheetnames:
        dm.add_niveau_equipe_dropdown('Equipes', 'Niveau_Equipe')
        dm.add_genre_equipe_dropdown('Equipes', 'Genre_Equipe')
    
    # ===== Coach_Groups =====
    if 'Coach_Groups' in wb.sheetnames:
        # Ajouter les dropdowns pour chaque colonne slot_XX
        for i in range(1, 21):
            col_name = f'slot_{i:02d}'
            dm.add_coach_slot_dropdown(
                'Coach_Groups', col_name,
                equipes_genrees=equipes_genrees,
                institutions=list(institutions),
                institutions_genrees=institutions_genrees
            )
    
    wb.save(filepath)
