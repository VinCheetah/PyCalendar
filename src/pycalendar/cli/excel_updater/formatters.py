"""
Excel formatting and styling utilities.

Provides visual formatting for Excel configuration files:
- Headers styling
- Alternating row colors
- Column width auto-adjustment
- Freeze panes
- Conditional formatting
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from typing import Dict, List, Optional, Tuple


# Color palette - Thème cohérent et élégant
class Colors:
    """Palette de couleurs pour le formatage Excel - Thème unifié bleu professionnel."""
    # Headers - Couleur unique pour cohérence
    HEADER_BG = "2C3E50"  # Bleu-gris foncé élégant (unique pour toutes les feuilles)
    HEADER_FG = "FFFFFF"  # Blanc
    
    # Alternating rows - Subtil et élégant
    ROW_EVEN = "FFFFFF"   # Blanc
    ROW_ODD = "F8F9FA"    # Gris très clair
    
    # Status colors
    ERROR_BG = "F8D7DA"   # Rouge clair
    WARNING_BG = "FFF3CD"  # Jaune/ambre clair
    SUCCESS_BG = "D4EDDA"  # Vert clair
    INFO_BG = "D1ECF1"    # Bleu info clair
    
    # Extra columns
    EXTRA_BG = "FFC107"   # Ambre/Or
    EXTRA_FG = "212529"   # Gris foncé
    
    # Coach groups alternating
    COACH_ROW_1 = "FFFFFF"
    COACH_ROW_2 = "F0F4F8"


# Largeurs de colonnes par type
COLUMN_WIDTHS = {
    # Colonnes courtes
    'Genre': 8,
    'Semaine': 12,
    'Capacite': 12,
    'Capacite_Occupee': 16,
    'Score': 10,
    'Type': 14,
    'Niveau': 14,
    'Genre_Prioritaire': 18,
    
    # Colonnes moyennes
    'Equipe': 20,
    'Equipe_1': 20,
    'Equipe_2': 20,
    'Institution': 18,
    'Institution_1': 18,
    'Institution_2': 18,
    'Gymnase': 18,
    'Poule': 14,
    'Date': 12,
    'Horaire': 10,
    'Horaire_Debut': 14,
    'Horaire_Fin': 12,
    'Horaire_Dispo': 14,
    'Horaire_Prefere': 16,
    'Type_Contrainte': 18,
    'Type_Competition': 18,
    'Type_Championnat': 18,
    
    # Colonnes de préférences gymnases
    'Gymnase_Pref_1': 16,
    'Gymnase_Pref_2': 16,
    'Gymnase_Pref_3': 16,
    'Gymnase_Pref_4': 16,
    'Gymnase_Pref_5': 16,
    'Gymnase_1': 16,
    'Gymnase_2': 16,
    'Gymnase_3': 16,
    'Gymnase_4': 16,
    'Gymnase_5': 16,
    
    # Colonnes longues
    'Adresse': 35,
    'Remarques': 30,
    'Remarque': 30,
    'Motif': 25,
    'Creneaux': 25,
    'Arbitres': 25,
    
    # Coach groups
    'Coach': 16,
    'ID': 8,
    
    # Contacts
    'Responsable_Nom': 20,
    'Responsable_Email': 28,
    'Responsable_Telephone': 18,
}

# Largeur par défaut pour colonnes non définies
DEFAULT_COLUMN_WIDTH = 15
MIN_COLUMN_WIDTH = 10
MAX_COLUMN_WIDTH = 45


class ExcelFormatter:
    """Gestionnaire de formatage Excel."""
    
    def __init__(self, workbook: openpyxl.Workbook):
        self.wb = workbook
        
        # Styles prédéfinis - Police plus grande et lisible
        self.header_font = Font(bold=True, color=Colors.HEADER_FG, size=11, name='Calibri')
        self.header_fill = PatternFill(
            start_color=Colors.HEADER_BG, 
            end_color=Colors.HEADER_BG, 
            fill_type="solid"
        )
        self.header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Alignement centré pour les données
        self.cell_font = Font(size=11, name='Calibri')
        self.cell_alignment = Alignment(horizontal='center', vertical='center')
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.left_alignment = Alignment(horizontal='left', vertical='center')
        
        self.thin_border = Border(
            left=Side(style='thin', color='BDC3C7'),
            right=Side(style='thin', color='BDC3C7'),
            top=Side(style='thin', color='BDC3C7'),
            bottom=Side(style='thin', color='BDC3C7')
        )
        
        self.row_fills = [
            PatternFill(start_color=Colors.ROW_EVEN, end_color=Colors.ROW_EVEN, fill_type="solid"),
            PatternFill(start_color=Colors.ROW_ODD, end_color=Colors.ROW_ODD, fill_type="solid"),
        ]
        
        self.extra_fill = PatternFill(
            start_color=Colors.EXTRA_BG, 
            end_color=Colors.EXTRA_BG, 
            fill_type="solid"
        )
        self.extra_font = Font(italic=True, color=Colors.EXTRA_FG, size=11, name='Calibri')
        
        # Colonnes à aligner à gauche (texte long)
        self.left_align_columns = {'Adresse', 'Remarques', 'Remarque', 'Motif', 'Notes', 'Creneaux', 'Arbitres'}
    
    def _get_header_color(self, sheet_name: str) -> str:
        """Retourne la couleur du header - couleur unique pour cohérence."""
        # Couleur unique pour toutes les feuilles = cohérence visuelle
        return Colors.HEADER_BG
    
    def format_sheet(self, sheet_name: str, 
                     freeze_header: bool = True,
                     auto_width: bool = True,
                     alternating_rows: bool = True,
                     extra_columns: Optional[List[str]] = None):
        """
        Applique le formatage complet à une feuille.
        
        Args:
            sheet_name: Nom de la feuille
            freeze_header: Geler la ligne d'en-tête
            auto_width: Ajuster automatiquement la largeur des colonnes
            alternating_rows: Appliquer des couleurs alternées aux lignes
            extra_columns: Liste des colonnes "extra" à mettre en évidence
        """
        if sheet_name not in self.wb.sheetnames:
            return
        
        ws = self.wb[sheet_name]
        extra_columns = extra_columns or []
        
        # Déterminer la couleur du header selon le type de feuille
        header_color = self._get_header_color(sheet_name)
        
        # Format headers avec couleur spécifique
        self._format_headers(ws, extra_columns, header_color)
        
        # Format data rows
        if alternating_rows:
            self._format_alternating_rows(ws, extra_columns)
        
        # Auto-adjust column widths
        if auto_width:
            self._auto_column_widths(ws)
        
        # Freeze header row
        if freeze_header:
            ws.freeze_panes = 'A2'
    
    def _format_headers(self, ws, extra_columns: List[str], header_color: Optional[str] = None):
        """Formate la ligne d'en-tête avec couleur personnalisable."""
        if header_color is None:
            header_color = Colors.HEADER_BG
        
        header_fill = PatternFill(
            start_color=header_color,
            end_color=header_color,
            fill_type="solid"
        )
        
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value:
                col_name = str(cell.value)
                
                if col_name.startswith('EXTRA_') or col_name in extra_columns:
                    cell.fill = self.extra_fill
                    cell.font = Font(bold=True, italic=True, color=Colors.EXTRA_FG)
                else:
                    cell.fill = header_fill
                    cell.font = self.header_font
                
                cell.alignment = self.header_alignment
                cell.border = self.thin_border
    
    def _format_alternating_rows(self, ws, extra_columns: List[str]):
        """Applique des couleurs alternées aux lignes de données avec alignement centré."""
        extra_col_indices = set()
        col_names = {}
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value:
                col_name = str(cell.value)
                col_names[col_idx] = col_name
                if col_name.startswith('EXTRA_') or col_name in extra_columns:
                    extra_col_indices.add(col_idx)
        
        for row_idx in range(2, ws.max_row + 1):
            fill = self.row_fills[(row_idx - 2) % 2]
            
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row_idx, col_idx)
                
                if col_idx in extra_col_indices:
                    # Extra columns get special treatment
                    cell.font = Font(italic=True, color='666666')
                else:
                    cell.fill = fill
                
                cell.border = self.thin_border
                
                # Alignement: gauche pour les longues colonnes, centré pour le reste
                col_name = col_names.get(col_idx, '')
                if col_name in self.left_align_columns:
                    cell.alignment = self.left_alignment
                else:
                    cell.alignment = self.cell_alignment  # Centré
    
    def _auto_column_widths(self, ws):
        """Ajuste la largeur des colonnes - largeurs prédéfinies et uniformes."""
        # Récupérer les noms de colonnes
        col_names = {}
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value:
                col_names[col_idx] = str(cell.value)
        
        for col_idx in range(1, ws.max_column + 1):
            column_letter = get_column_letter(col_idx)
            col_name = col_names.get(col_idx, '')
            
            # Utiliser la largeur prédéfinie si disponible
            if col_name in COLUMN_WIDTHS:
                width = COLUMN_WIDTHS[col_name]
            else:
                # Calculer une largeur basée sur le contenu mais avec min/max
                max_length = len(col_name) if col_name else 0
                for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 50), 
                                         min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            max_length = max(max_length, cell_length)
                
                # Ajouter un padding et appliquer les limites
                width = min(max(max_length + 4, MIN_COLUMN_WIDTH), MAX_COLUMN_WIDTH)
            
            ws.column_dimensions[column_letter].width = width
        
        # Hauteur de ligne uniforme pour l'élégance
        ws.row_dimensions[1].height = 22  # Header row légèrement plus haute
    
    def format_coach_groups(self, sheet_name: str = 'Coach_Groups'):
        """Formatage spécial pour la feuille Coach_Groups."""
        if sheet_name not in self.wb.sheetnames:
            return
        
        ws = self.wb[sheet_name]
        
        # Format headers
        self._format_headers(ws, [])
        
        # Special alternating for coach groups (by group)
        coach_fills = [
            PatternFill(start_color=Colors.COACH_ROW_1, end_color=Colors.COACH_ROW_1, fill_type="solid"),
            PatternFill(start_color=Colors.COACH_ROW_2, end_color=Colors.COACH_ROW_2, fill_type="solid"),
        ]
        
        for row_idx in range(2, ws.max_row + 1):
            fill = coach_fills[(row_idx - 2) % 2]
            
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row_idx, col_idx)
                cell.fill = fill
                cell.border = self.thin_border
        
        self._auto_column_widths(ws)
        ws.freeze_panes = 'C2'  # Freeze ID and coach name columns
    
    def add_error_highlighting(self, sheet_name: str, error_cells: List[Tuple[int, int]]):
        """
        Ajoute une mise en évidence des cellules en erreur.
        
        Args:
            sheet_name: Nom de la feuille
            error_cells: Liste de tuples (row, col) des cellules en erreur
        """
        if sheet_name not in self.wb.sheetnames:
            return
        
        ws = self.wb[sheet_name]
        error_fill = PatternFill(
            start_color=Colors.ERROR_BG, 
            end_color=Colors.ERROR_BG, 
            fill_type="solid"
        )
        
        for row, col in error_cells:
            cell = ws.cell(row, col)
            cell.fill = error_fill
    
    def add_warning_highlighting(self, sheet_name: str, warning_cells: List[Tuple[int, int]]):
        """Ajoute une mise en évidence des cellules avec avertissement."""
        if sheet_name not in self.wb.sheetnames:
            return
        
        ws = self.wb[sheet_name]
        warning_fill = PatternFill(
            start_color=Colors.WARNING_BG, 
            end_color=Colors.WARNING_BG, 
            fill_type="solid"
        )
        
        for row, col in warning_cells:
            cell = ws.cell(row, col)
            cell.fill = warning_fill


def format_workbook(filepath: str, sheet_configs: Optional[Dict[str, dict]] = None):
    """
    Formate un classeur Excel complet.
    
    Args:
        filepath: Chemin du fichier Excel
        sheet_configs: Configuration par feuille (optionnel)
    """
    wb = openpyxl.load_workbook(filepath)
    formatter = ExcelFormatter(wb)
    
    sheet_configs = sheet_configs or {}
    
    for sheet_name in wb.sheetnames:
        if sheet_name.startswith('_'):
            continue  # Skip hidden/system sheets
        
        config = sheet_configs.get(sheet_name, {})
        
        if sheet_name == 'Coach_Groups':
            formatter.format_coach_groups(sheet_name)
        else:
            formatter.format_sheet(
                sheet_name,
                freeze_header=config.get('freeze_header', True),
                auto_width=config.get('auto_width', True),
                alternating_rows=config.get('alternating_rows', True),
                extra_columns=config.get('extra_columns', [])
            )
    
    wb.save(filepath)
