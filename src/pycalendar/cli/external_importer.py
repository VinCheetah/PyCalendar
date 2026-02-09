#!/usr/bin/env python3
"""
Script d'importation de matchs depuis un fichier Excel partagé en ligne.

Ce script permet d'importer des matchs déjà joués ou planifiés depuis un
tableau Excel partagé (Google Sheets, OneDrive, SharePoint, etc.) vers la feuille
Matchs_Fixes d'une configuration PyCalendar.

Sports supportés:
    VB - Volleyball 🏐
    HB - Handball 🤾
    BB - Basketball 🏀
    FB - Football ⚽
    FS - Futsal ⚽
    RG - Rugby 🏉
    BD - Badminton 🏸
    TT - Tennis de Table 🏓

Utilisation:
    python src/pycalendar/cli/external_importer.py --config CONFIG_YAML --url URL [OPTIONS]

Exemples:
    # Importer des matchs de volleyball
    python src/pycalendar/cli/external_importer.py \\
        --config configs/config_volley.yaml \\
        --url "https://docs.google.com/spreadsheets/d/.../export?format=xlsx" \\
        --sport VB \\
        --journee 3 \\
        --avec-score

    # Importer des matchs de handball
    python src/pycalendar/cli/external_importer.py \\
        --config configs/config_hand.yaml \\
        --url "https://..." \\
        --sport HB

    # Importer depuis un fichier local
    python src/pycalendar/cli/external_importer.py \\
        --config configs/config_volley.yaml \\
        --fichier-local "/path/to/file.xlsx" \\
        --sport VB \\
        --tous

Note: Le paramètre --config doit pointer vers un fichier YAML de configuration,
      pas vers le fichier Excel directement.
"""

import argparse
import sys
import os
import pandas as pd
import requests
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Optional, Tuple
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import tempfile
import urllib.parse

from pycalendar.core.constants import (
    DATE_USER_FORMAT_LABEL,
    format_user_date,
    parse_user_date,
)

# Imports optionnels pour l'authentification SharePoint/Microsoft
try:
    import msal
    MSAL_AVAILABLE = True
except ImportError:
    MSAL_AVAILABLE = False

try:
    from requests_oauthlib import OAuth2Session
    OAUTH_AVAILABLE = True
except ImportError:
    OAUTH_AVAILABLE = False

import json
import re
import base64


class ImporteurMatchsExternes:
    """Classe pour importer des matchs depuis une source externe."""
    
    def __init__(
        self,
        config_path: str,
        url_externe: Optional[str] = None,
        fichier_local: Optional[str] = None,
        sport: str = "VB",
        journee: Optional[int] = None,
        date_limite: Optional[str] = None,
        avec_score: bool = False,
        sans_score: bool = False,
        tous: bool = False,
        dry_run: bool = False,
        ignorer_annules: bool = True,
        doublon_priorite: str = "ancien"
    ):
        """
        Initialise l'importeur.
        
        Args:
            config_path: Chemin vers le fichier de configuration Excel
            url_externe: URL du fichier Excel partagé en ligne (optionnel)
            fichier_local: Chemin vers un fichier Excel local (optionnel)
            sport: Code du sport (VB, HB, etc.)
            journee: Numéro de journée à importer (optionnel)
            date_limite: Date limite pour l'import (format DD/MM/YYYY)
            avec_score: Importer uniquement les matchs avec score
            sans_score: Importer uniquement les matchs sans score
            tous: Importer tous les matchs (défaut si rien spécifié)
            dry_run: Mode simulation (ne modifie pas le fichier)
            ignorer_annules: Ignorer les matchs avec 'annule' dans les remarques
            doublon_priorite: "ancien" pour garder les matchs existants, "nouveau" pour remplacer par l'import
        """
        self.config_yaml_path = Path(config_path)
        self._config_yaml_cache: Optional[dict] = None
        self.url_externe = url_externe
        self.fichier_local = Path(fichier_local) if fichier_local else None
        self.sport = sport.upper()
        self.journee = journee
        self.date_limite = self._parse_date(date_limite) if date_limite else None
        self.avec_score = avec_score
        self.sans_score = sans_score
        self.tous = tous or (not avec_score and not sans_score)
        self.dry_run = dry_run
        self.ignorer_annules = ignorer_annules
        self.doublon_priorite = (doublon_priorite or "ancien").strip().lower()
        if self.doublon_priorite not in {"ancien", "nouveau"}:
            raise ValueError("doublon_priorite doit valoir 'ancien' ou 'nouveau'")
        
        # Paramètres SharePoint (à configurer via variables d'environnement)
        self.sharepoint_client_id = os.getenv("SHAREPOINT_CLIENT_ID", "")
        self.sharepoint_tenant_id = os.getenv("SHAREPOINT_TENANT_ID", "")
        self.sharepoint_client_secret = os.getenv("SHAREPOINT_CLIENT_SECRET", "")
        
        # Charger le fichier Excel depuis le YAML
        self.config_excel_path = self._charger_chemin_excel()
        self.calendrier_date_debut, self.calendrier_jour_match = self._charger_parametres_calendrier()
        
        # Validation
        if not url_externe and not fichier_local:
            raise ValueError("Vous devez spécifier soit --url soit --fichier-local")
        
        # DataFrame du fichier externe
        self.df_externe: Optional[pd.DataFrame] = None
        
        # DataFrame de la configuration
        self.df_matchs_fixes: Optional[pd.DataFrame] = None
        
        # Mapping des gymnases et équipes mixtes (chargé à la demande)
        self._mapping_gymnases: Optional[dict] = None
        self._equipes_mixtes: Optional[set] = None
        
        # Cache pour l'authentification SharePoint
        self._sharepoint_token: Optional[str] = None

        # Tag utilisé pour annoter les remarques lors de l'import
        self.import_comment_tag = self._build_import_comment_tag()
    
    def _parse_sharepoint_url(self, url: str) -> Dict[str, Optional[str]]:
        """
        Parse une URL SharePoint pour extraire les informations nécessaires.
        
        Args:
            url: URL SharePoint
            
        Returns:
            Dictionnaire avec les informations extraites
        """
        # URL format: https://tenant.sharepoint.com/sites/site/_layouts/15/Doc.aspx?sourcedoc={item_id}&file=filename.xlsx&action=default&mobileredirect=true
        # or: https://tenant.sharepoint.com/:x:/g/personal/user_domain_com/E...
        
        parsed = urllib.parse.urlparse(url)
        
        if 'sharepoint.com' not in parsed.netloc:
            raise ValueError("URL n'est pas une URL SharePoint valide")
        
        # Extraire le tenant
        tenant = parsed.netloc.split('.')[0]
        
        # Pour les URLs de type /:x:/g/personal/...
        if '/:x:/' in parsed.path:
            # Format: /:x:/g/personal/user_domain_com/encoded_path?e=sharing_param
            path_parts = parsed.path.split('/')
            if len(path_parts) >= 4:
                resource_type = path_parts[2]  # 'g' for group, 'personal' for personal
                # Reconstruct the full container path (everything after resource_type)
                container_path = '/'.join(path_parts[3:])  # 'personal/klucediniz_sport-u_com/EXeOs9uh4dxGkxsKLI0B_tABH0LSZdk2qWOukF8kVmNH_g'
                
                # L'ID de l'item est encodé dans l'URL
                query_params = urllib.parse.parse_qs(parsed.query)
                if 'e' in query_params:
                    sharing_token = query_params['e'][0]
                    
                    return {
                        'tenant': tenant,
                        'resource_type': resource_type,
                        'container': container_path,  # Now includes the full path
                        'sharing_token': sharing_token,
                        'url_type': 'sharing'
                    }
        
        # Pour les URLs classiques avec sourcedoc
        query_params = urllib.parse.parse_qs(parsed.query)
        if 'sourcedoc' in query_params:
            item_id = query_params['sourcedoc'][0]
            
            # Extraire le site depuis l'URL
            path_parts = parsed.path.split('/')
            site_name = None
            for i, part in enumerate(path_parts):
                if part == 'sites' and i + 1 < len(path_parts):
                    site_name = path_parts[i + 1]
                    break
            
            return {
                'tenant': tenant,
                'site_name': site_name,
                'item_id': item_id,
                'url_type': 'classic'
            }
        
        raise ValueError("Format d'URL SharePoint non reconnu")
    
    def _authenticate_sharepoint(self) -> str:
        """
        Authentifie auprès de Microsoft et retourne un token d'accès.
        
        Returns:
            Token d'accès pour Microsoft Graph API
        """
        if not all([self.sharepoint_client_id, self.sharepoint_tenant_id, self.sharepoint_client_secret]):
            raise ValueError(
                "Authentification SharePoint requise. Fournissez:\n"
                "  • Client ID Azure AD\n"
                "  • Tenant ID Azure AD\n"
                "  • Client Secret Azure AD"
            )
        
        return self._authenticate_client_credentials()
    
    def _authenticate_client_credentials(self) -> str:
        """
        Authentification utilisant client credentials (client secret).
        
        Returns:
            Token d'accès
        """
        authority = f"https://login.microsoftonline.com/{self.sharepoint_tenant_id}"
        
        app = msal.ConfidentialClientApplication(
            self.sharepoint_client_id,
            client_credential=self.sharepoint_client_secret,
            authority=authority
        )
        
        result = app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
        
        if "access_token" in result:
            return result["access_token"]
        else:
            error = result.get("error_description", result.get("error", "Erreur inconnue"))
            raise ValueError(f"Échec de l'authentification client credentials: {error}")
    
    def _download_sharepoint_file(self, url: str) -> bytes:
        """
        Télécharge un fichier depuis SharePoint en utilisant Microsoft Graph API.
        
        Args:
            url: URL SharePoint du fichier
            
        Returns:
            Contenu du fichier en bytes
        """
        # Parser l'URL
        sharepoint_info = self._parse_sharepoint_url(url)
        
        # Authentifier
        if not self._sharepoint_token:
            self._sharepoint_token = self._authenticate_sharepoint()
        
        headers = {
            'Authorization': f'Bearer {self._sharepoint_token}',
            'Accept': 'application/octet-stream'
        }
        
        if sharepoint_info['url_type'] == 'sharing':
            # Pour les URLs de partage, utiliser l'API de partage
            sharing_url = f"https://{sharepoint_info['tenant']}.sharepoint.com/:x:/{sharepoint_info['resource_type']}/{sharepoint_info['container']}?e={sharepoint_info['sharing_token']}"
            
            # Encoder l'URL pour l'API Graph
            encoded_url = base64.b64encode(sharing_url.encode()).decode()
            
            graph_url = f"https://graph.microsoft.com/v1.0/shares/u!{encoded_url}/driveItem/content"
            
        else:
            # Pour les URLs classiques, construire l'URL Graph
            graph_url = (
                f"https://graph.microsoft.com/v1.0/sites/{sharepoint_info['tenant']}.sharepoint.com:/sites/{sharepoint_info['site_name']}:/"
                f"drive/items/{sharepoint_info['item_id']}/content"
            )
        
        response = requests.get(graph_url, headers=headers)
        response.raise_for_status()
        
        return response.content
    
    def _charger_chemin_excel(self) -> Path:
        """Charge le chemin du fichier Excel depuis le YAML."""
        import yaml
        
        # Essayer différents encodages
        encodings_to_try = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
        config = None
        
        for encoding in encodings_to_try:
            try:
                with open(self.config_yaml_path, 'r', encoding=encoding) as f:
                    config = yaml.safe_load(f)
                break  # Si ça marche, on sort de la boucle
            except UnicodeDecodeError:
                continue  # Essayer le prochain encodage
        
        if config is None:
            raise ValueError(f"Impossible de lire le fichier YAML {self.config_yaml_path} avec les encodages testés: {encodings_to_try}")
        self._config_yaml_cache = config if isinstance(config, dict) else None
        
        # Essayer plusieurs chemins possibles
        fichier_excel = None
        if 'fichiers' in config and 'donnees' in config['fichiers']:
            fichier_excel = config['fichiers']['donnees']
        elif 'fichier_excel' in config:
            fichier_excel = config['fichier_excel']
        
        if not fichier_excel:
            raise ValueError(f"Aucun fichier Excel trouvé dans {self.config_yaml_path} (cherché: fichiers.donnees ou fichier_excel)")
        
        # Résoudre le chemin relatif par rapport à la racine du projet (où se trouve le YAML, généralement on remonte d'un niveau)
        chemin = Path(fichier_excel)
        if not chemin.is_absolute():
            # Le YAML est dans configs/, les données sont à la racine
            # Donc on prend le parent du YAML (configs/) puis on remonte encore (racine)
            racine_projet = self.config_yaml_path.parent.parent
            chemin = racine_projet / fichier_excel
        
        return chemin
    
    def _charger_parametres_calendrier(self) -> Tuple[Optional[datetime], str]:
        """Extrait date de début et jour officiel des matchs depuis le YAML."""

        config = self._config_yaml_cache or {}
        calendrier = config.get('calendrier') if isinstance(config, dict) else None
        date_debut = None
        jour_match = 'jeudi'

        if isinstance(calendrier, dict):
            jour_match = str(calendrier.get('jour_match', jour_match)).strip().lower() or 'jeudi'
            brute = calendrier.get('date_debut')
            if brute:
                brute_str = str(brute).strip()
                for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d/%m/%y'):
                    try:
                        date_debut = datetime.strptime(brute_str, fmt)
                        break
                    except ValueError:
                        continue

        return date_debut, jour_match

    def _jour_match_weekday_index(self) -> Optional[int]:
        """Retourne l'index weekday Python (0=lundi) pour le jour officiel."""

        mapping = {
            'lundi': 0,
            'mardi': 1,
            'mercredi': 2,
            'jeudi': 3,
            'vendredi': 4,
            'samedi': 5,
            'dimanche': 6,
            'monday': 0,
            'tuesday': 1,
            'wednesday': 2,
            'thursday': 3,
            'friday': 4,
            'saturday': 5,
            'sunday': 6,
        }

        jour = (self.calendrier_jour_match or 'jeudi').strip().lower()
        return mapping.get(jour)

    def _calculer_semaine_depuis_timestamp(self, valeur: Optional[pd.Timestamp]) -> Optional[int]:
        """Calcule la semaine relative à la date de début configurée."""

        if valeur is None or pd.isna(valeur) or not self.calendrier_date_debut:
            return None
        delta = valeur.to_pydatetime().date() - self.calendrier_date_debut.date()
        if delta.days < 0:
            return None
        return delta.days // 7 + 1
        
    def _parser_score(self, score_raw) -> str:
        """
        Parse et normalise un score en format 'score_E1 - score_E2'.
        
        Args:
            score_raw: Score brut depuis le fichier externe
            
        Returns:
            Score normalisé ou score original si parsing impossible
        """
        if pd.isna(score_raw):
            return ''
        
        score_str = str(score_raw).strip()
        if not score_str:
            return ''
        
        # Essayer différents formats courants
        import re
        
        # Format: "3-1" ou "3 - 1"
        match = re.match(r'^(\d+)\s*-\s*(\d+)$', score_str)
        if match:
            return f"{match.group(1)}-{match.group(2)}"
        
        # Format: "3/1" ou "3 / 1"
        match = re.match(r'^(\d+)\s*/\s*(\d+)$', score_str)
        if match:
            return f"{match.group(1)}-{match.group(2)}"
        
        # Format: "3-1 (25-20, 25-18, 25-15)" - garder tel quel
        if re.match(r'^\d+-\d+\s*\(.+\)$', score_str):
            return score_str
        
        # Format: "3-1, 3-0" (match en 5 sets) - garder tel quel
        if re.match(r'^\d+-\d+,\s*\d+-\d+$', score_str):
            return score_str
        
        # Si on ne peut pas parser, garder tel quel
        return score_str
    
    def _parse_date(self, date_str: str) -> datetime:
        """Parse une date utilisateur (DD/MM/YY)."""
        parsed = parse_user_date(date_str)
        if not parsed:
            raise ValueError(f"Format de date invalide: {date_str}. Attendu: {DATE_USER_FORMAT_LABEL}")
        return parsed

    def _build_import_comment_tag(self) -> str:
        """Construit une étiquette cohérente pour annoter les remarques importées."""
        parts = []
        if self.journee is not None:
            parts.append(f"J{self.journee}")
        if self.fichier_local:
            parts.append(self.fichier_local.name)
        elif self.url_externe:
            try:
                host = urllib.parse.urlparse(self.url_externe).netloc
            except Exception:
                host = ''
            if host:
                parts.append(host)
        if not parts:
            return "Import externe"
        return "Importé " + " | ".join(parts)

    def _merge_import_comment_tag(self, remark: Optional[str]) -> str:
        """Ajoute l'étiquette d'import aux remarques sans écraser le contenu existant."""
        remark_str = '' if remark is None else str(remark).strip()
        if remark_str.lower() == 'nan':
            remark_str = ''
        tag = self.import_comment_tag
        if not tag:
            return remark_str
        if not remark_str:
            return tag
        if tag in remark_str:
            return remark_str
        return f"{remark_str} | {tag}"
    
    def mapper_gymnases(self) -> dict:
        """
        Retourne le mapping des gymnases externes vers gymnases de la config.
        
        Returns:
            Dictionnaire {nom_externe: nom_config}
        """
        if self._mapping_gymnases is None:
            self._mapping_gymnases = {
                # Mappings LAURASU → Config
                'HALLE - C.BESSON': 'BESSON',
                'HALLE - C. BESSON': 'BESSON',
                'GYMNASE DESCARTES': 'DESCARTES',
                'ENS DESCARTES': 'DESCARTES',
                'DESCARTES': 'DESCARTES',
                'CENTRALE': 'ECL',
                'GYMNASE CENTRALE': 'ECL',
                'GYMNASE CENTRALE LYON': 'ECL',
                'GYMNASE ESA': 'ESA',
                'COMPET C (HAUT) - LEON JOUHAUX': 'L. J. HAUT',
                'HALLE LYON 2': 'LYON 2 HC',
                'HALLE - 3D': 'LAENNEC',
                'LAENNEC': 'LAENNEC',
                'CSU GRENOBLE': 'GRENOBLE',
            }
        return self._mapping_gymnases
    
    def charger_equipes_mixtes(self) -> set:
        """
        Charge la liste des équipes présentes en F et M depuis la config.
        
        Returns:
            Set des noms d'équipes mixtes
        """
        if self._equipes_mixtes is None:
            from collections import defaultdict
            
            # Lire les équipes de la config
            df_equipes = pd.read_excel(self.config_excel_path, sheet_name='Equipes')
            
            # Grouper par équipe et compter les genres
            equipes_genres = defaultdict(set)
            for _, row in df_equipes.iterrows():
                equipe = row['Equipe']
                poule = row['Poule']
                # Le 3ème caractère de la poule indique le genre (F ou M)
                genre = poule[2] if len(poule) > 2 else None
                if genre in ['F', 'M']:
                    equipes_genres[equipe].add(genre)
            
            # Garder uniquement celles avec F et M
            self._equipes_mixtes = {
                equipe for equipe, genres in equipes_genres.items()
                if len(genres) == 2
            }
        
        return self._equipes_mixtes
    
    def ajouter_genre_equipe(self, nom_equipe: str, genre: str) -> str:
        """
        Ajoute le genre entre crochets si l'équipe existe en F et M.
        
        Args:
            nom_equipe: Nom de l'équipe
            genre: Genre ('F' ou 'M')
            
        Returns:
            Nom de l'équipe avec [F] ou [M] si mixte
        """
        equipes_mixtes = self.charger_equipes_mixtes()
        
        if nom_equipe in equipes_mixtes and genre in ['F', 'M']:
            return f"{nom_equipe} [{genre}]"
        
        return nom_equipe
    
    def normaliser_gymnase(self, gymnase_externe: str) -> str:
        """
        Normalise le nom d'un gymnase selon le mapping.
        
        Args:
            gymnase_externe: Nom du gymnase dans le fichier externe
            
        Returns:
            Nom normalisé du gymnase
        """
        if pd.isna(gymnase_externe):
            return ''
        
        mapping = self.mapper_gymnases()
        gymnase_clean = str(gymnase_externe).strip()
        
        # Chercher dans le mapping
        return mapping.get(gymnase_clean, gymnase_clean)
    
    def telecharger_fichier_externe(self) -> pd.DataFrame:
        """
        Télécharge et charge le fichier Excel depuis l'URL ou charge depuis un fichier local.
        
        Returns:
            DataFrame contenant les données du fichier externe
        """
        # Cas 1: Fichier local
        if self.fichier_local:
            print(f"📂 Chargement du fichier local: {self.fichier_local}")
            
            if not self.fichier_local.exists():
                raise FileNotFoundError(f"Fichier introuvable: {self.fichier_local}")
            
            try:
                # Charger sans en-tête d'abord pour trouver la ligne de début
                df_raw = pd.read_excel(self.fichier_local, sheet_name=0, header=None)
                
                # Trouver la ligne avec "Date" et "Sport" (en-têtes)
                header_row = None
                for i, row in df_raw.iterrows():
                    if pd.notna(row[0]) and str(row[0]).strip() == 'Date':
                        if pd.notna(row[1]) and str(row[1]).strip() == 'Sport':
                            header_row = i
                            break
                
                if header_row is None:
                    raise ValueError("Impossible de trouver les en-têtes dans le fichier")
                
                print(f"   → En-têtes trouvés à la ligne {header_row}")
                
                # Recharger avec le bon header
                df = pd.read_excel(self.fichier_local, sheet_name=0, header=header_row)
                
                # Nettoyer les noms de colonnes (enlever espaces)
                df.columns = df.columns.str.strip()
                
                print(f"✓ Fichier chargé: {len(df)} lignes")
                print(f"   Colonnes: {list(df.columns)}")
                return df
            except Exception as e:
                raise RuntimeError(f"Erreur lors du chargement: {e}")
        
        # Cas 2: URL externe
        if self.url_externe:
            print(f"📥 Téléchargement du fichier depuis: {self.url_externe[:50]}...")
            
            try:
                temp_file = None
                # Télécharger le fichier
                response = requests.get(self.url_externe, timeout=30)
                response.raise_for_status()
                
                # Vérifier si la réponse contient du HTML (page d'authentification SharePoint)
                if response.headers.get('content-type', '').startswith('text/html'):
                    print(f"   → Détection d'une page d'authentification SharePoint, tentative d'authentification...")
                    
                    try:
                        # Essayer l'authentification SharePoint
                        file_content = self._download_sharepoint_file(self.url_externe)
                        temp_file = Path("temp_externe.xlsx")
                        temp_file.write_bytes(file_content)
                        
                    except Exception as auth_error:
                        raise ValueError(
                            f"L'URL SharePoint nécessite une authentification. "
                            f"Assurez-vous que les paramètres suivants sont correctement configurés dans le code :\n"
                            f"  • Client ID Azure AD\n"
                            f"  • Tenant ID Azure AD\n"
                            f"  • Client Secret Azure AD\n"
                            f"Erreur d'authentification : {auth_error}"
                        )
                
                else:
                    # Sauvegarder temporairement
                    temp_file = Path("temp_externe.xlsx")
                    temp_file.write_bytes(response.content)
                
                # Charger avec pandas (essayer différents engines Excel)
                df_raw = None
                engine_used = None
                
                # Charger avec openpyxl (format moderne .xlsx)
                try:
                    df_raw = pd.read_excel(temp_file, sheet_name=0, header=None, engine='openpyxl')
                    engine_used = 'openpyxl'
                except Exception as e:
                    raise ValueError(f"Impossible de lire le fichier Excel avec openpyxl: {e}")
                
                # Trouver la ligne avec "Date" et "Sport"
                header_row = None
                for i, row in df_raw.iterrows():
                    if pd.notna(row[0]) and str(row[0]).strip() == 'Date':
                        if pd.notna(row[1]) and str(row[1]).strip() == 'Sport':
                            header_row = i
                            break
                
                if header_row is None:
                    raise ValueError("Impossible de trouver les en-têtes dans le fichier")
                
                # Recharger avec le bon header
                df = pd.read_excel(temp_file, sheet_name=0, header=header_row, engine=engine_used)
                df.columns = df.columns.str.strip()
                
                # Nettoyer
                temp_file.unlink()
                
                print(f"✓ Fichier téléchargé: {len(df)} lignes")
                return df
                
            except requests.RequestException as e:
                raise RuntimeError(f"Erreur lors du téléchargement: {e}")
            except Exception as e:
                # Nettoyer le fichier temporaire si il existe
                try:
                    if 'temp_file' in locals():
                        temp_file_obj = locals()['temp_file']
                        if temp_file_obj is not None and hasattr(temp_file_obj, 'exists') and temp_file_obj.exists():
                            temp_file_obj.unlink(missing_ok=True)
                except:
                    pass
                
                # Message d'erreur plus informatif
                error_msg = str(e)
                if "Expected BOF record" in error_msg and "<!DOCTYPE" in error_msg:
                    error_msg = (
                        "Le fichier téléchargé n'est pas un Excel valide. Cela indique probablement :\n"
                        "  • L'URL SharePoint nécessite une authentification\n"
                        "  • Le fichier n'est pas accessible publiquement\n"
                        "  • L'URL est une page de connexion plutôt qu'un lien direct vers le fichier\n"
                        "\nSolutions :\n"
                        "  1. Téléchargez manuellement le fichier Excel depuis SharePoint\n"
                        "  2. Utilisez --fichier-local avec le fichier téléchargé\n"
                        "  3. Vérifiez les permissions de partage du fichier SharePoint\n"
                        "  4. Utilisez l'option 'Télécharger une copie' dans SharePoint pour obtenir l'URL directe"
                    )
                elif "Excel file format cannot be determined" in error_msg:
                    error_msg = (
                        "Format de fichier Excel non reconnu. Essayez :\n"
                        "  • De télécharger manuellement le fichier\n"
                        "  • De vérifier que c'est bien un fichier .xlsx ou .xls"
                    )
                
                raise RuntimeError(f"Erreur lors du chargement du fichier: {error_msg}")
        
        raise ValueError("Aucune source de fichier spécifiée")
    
    def explorer_structure(self):
        """
        Explore et affiche la structure du fichier externe.
        Utile pour comprendre le format des données.
        """
        if self.df_externe is None:
            raise ValueError("Fichier externe non chargé. Appelez telecharger_fichier_externe() d'abord.")
        
        print("\n" + "="*70)
        print("📊 EXPLORATION DU FICHIER EXTERNE")
        print("="*70)
        
        print(f"\n📋 Colonnes détectées ({len(self.df_externe.columns)}):")
        for i, col in enumerate(self.df_externe.columns, 1):
            print(f"   {i}. {col}")
        
        print(f"\n📏 Dimensions:")
        print(f"   - Lignes: {len(self.df_externe)}")
        print(f"   - Colonnes: {len(self.df_externe.columns)}")
        
        print(f"\n🔍 Aperçu des 5 premières lignes:")
        print(self.df_externe.head(5).to_string(index=False))
        
        print(f"\n📊 Types de données:")
        for col in self.df_externe.columns:
            dtype = self.df_externe[col].dtype
            non_null = self.df_externe[col].notna().sum()
            print(f"   - {col}: {dtype} ({non_null} valeurs non-nulles)")
        
        print("\n" + "="*70)
    
    def mapper_colonnes(self) -> Dict[str, str]:
        """
        Mappe les colonnes du fichier externe vers le format Matchs_Fixes.
        
        Format fichier externe (LAURASU):
        - Date, Sport, Sexe, Poule, Equipe 1, Equipe 2, Hre Déb, Lieu, Commentaire, Arbitres, Résultats
        
        Format Matchs_Fixes:
        - Equipe_1, Equipe_2, Poule, Semaine, Horaire, Gymnase, Score, Type_Competition, Remarques, Arbitres
        
        Returns:
            Dictionnaire de mapping {colonne_externe: colonne_config}
        """
        # Mapping flexible qui gère différents formats de colonnes
        mapping = {
            # Format LAURASU (avec espaces)
            'Equipe 1': 'Equipe_1',
            'Equipe 2': 'Equipe_2',
            'Poule': 'Poule',
            'Sexe': 'Genre',
            'Hre Déb': 'Horaire',
            'Lieu': 'Gymnase',
            'Résultats': 'Score',
            'Commentaire': 'Remarques',
            'Arbitres': 'Arbitres',
            # Format alternatif (underscores)
            'Résultat': 'Score',
            'Equipe_1': 'Equipe_1',
            'Equipe_2': 'Equipe_2',
            'Genre': 'Genre',
            'Score': 'Score',
            'Remarques': 'Remarques',
            'Gymnase': 'Gymnase',
            'Horaire': 'Horaire',
            'Arbitre': 'Arbitres',  # Singulier aussi supporté
        }
        return mapping
    
    def filtrer_matchs(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Filtre les matchs selon les critères de base (sport, date, score).
        
        Note: Cette méthode n'est plus utilisée directement par executer(),
        qui utilise maintenant identifier_matchs_fixes() et identifier_matchs_annules().
        Elle est conservée pour compatibilité.
        
        Args:
            df: DataFrame source
            
        Returns:
            DataFrame filtré
        """
        df_filtre = df.copy()
        
        print(f"\n🔍 Filtrage des matchs (départ: {len(df_filtre)} matchs)...")
        
        # Filtre par sport
        if 'Sport' in df_filtre.columns:
            df_filtre = df_filtre[df_filtre['Sport'].str.upper() == self.sport]
            print(f"   → Filtre sport '{self.sport}': {len(df_filtre)} matchs")
        
        # Appliquer les filtres optionnels
        df_filtre = self._appliquer_filtres_optionnels(df_filtre)
        
        # Appliquer le tri fixes/annulés si demandé
        if self.ignorer_annules:
            df_filtre = self.identifier_matchs_fixes(df_filtre)
            print(f"   → Après exclusion annulations: {len(df_filtre)} matchs")
        
        return df_filtre
    
    def identifier_matchs_annules(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Identifie les matchs annulés pour les stocker dans la feuille Matchs_Annules.
        
        Un match est considéré comme annulé s'il a une mention d'annulation
        dans les remarques/commentaires ET qu'il n'a PAS de score.
        Les forfaits avec score vont dans Matchs_Fixes.
        
        Args:
            df: DataFrame source (déjà filtré par sport)
            
        Returns:
            DataFrame contenant uniquement les matchs annulés sans score
        """
        df_copie = df.copy()
        
        # Vérifier si le match a un score
        score_col = 'Résultats' if 'Résultats' in df_copie.columns else 'Score'
        if score_col in df_copie.columns:
            has_score = df_copie[score_col].notna() & (df_copie[score_col].astype(str).str.strip() != '')
        else:
            has_score = pd.Series([False] * len(df_copie), index=df_copie.index)
        
        # Identifier les mentions d'annulation dans les remarques
        remarques_col = 'Remarques' if 'Remarques' in df_copie.columns else ('Commentaire' if 'Commentaire' in df_copie.columns else None)
        
        if remarques_col is None:
            return pd.DataFrame()
        
        # Regex pour détecter différents types d'annulation
        pattern_annule = r'annul|report|forfait|blessure|maladie|erreur'
        mask_annule = df_copie[remarques_col].astype(str).str.contains(pattern_annule, case=False, na=False, regex=True)
        
        # Matchs_Annules = SANS score ET AVEC mention d'annulation
        mask_matchs_annules = ~has_score & mask_annule
        
        return df_copie[mask_matchs_annules]
    
    def identifier_matchs_fixes(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Identifie les matchs qui vont dans Matchs_Fixes.
        
        Un match va dans Matchs_Fixes si :
        - Il a un score (match joué, y compris forfait avec score)
        - OU il n'a PAS de mention d'annulation dans les remarques
        
        Les forfaits AVEC score vont dans Matchs_Fixes.
        Seuls les matchs SANS score ET AVEC annulation vont dans Matchs_Annules.
        
        Args:
            df: DataFrame source (déjà filtré par sport)
            
        Returns:
            DataFrame contenant les matchs pour Matchs_Fixes
        """
        df_copie = df.copy()
        
        # Déterminer quels matchs ont un score valide
        score_col = 'Résultats' if 'Résultats' in df_copie.columns else 'Score'
        if score_col in df_copie.columns:
            has_score = df_copie[score_col].notna() & (df_copie[score_col].astype(str).str.strip() != '')
        else:
            has_score = pd.Series([False] * len(df_copie), index=df_copie.index)
        
        # Identifier les mentions d'annulation dans les remarques
        remarques_col = 'Remarques' if 'Remarques' in df_copie.columns else ('Commentaire' if 'Commentaire' in df_copie.columns else None)
        
        if remarques_col is None:
            # Pas de colonne remarques, tous les matchs vont dans Matchs_Fixes
            return df_copie
        
        # Regex pour détecter différents types d'annulation
        pattern_annule = r'annul|report|forfait|blessure|maladie|erreur'
        mask_annule = df_copie[remarques_col].astype(str).str.contains(pattern_annule, case=False, na=False, regex=True)
        
        # Matchs_Fixes = matchs avec score OU matchs sans mention d'annulation
        mask_fixes = has_score | ~mask_annule
        
        return df_copie[mask_fixes]
    
    def convertir_vers_format_config(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Convertit le DataFrame externe vers le format Matchs_Fixes.
        
        Args:
            df: DataFrame source
            
        Returns:
            DataFrame au format Matchs_Fixes
        """
        mapping = self.mapper_colonnes()
        computed_weeks = None
        import_tag = self.import_comment_tag or 'Import externe'
        
        # Créer le DataFrame cible
        df_config = pd.DataFrame()
        
        # Mapper les colonnes directes
        for col_externe, col_config in mapping.items():
            if col_externe in df.columns:
                df_config[col_config] = df[col_externe]
        
        dates_parsed = None
        if 'Date' in df.columns:
            dates_parsed = pd.to_datetime(df['Date'], errors='coerce', dayfirst=True)
            df_config['Date'] = dates_parsed.dt.normalize()
        else:
            df_config['Date'] = pd.Series([''] * len(df_config), dtype='object')

        official_idx = self._jour_match_weekday_index()
        is_official_day = pd.Series([False] * len(df_config), index=df_config.index)
        week_from_dates = None
        if dates_parsed is not None:
            if official_idx is not None:
                is_official_day = dates_parsed.dt.weekday.eq(official_idx).fillna(False)
            week_from_dates = dates_parsed.apply(self._calculer_semaine_depuis_timestamp)

        df_config['Semaine'] = ''
        if week_from_dates is not None:
            mask_valid = is_official_day & week_from_dates.notna()
            if mask_valid.any():
                df_config.loc[mask_valid, 'Semaine'] = week_from_dates[mask_valid].astype(int).astype(str)

        if self.journee is not None:
            journee_value = int(self.journee)
            mask_no_date = pd.Series([True] * len(df_config), index=df_config.index)
            if dates_parsed is not None:
                mask_no_date = dates_parsed.isna()

            mask_missing_week = df_config['Semaine'].astype(str).str.strip() == ''
            mask_apply = mask_no_date & mask_missing_week
            if mask_apply.any():
                df_config.loc[mask_apply, 'Semaine'] = str(journee_value)

            if week_from_dates is not None:
                mismatches = is_official_day & week_from_dates.notna() & (week_from_dates.astype(float) != journee_value)
                mismatches_count = int(mismatches.sum())
                if mismatches_count > 0:
                    print(f"   ℹ️  {mismatches_count} match(s) datés hors de la journée J{journee_value} (semaine estimée conservée depuis la date)")
        
        # Traitement spécial pour l'horaire (convertir format time en HH:MM)
        if 'Horaire' in df_config.columns:
            def formater_horaire(horaire):
                if pd.isna(horaire):
                    return ''
                # Si c'est déjà un string HH:MM
                if isinstance(horaire, str):
                    return horaire.strip()
                # Si c'est un time object
                if hasattr(horaire, 'hour'):
                    return f"{horaire.hour:02d}:{horaire.minute:02d}"
                # Essayer de parser
                try:
                    horaire_str = str(horaire)
                    if ':' in horaire_str:
                        parts = horaire_str.split(':')
                        return f"{int(parts[0]):02d}:{int(parts[1]):02d}"
                except:
                    pass
                return str(horaire)
            
            df_config['Horaire'] = df_config['Horaire'].apply(formater_horaire)
        
        # Formater les dates en chaîne DD/MM/YY pour Matchs_Fixes
        if 'Date' in df_config.columns:
            dates_normalises = pd.to_datetime(df_config['Date'], errors='coerce')
            df_config['Date'] = dates_normalises.apply(
                lambda value: format_user_date(value.to_pydatetime()) if pd.notna(value) else ''
            )

        # Nettoyer et normaliser les équipes
        def normaliser_nom_equipe(nom_equipe: str) -> str:
            """
            Normalise le nom d'une équipe :
            - Enlève les espaces superflus
            - Corrige "INP G" en "INPG"
            - Ajoute "(1)" si l'équipe n'a pas de numéro entre parenthèses
            """
            if pd.isna(nom_equipe):
                return ''
            
            # Nettoyer les espaces
            nom = str(nom_equipe).strip()
            
            # Correction spécifique : "INP G" -> "INPG"
            nom = nom.replace('INP G', 'INPG')
            
            # Vérifier si l'équipe a déjà un numéro entre parenthèses
            # Pattern: (1), (2), etc.
            import re
            if not re.search(r'\(\d+\)', nom):
                # Pas de numéro trouvé, ajouter "(1)"
                nom = f"{nom} (1)"
            
            return nom
        
        for col in ['Equipe_1', 'Equipe_2']:
            if col in df_config.columns:
                df_config[col] = df_config[col].apply(normaliser_nom_equipe)
        
        # Nettoyer les poules
        if 'Poule' in df_config.columns:
            df_config['Poule'] = df_config['Poule'].astype(str).str.strip()
        
        # Gérer le genre : priorité à la colonne Genre (Sexe), avec vérification via la poule
        def determiner_genre(row):
            """
            Détermine le genre en priorité depuis la colonne Genre,
            avec vérification de cohérence avec la poule si disponible.
            """
            genre_colonne = None
            genre_poule = None
            
            # 1. Extraire le genre depuis la colonne Genre si elle existe
            if 'Genre' in row and pd.notna(row['Genre']):
                genre_str = str(row['Genre']).strip().upper()
                if genre_str in ['F', 'FEMININ', 'FÉMININ', 'FEMME', 'FILLE']:
                    genre_colonne = 'F'
                elif genre_str in ['M', 'MASCULIN', 'HOMME', 'GARCON', 'GARÇON']:
                    genre_colonne = 'M'
            
            # 2. Extraire le genre depuis la poule si elle existe (3ème caractère)
            if 'Poule' in row and pd.notna(row['Poule']):
                poule_str = str(row['Poule']).strip()
                if len(poule_str) >= 3:
                    genre_char = poule_str[2]
                    if genre_char in ['F', 'M']:
                        genre_poule = genre_char
            
            # 3. Décision finale avec vérification de cohérence
            if genre_colonne and genre_poule:
                # Les deux sont disponibles : vérifier la cohérence
                if genre_colonne != genre_poule:
                    print(f"   ⚠️  Incohérence de genre détectée pour poule '{row.get('Poule', '')}': "
                          f"colonne='{genre_colonne}' vs poule='{genre_poule}' - Utilisation de la colonne")
                return genre_colonne
            elif genre_colonne:
                # Priorité à la colonne
                return genre_colonne
            elif genre_poule:
                # Fallback sur la poule
                return genre_poule
            else:
                # Aucune information disponible
                return ''
        
        # Appliquer la fonction de détermination du genre
        if 'Genre' in df_config.columns or 'Poule' in df_config.columns:
            df_config['Genre'] = df_config.apply(determiner_genre, axis=1)
        else:
            df_config['Genre'] = ''
        
        # Enlever le [F] ou [M] des noms d'équipes (déjà fait dans ajouter_genre_equipe, mais on nettoie)
        for col in ['Equipe_1', 'Equipe_2']:
            if col in df_config.columns:
                df_config[col] = df_config[col].str.replace(' [F]', '', regex=False)
                df_config[col] = df_config[col].str.replace(' [M]', '', regex=False)
                df_config[col] = df_config[col].str.strip()
        
        # Normaliser les gymnases
        if 'Gymnase' in df_config.columns:
            df_config['Gymnase'] = df_config['Gymnase'].apply(self.normaliser_gymnase)
        
        # Parser et normaliser les scores
        if 'Score' in df_config.columns:
            df_config['Score'] = df_config['Score'].apply(self._parser_score)
        
        # Déterminer le type de compétition depuis la poule
        # et ajuster la poule pour CFE/CFU (vider la poule, garder juste le type)
        def determiner_type_competition(poule):
            """
            Détermine le type de compétition depuis le code de poule.
            - CFE* ou CFU* -> CFE ou CFU
            - Poule valide Acad (format: SPORT+GENRE+NIVEAU+..., ex: VBFA1PA) -> Acad
            - Sinon (poule vide ou format inconnu) -> Autre
            """
            if pd.isna(poule) or not str(poule).strip():
                return 'Autre'
            
            poule_str = str(poule).strip().upper()
            
            if poule_str.startswith('CFE'):
                return 'CFE'
            elif poule_str.startswith('CFU'):
                return 'CFU'
            else:
                # Vérifier si c'est une poule Acad valide
                # Format attendu: CODE_SPORT (2 cars) + GENRE (F/M) + autres chars
                # Exemples: VBFA1PA, VBMA2PB, HBFA1, BBMA1
                import re
                # Pattern: 2 lettres (sport) + F ou M (genre) + au moins 1 char supplémentaire
                if re.match(r'^[A-Z]{2}[FM].+$', poule_str):
                    return 'Acad'
                else:
                    return 'Autre'
        
        if 'Poule' in df_config.columns:
            # D'abord déterminer le type de compétition depuis la poule originale
            df_config['Type_Competition'] = df_config['Poule'].apply(determiner_type_competition)
            
            # Ensuite, pour les matchs CFE et CFU, vider la poule
            # (seul Type_Competition compte pour ces compétitions)
            def ajuster_poule(row):
                if row['Type_Competition'] in ['CFE', 'CFU']:
                    return ''
                else:
                    return row['Poule']
            
            df_config['Poule'] = df_config.apply(ajuster_poule, axis=1)
        else:
            df_config['Type_Competition'] = 'Autre'
        
        # Ajouter les colonnes manquantes avec valeurs par défaut
        colonnes_requises = [
            'Equipe_1', 'Equipe_2', 'Genre', 'Poule', 'Semaine', 'Date', 'Horaire',
            'Gymnase', 'Score', 'Type_Competition', 'Remarques', 'Arbitres'
        ]
        
        for col in colonnes_requises:
            if col not in df_config.columns:
                if col == 'Remarques':
                    df_config[col] = import_tag
                elif col == 'Genre':
                    df_config[col] = ''
                elif col == 'Date':
                    df_config[col] = ''
                elif col == 'Type_Competition':
                    # Type_Competition devrait déjà être défini plus haut
                    # Si ce n'est pas le cas, on met Autre par défaut
                    df_config[col] = 'Autre'
                elif col == 'Arbitres':
                    df_config[col] = ''
                else:
                    df_config[col] = ''
        
        # Filtrer les matchs annulés si demandé (après avoir ajouté les remarques par défaut)
        # IMPORTANT: les matchs avec score (forfaits) ne sont PAS filtrés
        if self.ignorer_annules:
            # Déterminer quels matchs ont un score valide
            has_score = df_config['Score'].notna() & (df_config['Score'].astype(str).str.strip() != '')
            
            # Recherche insensible à la casse pour 'annul' ou 'erreur' (annulé, annulé, erreur, etc.)
            remarques_annules = df_config['Remarques'].astype(str).str.lower().str.contains(r'annul|erreur', regex=True)
            
            # On ne filtre QUE les matchs annulés SANS score
            mask_vraiment_annule = remarques_annules & ~has_score
            nb_annules = mask_vraiment_annule.sum()
            if nb_annules > 0:
                print(f"   ⚠️  {nb_annules} match(s) annulé(s) ou avec erreur (sans score) ignoré(s)")
                df_config = df_config[~mask_vraiment_annule]
        
        # Compléter les remarques existantes
        if 'Remarques' in df_config.columns:
            df_config['Remarques'] = df_config['Remarques'].apply(self._merge_import_comment_tag)
        
        # Réordonner les colonnes
        df_config = df_config[colonnes_requises]
        
        return df_config
    
    def charger_matchs_fixes_existants(self) -> pd.DataFrame:
        """
        Charge les matchs fixes existants depuis la configuration.
        
        Returns:
            DataFrame des matchs fixes existants
        """
        if not self.config_excel_path.exists():
            raise FileNotFoundError(f"Fichier de configuration introuvable: {self.config_excel_path}")
        
        try:
            df = pd.read_excel(self.config_excel_path, sheet_name='Matchs_Fixes')
            print(f"✓ Matchs fixes existants chargés: {len(df)} matchs")
            return df
        except Exception as e:
            print(f"⚠️  Aucun match fixe existant (feuille vide ou inexistante)")
            # Créer un DataFrame vide avec les bonnes colonnes
            return pd.DataFrame(columns=[
                'Equipe_1', 'Equipe_2', 'Genre', 'Poule', 'Semaine', 'Horaire',
                'Gymnase', 'Score', 'Type_Competition', 'Remarques', 'Arbitres'
            ])
    
    def fusionner_matchs(self, df_nouveaux: pd.DataFrame, df_existants: pd.DataFrame) -> pd.DataFrame:
        """
        Fusionne les nouveaux matchs avec les existants (évite les doublons).
        Gère la cohérence des scores entre doublons.
        
        Args:
            df_nouveaux: Nouveaux matchs à ajouter
            df_existants: Matchs déjà présents
            
        Returns:
            DataFrame fusionné
        """
        # Créer une clé unique pour détecter les doublons
        # IMPORTANT: Inclure le Genre et Type_Competition pour différencier les matchs
        # avec les mêmes équipes mais de genre ou compétition différents
        def creer_cle(df):
            if 'Date' in df.columns:
                date_component = pd.to_datetime(df['Date'], errors='coerce', dayfirst=True).dt.strftime('%Y-%m-%d').fillna('')
            else:
                date_component = pd.Series([''] * len(df), index=df.index)
            return (
                df['Equipe_1'].astype(str) + '|' +
                df['Equipe_2'].astype(str) + '|' +
                (df['Genre'].astype(str) if 'Genre' in df.columns else '') + '|' +
                (df['Type_Competition'].astype(str) if 'Type_Competition' in df.columns else '') + '|' +
                (df['Semaine'].astype(str) if 'Semaine' in df.columns else '') + '|' +
                date_component
            )
        
        df_nouveaux['_cle'] = creer_cle(df_nouveaux)
        df_existants['_cle'] = creer_cle(df_existants)
        
        # Identifier les doublons
        doublons = df_nouveaux[df_nouveaux['_cle'].isin(df_existants['_cle'])]
        nouveaux_uniques = df_nouveaux[~df_nouveaux['_cle'].isin(df_existants['_cle'])]
        
        print(f"\n📊 Fusion des matchs:")
        print(f"   - Matchs à ajouter: {len(df_nouveaux)}")
        print(f"   - Doublons détectés: {len(doublons)}")
        print(f"   - Nouveaux matchs uniques: {len(nouveaux_uniques)}")
        strategie_label = 'nouvelle version' if self.doublon_priorite == 'nouveau' else 'version existante'
        print(f"   - Stratégie doublons: {strategie_label}")
        
        # Traiter les doublons pour gérer les scores
        df_doublons_traite = pd.DataFrame()
        changements_effectues = 0

        if not doublons.empty:
            print(f"   🔍 Analyse des {len(doublons)} doublon(s)...")

            prefer_new = self.doublon_priorite == 'nouveau'

            for cle in doublons['_cle'].unique():
                # Récupérer les versions nouveau et existant
                nouveau = df_nouveaux[df_nouveaux['_cle'] == cle].iloc[0]
                existant = df_existants[df_existants['_cle'] == cle].iloc[0]

                gagnant = nouveau if prefer_new else existant
                perdant = existant if prefer_new else nouveau

                # Comparer les scores
                score_gagnant = str(gagnant.get('Score', '')).strip()
                score_perdant = str(perdant.get('Score', '')).strip()

                # Considérer comme vide/invalide: '', 'nan', 'NaN', ou seulement des espaces
                def score_est_valide(score):
                    return score and score.lower() not in ['nan', ''] and score.strip() != ''

                score_gagnant_valide = score_est_valide(score_gagnant)
                score_perdant_valide = score_est_valide(score_perdant)

                match_final = gagnant.copy()

                equipes = f"{cle.split('|')[0]} vs {cle.split('|')[1]}"

                if score_gagnant_valide and score_perdant_valide:
                    if score_gagnant != score_perdant:
                        source = 'nouvelle version' if prefer_new else 'ancienne version'
                        print(f"   ⚠️  Doublon {equipes}: scores différents!")
                        print(f"      Version conservée ({source}): '{score_gagnant}' | Autre: '{score_perdant}'")
                elif not score_gagnant_valide and score_perdant_valide:
                    source = 'existant' if prefer_new else 'nouveau'
                    print(f"   📝 Doublon {equipes}: ajout score depuis le {source} '{score_perdant}'")
                    match_final['Score'] = score_perdant
                    changements_effectues += 1

                df_doublons_traite = pd.concat([df_doublons_traite, match_final.to_frame().T], ignore_index=True)

        # Afficher un résumé des changements seulement s'il y en a eu
        if changements_effectues > 0:
            print(f"   ✅ {changements_effectues} doublon(s) mis à jour avec de nouveaux scores")
        
        # Fusionner: existants (sans les doublons traités) + doublons traités + nouveaux uniques
        existants_sans_doublons = df_existants[~df_existants['_cle'].isin(doublons['_cle'])]
        df_fusionne = pd.concat([
            existants_sans_doublons.drop(columns=['_cle']),
            df_doublons_traite.drop(columns=['_cle']) if not df_doublons_traite.empty else pd.DataFrame(),
            nouveaux_uniques.drop(columns=['_cle'])
        ], ignore_index=True)
        
        return df_fusionne
    
    def sauvegarder_configuration(self, df_matchs: pd.DataFrame, df_annules: Optional[pd.DataFrame] = None):
        """
        Sauvegarde les matchs dans la feuille Matchs_Fixes de la configuration.
        Optionnellement sauvegarde aussi les matchs annulés dans Matchs_Annules.
        
        Args:
            df_matchs: DataFrame des matchs à sauvegarder
            df_annules: DataFrame optionnel des matchs annulés (sans score)
        """
        if self.dry_run:
            print("\n🔍 MODE SIMULATION - Aucune modification effectuée")
            print(f"   {len(df_matchs)} matchs seraient sauvegardés dans Matchs_Fixes")
            if df_annules is not None and len(df_annules) > 0:
                print(f"   {len(df_annules)} matchs seraient sauvegardés dans Matchs_Annules")
            return
        
        print(f"\n💾 Sauvegarde dans {self.config_excel_path}...")
        
        try:
            # Charger le workbook existant
            wb = openpyxl.load_workbook(self.config_excel_path)
            
            # Supprimer la feuille Matchs_Fixes si elle existe
            if 'Matchs_Fixes' in wb.sheetnames:
                del wb['Matchs_Fixes']
            
            # Créer une nouvelle feuille
            ws = wb.create_sheet('Matchs_Fixes')
            
            # Écrire les données
            for r in dataframe_to_rows(df_matchs, index=False, header=True):
                ws.append(r)
            
            print(f"✓ {len(df_matchs)} matchs dans Matchs_Fixes")
            
            # Sauvegarder les matchs annulés si fournis
            if df_annules is not None and len(df_annules) > 0:
                # Supprimer la feuille Matchs_Annules si elle existe
                if 'Matchs_Annules' in wb.sheetnames:
                    del wb['Matchs_Annules']
                
                # Créer une nouvelle feuille pour les matchs annulés
                ws_annules = wb.create_sheet('Matchs_Annules')
                
                # Écrire les données
                for r in dataframe_to_rows(df_annules, index=False, header=True):
                    ws_annules.append(r)
                
                print(f"✓ {len(df_annules)} matchs dans Matchs_Annules")
            
            # Sauvegarder
            wb.save(self.config_excel_path)
            
            print(f"✓ Configuration sauvegardée avec succès")
            
        except Exception as e:
            raise RuntimeError(f"Erreur lors de la sauvegarde: {e}")
    
    def charger_matchs_annules_existants(self) -> pd.DataFrame:
        """
        Charge les matchs annulés existants depuis la configuration.
        
        Returns:
            DataFrame des matchs annulés existants
        """
        if not self.config_excel_path.exists():
            return pd.DataFrame()
        
        try:
            df = pd.read_excel(self.config_excel_path, sheet_name='Matchs_Annules')
            print(f"✓ Matchs annulés existants chargés: {len(df)} matchs")
            return df
        except Exception as e:
            print(f"   ℹ️  Aucun match annulé existant (feuille vide ou inexistante)")
            # Créer un DataFrame vide avec les bonnes colonnes
            return pd.DataFrame(columns=[
                'Equipe_1', 'Equipe_2', 'Genre', 'Poule', 'Semaine', 'Date', 'Horaire',
                'Gymnase', 'Score', 'Type_Competition', 'Remarques', 'Arbitres'
            ])
    
    def executer(self):
        """Exécute le processus complet d'importation."""
        print("="*70)
        print("🔄 IMPORTATION DE MATCHS EXTERNES")
        print("="*70)
        
        print(f"\n⚙️  Configuration:")
        print(f"   - Fichier config: {self.config_yaml_path}")
        print(f"   - Fichier Excel: {self.config_excel_path}")
        print(f"   - Sport: {self.sport}")
        if self.journee:
            print(f"   - Journée: {self.journee}")
        if self.date_limite:
            print(f"   - Date limite: {format_user_date(self.date_limite)}")
        print(f"   - Filtre: {'Avec score' if self.avec_score else 'Sans score' if self.sans_score else 'Tous'}")
        if self.dry_run:
            print(f"   - Mode: SIMULATION (dry-run)")
        
        # 1. Télécharger le fichier externe (ou réutiliser un DataFrame préchargé)
        if self.df_externe is None:
            self.df_externe = self.telecharger_fichier_externe()
        else:
            try:
                nb_lignes = len(self.df_externe)
            except Exception:
                nb_lignes = 'N/A'
            print(f"\n📂 Fichier externe déjà chargé (réutilisation, {nb_lignes} lignes)")
        
        # 2. Pré-filtrer par sport pour avoir tous les matchs du sport
        df_sport = self.df_externe.copy()
        if 'Sport' in df_sport.columns:
            df_sport = df_sport[df_sport['Sport'].str.upper() == self.sport]
            print(f"\n📊 Matchs du sport '{self.sport}': {len(df_sport)} matchs au total")
        
        if len(df_sport) == 0:
            print("⚠️  Aucun match pour ce sport. Arrêt.")
            return
        
        # 3. Séparer les matchs en deux catégories :
        #    - Matchs_Fixes : avec score OU sans mention d'annulation
        #    - Matchs_Annules : avec mention d'annulation dans les remarques
        print(f"\n📋 Séparation des matchs...")
        
        df_fixes_bruts = self.identifier_matchs_fixes(df_sport)
        df_annules_bruts = self.identifier_matchs_annules(df_sport)
        
        print(f"   → {len(df_fixes_bruts)} match(s) pour Matchs_Fixes")
        print(f"   → {len(df_annules_bruts)} match(s) pour Matchs_Annules")
        
        # 4. Appliquer les filtres supplémentaires (date_limite, avec_score, sans_score)
        if self.date_limite or self.avec_score or self.sans_score:
            print(f"\n🔍 Application des filtres supplémentaires...")
            df_fixes_bruts = self._appliquer_filtres_optionnels(df_fixes_bruts)
            print(f"   → {len(df_fixes_bruts)} match(s) après filtres pour Matchs_Fixes")
        
        # 5. Convertir vers le format config
        df_convertis = pd.DataFrame()
        if len(df_fixes_bruts) > 0:
            print(f"\n🔄 Conversion vers le format Matchs_Fixes...")
            df_convertis = self.convertir_vers_format_config(df_fixes_bruts.copy())
        
        df_annules_convertis = pd.DataFrame()
        if len(df_annules_bruts) > 0:
            print(f"\n🔄 Conversion vers le format Matchs_Annules...")
            # Désactiver temporairement le filtrage des annulations pour la conversion
            old_ignorer_annules = self.ignorer_annules
            self.ignorer_annules = False
            df_annules_convertis = self.convertir_vers_format_config(df_annules_bruts.copy())
            self.ignorer_annules = old_ignorer_annules
        
        # 6. Charger les matchs existants
        print(f"\n📂 Chargement des matchs existants...")
        df_existants = self.charger_matchs_fixes_existants()
        
        # 7. Fusionner les matchs fixes
        if len(df_convertis) > 0:
            df_final = self.fusionner_matchs(df_convertis, df_existants)
        else:
            df_final = df_existants
        
        # 8. Charger et fusionner les matchs annulés existants
        df_annules_final = None
        if len(df_annules_convertis) > 0:
            df_annules_existants = self.charger_matchs_annules_existants()
            if len(df_annules_existants) > 0:
                df_annules_final = self.fusionner_matchs(df_annules_convertis, df_annules_existants)
            else:
                df_annules_final = df_annules_convertis
        
        # 9. Sauvegarder
        self.sauvegarder_configuration(df_final, df_annules_final)
        
        print("\n" + "="*70)
        print("✅ IMPORTATION TERMINÉE")
        print("="*70)
    
    def _appliquer_filtres_optionnels(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Applique les filtres optionnels (date_limite, avec_score, sans_score).
        
        Args:
            df: DataFrame source
            
        Returns:
            DataFrame filtré
        """
        df_filtre = df.copy()
        
        # Filtre par date
        if self.date_limite and 'Date' in df_filtre.columns:
            df_filtre['_Date_parsed'] = pd.to_datetime(df_filtre['Date'], errors='coerce')
            df_filtre = df_filtre[df_filtre['_Date_parsed'] <= self.date_limite]
            df_filtre = df_filtre.drop(columns=['_Date_parsed'])
            print(f"   → Filtre date ≤ {format_user_date(self.date_limite)}: {len(df_filtre)} matchs")
        
        # Filtre par score
        score_col = 'Résultats' if 'Résultats' in df_filtre.columns else 'Score'
        if score_col in df_filtre.columns:
            if self.avec_score:
                df_filtre = df_filtre[df_filtre[score_col].notna() & (df_filtre[score_col].astype(str).str.strip() != '')]
                print(f"   → Filtre avec score: {len(df_filtre)} matchs")
            elif self.sans_score:
                df_filtre = df_filtre[df_filtre[score_col].isna() | (df_filtre[score_col].astype(str).str.strip() == '')]
                print(f"   → Filtre sans score: {len(df_filtre)} matchs")
        
        return df_filtre


def main():
    """Point d'entrée principal."""
    parser = argparse.ArgumentParser(
        description="Importer des matchs depuis un fichier Excel partagé en ligne",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__
    )
    
    # Arguments obligatoires
    parser.add_argument(
        '--config',
        required=True,
        help="Chemin vers le fichier de configuration Excel"
    )
    
    # Source du fichier (l'un des deux requis)
    source_group = parser.add_mutually_exclusive_group(required=True)
    source_group.add_argument(
        '--url',
        help="URL du fichier Excel partagé en ligne"
    )
    source_group.add_argument(
        '--fichier-local',
        help="Chemin vers un fichier Excel local"
    )
    
    # Arguments optionnels
    parser.add_argument(
        '--sport',
        default='VB',
        choices=['VB', 'HB', 'BB', 'FB', 'FS', 'RG', 'BD', 'TT'],
        help="Code du sport: VB (Volleyball), HB (Handball), BB (Basketball), FB (Football), FS (Futsal), RG (Rugby), BD (Badminton), TT (Tennis Table). Défaut: VB"
    )
    parser.add_argument(
        '--journee',
        type=int,
        help="Numéro de journée à importer"
    )
    parser.add_argument(
        '--date-limite',
        help="Date limite pour l'import (format DD/MM/YYYY)"
    )
    
    # Filtres
    filtre_group = parser.add_mutually_exclusive_group()
    filtre_group.add_argument(
        '--avec-score',
        action='store_true',
        help="Importer uniquement les matchs avec score"
    )
    filtre_group.add_argument(
        '--sans-score',
        action='store_true',
        help="Importer uniquement les matchs sans score (planifiés)"
    )
    filtre_group.add_argument(
        '--tous',
        action='store_true',
        help="Importer tous les matchs (défaut)"
    )
    
    # Options
    parser.add_argument(
        '--dry-run',
        action='store_true',
        help="Mode simulation (ne modifie pas le fichier)"
    )
    parser.add_argument(
        '--explorer',
        action='store_true',
        help="Explorer la structure du fichier externe sans importer"
    )
    parser.add_argument(
        '--ignorer-annules',
        action='store_true',
        default=True,
        help="Ignorer les matchs avec 'annule' ou 'erreur' dans les remarques (défaut: activé)"
    )
    parser.add_argument(
        '--garder-annules',
        action='store_false',
        dest='ignorer_annules',
        help="Garder les matchs avec 'annule' ou 'erreur' dans les remarques"
    )
    parser.add_argument(
        '--priorite-doublons',
        choices=['ancien', 'nouveau'],
        default='ancien',
        help="Détermine quelle version garder en cas de doublon (défaut: ancien)"
    )
    
    args = parser.parse_args()
    
    try:
        importeur = ImporteurMatchsExternes(
            config_path=args.config,
            url_externe=args.url,
            fichier_local=args.fichier_local,
            sport=args.sport,
            journee=args.journee,
            date_limite=args.date_limite,
            avec_score=args.avec_score,
            sans_score=args.sans_score,
            tous=args.tous,
            dry_run=args.dry_run,
            ignorer_annules=args.ignorer_annules,
            doublon_priorite=args.priorite_doublons
        )
        
        if args.explorer:
            importeur.df_externe = importeur.telecharger_fichier_externe()
            importeur.explorer_structure()
        else:
            importeur.executer()
        
        return 0
        
    except Exception as e:
        print(f"\n❌ Erreur: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())
