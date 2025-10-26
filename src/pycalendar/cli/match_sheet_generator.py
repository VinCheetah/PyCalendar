#!/usr/bin/env python3
"""
Script pour générer une feuille de matchs formatée à partir du calendrier volley.

Usage:
    python generer_feuille_matchs.py --semaine 1 --date-depart "16/10/2025"
    python generer_feuille_matchs.py -s 2 -d "23/10/2025"
"""

import pandas as pd
import argparse
from datetime import datetime, timedelta
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# Dictionnaire de mapping des gymnases
GYMNASES_SIUAPS = {
    "DESCARTES": "GYMNASE DESCARTES",
    "ECL": "GYMNASE CENTRALE LYON",
    "ESA": "GYMNASE ESA",
    "LYON 2 HC": "HALLE LYON 2",
    "LAENNEC": "HALLE - 3D",
    "BESSON": "HALLE - C.BESSON",
    "L. J. HAUT": "COMPET C (HAUT) - LEON JOUHAUX"
}


def extraire_genre_niveau(code_poule):
    """
    Extrait le genre et le niveau depuis le code de poule.
    Format: VB + (F/M) + (A1/A2/A3/A4) + P + (A/B/C/...)
    
    Exemples:
        VBFA1PA -> (F, A1)
        VBMA3PB -> (M, A3)
    """
    # Gérer les cas où code_poule est NaN ou None
    if pd.isna(code_poule) or not isinstance(code_poule, str):
        return 'M', 'A1'  # Valeurs par défaut
    
    # Vérifier que le code fait au moins 5 caractères
    if len(code_poule) < 5:
        return 'M', 'A1'  # Valeurs par défaut
    
    # Le genre est en position 2 (après VB)
    genre = code_poule[2] if len(code_poule) > 2 else 'M'  # 'F' ou 'M'
    
    # Le niveau est en position 3-4 (A + chiffre)
    niveau = code_poule[3:5] if len(code_poule) > 4 else 'A1'  # 'A1', 'A2', 'A3', ou 'A4'
    
    return genre, niveau


def mapper_gymnase(gymnase_code):
    """Convertit le code gymnase en nom complet selon le dictionnaire."""
    return GYMNASES_SIUAPS.get(gymnase_code, gymnase_code)


def calculer_date_semaine(date_depart, numero_semaine):
    """
    Calcule la date correspondant à une semaine donnée.
    
    Args:
        date_depart: Date de départ (semaine 1) au format "JJ/MM/AAAA"
        numero_semaine: Numéro de la semaine (1, 2, 3, ...)
    
    Returns:
        Date au format "JJ/MM/AAAA"
    """
    date_obj = datetime.strptime(date_depart, "%d/%m/%Y")
    # Ajouter (numero_semaine - 1) semaines
    date_cible = date_obj + timedelta(weeks=(numero_semaine - 1))
    return date_cible.strftime("%d/%m/%Y")


def charger_et_filtrer_matchs(fichier_calendrier, numero_semaine, date_str):
    """
    Charge le calendrier et filtre les matchs pour une semaine donnée.
    
    Args:
        fichier_calendrier: Chemin vers le fichier Excel du calendrier
        numero_semaine: Numéro de la semaine à extraire
        date_str: Date au format "JJ/MM/AAAA" pour affichage
    
    Returns:
        DataFrame avec les matchs filtrés et formatés
    """
    # Charger le calendrier
    df = pd.read_excel(fichier_calendrier)
    
    # Filtrer par semaine
    df_semaine = df[df['Semaine'] == numero_semaine].copy()
    
    if df_semaine.empty:
        print(f"⚠️  Aucun match trouvé pour la semaine {numero_semaine}")
        return None
    
    # NE PAS filtrer les matchs sans poule - inclure tous les matchs
    # df_semaine = df_semaine[df_semaine['Poule'].notna()].copy()
    
    # Remplir les poules manquantes avec une valeur par défaut
    df_semaine['Poule'] = df_semaine['Poule'].fillna('HORS_CHAMPIONNAT')
    
    # Extraire genre et niveau depuis le code de poule OU utiliser les colonnes Genre directement
    def get_genre(row):
        # Pour les matchs avec poule, extraire depuis le code
        if row['Poule'] != 'HORS_CHAMPIONNAT':
            return extraire_genre_niveau(row['Poule'])[0]
        # Pour les matchs hors championnat, déterminer le genre approprié
        else:
            genre1 = str(row.get('Genre_1', '')).upper()
            genre2 = str(row.get('Genre_2', '')).upper()
            equipe1 = str(row.get('Equipe_1', ''))
            equipe2 = str(row.get('Equipe_2', ''))
            
            # Normaliser les genres
            genre1 = genre1 if genre1 in ['F', 'M'] else None
            genre2 = genre2 if genre2 in ['F', 'M'] else None
            
            # Si les deux genres sont identiques et valides, utiliser ce genre
            if genre1 and genre2 and genre1 == genre2:
                return genre1
            
            # Si un genre est manquant, utiliser l'autre
            if genre1 and not genre2:
                return genre1
            if genre2 and not genre1:
                return genre2
            
            # Pour les matchs avec EXTERNE, privilégier le genre de l'équipe non-EXTERNE
            if 'EXTERNE' in equipe1 and genre2:
                return genre2
            if 'EXTERNE' in equipe2 and genre1:
                return genre1
            
            # Pour les autres cas (genres différents ou mixtes), privilégier F si présent, sinon M
            if genre1 == 'F' or genre2 == 'F':
                return 'F'
            elif genre1 == 'M' or genre2 == 'M':
                return 'M'
            else:
                # Fallback par défaut
                return 'M'
    
    def get_niveau(row):
        # Pour les matchs avec poule, extraire depuis le code
        if row['Poule'] != 'HORS_CHAMPIONNAT':
            return extraire_genre_niveau(row['Poule'])[1]
        # Pour les matchs hors championnat, niveau par défaut
        else:
            return 'EXT'
    
    df_semaine['Genre'] = df_semaine.apply(get_genre, axis=1)
    df_semaine['Niveau'] = df_semaine.apply(get_niveau, axis=1)
    
    # Mapper les gymnases
    df_semaine['Lieu'] = df_semaine['Gymnase'].apply(mapper_gymnase)
    
    # Créer le DataFrame de sortie avec les colonnes du format cible
    df_sortie = pd.DataFrame({
        'Date': date_str,
        'Sport': 'VB',
        'Sexe': df_semaine['Genre'],
        'Poule': df_semaine['Poule'],
        'Equipe 1': df_semaine['Equipe_1'],
        'Equipe 2': df_semaine['Equipe_2'],
        'Hre Déb': df_semaine['Horaire'],
        'Lieu': df_semaine['Lieu']
    })
    
    # Définir l'ordre de tri pour le genre (F avant M)
    genre_order = {'F': 0, 'M': 1}
    df_sortie['_genre_order'] = df_sortie['Sexe'].map(genre_order)
    
    # Définir l'ordre de tri pour le niveau (A1 < A2 < A3 < A4)
    niveau_order = {'A1': 0, 'A2': 1, 'A3': 2, 'A4': 3}
    df_sortie['_niveau_order'] = df_semaine['Niveau'].map(niveau_order)
    
    # Trier par genre, niveau, puis horaire
    df_sortie = df_sortie.sort_values(
        by=['_genre_order', '_niveau_order', 'Hre Déb'],
        ascending=[True, True, True]
    )
    
    # Supprimer les colonnes temporaires de tri
    df_sortie = df_sortie.drop(columns=['_genre_order', '_niveau_order'])
    
    # Réinitialiser l'index
    df_sortie = df_sortie.reset_index(drop=True)
    
    return df_sortie


def appliquer_mise_en_forme(workbook, worksheet, df):
    """
    Applique la mise en forme au fichier Excel pour correspondre à l'exemple.
    
    Args:
        workbook: Objet Workbook d'openpyxl
        worksheet: Feuille de calcul active
        df: DataFrame contenant les données
    """
    # Définir les styles
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    cell_alignment = Alignment(horizontal="left", vertical="center")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    border_style = Side(style="thin", color="000000")
    border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
    
    # Appliquer le style aux en-têtes (ligne 1)
    for col_num, column_title in enumerate(df.columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Appliquer les bordures et alignement aux cellules de données
    for row_num in range(2, len(df) + 2):
        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.border = border
            
            # Alignement centré pour Date, Sport, Sexe, Hre Déb
            if col_num in [1, 2, 3, 7]:  # Date, Sport, Sexe, Hre Déb
                cell.alignment = center_alignment
            else:
                cell.alignment = cell_alignment
    
    # Ajuster la largeur des colonnes
    column_widths = {
        'A': 12,   # Date
        'B': 8,    # Sport
        'C': 6,    # Sexe
        'D': 12,   # Poule
        'E': 15,   # Equipe 1
        'F': 15,   # Equipe 2
        'G': 10,   # Hre Déb
        'H': 35    # Lieu
    }
    
    for col, width in column_widths.items():
        worksheet.column_dimensions[col].width = width


def generer_feuille_excel(df, fichier_sortie):
    """
    Génère un fichier Excel avec mise en forme à partir du DataFrame.
    
    Args:
        df: DataFrame contenant les matchs
        fichier_sortie: Chemin du fichier Excel de sortie
    """
    # Créer un workbook avec openpyxl
    wb = Workbook()
    ws = wb.active
    ws.title = "Matchs"
    
    # Écrire les en-têtes
    for col_num, column_title in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_num, value=column_title)
    
    # Écrire les données
    for row_num, row_data in enumerate(df.itertuples(index=False), 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Appliquer la mise en forme
    appliquer_mise_en_forme(wb, ws, df)
    
    # Sauvegarder le fichier
    wb.save(fichier_sortie)
    print(f"✅ Fichier généré : {fichier_sortie}")
    print(f"   Nombre de matchs : {len(df)}")


def main():
    parser = argparse.ArgumentParser(
        description="Génère une feuille de matchs formatée pour une semaine donnée.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Exemples:
  %(prog)s --semaine 1 --date-depart "16/10/2025"
  %(prog)s -s 2 -d "23/10/2025" -o matchs_semaine_2.xlsx
  %(prog)s -s 3 -d "16/10/2025" --auto-date
        """
    )
    
    parser.add_argument(
        '-s', '--semaine',
        type=int,
        required=True,
        help='Numéro de la semaine (1, 2, 3, ...)'
    )
    
    parser.add_argument(
        '-d', '--date-depart',
        type=str,
        default="16/10/2025",
        help='Date de départ pour la semaine 1 (format: JJ/MM/AAAA). Par défaut: 16/10/2025'
    )
    
    parser.add_argument(
        '--auto-date',
        action='store_true',
        help='Calcule automatiquement la date en fonction de la semaine et de la date de départ'
    )
    
    parser.add_argument(
        '-o', '--output',
        type=str,
        help='Nom du fichier de sortie (par défaut: Matchs_Semaine_X_DATE.xlsx)'
    )
    
    parser.add_argument(
        '-c', '--calendrier',
        type=str,
        default='data_volley/calendrier_volley.xlsx',
        help='Chemin vers le fichier calendrier (par défaut: data_volley/calendrier_volley.xlsx)'
    )
    
    args = parser.parse_args()
    
    # Calculer la date si demandé
    if args.auto_date:
        date_str = calculer_date_semaine(args.date_depart, args.semaine)
    else:
        date_str = args.date_depart
    
    # Construire le chemin du fichier calendrier
    script_dir = Path(__file__).parent
    fichier_calendrier = script_dir / args.calendrier
    
    if not fichier_calendrier.exists():
        print(f"❌ Erreur : Le fichier {fichier_calendrier} n'existe pas.")
        return 1
    
    # Charger et filtrer les matchs
    print(f"📅 Génération de la feuille de matchs")
    print(f"   Semaine : {args.semaine}")
    print(f"   Date : {date_str}")
    print(f"   Calendrier : {fichier_calendrier}")
    print()
    
    df_matchs = charger_et_filtrer_matchs(fichier_calendrier, args.semaine, date_str)
    
    if df_matchs is None or df_matchs.empty:
        return 1
    
    # Définir le nom du fichier de sortie
    if args.output:
        fichier_sortie = script_dir / args.output
    else:
        date_clean = date_str.replace('/', '-')
        fichier_sortie = script_dir / f"Matchs_Semaine_{args.semaine}_{date_clean}.xlsx"
    
    # Générer le fichier Excel
    generer_feuille_excel(df_matchs, fichier_sortie)
    
    print()
    print(f"📊 Résumé:")
    print(f"   Matchs féminins : {len(df_matchs[df_matchs['Sexe'] == 'F'])}")
    print(f"   Matchs masculins : {len(df_matchs[df_matchs['Sexe'] == 'M'])}")
    print(f"   Lieux uniques : {df_matchs['Lieu'].nunique()}")
    
    return 0


if __name__ == "__main__":
    exit(main())
