#!/usr/bin/env python3
"""
Synchronisation complète des données du Pool Editor vers Excel.

Ce module permet d'actualiser un fichier de configuration Excel à partir
d'un fichier JSON exporté par l'éditeur de poules (pool editor).

Feuilles synchronisées:
1. **Equipes** - Données des équipes:
   - Nom de l'équipe (institution + numéro)
   - Genre (F/M/X)
   - Niveau (A1/A2/A3/A4)
   - Poule
   - Horaire préféré
   - Les données supplémentaires (Responsable_*) sont préservées

2. **Types_Poules** - Types de championnat par poule:
   - Poule (nom de la poule)
   - Type (Classique ou Aller-Retour)

3. **Dispos_Gymnases_Equipes** - Équipes avec horaires aménagés:
   - Equipe (nom de l'équipe)
   - Genre (F/M)
   - Horaire_Dispo (horaire disponible plus tôt)
   - Gymnase_1 à Gymnase_5 (gymnases où l'horaire aménagé s'applique)

Usage:
    from pycalendar.cli.pool_editor_sync import synchroniser_depuis_json
    
    synchroniser_depuis_json(
        json_path="poules_export.json",
        excel_path="data/volleyball/config_volley.xlsx",
        backup=True,
        sync_equipes=True,
        sync_poules=True,
        sync_dispos=True
    )
"""

import json
import logging
from pathlib import Path
from typing import Dict, List, Set, Tuple, Optional, Any, Union
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)


class PoolEditorSyncError(Exception):
    """Exception levée lors d'erreurs de synchronisation."""
    pass


class EquipeData:
    """Représente les données d'une équipe."""
    
    def __init__(self, nom: str, niveau: str, genre: str, poule: Optional[str], 
                 horaire: Optional[str], institution: str,
                 horaire_amenage: Optional[str] = None,
                 gymnases_amenages: Optional[List[str]] = None):
        self.nom = nom  # Format: "LYON 1 (1)"
        self.niveau = niveau  # A1, A2, A3, A4
        self.genre = genre  # F, M, X
        self.poule = poule  # VBFA1PA, etc. (peut être None)
        self.horaire = horaire  # 14H, 16H, 18H, 20H
        self.institution = institution  # LYON 1
        # Nouveaux champs pour horaires aménagés
        self.horaire_amenage = horaire_amenage  # Horaire plus tôt (ex: 18H)
        self.gymnases_amenages = gymnases_amenages or []  # Gymnases où l'équipe peut jouer plus tôt
        
    def __repr__(self):
        amenage_str = f" [Aménagé: {self.horaire_amenage} @ {','.join(self.gymnases_amenages)}]" if self.horaire_amenage else ""
        return f"<EquipeData {self.nom} [{self.genre}] {self.niveau} - {self.poule or 'Non assignée'}{amenage_str}>"
    
    @property
    def has_amenaged_schedule(self) -> bool:
        """Retourne True si l'équipe a un horaire aménagé."""
        return bool(self.horaire_amenage and self.gymnases_amenages)
    
    def to_dict(self) -> Dict:
        """Convertit en dictionnaire pour l'Excel."""
        # Convertir l'horaire du format 14H en 14:00
        horaire_excel = None
        if self.horaire:
            horaire_str = self.horaire.upper()
            if horaire_str.endswith('H'):
                heure = horaire_str[:-1]
                horaire_excel = f"{heure}:00"
            else:
                horaire_excel = self.horaire
        
        return {
            'Equipe': self.nom,
            'Niveau_Equipe': self.niveau,
            'Genre_Equipe': self.genre,
            'Poule': self.poule or '',
            'Horaire_Prefere': horaire_excel or ''
        }


class PouleData:
    """Représente les données d'une poule."""
    
    def __init__(self, nom: str, genre: str, niveau: str, 
                 type_championnat: str = 'classique', lettre: str = 'A'):
        self.nom = nom  # VBFA1PA
        self.genre = genre  # F, M
        self.niveau = niveau  # A1, A2, A3, A4
        self.type_championnat = type_championnat  # 'classique' ou 'aller-retour'
        self.lettre = lettre  # A, B, C, ...
        
    def __repr__(self):
        return f"<PouleData {self.nom} [{self.genre}] {self.niveau} - {self.type_championnat}>"
    
    def to_excel_type(self) -> str:
        """Convertit le type en format Excel."""
        if self.type_championnat.lower() == 'aller-retour':
            return 'Aller-Retour'
        return 'Classique'


class DispoGymnaseData:
    """Représente les données de disponibilité par gymnase d'une équipe."""
    
    def __init__(self, equipe_nom: str, genre: str, horaire_dispo: str, 
                 gymnases: List[str], remarques: str = ''):
        self.equipe_nom = equipe_nom
        self.genre = genre
        self.horaire_dispo = horaire_dispo  # Format: "18:00"
        self.gymnases = gymnases  # Liste de max 5 gymnases
        self.remarques = remarques
        
    def __repr__(self):
        return f"<DispoGymnaseData {self.equipe_nom} [{self.genre}] {self.horaire_dispo} @ {','.join(self.gymnases)}>"
    
    def to_dict(self) -> Dict:
        """Convertit en dictionnaire pour l'Excel."""
        result = {
            'Equipe': self.equipe_nom,
            'Genre': self.genre,
            'Horaire_Dispo': self.horaire_dispo,
            'Remarques': self.remarques
        }
        # Ajouter les gymnases
        for i, gym in enumerate(self.gymnases[:5], start=1):
            result[f'Gymnase_{i}'] = gym
        return result


def charger_equipes_depuis_json(json_path: Path) -> Tuple[List[EquipeData], List[PouleData]]:
    """
    Charge les équipes et les poules depuis un fichier JSON exporté par le pool editor.
    
    Format attendu du JSON:
    {
        "teams": [
            {
                "nom": "LYON 1 (1)",
                "genre": "F",
                "niveau": "A1",
                "horaire": "14H",
                "institution": "LYON 1",
                "poule": "VBFA1PA",
                "horaireAmenage": "18H",  // Optionnel
                "gymnasesAmenages": ["BESSON", "LAENNEC"]  // Optionnel
            },
            ...
        ],
        "pools": [
            {
                "id": "VBFA1PA",
                "name": "VBFA1PA",
                "gender": "F",
                "level": "A1",
                "letter": "A",
                "type": "classique"
            },
            ...
        ],
        "settings": {...}
    }
    
    Args:
        json_path: Chemin vers le fichier JSON
        
    Returns:
        Tuple (équipes, poules)
        
    Raises:
        PoolEditorSyncError: Si le fichier JSON est invalide
    """
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except json.JSONDecodeError as e:
        raise PoolEditorSyncError(f"Fichier JSON invalide: {e}")
    except Exception as e:
        raise PoolEditorSyncError(f"Erreur lors de la lecture du JSON: {e}")
    
    if 'teams' not in data:
        raise PoolEditorSyncError("Le fichier JSON ne contient pas de clé 'teams'")
    
    # Charger les équipes
    equipes = []
    for team_data in data['teams']:
        try:
            equipe = EquipeData(
                nom=team_data['nom'],
                niveau=team_data.get('niveau'),
                genre=team_data.get('genre'),
                poule=team_data.get('poule'),
                horaire=team_data.get('horaire'),
                institution=team_data.get('institution', ''),
                horaire_amenage=team_data.get('horaireAmenage'),
                gymnases_amenages=team_data.get('gymnasesAmenages', [])
            )
            equipes.append(equipe)
        except KeyError as e:
            logger.warning(f"Équipe ignorée (champ manquant {e}): {team_data}")
            continue
    
    # Charger les poules
    poules = []
    if 'pools' in data:
        for pool_data in data['pools']:
            try:
                poule = PouleData(
                    nom=pool_data.get('id') or pool_data.get('name', ''),
                    genre=pool_data.get('gender', 'M'),
                    niveau=pool_data.get('level', 'A1'),
                    type_championnat=pool_data.get('type', 'classique'),
                    lettre=pool_data.get('letter', 'A')
                )
                if poule.nom:  # Ignorer les poules sans nom
                    poules.append(poule)
            except KeyError as e:
                logger.warning(f"Poule ignorée (champ manquant {e}): {pool_data}")
                continue
    
    logger.info(f"✅ {len(equipes)} équipes et {len(poules)} poules chargées depuis {json_path.name}")
    return equipes, poules


def charger_equipes_depuis_excel(excel_path: Path, sheet_name: str = 'Equipes') -> pd.DataFrame:
    """
    Charge la feuille Equipes depuis un fichier Excel.
    
    Args:
        excel_path: Chemin vers le fichier Excel
        sheet_name: Nom de la feuille (par défaut 'Equipes')
        
    Returns:
        DataFrame contenant les équipes existantes
        
    Raises:
        PoolEditorSyncError: Si le fichier ou la feuille n'existe pas
    """
    if not excel_path.exists():
        raise PoolEditorSyncError(f"Fichier Excel introuvable: {excel_path}")
    
    try:
        df = pd.read_excel(excel_path, sheet_name=sheet_name)
    except ValueError:
        raise PoolEditorSyncError(f"Feuille '{sheet_name}' introuvable dans {excel_path.name}")
    except Exception as e:
        raise PoolEditorSyncError(f"Erreur lors de la lecture de l'Excel: {e}")
    
    logger.info(f"✅ {len(df)} équipes existantes dans {excel_path.name}")
    return df


def comparer_equipes(
    equipes_json: List[EquipeData],
    df_excel: pd.DataFrame
) -> Tuple[List[EquipeData], List[str], List[Tuple[str, EquipeData]]]:
    """
    Compare les équipes du JSON avec celles de l'Excel.
    
    Args:
        equipes_json: Équipes depuis le JSON
        df_excel: DataFrame des équipes existantes
        
    Returns:
        Tuple (à_ajouter, à_supprimer, à_modifier)
        - à_ajouter: Liste des nouvelles équipes
        - à_supprimer: Liste des noms d'équipes à supprimer
        - à_modifier: Liste de tuples (nom_equipe, nouvelles_donnees)
    """
    # Construire un index des équipes JSON par nom
    equipes_json_dict = {eq.nom: eq for eq in equipes_json}
    noms_json = set(equipes_json_dict.keys())
    
    # Extraire les noms d'équipes existantes dans l'Excel
    if 'Equipe' not in df_excel.columns:
        raise PoolEditorSyncError("La feuille Equipes ne contient pas de colonne 'Equipe'")
    
    noms_excel = set(df_excel['Equipe'].dropna().astype(str).str.strip())
    
    # Déterminer les actions
    a_ajouter = [equipes_json_dict[nom] for nom in (noms_json - noms_excel)]
    a_supprimer = list(noms_excel - noms_json)
    
    # Pour les équipes communes, vérifier si des modifications sont nécessaires
    a_modifier = []
    for nom in (noms_json & noms_excel):
        equipe_json = equipes_json_dict[nom]
        # Récupérer la ligne correspondante dans l'Excel
        ligne_excel = df_excel[df_excel['Equipe'] == nom].iloc[0]
        
        # Vérifier si des champs ont changé
        modifications_necessaires = False
        
        # Comparer les champs synchronisables
        if pd.notna(ligne_excel.get('Niveau_Equipe')):
            if str(ligne_excel['Niveau_Equipe']).strip() != equipe_json.niveau:
                modifications_necessaires = True
        
        if pd.notna(ligne_excel.get('Genre_Equipe')):
            if str(ligne_excel['Genre_Equipe']).strip() != equipe_json.genre:
                modifications_necessaires = True
        
        if pd.notna(ligne_excel.get('Poule')):
            poule_excel = str(ligne_excel['Poule']).strip() if ligne_excel['Poule'] else ''
            poule_json = equipe_json.poule or ''
            if poule_excel != poule_json:
                modifications_necessaires = True
        
        if pd.notna(ligne_excel.get('Horaire_Prefere')):
            horaire_excel = str(ligne_excel['Horaire_Prefere']).strip()
            horaire_json_excel = equipe_json.to_dict()['Horaire_Prefere']
            if horaire_excel != horaire_json_excel:
                modifications_necessaires = True
        
        if modifications_necessaires:
            a_modifier.append((nom, equipe_json))
    
    logger.info(f"📊 Analyse: {len(a_ajouter)} à ajouter, {len(a_supprimer)} à supprimer, {len(a_modifier)} à modifier")
    return a_ajouter, a_supprimer, a_modifier


# ============================================================================
# SYNCHRONISATION DES ÉQUIPES
# ============================================================================

def synchroniser_equipes(
    equipes_json: List[EquipeData],
    excel_path: Path,
    wb: openpyxl.Workbook,
    mode: str = 'replace'
) -> Dict[str, int]:
    """
    Synchronise la feuille Equipes avec les données du JSON.
    
    Args:
        equipes_json: Liste des équipes depuis le JSON
        excel_path: Chemin du fichier Excel
        wb: Workbook openpyxl
        mode: 'replace' ou 'update'
        
    Returns:
        Statistiques {ajoutees, modifiees, supprimees, conservees}
    """
    sheet_name = 'Equipes'
    
    if sheet_name not in wb.sheetnames:
        raise PoolEditorSyncError(f"Feuille '{sheet_name}' introuvable")
    
    ws = wb[sheet_name]
    
    # Obtenir les en-têtes (filtrés et convertis en str)
    headers = [str(cell.value) for cell in ws[1] if cell.value is not None]
    
    # Vérifier les colonnes requises
    colonnes_requises = {'Equipe', 'Niveau_Equipe', 'Genre_Equipe', 'Poule', 'Horaire_Prefere'}
    colonnes_manquantes = colonnes_requises - set(headers)
    if colonnes_manquantes:
        raise PoolEditorSyncError(
            f"Colonnes manquantes dans la feuille Equipes: {', '.join(colonnes_manquantes)}"
        )
    
    # Créer un index des colonnes
    col_index = {header: idx for idx, header in enumerate(headers, start=1) if header}
    
    # Créer un dictionnaire des équipes existantes avec leurs données supplémentaires
    equipes_existantes = {}
    for row_idx in range(2, ws.max_row + 1):
        nom_equipe = ws.cell(row=row_idx, column=col_index['Equipe']).value
        if nom_equipe:
            nom_equipe = str(nom_equipe).strip()
            donnees_supp = {}
            for header, col_idx in col_index.items():
                if header not in colonnes_requises:
                    valeur = ws.cell(row=row_idx, column=col_idx).value
                    if valeur is not None:
                        donnees_supp[header] = valeur
            equipes_existantes[nom_equipe] = donnees_supp
    
    # Statistiques
    ajoutees = 0
    modifiees = 0
    supprimees = 0
    conservees = 0
    
    if mode == 'update':
        noms_json = {eq.nom for eq in equipes_json}
        noms_excel = set(equipes_existantes.keys())
        conservees = len(noms_excel - noms_json)
        
        for equipe_json in equipes_json:
            ligne_trouvee = False
            for row_idx in range(2, ws.max_row + 1):
                nom_cell = ws.cell(row=row_idx, column=col_index['Equipe']).value
                if nom_cell and str(nom_cell).strip() == equipe_json.nom:
                    donnees_dict = equipe_json.to_dict()
                    for col_name, valeur in donnees_dict.items():
                        if col_name in col_index:
                            ws.cell(row=row_idx, column=col_index[col_name]).value = valeur
                    modifiees += 1
                    ligne_trouvee = True
                    break
            
            if not ligne_trouvee:
                new_row = ws.max_row + 1
                donnees_dict = equipe_json.to_dict()
                for col_name, valeur in donnees_dict.items():
                    if col_name in col_index:
                        ws.cell(row=new_row, column=col_index[col_name]).value = valeur
                ajoutees += 1
    
    else:  # mode 'replace'
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
        
        for equipe_json in equipes_json:
            new_row = ws.max_row + 1
            donnees_dict = equipe_json.to_dict()
            
            for col_name, valeur in donnees_dict.items():
                if col_name in col_index:
                    ws.cell(row=new_row, column=col_index[col_name]).value = valeur
            
            if equipe_json.nom in equipes_existantes:
                donnees_supp = equipes_existantes[equipe_json.nom]
                for col_name, valeur in donnees_supp.items():
                    if col_name in col_index:
                        ws.cell(row=new_row, column=col_index[col_name]).value = valeur
                modifiees += 1
            else:
                ajoutees += 1
        
        noms_json = {eq.nom for eq in equipes_json}
        supprimees = len(set(equipes_existantes.keys()) - noms_json)
    
    logger.info(f"   📋 Equipes: +{ajoutees} ✏️{modifiees} 🗑️{supprimees}")
    
    return {
        'ajoutees': ajoutees,
        'modifiees': modifiees,
        'supprimees': supprimees,
        'conservees': conservees
    }


# ============================================================================
# SYNCHRONISATION DES TYPES DE POULES
# ============================================================================

def synchroniser_types_poules(
    poules_json: List[PouleData],
    excel_path: Path,
    wb: openpyxl.Workbook
) -> Dict[str, int]:
    """
    Synchronise la feuille Types_Poules avec les données du JSON.
    
    Args:
        poules_json: Liste des poules depuis le JSON
        excel_path: Chemin du fichier Excel
        wb: Workbook openpyxl
        
    Returns:
        Statistiques {ajoutees, modifiees, supprimees}
    """
    sheet_name = 'Types_Poules'
    
    if sheet_name not in wb.sheetnames:
        logger.warning(f"⚠️ Feuille '{sheet_name}' introuvable - création automatique")
        ws = wb.create_sheet(sheet_name)
        # Créer les en-têtes
        ws.cell(row=1, column=1, value='Poule')
        ws.cell(row=1, column=2, value='Type')
        ws.cell(row=1, column=3, value='Remarques')
    else:
        ws = wb[sheet_name]
    
    # Obtenir les en-têtes
    headers = [cell.value for cell in ws[1]]
    col_index = {header: idx for idx, header in enumerate(headers, start=1) if header}
    
    if 'Poule' not in col_index or 'Type' not in col_index:
        raise PoolEditorSyncError("La feuille Types_Poules doit contenir les colonnes 'Poule' et 'Type'")
    
    # Charger les types existants avec leurs remarques
    types_existants = {}
    for row_idx in range(2, ws.max_row + 1):
        nom_poule = ws.cell(row=row_idx, column=col_index['Poule']).value
        if nom_poule:
            nom_poule = str(nom_poule).strip()
            remarques = ''
            if 'Remarques' in col_index:
                val = ws.cell(row=row_idx, column=col_index['Remarques']).value
                remarques = str(val) if val else ''
            types_existants[nom_poule] = remarques
    
    # Supprimer toutes les lignes de données
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    
    # Statistiques
    ajoutees = 0
    modifiees = 0
    supprimees = 0
    
    noms_json = set()
    
    # Insérer les poules du JSON
    for poule in poules_json:
        noms_json.add(poule.nom)
        new_row = ws.max_row + 1
        
        ws.cell(row=new_row, column=col_index['Poule']).value = poule.nom
        ws.cell(row=new_row, column=col_index['Type']).value = poule.to_excel_type()
        
        # Restaurer les remarques si existantes
        if 'Remarques' in col_index and poule.nom in types_existants:
            ws.cell(row=new_row, column=col_index['Remarques']).value = types_existants[poule.nom]
        
        if poule.nom in types_existants:
            modifiees += 1
        else:
            ajoutees += 1
    
    supprimees = len(set(types_existants.keys()) - noms_json)
    
    logger.info(f"   🏆 Types_Poules: +{ajoutees} ✏️{modifiees} 🗑️{supprimees}")
    
    return {
        'ajoutees': ajoutees,
        'modifiees': modifiees,
        'supprimees': supprimees
    }


# ============================================================================
# SYNCHRONISATION DES DISPONIBILITÉS GYMNASES
# ============================================================================

def synchroniser_dispos_gymnases(
    equipes_json: List[EquipeData],
    excel_path: Path,
    wb: openpyxl.Workbook
) -> Dict[str, int]:
    """
    Synchronise la feuille Dispos_Gymnases_Equipes avec les données du JSON.
    
    Seules les équipes ayant un horaire aménagé sont incluses.
    
    Args:
        equipes_json: Liste des équipes depuis le JSON
        excel_path: Chemin du fichier Excel
        wb: Workbook openpyxl
        
    Returns:
        Statistiques {ajoutees, modifiees, supprimees}
    """
    sheet_name = 'Dispos_Gymnases_Equipes'
    
    if sheet_name not in wb.sheetnames:
        logger.warning(f"⚠️ Feuille '{sheet_name}' introuvable - création automatique")
        ws = wb.create_sheet(sheet_name)
        # Créer les en-têtes
        headers = ['Equipe', 'Genre', 'Horaire_Dispo', 'Gymnase_1', 'Gymnase_2', 
                   'Gymnase_3', 'Gymnase_4', 'Gymnase_5', 'Remarques']
        for idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=idx, value=header)
    else:
        ws = wb[sheet_name]
    
    # Obtenir les en-têtes (filtrés et convertis en str)
    raw_headers = [cell.value for cell in ws[1]]
    col_index = {str(h): idx for idx, h in enumerate(raw_headers, start=1) if h is not None}
    
    colonnes_requises = {'Equipe', 'Genre', 'Horaire_Dispo'}
    colonnes_manquantes = colonnes_requises - set(col_index.keys())
    if colonnes_manquantes:
        raise PoolEditorSyncError(
            f"Colonnes manquantes dans Dispos_Gymnases_Equipes: {', '.join(colonnes_manquantes)}"
        )
    
    # Charger les remarques existantes
    remarques_existantes = {}
    for row_idx in range(2, ws.max_row + 1):
        nom_equipe = ws.cell(row=row_idx, column=col_index['Equipe']).value
        genre = ws.cell(row=row_idx, column=col_index['Genre']).value
        if nom_equipe:
            cle = f"{str(nom_equipe).strip()}|{str(genre or '').strip()}"
            if 'Remarques' in col_index:
                val = ws.cell(row=row_idx, column=col_index['Remarques']).value
                remarques_existantes[cle] = str(val) if val else ''
    
    # Supprimer toutes les lignes de données
    nb_existantes = len(remarques_existantes)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    
    # Filtrer les équipes avec horaire aménagé
    equipes_amenagees = [eq for eq in equipes_json if eq.has_amenaged_schedule]
    
    # Statistiques
    ajoutees = 0
    modifiees = 0
    
    cles_json = set()
    
    for equipe in equipes_amenagees:
        cle = f"{equipe.nom}|{equipe.genre or ''}"
        cles_json.add(cle)
        
        new_row = ws.max_row + 1
        
        # Convertir l'horaire aménagé au format Excel (18H -> 18:00)
        horaire_excel = equipe.horaire_amenage
        if horaire_excel and horaire_excel.upper().endswith('H'):
            heure = horaire_excel[:-1]
            horaire_excel = f"{heure}:00"
        
        ws.cell(row=new_row, column=col_index['Equipe']).value = equipe.nom
        ws.cell(row=new_row, column=col_index['Genre']).value = equipe.genre
        ws.cell(row=new_row, column=col_index['Horaire_Dispo']).value = horaire_excel
        
        # Ajouter les gymnases
        for i, gym in enumerate(equipe.gymnases_amenages[:5], start=1):
            col_name = f'Gymnase_{i}'
            if col_name in col_index:
                ws.cell(row=new_row, column=col_index[col_name]).value = gym
        
        # Restaurer les remarques si existantes
        if 'Remarques' in col_index and cle in remarques_existantes:
            ws.cell(row=new_row, column=col_index['Remarques']).value = remarques_existantes[cle]
        
        if cle in remarques_existantes:
            modifiees += 1
        else:
            ajoutees += 1
    
    supprimees = nb_existantes - modifiees
    
    logger.info(f"   ⏰ Dispos_Gymnases: +{ajoutees} ✏️{modifiees} 🗑️{supprimees}")
    
    return {
        'ajoutees': ajoutees,
        'modifiees': modifiees,
        'supprimees': max(0, supprimees)
    }


# ============================================================================
# FONCTION PRINCIPALE DE SYNCHRONISATION
# ============================================================================

def synchroniser_depuis_json(
    json_path: Union[str, Path],
    excel_path: Union[str, Path],
    backup: bool = True,
    mode: str = 'replace',
    sync_equipes: bool = True,
    sync_poules: bool = True,
    sync_dispos: bool = True
) -> Dict[str, Any]:
    """
    Synchronise un fichier Excel avec les données d'un JSON du Pool Editor.
    
    Cette fonction est le point d'entrée principal pour la synchronisation complète.
    Elle gère:
    1. Feuille Equipes - Données des équipes
    2. Feuille Types_Poules - Types de championnat par poule
    3. Feuille Dispos_Gymnases_Equipes - Équipes avec horaires aménagés
    
    Args:
        json_path: Chemin vers le fichier JSON exporté par le pool editor
        excel_path: Chemin vers le fichier Excel de configuration
        backup: Si True, crée une sauvegarde avant modification
        mode: Mode de synchronisation ('replace' ou 'update')
        sync_equipes: Synchroniser la feuille Equipes
        sync_poules: Synchroniser la feuille Types_Poules
        sync_dispos: Synchroniser la feuille Dispos_Gymnases_Equipes
            
    Returns:
        Dictionnaire avec les statistiques de synchronisation:
        {
            'equipes': {ajoutees, modifiees, supprimees, conservees},
            'poules': {ajoutees, modifiees, supprimees},
            'dispos': {ajoutees, modifiees, supprimees},
            'backup_path': str (si backup=True)
        }
        
    Raises:
        PoolEditorSyncError: En cas d'erreur lors de la synchronisation
    """
    json_file = Path(json_path)
    excel_file = Path(excel_path)
    
    if mode == 'sync':
        mode = 'replace'
    
    logger.info(f"🔄 Synchronisation complète {excel_file.name} depuis {json_file.name}")
    logger.info(f"   Mode: {mode.upper()}")
    
    # Créer une sauvegarde si demandé
    backup_path = None
    if backup:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_path = excel_file.parent / f"{excel_file.stem}.backup_{timestamp}{excel_file.suffix}"
        import shutil
        shutil.copy2(excel_file, backup_path)
        logger.info(f"💾 Sauvegarde créée: {backup_path.name}")
    
    # Charger les données du JSON
    equipes_json, poules_json = charger_equipes_depuis_json(json_file)
    
    # Charger le workbook
    wb = openpyxl.load_workbook(excel_file)
    
    stats = {
        'equipes': {'ajoutees': 0, 'modifiees': 0, 'supprimees': 0, 'conservees': 0},
        'poules': {'ajoutees': 0, 'modifiees': 0, 'supprimees': 0},
        'dispos': {'ajoutees': 0, 'modifiees': 0, 'supprimees': 0},
        'backup_path': str(backup_path) if backup_path else None
    }
    
    # Synchroniser les équipes
    if sync_equipes:
        stats['equipes'] = synchroniser_equipes(equipes_json, excel_file, wb, mode)
    
    # Synchroniser les types de poules
    if sync_poules and poules_json:
        stats['poules'] = synchroniser_types_poules(poules_json, excel_file, wb)
    
    # Synchroniser les disponibilités gymnases
    if sync_dispos:
        stats['dispos'] = synchroniser_dispos_gymnases(equipes_json, excel_file, wb)
    
    # Sauvegarder
    try:
        wb.save(excel_file)
        logger.info(f"✅ Fichier Excel mis à jour: {excel_file.name}")
    except Exception as e:
        raise PoolEditorSyncError(f"Erreur lors de la sauvegarde: {e}")
    
    return stats


# Alias pour rétrocompatibilité
def synchroniser_equipes_depuis_json(
    json_path: str,
    excel_path: str,
    sheet_name: str = 'Equipes',
    backup: bool = True,
    mode: str = 'replace'
) -> Dict[str, Any]:
    """
    Alias de synchroniser_depuis_json pour rétrocompatibilité.
    
    DÉPRÉCIÉ: Utilisez synchroniser_depuis_json() à la place.
    """
    logger.warning("⚠️ synchroniser_equipes_depuis_json est déprécié, utilisez synchroniser_depuis_json")
    
    stats = synchroniser_depuis_json(
        json_path=json_path,
        excel_path=excel_path,
        backup=backup,
        mode=mode,
        sync_equipes=True,
        sync_poules=True,
        sync_dispos=True
    )
    
    # Retourner au format ancien pour compatibilité
    return {
        'ajoutees': stats['equipes']['ajoutees'],
        'modifiees': stats['equipes']['modifiees'],
        'supprimees': stats['equipes']['supprimees'],
        'conservees': stats['equipes']['conservees'],
        'backup_path': stats['backup_path']
    }


def afficher_rapport(stats: Dict[str, Any], complet: bool = True):
    """
    Affiche un rapport de synchronisation.
    
    Args:
        stats: Statistiques de synchronisation
        complet: Si True, affiche le rapport complet (nouveau format)
    """
    print("\n" + "="*70)
    print("📊 RAPPORT DE SYNCHRONISATION")
    print("="*70)
    
    # Détecter l'ancien format vs nouveau format
    if 'equipes' in stats:
        # Nouveau format
        eq = stats['equipes']
        print(f"\n📋 ÉQUIPES")
        print(f"   ➕ Ajoutées     : {eq['ajoutees']}")
        print(f"   ✏️  Modifiées    : {eq['modifiees']}")
        print(f"   🗑️  Supprimées  : {eq['supprimees']}")
        if eq.get('conservees', 0) > 0:
            print(f"   ℹ️  Conservées   : {eq['conservees']}")
        
        if 'poules' in stats and stats['poules']:
            po = stats['poules']
            total_po = po['ajoutees'] + po['modifiees'] + po['supprimees']
            if total_po > 0:
                print(f"\n🏆 TYPES DE POULES")
                print(f"   ➕ Ajoutées     : {po['ajoutees']}")
                print(f"   ✏️  Modifiées    : {po['modifiees']}")
                print(f"   🗑️  Supprimées  : {po['supprimees']}")
        
        if 'dispos' in stats and stats['dispos']:
            di = stats['dispos']
            total_di = di['ajoutees'] + di['modifiees'] + di['supprimees']
            if total_di > 0 or di['ajoutees'] > 0:
                print(f"\n⏰ DISPONIBILITÉS GYMNASES (horaires aménagés)")
                print(f"   ➕ Ajoutées     : {di['ajoutees']}")
                print(f"   ✏️  Modifiées    : {di['modifiees']}")
                print(f"   🗑️  Supprimées  : {di['supprimees']}")
    else:
        # Ancien format (rétrocompatibilité)
        print(f"✅ Équipes ajoutées     : {stats.get('ajoutees', 0)}")
        print(f"✏️  Équipes modifiées    : {stats.get('modifiees', 0)}")
        print(f"🗑️  Équipes supprimées  : {stats.get('supprimees', 0)}")
        
        if stats.get('conservees', 0) > 0:
            print(f"ℹ️  Équipes conservées   : {stats['conservees']}")
    
    if stats.get('backup_path'):
        print(f"\n💾 Sauvegarde créée: {Path(stats['backup_path']).name}")
    
    # Calculer le total des changements
    if 'equipes' in stats:
        total = (stats['equipes']['ajoutees'] + stats['equipes']['modifiees'] + 
                 stats['equipes']['supprimees'])
        total += stats.get('poules', {}).get('ajoutees', 0) + stats.get('poules', {}).get('modifiees', 0)
        total += stats.get('dispos', {}).get('ajoutees', 0) + stats.get('dispos', {}).get('modifiees', 0)
    else:
        total = stats.get('ajoutees', 0) + stats.get('modifiees', 0) + stats.get('supprimees', 0)
    
    if total == 0:
        print("\n✨ Aucune modification nécessaire - Les données sont déjà synchronisées")
    
    print("="*70 + "\n")


if __name__ == '__main__':
    # Configuration du logging
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )
    
    # Test
    print("🧪 Module de synchronisation Pool Editor → Excel")
    print("   Utilisez scripts/update_teams_from_pool_editor.py pour l'utilisation interactive")
