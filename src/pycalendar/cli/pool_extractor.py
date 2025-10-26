#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script d'extraction des poules depuis un fichier Excel et génération d'un fichier 
de configuration compatible avec l'algorithme de calendrier PyCalendar.

USAGE:
    # Volleyball (défaut)
    python extract_poules.py
    python extract_poules.py --sport VB
    
    # Handball
    python extract_poules.py --sport VB --input exemple/poules.xlsx
    python extract_poules.py --sport HB --input exemple/poules.xlsx
    
    # Personnalisé
    python extract_poules.py -i mon_fichier.xlsx -o ma_config.xlsx -s VB

FORMAT D'ENTRÉE:
    Fichier Excel avec poules disposées en colonnes, format attendu:
    - Chaque poule identifiée par un code (ex: VBFA1PA, HBMA2PB)
    - Les équipes listées sous chaque poule avec leur numéro
    - Optionnel: horaires préférés dans la colonne suivante

FORMAT DE SORTIE:
    Fichier Excel multi-feuilles prêt pour PyCalendar:
    - Equipes: [Equipe, Poule, Horaire_Prefere]
    - Gymnases: [Gymnase, Adresse, Capacite, Creneaux] (à remplir)
    - Indispos_Gymnases: [Gymnase, Semaine, Horaire_Debut, Horaire_Fin, Remarques]
    - Indispos_Equipes: [Equipe, Semaine, Horaire_Debut, Horaire_Fin, Remarques]
    - Indispos_Institutions: [Institution, Semaine, Horaire_Debut, Horaire_Fin, Remarques]
    - Preferences_Gymnases: [Institution, Gymnase, Rang_Preference, Remarques]
    - Obligation_Presence: [Gymnase, Institution_Obligatoire, Remarques]
    - Groupes_Non_Simultaneite: [Nom_Groupe, Entites, Remarques]

APRÈS GÉNÉRATION:
    1. Ouvrez le fichier généré
    2. Complétez la feuille 'Gymnases' (obligatoire)
    3. Complétez les autres feuilles selon vos besoins (optionnel)
    4. Utilisez actualiser_config.py pour valider la configuration
    5. Lancez la planification avec main.py
"""

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path
import re
import sys
from typing import List, Dict, Tuple, Optional

# Ajouter le répertoire courant au path pour importer core.utils
sys.path.insert(0, str(Path(__file__).parent))

try:
    from pycalendar.core.utils import extraire_genre_depuis_poule
except ImportError:
    # Fonction de secours si l'import échoue
    def extraire_genre_depuis_poule(nom_poule: str) -> str:
        """Extrait le genre depuis le code de la poule."""
        if not nom_poule:
            return ''
        nom_poule = nom_poule.strip().upper()
        match = re.match(r'^[A-Z]{2}([FM]).*$', nom_poule)
        return match.group(1) if match else ''


def normalize_team_name(team_name: str) -> str:
    """
    Normalise le nom d'une équipe au format 'Institution (numéro)'.
    
    Gère différents formats :
    - 'INSA 1' -> 'INSA (1)'
    - 'LYON 1(11)' -> 'LYON 1 (11)' (sans espace avant parenthèse)
    - 'LYON 1 - 1' -> 'LYON 1 (1)'
    - 'LYON 1 APS 1' -> 'LYON 1 APS (1)'
    - 'INSA-4' -> 'INSA (4)'
    - 'EML-2' -> 'EML (2)'
    - 'LYON 2 IEP' -> 'LYON 2 IEP (1)'
    - 'ECAM' -> 'ECAM (1)'
    - 'ESA (2)' -> 'ESA (2)' (déjà au bon format)
    
    Args:
        team_name: Nom de l'équipe à normaliser
        
    Returns:
        Nom normalisé au format 'Institution (numéro)'
    """
    team_name = team_name.strip()
    
    # Cas 1 : Déjà au bon format avec parenthèses "Institution (numéro)"
    # Mais on normalise l'espace avant la parenthèse
    match = re.match(r'^(.+?)\s*\((\d+)\)$', team_name)
    if match:
        institution = match.group(1).strip()
        numero = match.group(2)
        return f"{institution} ({numero})"
    
    # Cas 2 : Format avec tiret "Institution - numéro" (avec ou sans espaces)
    match = re.match(r'^(.+?)\s*-\s*(\d+)$', team_name)
    if match:
        institution = match.group(1).strip()
        numero = match.group(2)
        return f"{institution} ({numero})"
    
    # Cas 3 : Numéro à la fin après un espace "Institution numéro"
    # On doit être prudent pour ne pas confondre avec des noms comme "LYON 1" où "1" fait partie du nom
    # On cherche un pattern où le dernier mot est un seul chiffre
    match = re.match(r'^(.+)\s+(\d+)$', team_name)
    if match:
        potential_institution = match.group(1).strip()
        numero = match.group(2)
        
        # Vérifier si c'est vraiment un numéro d'équipe ou partie du nom
        # Si l'institution contient déjà un chiffre en début (LYON 1, LYON 2, etc.)
        # et qu'il y a encore du texte après, c'est probablement un numéro d'équipe
        # Exemples: "LYON 1 APS 1" -> "LYON 1 APS" est l'institution, "1" est le numéro
        #           "INSA 1" -> "INSA" est l'institution, "1" est le numéro
        #           "LYON 2 IEP" -> "LYON 2 IEP" est l'institution, on attribuera (1) plus tard
        
        # Si le dernier mot avant le numéro est une suite de lettres majuscules (acronyme)
        # ou si l'institution contient déjà un chiffre, c'est probablement un numéro d'équipe
        words = potential_institution.split()
        if len(words) > 0:
            last_word = words[-1]
            # Si le dernier mot est un acronyme (lettres majuscules) ou "APS", c'est un numéro d'équipe
            if last_word.isupper() or potential_institution.count(' ') >= 1:
                return f"{potential_institution} ({numero})"
            # Si l'institution est simple (un seul mot), c'est un numéro d'équipe
            elif len(words) == 1:
                return f"{potential_institution} ({numero})"
    
    # Cas 4 : Pas de numéro détecté, attribuer (1) par défaut
    # Car c'est probablement la première équipe de cette institution
    return f"{team_name} (1)"


def extract_poules_from_excel(input_file: str, output_file: str, sport_prefix: str = 'HB'):
    """
    Extrait les poules depuis un fichier Excel et génère un fichier de configuration multi-feuilles.
    
    Args:
        input_file: Chemin vers le fichier Excel d'entrée
        output_file: Chemin vers le fichier Excel de sortie
        sport_prefix: Préfixe du sport (HB pour handball, VB pour volleyball, etc.)
    """
    print(f"📖 Lecture du fichier: {input_file}")
    print(f"🏐 Sport: {sport_prefix}")
    
    # Lire le fichier Excel sans header
    df = pd.read_excel(input_file, sheet_name=0, header=None)
    
    # Liste pour stocker toutes les équipes extraites
    all_teams = []
    
    # D'abord, trouver toutes les poules dans le fichier
    poules_found = []
    # Pattern modifié pour accepter toutes les lettres de poule (A, B, C, D, E, F, etc.)
    pattern = rf'^{sport_prefix}[FM]A\d+P[A-Z]$'
    
    for col_idx in range(len(df.columns)):
        for row_idx in range(len(df)):
            cell_value = df.iloc[row_idx, col_idx]
            
            # Ignorer les valeurs NaN
            if pd.isna(cell_value):
                continue
            
            cell_str = str(cell_value).strip()
            
            # Chercher un pattern de poule (ex: HBFA1PA, VBMA2PB, VBFA3PC, etc.)
            if re.match(pattern, cell_str):
                poules_found.append({
                    'name': cell_str,
                    'col': col_idx,
                    'row': row_idx
                })
    
    print(f"🔍 Nombre de poules trouvées: {len(poules_found)}")
    
    if not poules_found:
        print(f"❌ Aucune poule trouvée avec le pattern {pattern}")
        print("💡 Vérifie le préfixe du sport (HB, VB, etc.)")
        return False
    
    # Extraire les équipes pour chaque poule trouvée
    for poule_info in poules_found:
        poule_name = poule_info['name']
        col_idx = poule_info['col']
        poule_row = poule_info['row']
        
        print(f"\n✓ Poule: {poule_name} (colonne {col_idx}, ligne {poule_row})")
        
        # Extraire les équipes de cette poule
        # Les équipes commencent à la ligne suivante
        row = poule_row + 1
        while row < len(df):
            # Le numéro d'équipe est dans la colonne précédente (col_idx - 1)
            # sauf si on est en colonne 0 où le numéro n'existe pas
            num_col = max(0, col_idx - 1)
            
            # Vérifier si c'est une ligne d'équipe (commence par un numéro dans la première colonne du groupe)
            num_cell = df.iloc[row, num_col]
            
            if pd.isna(num_cell):
                break
            
            try:
                # Essayer de convertir en entier
                num_str = str(num_cell)
                num = int(float(num_str))
            except (ValueError, TypeError):
                # Si ce n'est pas un numéro, on arrête
                break
            
            # Récupérer le nom de l'équipe (dans la colonne de la poule)
            team_name = df.iloc[row, col_idx]
            
            if pd.notna(team_name):
                team_name = str(team_name).strip()
                
                # Ignorer les lignes qui ne sont pas des noms d'équipe
                if team_name in ['CHAMPIONNAT', 'CHAMPIONNAT A/R']:
                    row += 1
                    continue
                
                # Normaliser le nom de l'équipe au format 'Institution (numéro)'
                team_name_normalized = normalize_team_name(team_name)
                
                # Extraire le genre depuis le nom de la poule (pour info seulement)
                genre = extraire_genre_depuis_poule(poule_name)
                
                # Récupérer l'horaire si disponible (colonne suivante)
                horaire = None
                if col_idx + 1 < len(df.columns):
                    horaire_cell = df.iloc[row, col_idx + 1]
                    if pd.notna(horaire_cell):
                        # Convertir l'horaire en format texte
                        if isinstance(horaire_cell, pd.Timestamp):
                            horaire = horaire_cell.strftime('%H:%M')
                        elif isinstance(horaire_cell, str):
                            # Nettoyer les formats 14H, 16H, etc.
                            horaire_clean = horaire_cell.strip().upper().replace('H', ':00')
                            # Ajouter :00 si seulement l'heure
                            if ':' not in horaire_clean:
                                horaire_clean = horaire_clean + ':00'
                            horaire = horaire_clean
                        else:
                            try:
                                # Essayer de convertir en timestamp
                                horaire = pd.to_datetime(str(horaire_cell)).strftime('%H:%M')
                            except:
                                horaire = str(horaire_cell)
                
                # Ajouter l'équipe à la liste
                all_teams.append({
                    'Equipe': team_name_normalized,
                    'Poule': poule_name,
                    'Horaire_Prefere': horaire if horaire else None
                })
                
                # Afficher avec indication du genre (info) et si le nom a été modifié
                genre_str = f" (genre: {genre})" if genre else ""
                if team_name != team_name_normalized:
                    print(f"  → Équipe {num}: {team_name} → {team_name_normalized}{genre_str} (Horaire: {horaire})")
                else:
                    print(f"  → Équipe {num}: {team_name_normalized}{genre_str} (Horaire: {horaire})")
            
            row += 1
    
    # Créer le DataFrame de sortie
    if not all_teams:
        print("❌ Aucune équipe trouvée dans le fichier!")
        return False
    
    df_equipes = pd.DataFrame(all_teams)
    
    # Réorganiser les colonnes dans le bon ordre
    df_equipes = df_equipes[['Equipe', 'Poule', 'Horaire_Prefere']]
    
    # Créer le dossier de sortie s'il n'existe pas
    output_path = Path(output_file)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Extraire les institutions uniques depuis les noms d'équipes
    institutions = set()
    for equipe in df_equipes['Equipe']:
        # Extraire l'institution du format "Institution (numéro)"
        match = re.match(r'^(.+?)\s*\(\d+\)$', equipe)
        if match:
            institutions.add(match.group(1).strip())
    
    print(f"\n📊 {len(institutions)} institutions détectées: {', '.join(sorted(institutions))}")
    
    # Créer les feuilles vides avec la structure correcte
    sheets_data = create_empty_config_sheets(list(institutions))
    sheets_data['Equipes'] = df_equipes
    
    # Créer le fichier Excel avec toutes les feuilles
    _write_multisheet_excel(output_file, sheets_data)
    
    print(f"\n✅ Extraction terminée!")
    print(f"📊 Nombre total d'équipes extraites: {len(all_teams)}")
    print(f"📁 Fichier de sortie: {output_file}")
    
    # Afficher un résumé par poule
    print("\n📋 Résumé par poule:")
    poule_counts = df_equipes['Poule'].value_counts().sort_index()
    for poule, count in poule_counts.items():
        print(f"  • {poule}: {count} équipes")
    
    print("\n📝 Feuilles créées dans le fichier Excel:")
    for sheet_name in sheets_data.keys():
        print(f"  • {sheet_name}")
    
    return True


def create_empty_config_sheets(institutions: List[str]) -> Dict[str, pd.DataFrame]:
    """
    Crée les feuilles vides avec la structure de configuration requise.
    
    Args:
        institutions: Liste des institutions détectées
        
    Returns:
        Dictionnaire {nom_feuille: DataFrame}
    """
    sheets = {}
    
    # Feuille Gymnases (vide - à remplir manuellement)
    sheets['Gymnases'] = pd.DataFrame(columns=['Gymnase', 'Adresse', 'Capacite', 'Creneaux'])
    
    # Feuille Indispos_Gymnases (vide)
    sheets['Indispos_Gymnases'] = pd.DataFrame(
        columns=['Gymnase', 'Semaine', 'Horaire_Debut', 'Horaire_Fin', 'Remarques']
    )
    
    # Feuille Indispos_Equipes (vide)
    sheets['Indispos_Equipes'] = pd.DataFrame(
        columns=['Equipe', 'Semaine', 'Horaire_Debut', 'Horaire_Fin', 'Remarques']
    )
    
    # Feuille Indispos_Institutions (vide)
    sheets['Indispos_Institutions'] = pd.DataFrame(
        columns=['Institution', 'Semaine', 'Horaire_Debut', 'Horaire_Fin', 'Remarques']
    )
    
    # Feuille Preferences_Gymnases (vide)
    sheets['Preferences_Gymnases'] = pd.DataFrame(
        columns=['Institution', 'Gymnase', 'Rang_Preference', 'Remarques']
    )
    
    # Feuille Obligation_Presence (vide)
    sheets['Obligation_Presence'] = pd.DataFrame(
        columns=['Gymnase', 'Institution_Obligatoire', 'Remarques']
    )
    
    # Feuille Groupes_Non_Simultaneite (vide)
    sheets['Groupes_Non_Simultaneite'] = pd.DataFrame(
        columns=['Nom_Groupe', 'Entites', 'Remarques']
    )
    
    return sheets


def _write_multisheet_excel(output_file: str, sheets_data: Dict[str, pd.DataFrame]):
    """
    Écrit un fichier Excel multi-feuilles avec formatage.
    
    Args:
        output_file: Chemin du fichier de sortie
        sheets_data: Dictionnaire {nom_feuille: DataFrame}
    """
    # Ordre des feuilles
    sheet_order = [
        'Equipes', 'Gymnases', 'Indispos_Gymnases', 'Indispos_Equipes',
        'Indispos_Institutions', 'Preferences_Gymnases', 
        'Obligation_Presence', 'Groupes_Non_Simultaneite'
    ]
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for sheet_name in sheet_order:
            if sheet_name in sheets_data:
                df = sheets_data[sheet_name]
                df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    # Formatter le fichier
    wb = openpyxl.load_workbook(output_file)
    
    # Styles
    header_font = Font(bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Formater les en-têtes (ligne 1)
        if ws.max_row > 0:
            for cell in ws[1]:
                if cell.value:  # Only format cells with values
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
        
        # Ajuster la largeur des colonnes
        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row_idx, col_idx)
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_file)


def main():
    """Fonction principale"""
    import argparse
    
    parser = argparse.ArgumentParser(
        description='Extrait les poules depuis un fichier Excel et génère un fichier de configuration PyCalendar'
    )
    parser.add_argument(
        '--input', '-i',
        type=str,
        help='Fichier Excel d\'entrée (défaut: exemple/poules.xlsx)'
    )
    parser.add_argument(
        '--output', '-o',
        type=str,
        help='Fichier Excel de sortie (défaut: exemple/config_exemple.xlsx)'
    )
    parser.add_argument(
        '--sport', '-s',
        type=str,
        default='VB',
        help='Préfixe du sport: VB (volleyball), HB (handball), etc. (défaut: VB)'
    )
    
    args = parser.parse_args()
    
    # Chemins des fichiers
    script_dir = Path(__file__).parent
    
    # Utiliser le dossier exemple par défaut et conserver sport_prefix
    sport_prefix = args.sport.upper()
    data_dir = 'exemple'
    default_output = 'config_exemple.xlsx'
    
    # Chemins par défaut
    input_file = Path(args.input) if args.input else script_dir / data_dir / "poules.xlsx"
    output_file = Path(args.output) if args.output else script_dir / data_dir / default_output
    
    # Vérifier que le fichier d'entrée existe
    if not input_file.exists():
        print(f"❌ Erreur: Le fichier {input_file} n'existe pas!")
        print(f"💡 Utilisation: python extract_poules.py --input FICHIER [--output SORTIE] [--sport {sport_prefix}]")
        return
    
    print("="*80)
    print("🔄 EXTRACTION DES POULES ET GÉNÉRATION DE CONFIGURATION")
    print("="*80)
    print(f"📥 Fichier d'entrée: {input_file}")
    print(f"📤 Fichier de sortie: {output_file}")
    print(f"🏐 Sport: {sport_prefix}")
    print("="*80 + "\n")
    
    # Extraire les poules
    success = extract_poules_from_excel(str(input_file), str(output_file), sport_prefix)
    
    if success:
        print("\n" + "="*80)
        print("🎉 EXTRACTION TERMINÉE AVEC SUCCÈS!")
        print("="*80)
        print("\n📝 Prochaines étapes:")
        print("  1. Ouvrez le fichier Excel généré et complétez la feuille 'Gymnases'")
        print("     (ajoutez les gymnases, leurs créneaux disponibles et capacités)")
        print("  2. Optionnel: Complétez les autres feuilles selon vos besoins")
        print("     (indisponibilités, préférences, contraintes)")
        print("  3. Utilisez ce fichier comme configuration pour l'algorithme de calendrier")
        print(f"     Exemple: python main.py configs/config_hand.yaml")
        print("\n💡 Pour actualiser/valider la configuration:")
        print(f"     python actualiser_config.py {output_file}")
    else:
        print("\n❌ Le script a échoué.")


if __name__ == "__main__":
    main()
