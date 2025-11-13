"""
Gestionnaire du fichier de configuration central.

Ce module gère un fichier Excel unique contenant toutes        'Groupes_Non_Simultaneite': {
            'colonnes': ['Nom_Groupe', 'Entites', 'Remarques'],
            'description': 'Groupes d\'équipes/institutions à ne pas faire jouer simultanément',
            'type': 'manuel',
            'exemple': {
                'Nom_Groupe': 'Grandes Écoles Lyon',
                'Entites': 'ECL, EML, CENTRALE',
                'Remarques': 'Grandes écoles d\'ingénieurs qui ne doivent pas jouer simultanément'
            }
        }tions nécessaires
à la planification : équipes, gymnases, indisponibilités, préférences, contraintes.

Structure du fichier Excel :
- Equipes : Liste des équipes avec poule et horaire préféré
- Gymnases : Gymnases avec créneaux disponibles
- Indispos_Equipes : Indisponibilités spécifiques par équipe
- Indispos_Institutions : Indisponibilités par institution (appliquées à toutes les équipes)
- Preferences_Institutions : Lieux préférés par institution avec classement
- Contraintes_Specifiques : Contraintes particulières (anti-collisions, etc.)
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import logging

logger = logging.getLogger(__name__)


class ConfigManager:
    """Gestionnaire du fichier de configuration central."""
    
    # Définition des structures de chaque feuille
    STRUCTURES = {
        'Equipes_Hors_Championnat': {
            'colonnes': ['Equipe', 'Institution', 'Genre', 'Type_Championnat', 'Motif', 'Remarques'],
            'description': 'Équipes autorisées hors championnat académique pour les matchs fixés',
            'type': 'manuel',
            'exemple': {
                'Equipe': 'EXTERIEUR (1)',
                'Institution': 'EXTERIEUR',
                'Genre': 'M',
                'Type_Championnat': 'CFE',
                'Motif': 'Match amical',
                'Remarques': 'Équipe invitée pour tournoi'
            },
            'notes': [
                'Equipe: Nom complet de l\'équipe (format: Institution (numéro))',
                'Institution: Institution de rattachement',
                'Genre: M (masculin) ou F (féminin)',
                'Type_Championnat: Type de championnat (CFE, CFU, Autre)',
                'Motif: Raison de l\'autorisation (match amical, tournoi, etc.)',
                'Remarques: Informations complémentaires'
            ]
        },
        'Gymnases': {
            'colonnes': ['Gymnase', 'Adresse', 'Capacite', 'Creneaux'],
            'description': 'Gymnases avec capacité et créneaux disponibles (À REMPLIR MANUELLEMENT)',
            'type': 'manuel',
            'exemple': {
                'Gymnase': 'INSA C',
                'Adresse': '20 Avenue Albert Einstein, 69100 Villeurbanne',
                'Capacite': 2,
                'Creneaux': '09:00, 14:00, 18:00'
            }
        },
        'Indispos_Gymnases': {
            'colonnes': ['Gymnase', 'Semaine', 'Horaire_Debut', 'Horaire_Fin', 'Capacite_Occupee', 'Remarques'],
            'description': 'Indisponibilités des gymnases avec capacité partielle (GÉNÉRÉE AUTOMATIQUEMENT)',
            'type': 'auto',
            'exemple': {
                'Gymnase': 'PARC DES SPORTS',
                'Semaine': 3,
                'Horaire_Debut': '09:00',
                'Horaire_Fin': '12:00',
                'Capacite_Occupee': 1,
                'Remarques': 'Maintenance'
            },
            'notes': [
                'Capacite_Occupee: Nombre de terrains/créneaux occupés (défaut = capacité totale si vide)',
                '  → Si gymnase a capacité 3 et Capacite_Occupee = 1, il reste 2 créneaux disponibles',
                '  → Si vide ou >= capacité totale, le gymnase est totalement indisponible',
                'Horaire_Debut/Fin: Plage horaire concernée (format HH:MM)'
            ]
        },
        'Indispos_Equipes': {
            'colonnes': ['Equipe', 'Semaine', 'Horaire_Debut', 'Horaire_Fin', 'Remarques'],
            'description': 'Indisponibilités des équipes (GÉNÉRÉE AUTOMATIQUEMENT)',
            'type': 'auto',
            'exemple': {
                'Equipe': 'CENTRALE 1 (1)',
                'Semaine': 5,
                'Horaire_Debut': '14:00',
                'Horaire_Fin': '18:00',
                'Remarques': 'Compétition nationale'
            }
        },
        'Indispos_Institutions': {
            'colonnes': ['Institution', 'Semaine', 'Horaire_Debut', 'Horaire_Fin', 'Remarques'],
            'description': 'Indisponibilités par institution - S\'applique à toutes les équipes de l\'institution',
            'type': 'auto',
            'exemple': {
                'Institution': 'EDP 1',
                'Semaine': 2,
                'Horaire_Debut': '08:00',
                'Horaire_Fin': '10:00',
                'Remarques': 'Réunion BDE'
            },
            'notes': [
                'Institution: Nom exact de l\'institution concernée',
                'Semaine: Numéro de la semaine (1 à N)',
                'Horaire_Debut: Heure de début de l\'indisponibilité (format HH:MM)',
                'Horaire_Fin: Heure de fin de l\'indisponibilité (format HH:MM)',
                '  → Si vide ou non renseigné: considérer la journée complète',
                'Remarques: Informations complémentaires (optionnel)'
            ]
        },
        'Preferences_Gymnases': {
            'colonnes': ['Institution', 'Gymnase_Pref_1', 'Gymnase_Pref_2', 'Gymnase_Pref_3', 'Gymnase_Pref_4', 'Gymnase_Pref_5'],
            'description': 'Préférences de gymnases par institution - Remplir du plus préféré (1) au moins préféré (5)',
            'type': 'auto',
            'exemple': {
                'Institution': 'EDP 1',
                'Gymnase_Pref_1': 'PARC DES PRINCES',
                'Gymnase_Pref_2': 'GYMNASE CENTRAL',
                'Gymnase_Pref_3': '',
                'Gymnase_Pref_4': '',
                'Gymnase_Pref_5': ''
            },
            'notes': [
                'Institution: Nom exact de l\'institution (automatiquement pré-rempli)',
                'Gymnase_Pref_1 à 5: Gymnases préférés par ordre de préférence',
                '  → Gymnase_Pref_1 = premier choix (bonus maximal)',
                '  → Gymnase_Pref_5 = cinquième choix (bonus minimal)',
                '  → Laisser vide si moins de 5 préférences',
                'Bonus configurables dans le fichier YAML (bonus_preferences_gymnases)'
            ]
        },
        'Dispos_Gymnases_Equipes': {
            'colonnes': ['Equipe', 'Genre', 'Horaire_Dispo', 'Gymnase_1', 'Gymnase_2', 'Gymnase_3', 'Gymnase_4', 'Gymnase_5', 'Remarques'],
            'description': 'Disponibilités anticipées d\'équipes sur gymnases spécifiques',
            'type': 'manuel',
            'exemple': {
                'Equipe': 'LYON 1 (1)',
                'Genre': 'M',
                'Horaire_Dispo': '18:00',
                'Gymnase_1': 'PARC DES SPORTS',
                'Gymnase_2': 'INSA C',
                'Gymnase_3': '',
                'Gymnase_4': '',
                'Gymnase_5': '',
                'Remarques': 'Équipe disponible plus tôt sur ces gymnases'
            },
            'notes': [
                'Equipe: Nom exact de l\'équipe sans le genre (Institution (numéro))',
                'Genre: M (masculin) ou F (féminin) - obligatoire',
                'Horaire_Dispo: Horaire de disponibilité anticipée (format HH:MM)',
                '  → Doit être ANTÉRIEUR à l\'horaire préféré général de l\'équipe',
                '  → Permet à l\'équipe d\'être disponible plus tôt sur les gymnases listés',
                'Gymnase_1 à 5: Gymnases où la disponibilité anticipée s\'applique',
                '  → Doivent exister dans la feuille Gymnases',
                '  → Laisser vide si moins de 5 gymnases concernés',
                'Remarques: Informations complémentaires (optionnel)',
                '',
                '⚠️ IMPORTANT: Cette disponibilité peut être invalidée par:',
                '  → Les indisponibilités de la feuille Indispos_Equipes',
                '  → Les indisponibilités de la feuille Indispos_Institutions',
                '  → Les indisponibilités des gymnases (Indispos_Gymnases)',
                'Exemple: Une équipe normalement disponible à 20:00 peut jouer à 18:00',
                '         uniquement sur PARC DES SPORTS et INSA C'
            ]
        },
        'Obligation_Presence': {
            'colonnes': ['Gymnase', 'Institution_Obligatoire', 'Remarques'],
            'description': 'Obligation de présence institutionnelle par gymnase (GÉNÉRÉE AUTOMATIQUEMENT)',
            'type': 'auto',
            'exemple': {
                'Gymnase': 'PARC DES PRINCES',
                'Institution_Obligatoire': 'ENS',
                'Remarques': 'Au moins une équipe ENS dans chaque match'
            }
        },
        'Groupes_Non_Simultaneite': {
            'colonnes': ['Nom_Groupe', 'Entites', 'Remarques'],
            'description': 'Groupes d\'équipes/institutions à ne pas faire jouer simultanément',
            'type': 'manuel',
            'exemple': {
                'Nom_Groupe': 'Grandes Écoles Lyon',
                'Entites': 'EDP, CENTRALE',
                'Remarques': 'Grandes écoles d\'ingénieurs qui ne doivent pas jouer simultanément'
            }
        },
        'Ententes': {
            'colonnes': ['Institution_1', 'Institution_2', 'Penalite_Non_Planif', 'Remarques'],
            'description': 'Paires d\'institutions avec pénalité réduite si match non planifié',
            'type': 'manuel',
            'exemple': {
                'Institution_1': 'LYON 1',
                'Institution_2': 'LYON 2',
                'Penalite_Non_Planif': '10000.0',
                'Remarques': 'Entente historique entre les deux universités'
            },
            'notes': [
                'Institution_1 et Institution_2: Paire d\'institutions (ordre non important)',
                'Penalite_Non_Planif: Pénalité si le match entente n\'est pas planifié (optionnel, défaut configuré dans YAML)',
                'Les matchs entre ces institutions ont une priorité plus faible et peuvent plus facilement ne pas être planifiés',
                'Exemple: LYON 1 ↔ LYON 2 signifie qu\'un match entre une équipe LYON 1 et une équipe LYON 2 est une entente'
            ]
        },
        'Contraintes_Temporelles': {
            'colonnes': ['Equipe_1', 'Equipe_2', 'Genre', 'Type_Contrainte', 'Semaine', 'Horaires_Possibles', 'Remarques'],
            'description': 'Contraintes temporelles sur matchs spécifiques (ex: CFE après semaine X)',
            'type': 'manuel',
            'exemple': {
                'Equipe_1': 'LYON 1 (1)',
                'Equipe_2': 'LYON 2 (1)',
                'Genre': 'M',
                'Type_Contrainte': 'Apres',
                'Semaine': '8',
                'Horaires_Possibles': 'Mercredi 18h00, Vendredi 16h00',
                'Remarques': 'Match CFE à planifier après la semaine 8'
            },
            'notes': [
                'Equipe_1 et Equipe_2: Noms des équipes sans genre (Institution (numéro))',
                'Genre: M (masculin) ou F (féminin) - même genre pour les deux équipes',
                'Type_Contrainte: "Avant" ou "Apres" (planifier avant/après la semaine indiquée)',
                'Semaine: Numéro de semaine limite (1-52)',
                'Horaires_Possibles: Liste d\'horaires autorisés pour ce match (optionnel, séparés par virgule)',
                'S\'applique à TOUS les matchs entre ces deux équipes (ordre bidirectionnel)',
                'Contrainte paramétrable: souple (pénalité) ou dure (bloquante) via YAML'
            ]
        },
        'Types_Poules': {
            'colonnes': ['Poule', 'Type', 'Remarques'],
            'description': 'Type de championnat pour chaque poule (Classique ou Aller-Retour)',
            'type': 'auto',
            'exemple': {
                'Poule': 'VBFA1PA',
                'Type': 'Classique',
                'Remarques': 'Championnat simple tour'
            },
            'notes': [
                'Poule: Nom exact de la poule (automatiquement pré-rempli)',
                'Type: "Classique" ou "Aller-Retour"',
                '  → Classique: Chaque paire d\'équipes joue 1 seul match (n×(n-1)/2 matchs)',
                '  → Aller-Retour: Chaque paire joue 2 matchs (aller ET retour, n×(n-1) matchs)',
                'Aller-Retour implique de jouer le match une fois chez chaque équipe (inversion ordre)',
                'L\'espacement entre matchs aller et retour est paramétrable via YAML'
            ]
        },
        'Matchs_Fixes': {
            'colonnes': ['Equipe_1', 'Equipe_2', 'Genre', 'Poule', 'Semaine', 'Horaire', 'Gymnase', 'Score', 'Type_Competition', 'Remarques'],
            'description': 'Matchs déjà joués ou planifiés à intégrer dans le calendrier',
            'type': 'manuel',
            'exemple': {
                'Equipe_1': 'LYON 1 (1)',
                'Equipe_2': 'LYON 2 (1)',
                'Genre': 'F',
                'Poule': 'VBFA1PA',
                'Semaine': '1',
                'Horaire': '18:00',
                'Gymnase': 'PARC DES SPORTS',
                'Score': '3-1',
                'Type_Competition': 'CFE',
                'Remarques': 'Match déjà joué'
            },
            'notes': [
                'Equipe_1 et Equipe_2: Noms exacts des équipes sans le genre (Institution (numéro) seulement)',
                'Genre: F ou M (obligatoire)',
                'Poule: Code de la poule (doit correspondre aux équipes)',
                'Semaine: Numéro de semaine où le match a été/sera joué (1 à nb_semaines)',
                'Horaire: Heure du match (format HH:MM)',
                'Gymnase: Nom exact du gymnase (doit exister dans la feuille Gymnases)',
                'Score: Score du match si déjà joué (optionnel, format: "X-Y")',
                'Type_Competition: Nature de la compétition',
                '  → CFE: Championnat de France des Écoles',
                '  → CFU: Championnat de France Universitaire',
                '  → Acad: Match de championnat régulier',
                '  → Autre: Autre type de match',
                'Remarques: Informations complémentaires (optionnel)',
                '',
                '⚠️ IMPORTANT: Ces matchs seront exclus de la planification automatique',
                'Ils apparaîtront dans le calendrier final aux créneaux indiqués'
            ]
        },
                'Niveaux_Gymnases': {
            'colonnes': ['Gymnase', 'Niveau', 'Remarque'],
            'description': 'Classification des gymnases par niveau (haut/bas) avec bonus par niveau de match',
            'type': 'manuel',
            'exemple': {
                'Gymnase': 'PARC DES SPORTS',
                'Niveau': 'Haut niveau',
                'Remarque': 'Gymnase principal de la ville'
            },
            'notes': [
                'Gymnase: Nom exact du gymnase (doit exister dans la feuille Gymnases)',
                'Niveau: "Haut niveau" ou "Bas niveau"',
                'Remarque: Explication ou commentaire optionnel',
                'Bonus par niveau de match configurés dans YAML: bonus_haut_niveau et bonus_bas_niveau',
                'Exemple: bonus_haut_niveau: [10, 8, 5, 1] signifie +10 pour A1, +8 pour A2, etc. dans gymnase haut niveau'
            ]
        }
    }
    
    def __init__(self, fichier_path: str):
        """
        Initialise le gestionnaire de configuration.
        
        Args:
            fichier_path: Chemin vers le fichier Excel de configuration
        """
        self.fichier_path = Path(fichier_path)
        self.fichier_path.parent.mkdir(parents=True, exist_ok=True)
    
    def fichier_existe(self) -> bool:
        """Vérifie si le fichier existe."""
        return self.fichier_path.exists()
    
    def lire_feuille(self, nom_feuille: str) -> Optional[pd.DataFrame]:
        """
        Lit une feuille du fichier Excel.
        
        Args:
            nom_feuille: Nom de la feuille à lire
            
        Returns:
            DataFrame contenant les données, ou None si la feuille n'existe pas
        """
        if not self.fichier_existe():
            logger.warning(f"Le fichier {self.fichier_path} n'existe pas")
            return None
        
        try:
            df = pd.read_excel(self.fichier_path, sheet_name=nom_feuille)
            logger.info(f"Feuille '{nom_feuille}' lue avec succès ({len(df)} lignes)")
            return df
        except ValueError:
            logger.warning(f"La feuille '{nom_feuille}' n'existe pas dans le fichier")
            return None
        except Exception as e:
            logger.error(f"Erreur lors de la lecture de '{nom_feuille}': {e}")
            return None
    
    def lire_toutes_feuilles(self) -> Dict[str, pd.DataFrame]:
        """
        Lit toutes les feuilles définies dans STRUCTURES.
        
        Returns:
            Dictionnaire {nom_feuille: DataFrame}
        """
        resultat = {}
        for nom_feuille in self.STRUCTURES.keys():
            df = self.lire_feuille(nom_feuille)
            if df is not None:
                resultat[nom_feuille] = df
        return resultat
    
    def valider_structure(self, nom_feuille: str, df: pd.DataFrame) -> Tuple[bool, List[str]]:
        """
        Valide qu'une feuille a la structure correcte.
        
        Args:
            nom_feuille: Nom de la feuille
            df: DataFrame à valider
            
        Returns:
            Tuple (valide, liste des erreurs)
        """
        if nom_feuille not in self.STRUCTURES:
            return False, [f"Feuille '{nom_feuille}' non définie dans STRUCTURES"]
        
        colonnes_attendues = set(self.STRUCTURES[nom_feuille]['colonnes'])
        colonnes_presentes = set(df.columns)
        
        erreurs = []
        
        # Colonnes manquantes
        colonnes_manquantes = colonnes_attendues - colonnes_presentes
        if colonnes_manquantes:
            erreurs.append(f"Colonnes manquantes: {', '.join(colonnes_manquantes)}")
        
        # Colonnes en trop (warning, pas erreur bloquante)
        colonnes_extra = colonnes_presentes - colonnes_attendues
        if colonnes_extra:
            logger.warning(f"Colonnes non standard dans '{nom_feuille}': {', '.join(colonnes_extra)}")
        
        return len(erreurs) == 0, erreurs
    
    def _preremplir_preferences_gymnases(self, ws):
        """
        Pré-remplit la feuille Preferences_Gymnases avec toutes les institutions uniques
        trouvées dans la feuille Equipes.
        
        Args:
            ws: Worksheet Preferences_Gymnases à remplir
        """
        # Lire la feuille Equipes pour extraire les institutions
        try:
            df_equipes = self.lire_feuille('Equipes')
            if df_equipes is None or df_equipes.empty or 'Equipe' not in df_equipes.columns:
                logger.warning("Impossible de pré-remplir Preferences_Gymnases : feuille Equipes vide ou invalide")
                return
            
            # Extraire les institutions depuis les noms d'équipes
            institutions = set()
            for equipe_nom in df_equipes['Equipe'].dropna():
                equipe_str = str(equipe_nom).strip()
                # Extraire l'institution (tout avant le dernier "(numéro)")
                if '(' in equipe_str:
                    institution = equipe_str.rsplit('(', 1)[0].strip()
                    if institution:
                        institutions.add(institution)
            
            if not institutions:
                logger.warning("Aucune institution trouvée dans la feuille Equipes")
                return
            
            # Trier les institutions par ordre alphabétique
            institutions_triees = sorted(institutions)
            
            # Remplir la colonne Institution (à partir de la ligne 2)
            for idx, institution in enumerate(institutions_triees, start=2):
                ws.cell(row=idx, column=1, value=institution)
            
            logger.info(f"Preferences_Gymnases pré-remplie avec {len(institutions_triees)} institutions")
            
        except Exception as e:
            logger.error(f"Erreur lors du pré-remplissage de Preferences_Gymnases : {e}")
    
    def creer_feuille_vide(self, nom_feuille: str, avec_exemple: bool = False) -> pd.DataFrame:
        """
        Crée une feuille vide avec la structure correcte.
        
        IMPORTANT: Les exemples ne doivent JAMAIS être ajoutés dans les fichiers de configuration.
        Ils sont définis uniquement pour la documentation et la validation.
        
        Args:
            nom_feuille: Nom de la feuille à créer
            avec_exemple: Si True, ajoute une ligne d'exemple (PAR DÉFAUT: False)
                         ⚠️ NE JAMAIS UTILISER avec_exemple=True dans les scripts de production
            
        Returns:
            DataFrame vide avec les bonnes colonnes (SANS ligne d'exemple)
        """
        if nom_feuille not in self.STRUCTURES:
            raise ValueError(f"Feuille '{nom_feuille}' non définie")
        
        structure = self.STRUCTURES[nom_feuille]
        colonnes = structure['colonnes']
        
        if avec_exemple:
            # ⚠️ MODE DEBUG UNIQUEMENT - Ne pas utiliser en production
            logger.warning(f"⚠️ Création de la feuille '{nom_feuille}' AVEC exemple (mode debug)")
            exemple = structure['exemple']
            df = pd.DataFrame([exemple], columns=colonnes)
        else:
            # Mode normal: créer DataFrame vide SANS exemple
            df = pd.DataFrame(columns=colonnes)
        
        return df
    
    def generer_feuilles_manquantes(self, conserver_existant: bool = True) -> Dict[str, str]:
        """
        Génère UNIQUEMENT les feuilles manquantes, conserve les existantes intactes.
        
        Args:
            conserver_existant: Si True, conserve les données existantes
            
        Returns:
            Dictionnaire {nom_feuille: 'créée' ou 'conservée'}
        """
        statuts = {}
        
        # Si le fichier n'existe pas, créer toutes les feuilles
        if not self.fichier_existe():
            logger.info("Création d'un nouveau fichier")
            with pd.ExcelWriter(self.fichier_path, engine='openpyxl') as writer:
                for nom_feuille in self.STRUCTURES.keys():
                    df = self.creer_feuille_vide(nom_feuille, avec_exemple=False)
                    df.to_excel(writer, sheet_name=nom_feuille, index=False)
                    statuts[nom_feuille] = 'créée'
                    logger.info(f"Feuille '{nom_feuille}' créée")
            
            # Formater le fichier
            self._formater_fichier()
            return statuts
        
        # Fichier existe : n'ajouter QUE les feuilles manquantes
        if conserver_existant:
            import openpyxl
            wb = openpyxl.load_workbook(self.fichier_path)
            feuilles_existantes = set(wb.sheetnames)
            logger.info(f"Fichier existant trouvé avec {len(feuilles_existantes)} feuilles")
            
            feuilles_a_creer = []
            for nom_feuille in self.STRUCTURES.keys():
                if nom_feuille in feuilles_existantes:
                    statuts[nom_feuille] = 'conservée'
                    logger.info(f"Feuille '{nom_feuille}' conservée")
                else:
                    feuilles_a_creer.append(nom_feuille)
            
            # Créer seulement les feuilles manquantes
            if feuilles_a_creer:
                for nom_feuille in feuilles_a_creer:
                    ws = wb.create_sheet(nom_feuille)
                    
                    # Écrire les en-têtes
                    colonnes = self.STRUCTURES[nom_feuille]['colonnes']
                    for idx, col in enumerate(colonnes, 1):
                        ws.cell(1, idx, col)
                    
                    # Pré-remplir les institutions pour Preferences_Gymnases
                    if nom_feuille == 'Preferences_Gymnases':
                        self._preremplir_preferences_gymnases(ws)
                    
                    statuts[nom_feuille] = 'créée'
                    logger.info(f"Feuille '{nom_feuille}' créée sans exemple")
                
                wb.save(self.fichier_path)
        else:
            # Mode écrasement : recréer tout
            with pd.ExcelWriter(self.fichier_path, engine='openpyxl') as writer:
                for nom_feuille in self.STRUCTURES.keys():
                    df = self.creer_feuille_vide(nom_feuille, avec_exemple=False)
                    df.to_excel(writer, sheet_name=nom_feuille, index=False)
                    statuts[nom_feuille] = 'créée'
                    logger.info(f"Feuille '{nom_feuille}' créée")
        
        # Formater le fichier (couleurs, largeurs de colonnes)
        self._formater_fichier()
        
        return statuts
    
    def _ajouter_validation_liste(self, ws, nom_colonne: str, valeurs: List[str], 
                                   ligne_debut: int, ligne_fin: int):
        """
        Ajoute une liste déroulante à une colonne.
        
        Pour les grandes listes (>100 éléments ou >200 caractères), utilise une plage
        nommée dans une feuille cachée au lieu d'une formule inline.
        """
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.utils import get_column_letter
        
        # Trouver l'index de la colonne
        col_idx = None
        for idx, cell in enumerate(ws[1], 1):
            if hasattr(cell, 'value') and cell.value == nom_colonne:
                col_idx = idx
                break
        
        if col_idx is None or not valeurs:
            return
        
        # Calculer la taille de la formule inline
        formule_inline = ",".join(valeurs)
        use_named_range = len(valeurs) > 50 or len(formule_inline) > 200
        
        if use_named_range:
            # Utiliser une plage nommée (pour grandes listes)
            nom_plage = f"Liste_{nom_colonne}_{ws.title}"
            nom_plage = nom_plage.replace(" ", "_").replace("-", "_")[:31]  # Limite Excel: 31 caractères
            
            # Créer ou récupérer feuille cachée pour les listes
            wb = ws.parent
            if '_Listes_Validation' not in wb.sheetnames:
                ws_listes = wb.create_sheet('_Listes_Validation')
                ws_listes.sheet_state = 'hidden'
            else:
                ws_listes = wb['_Listes_Validation']
            
            # Trouver la première colonne vide dans la feuille cachée
            col_liste = ws_listes.max_column + 1 if ws_listes.max_column > 0 else 1
            
            # Écrire les valeurs dans la feuille cachée
            for idx, valeur in enumerate(valeurs, 1):
                ws_listes.cell(row=idx, column=col_liste, value=valeur)
            
            # Créer la formule de référence à la plage
            col_letter_liste = get_column_letter(col_liste)
            formule = f"_Listes_Validation!${col_letter_liste}$1:${col_letter_liste}${len(valeurs)}"
            
            logger.debug(f"Validation avec plage nommée pour {nom_colonne} : {len(valeurs)} valeurs -> {formule}")
        else:
            # Utiliser formule inline (pour petites listes)
            formule = f'"{formule_inline}"'
            logger.debug(f"Validation inline pour {nom_colonne} : {len(valeurs)} valeurs")
        
        # Créer la validation
        dv = DataValidation(type="list", formula1=formule, allow_blank=True)
        dv.error = 'Valeur invalide'
        dv.errorTitle = 'Erreur de saisie'
        dv.prompt = f'Sélectionnez une valeur dans la liste ({len(valeurs)} choix)'
        dv.promptTitle = 'Liste déroulante'
        
        # Appliquer la validation à la plage
        col_letter = get_column_letter(col_idx)
        plage = f"{col_letter}{ligne_debut}:{col_letter}{ligne_fin}"
        dv.add(plage)
        ws.add_data_validation(dv)
        
        logger.info(f"✓ Validation ajoutée pour {nom_colonne} : {len(valeurs)} valeurs")
    
    def _extraire_liste_valeurs(self, wb, nom_feuille: str, nom_colonne: str) -> List[str]:
        """
        Extrait les valeurs uniques d'une colonne.
        
        Pour la colonne 'Equipe', génère automatiquement les variantes avec genre si nécessaire:
        - Si une équipe existe en plusieurs genres (M et F): génère "NOM [M]" et "NOM [F]"
        - Si une équipe existe en un seul genre ou sans genre: pas de suffixe
        """
        from pycalendar.core.utils import extraire_genre_depuis_poule, formater_nom_avec_genre
        
        if nom_feuille not in wb.sheetnames:
            return []
        
        ws = wb[nom_feuille]
        valeurs = []
        
        # Trouver l'index de la colonne
        col_idx = None
        for idx, cell in enumerate(ws[1], 1):
            if hasattr(cell, 'value') and cell.value == nom_colonne:
                col_idx = idx
                break
        
        if col_idx is None:
            return []
        
        # Cas spécial pour les équipes: détecter genres multiples
        if nom_colonne == 'Equipe' and nom_feuille == 'Equipes':
            # Détecter si la feuille a une colonne Poule pour extraire le genre
            poule_col_idx = None
            for idx, cell in enumerate(ws[1], 1):
                if hasattr(cell, 'value') and cell.value == 'Poule':
                    poule_col_idx = idx
                    break
            
            # Mapping: {nom_equipe: {genres}}
            equipes_par_nom = {}
            
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row, col_idx)
                if not (cell.value and str(cell.value).strip()):
                    continue
                
                nom_equipe = str(cell.value).strip()
                
                # Extraire le genre depuis la poule si disponible
                genre = ''
                if poule_col_idx:
                    poule_cell = ws.cell(row, poule_col_idx)
                    if poule_cell.value:
                        genre = extraire_genre_depuis_poule(str(poule_cell.value))
                
                # Stocker dans le mapping
                if nom_equipe not in equipes_par_nom:
                    equipes_par_nom[nom_equipe] = set()
                if genre:
                    equipes_par_nom[nom_equipe].add(genre)
            
            # Générer les variantes
            for nom_equipe, genres in equipes_par_nom.items():
                if len(genres) > 1:
                    # Plusieurs genres: créer variantes [M] et [F]
                    for genre in sorted(genres):
                        valeurs.append(formater_nom_avec_genre(nom_equipe, genre))
                else:
                    # Un seul genre ou pas de genre: pas de suffixe
                    valeurs.append(nom_equipe)
        else:
            # Extraction standard pour les autres colonnes
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row, col_idx)
                if cell.value and str(cell.value).strip():
                    valeur = str(cell.value).strip()
                    if valeur not in valeurs:
                        valeurs.append(valeur)
        
        return sorted(valeurs)
    
    def _extraire_institutions(self, wb) -> List[str]:
        """Extrait la liste des institutions depuis les équipes."""
        institutions = set()
        
        if 'Equipes' not in wb.sheetnames:
            return []
        
        ws = wb['Equipes']
        
        # Trouver la colonne Equipe
        col_idx = None
        for idx, cell in enumerate(ws[1], 1):
            if hasattr(cell, 'value') and cell.value == 'Equipe':
                col_idx = idx
                break
        
        if col_idx is None:
            return []
        
        # Extraire les institutions (partie avant le numéro)
        import re
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row, col_idx)
            if cell.value:
                equipe = str(cell.value).strip()
                # Format: "Institution (numéro)"
                match = re.match(r'^(.+?)\s*\(\d+\)\s*$', equipe)
                if match:
                    institutions.add(match.group(1).strip())
        
        return sorted(list(institutions))
    
    def _formater_fichier(self):
        """Applique le formatage au fichier Excel (en-têtes en gras, couleurs, listes déroulantes, centrage, bordures)."""
        try:
            from openpyxl.styles import Border, Side
            
            wb = openpyxl.load_workbook(self.fichier_path)
            
            # Extraire les listes pour les validations
            equipes_list = self._extraire_liste_valeurs(wb, 'Equipes', 'Equipe')
            gymnases_list = self._extraire_liste_valeurs(wb, 'Gymnases', 'Gymnase')
            institutions_list = self._extraire_institutions(wb)
            
            # Styles pour le formatage
            couleur_entete = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
            couleur_ligne_paire = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            couleur_ligne_impaire = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            
            police_entete = Font(bold=True, color="FFFFFF", size=11)
            police_normale = Font(size=10)
            
            alignement_centre = Alignment(horizontal="center", vertical="center", wrap_text=True)
            alignement_gauche = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            # Bordures fines
            bordure_fine = Border(
                left=Side(style='thin', color='D3D3D3'),
                right=Side(style='thin', color='D3D3D3'),
                top=Side(style='thin', color='D3D3D3'),
                bottom=Side(style='thin', color='D3D3D3')
            )
            
            for nom_feuille in self.STRUCTURES.keys():
                if nom_feuille not in wb.sheetnames:
                    continue
                
                ws = wb[nom_feuille]
                
                # Déterminer le nombre de lignes avec données
                max_row = ws.max_row
                max_col = ws.max_column
                
                # Formater la ligne d'en-tête (ligne 1)
                for cell in ws[1]:
                    cell.fill = couleur_entete
                    cell.font = police_entete
                    cell.alignment = alignement_centre
                    cell.border = bordure_fine
                
                # Formater toutes les lignes de données (alternance de couleurs + centrage)
                for row_idx in range(2, max_row + 1):
                    # Alterner les couleurs des lignes
                    couleur_ligne = couleur_ligne_paire if row_idx % 2 == 0 else couleur_ligne_impaire
                    
                    for col_idx in range(1, max_col + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        
                        # Appliquer la couleur de fond
                        cell.fill = couleur_ligne
                        
                        # Appliquer la police
                        cell.font = police_normale
                        
                        # Centrage pour toutes les cellules SAUF les colonnes de remarques/notes
                        col_letter = cell.column_letter
                        header_cell = ws.cell(row=1, column=col_idx)
                        col_name = header_cell.value if header_cell.value else ""
                        
                        # Alignement à gauche pour les colonnes de texte long
                        if any(keyword in str(col_name).lower() for keyword in ['remarque', 'note', 'commentaire', 'description', 'entite']):
                            cell.alignment = alignement_gauche
                        else:
                            cell.alignment = alignement_centre
                        
                        # Appliquer les bordures
                        cell.border = bordure_fine
                
                # Ajuster la largeur des colonnes de manière plus intelligente
                for column in ws.columns:
                    max_length = 0
                    column_letter = None
                    
                    for cell in column:
                        if column_letter is None and hasattr(cell, 'column_letter'):
                            column_letter = cell.column_letter
                        try:
                            if cell.value:
                                cell_length = len(str(cell.value))
                                if cell_length > max_length:
                                    max_length = cell_length
                        except:
                            pass
                    
                    if column_letter:
                        # Largeur minimale de 12, maximale de 50
                        adjusted_width = max(12, min(max_length + 3, 50))
                        ws.column_dimensions[column_letter].width = adjusted_width
                
                # Figer la première ligne
                ws.freeze_panes = "A2"
                
                # Nettoyer les validations existantes pour éviter les doublons
                ws.data_validations.dataValidation.clear()
                
                # Ajouter les listes déroulantes pour les feuilles auto-générées
                if nom_feuille in ['Indispos_Gymnases', 'Obligation_Presence'] and gymnases_list:
                    self._ajouter_validation_liste(ws, 'Gymnase', gymnases_list, 2, 1000)
                
                if nom_feuille == 'Indispos_Equipes' and equipes_list:
                    self._ajouter_validation_liste(ws, 'Equipe', equipes_list, 2, 1000)
                
                if nom_feuille in ['Indispos_Institutions', 'Preferences_Gymnases', 'Obligation_Presence']:
                    if nom_feuille != 'Obligation_Presence' and institutions_list:
                        self._ajouter_validation_liste(ws, 'Institution', institutions_list, 2, 1000)
                    if nom_feuille == 'Obligation_Presence' and institutions_list:
                        self._ajouter_validation_liste(ws, 'Institution_Obligatoire', institutions_list, 2, 1000)
                
                if nom_feuille == 'Preferences_Gymnases' and gymnases_list:
                    # Ajouter la validation pour toutes les colonnes Gymnase_Pref_*
                    colonnes_pref = [col for col in self.STRUCTURES['Preferences_Gymnases']['colonnes'] 
                                    if col.startswith('Gymnase_Pref_')]
                    for col_name in colonnes_pref:
                        self._ajouter_validation_liste(ws, col_name, gymnases_list, 2, 1000)
                
                # Validation pour Applique_A
                if nom_feuille == 'Indispos_Institutions':
                    self._ajouter_validation_liste(ws, 'Applique_A', 
                                                   ['Toutes', 'Masculin', 'Feminin'], 2, 1000)
                
                # Validation pour Groupes_Non_Simultaneite
                if nom_feuille == 'Groupes_Non_Simultaneite':
                    from openpyxl.comments import Comment
                    # Pour la colonne Entites, on met une note explicative dans la cellule B1
                    # (on ne peut pas faire de vraie validation multi-sélection dans Excel basique)
                    note_cell = ws['B1']
                    note_cell.comment = Comment(
                        "ASTUCE: Vous pouvez saisir plusieurs entités séparées par des virgules ou points-virgules.\n\n"
                        "Exemples:\n"
                        "• ECL, EML, ENS\n"
                        "• LYON 1; LYON 2; LYON 3\n"
                        "• LYON 1 (1), LYON 1 (2)\n"
                        "• INSA\n\n"
                        f"Institutions disponibles: {', '.join(institutions_list[:10])}{'...' if len(institutions_list) > 10 else ''}\n\n"
                        f"Équipes disponibles: {', '.join(equipes_list[:10])}{'...' if len(equipes_list) > 10 else ''}",
                        "PyCalendar"
                    )
                    # Augmenter la largeur de la colonne Entites
                    ws.column_dimensions['B'].width = 60
                
                # Validation pour Ententes (listes déroulantes institutions)
                if nom_feuille == 'Ententes' and institutions_list:
                    self._ajouter_validation_liste(ws, 'Institution_1', institutions_list, 2, 1000)
                    self._ajouter_validation_liste(ws, 'Institution_2', institutions_list, 2, 1000)
                
                # Validation pour Contraintes_Temporelles
                if nom_feuille == 'Contraintes_Temporelles':
                    if equipes_list:
                        self._ajouter_validation_liste(ws, 'Equipe_1', equipes_list, 2, 1000)
                        self._ajouter_validation_liste(ws, 'Equipe_2', equipes_list, 2, 1000)
                    self._ajouter_validation_liste(ws, 'Type_Contrainte', ['Avant', 'Apres'], 2, 1000)
            
            wb.save(self.fichier_path)
            logger.info("Formatage du fichier appliqué avec succès (avec listes déroulantes)")
            
        except Exception as e:
            logger.error(f"Erreur lors du formatage: {e}")
    
    def reorganiser_feuilles(self):
        """
        Réorganise les feuilles dans l'ordre logique défini par STRUCTURES.
        
        Ordre souhaité:
        1. Equipes (manuel)
        2. Gymnases (manuel)
        3. Indispos_Gymnases (auto)
        4. Indispos_Equipes (auto)
        5. Indispos_Institutions (auto)
        6. Preferences_Gymnases (auto)
        7. Obligation_Presence (auto)
        8. Groupes_Non_Simultaneite (manuel)
        """
        try:
            if not self.fichier_existe():
                logger.warning("Fichier inexistant, impossible de réorganiser")
                return
            
            wb = openpyxl.load_workbook(self.fichier_path)
            
            # Ordre désiré (selon STRUCTURES)
            ordre_desire = list(self.STRUCTURES.keys())
            
            # Feuilles existantes
            feuilles_existantes = wb.sheetnames.copy()
            
            # Réorganiser uniquement les feuilles qui existent
            ordre_final = [f for f in ordre_desire if f in feuilles_existantes]
            
            # Ajouter les feuilles non définies dans STRUCTURES à la fin
            feuilles_extra = [f for f in feuilles_existantes if f not in ordre_desire]
            ordre_final.extend(feuilles_extra)
            
            # Réorganiser les feuilles en les déplaçant une par une
            for idx_cible, nom_feuille in enumerate(ordre_final):
                idx_actuel = wb.sheetnames.index(nom_feuille)
                if idx_actuel != idx_cible:
                    # Déplacer la feuille à la position cible
                    # offset = position_cible - position_actuelle
                    offset = idx_cible - idx_actuel
                    wb.move_sheet(nom_feuille, offset=offset)
            
            wb.save(self.fichier_path)
            logger.info(f"Feuilles réorganisées dans l'ordre: {', '.join(ordre_final)}")
            
        except Exception as e:
            logger.error(f"Erreur lors de la réorganisation: {e}")
    
    def valider_fichier_complet(self) -> Tuple[bool, Dict[str, List[str]]]:
        """
        Valide que toutes les feuilles du fichier ont la structure correcte.
        
        Returns:
            Tuple (tout_valide, dictionnaire des erreurs par feuille)
        """
        if not self.fichier_existe():
            return False, {'global': ["Le fichier n'existe pas"]}
        
        feuilles = self.lire_toutes_feuilles()
        erreurs_par_feuille = {}
        tout_valide = True
        
        for nom_feuille in self.STRUCTURES.keys():
            if nom_feuille not in feuilles:
                erreurs_par_feuille[nom_feuille] = ["Feuille manquante"]
                tout_valide = False
            else:
                valide, erreurs = self.valider_structure(nom_feuille, feuilles[nom_feuille])
                if not valide:
                    erreurs_par_feuille[nom_feuille] = erreurs
                    tout_valide = False
        
        return tout_valide, erreurs_par_feuille
    
    def afficher_rapport(self):
        """Affiche un rapport sur l'état du fichier de configuration."""
        print("\n" + "="*80)
        print(f"📋 RAPPORT DE CONFIGURATION : {self.fichier_path.name}")
        print("="*80 + "\n")
        
        if not self.fichier_existe():
            print("❌ Le fichier n'existe pas encore.\n")
            print("💡 Utilisez generer_feuilles_manquantes() pour le créer.\n")
            return
        
        # Lire toutes les feuilles
        feuilles = self.lire_toutes_feuilles()
        
        # Valider
        tout_valide, erreurs = self.valider_fichier_complet()
        
        print(f"✅ Fichier trouvé : {self.fichier_path}")
        print(f"📊 Nombre de feuilles : {len(feuilles)}/{len(self.STRUCTURES)}\n")
        
        # Détails par feuille
        for nom_feuille, structure in self.STRUCTURES.items():
            print(f"📄 {nom_feuille}")
            print(f"   Description : {structure['description']}")
            
            if nom_feuille in feuilles:
                df = feuilles[nom_feuille]
                nb_lignes = len(df)
                nb_remplies = df.dropna(how='all').shape[0] - 1  # -1 pour exclure l'en-tête
                
                print(f"   ✅ Présente : {nb_lignes} lignes ({nb_remplies} remplies)")
                
                # Valider la structure
                valide, errs = self.valider_structure(nom_feuille, df)
                if not valide:
                    print(f"   ⚠️  Erreurs de structure :")
                    for err in errs:
                        print(f"      - {err}")
            else:
                print(f"   ❌ Manquante")
            
            print()
        
        # Résumé final
        print("="*80)
        if tout_valide:
            print("✅ Toutes les feuilles sont valides et prêtes à l'emploi !")
        else:
            print("⚠️  Certaines feuilles nécessitent des corrections.")
            print("\n💡 Utilisez generer_feuilles_manquantes() pour corriger automatiquement.")
        print("="*80 + "\n")


def migrer_depuis_ancien_format(ancien_fichier: str, nouveau_fichier: str):
    """
    Migre les données d'un ancien format vers le nouveau format de configuration central.
    
    Args:
        ancien_fichier: Chemin vers l'ancien fichier (ex: equipes_handball.xlsx)
        nouveau_fichier: Chemin vers le nouveau fichier de configuration
    """
    logger.info(f"Migration de {ancien_fichier} vers {nouveau_fichier}")
    
    # Créer le gestionnaire pour le nouveau fichier
    config = ConfigManager(nouveau_fichier)
    
    # Générer la structure de base
    config.generer_feuilles_manquantes(conserver_existant=False)
    
    # Lire l'ancien fichier
    try:
        df_ancien = pd.read_excel(ancien_fichier)
        logger.info(f"Ancien fichier lu : {len(df_ancien)} équipes")
        
        # Mapper les colonnes de l'ancien format vers le nouveau
        colonnes_mapping = {
            'Equipe': 'Equipe',
            'Poule': 'Poule',
            'Horaire_1': 'Horaire_1',
            'Horaire_2': 'Horaire_2',
            'Lieu_1': 'Lieu_1',
            'Lieu_2': 'Lieu_2',
            'Indispo_1': 'Indispo_1',
            'Indispo_2': 'Indispo_2'
        }
        
        # Créer le DataFrame pour la feuille Equipes
        df_nouveau = pd.DataFrame()
        for col_ancien, col_nouveau in colonnes_mapping.items():
            if col_ancien in df_ancien.columns:
                df_nouveau[col_nouveau] = df_ancien[col_ancien]
        
        # Écrire dans le nouveau fichier
        with pd.ExcelWriter(nouveau_fichier, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            df_nouveau.to_excel(writer, sheet_name='Equipes', index=False)
        
        logger.info(f"Migration réussie : {len(df_nouveau)} équipes migrées")
        
        # Reformater
        config._formater_fichier()
        
        print(f"✅ Migration terminée !")
        print(f"   {len(df_nouveau)} équipes migrées vers {nouveau_fichier}")
        
    except Exception as e:
        logger.error(f"Erreur lors de la migration : {e}")
        raise


if __name__ == "__main__":
    # Configuration du logging
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )
    
    # Test avec le fichier exemple
    config = ConfigManager("exemple/config_exemple.xlsx")
    
    print("🔧 Test du gestionnaire de configuration\n")
    
    # Générer les feuilles
    statuts = config.generer_feuilles_manquantes()
    
    print("\n📊 Statuts des feuilles :")
    for feuille, statut in statuts.items():
        emoji = "✅" if statut == "créée" else "📋"
        print(f"  {emoji} {feuille}: {statut}")
    
    # Afficher le rapport
    config.afficher_rapport()
